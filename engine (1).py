"""
engine.py — TimetableEngine

Pure-Python scheduling engine (no tkinter).
All scheduling/generation logic from the original app, adapted for Streamlit:
  • run_stage1()       — synchronous Stage 1 (HC1/HC2 placement)
  • run_stage3()       — synchronous Stage 3 (filler + repair)
  • validate_step3()   — returns dict instead of showing a window
  • get_excel_bytes()  — returns Excel bytes for st.download_button
"""
import random
import copy
from collections import defaultdict
from datetime import datetime


class TimetableEngine:
    """Holds all application state and scheduling logic."""

    # Bump this string whenever the class interface changes.
    # streamlit_app.py compares against this to detect stale pickled instances.
    ENGINE_VERSION = "4.2"

    def __init__(self):
        self.configuration        = {}
        self.class_config_data    = {}
        self.step3_data           = {}
        self.step3_unavailability = {}
        self._relaxed_consec_keys = set()
        self._relaxed_main_keys   = set()
        self._gen_stage           = 0
        self._progress_log        = []
        self._gen                 = None
        self._timetable           = None
        self._last_allocation     = None
        self._last_all_rows       = None
        self._last_group_slots    = None
        self._last_ta2_allocation = None
        self._stage1_status       = None
        self._stage2_status       = None

    def _check_unavailability_feasible(self, teacher, blocked_days, blocked_periods):
        """Two-part feasibility check for teacher unavailability.

        CHECK 1 — Direct slot conflicts:
          For every subject assigned to this teacher, if the subject has specific
          period preferences AND specific day preferences, check whether ANY of
          those (day, period) pairs fall inside the blocked slots.
          Class-teacher duty is also checked: if the CT period falls on a blocked
          day+period combination that is a conflict.

        CHECK 2 — Total slot availability:
          After removing blocked slots, the remaining available slots per week
          must be >= teacher's effective assigned periods.

        Returns (ok: bool, message: str)
          ok=False if either check fails; message explains what is wrong.
        """
        cfg       = self.configuration
        ppd       = cfg['periods_per_day']
        wdays     = cfg['working_days']
        day_names = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'][:wdays]

        blocked_days_set    = set(blocked_days)
        blocked_periods_set = set(int(p) for p in blocked_periods)

        # ── CHECK 1: Direct assignment conflicts ─────────────────────────
        slot_conflicts = []   # list of human-readable conflict strings

        for cls in range(6, 13):
            for si in range(cfg['classes'][cls]):
                cn = "{}{}".format(cls, chr(65 + si))
                if cn not in self.class_config_data:
                    continue
                cd  = self.class_config_data[cn]
                ct  = cd.get('teacher', '').strip()
                ct_per = cd.get('teacher_period', 1)

                # Class-teacher duty: fixed period, every working day
                if ct == teacher and ct_per in blocked_periods_set:
                    # Only a conflict on the blocked days
                    conflict_days = [d for d in day_names if d in blocked_days_set]
                    if conflict_days:
                        slot_conflicts.append(
                            "Class Teacher of {} (Period {}) conflicts on: {}".format(
                                cn, ct_per, ', '.join(conflict_days)))

                # Subject assignments
                for s in cd['subjects']:
                    t = s['teacher'].strip()
                    if t != teacher:
                        continue
                    s_periods = s.get('periods_pref', [])
                    s_days    = s.get('days_pref', [])

                    if not s_periods and not s_days:
                        # No specific preference — no direct conflict detectable
                        continue

                    # Determine which days are relevant
                    relevant_days = set(s_days) if s_days else set(day_names)
                    conflict_days = relevant_days & blocked_days_set

                    if not conflict_days:
                        continue

                    if s_periods:
                        # Check if any preferred period is in the blocked set
                        bad_periods = set(s_periods) & blocked_periods_set
                        if bad_periods:
                            slot_conflicts.append(
                                "'{}' in {} — Period(s) {} on {} are both "
                                "preferred and blocked".format(
                                    s['name'], cn,
                                    sorted(bad_periods),
                                    ', '.join(sorted(conflict_days))))
                    # If subject has day pref but no period pref: warn only (soft)
                    else:
                        # Soft: some teaching days overlap with blocked days
                        # but no period specified — flag as warning
                        slot_conflicts.append(
                            "'{}' in {} — preferred days {} overlap with "
                            "blocked days (no period preference set, "
                            "scheduler may still place it in a blocked slot)".format(
                                s['name'], cn,
                                ', '.join(sorted(conflict_days))))

        # ── CHECK 2: Total available slots ──────────────────────────────
        total_week    = ppd * wdays
        blocked_total = len(blocked_days_set) * len(blocked_periods_set)
        available     = total_week - blocked_total

        wl = getattr(self, '_step3_teacher_wl', {})
        assigned = self._effective_total(teacher) if teacher in wl else 0

        slot_ok  = available >= assigned
        free     = available - assigned

        # ── Build result message ─────────────────────────────────────────
        parts = []

        if slot_conflicts:
            parts.append(
                "SLOT CONFLICTS ({}):\n{}".format(
                    len(slot_conflicts),
                    "\n".join("  • " + c for c in slot_conflicts)))

        parts.append(
            "CAPACITY: {} assigned, {} available after blocking "
            "({} blocked, {} free).".format(
                assigned, available, blocked_total, free))

        message = "\n".join(parts)

        ok = (not slot_conflicts) and slot_ok
        if not ok:
            if slot_conflicts and not slot_ok:
                message = "Slot conflicts AND capacity problem.\n" + message
            elif slot_conflicts:
                message = "Slot conflicts found (capacity OK).\n" + message
            else:
                message = "Capacity problem.\n" + message

        return (ok, message)


    def _compute_teacher_workload(self):
        """Compute teacher workload.

        NOTE: Class teacher period is already included in the subject's period count.
        For example if teacher A is class teacher of 8A and teaches English (7 periods),
        those 7 periods already include the class teacher period — so we do NOT add
        extra periods for class teacher duty. We only record it as metadata (is_ct=True
        on the subject entry) so the fixed-period constraint is known.
        """
        cfg   = self.configuration
        wdays = cfg['working_days']
        result = {}

        def _add(t, entry):
            if not t: return
            result.setdefault(t, {'total': 0, 'entries': []})
            result[t]['entries'].append(entry)
            result[t]['total'] += entry['periods']

        for cls in range(6, 13):
            for si in range(cfg['classes'][cls]):
                cn = "{}{}".format(cls, chr(65 + si))
                if cn not in self.class_config_data:
                    continue
                cd = self.class_config_data[cn]
                ct = cd.get('teacher', '').strip()
                ct_per = cd.get('teacher_period', 1)

                for s in cd['subjects']:
                    t = s['teacher'].strip()
                    if t:
                        # Mark if this teacher is also the class teacher for this class
                        is_ct_subject = (t == ct)
                        ct_note = "  [incl. CT Period {}]".format(ct_per) if is_ct_subject else ""
                        _add(t, {
                            'class':    cn,
                            'subject':  s['name'],
                            'label':    "'{}' in {}  x{}/wk{}".format(
                                s['name'], cn, s['periods'], ct_note),
                            'periods':  s['periods'],
                            'is_ct':    is_ct_subject,
                            'ct_period': ct_per if is_ct_subject else None,
                        })
                    pt = s['parallel_teacher'].strip() if s['parallel'] else ''
                    if pt:
                        _add(pt, {
                            'class':   cn,
                            'subject': s.get('parallel_subject', '?'),
                            'label':   "Parallel '{}' in {}  x{}/wk".format(
                                s.get('parallel_subject', '?'), cn, s['periods']),
                            'periods': s['periods'],
                            'is_ct':   False,
                            'ct_period': None,
                        })
        return result

    def _effective_total(self, teacher):
        """Total periods after subtracting savings from combines."""
        wl    = self._step3_teacher_wl.get(teacher, {})
        total = wl.get('total', 0)
        for cb in self.step3_data.get(teacher, {}).get('combines', []):
            n   = len(cb.get('entry_indices', []))
            per = cb.get('periods_each', 0)
            if n > 1:
                total -= (n - 1) * per
        return total

    # ── Left panel: ALL teachers ──────────────────────────────────────────
    def prepare_step3_workload(self):
        """Compute workload AND set all step-3 attributes needed by validate_step3.

        Must be called before validate_step3() or _render_workload().
        Sets:
          _step3_teacher_wl   – {teacher: {'total': int, 'entries': [...]}}
          _step3_overloaded   – set of teacher names whose raw total > max_allowed
          _step3_max_allowed  – (ppd - 2) * wdays  (mirrors original formula)
          _step3_total_week   – ppd * wdays

        Returns the workload dict (same as _compute_teacher_workload).
        """
        cfg          = self.configuration
        ppd          = cfg['periods_per_day']
        wdays        = cfg['working_days']
        total_week   = ppd * wdays
        max_allowed  = (ppd - 2) * wdays   # P1: each teacher must have ≥1 free per half

        wl = self._compute_teacher_workload()

        self._step3_teacher_wl  = wl
        self._step3_total_week  = total_week
        self._step3_max_allowed = max_allowed
        self._step3_overloaded  = {
            t for t, info in wl.items() if info['total'] > max_allowed
        }
        return wl

    def validate_step3(self):
        """Return dict with overload status; no UI.

        Call prepare_step3_workload() first so that _step3_overloaded and
        _step3_max_allowed are properly initialised.
        """
        overloaded = getattr(self, '_step3_overloaded', set())
        max_all    = getattr(self, '_step3_max_allowed', 99999)
        issues, resolved = [], []
        for teacher in sorted(overloaded):
            s3d     = self.step3_data.get(teacher, {})
            skipped = s3d.get('skipped', False)
            eff     = self._effective_total(teacher)
            if skipped:
                resolved.append("{}: SKIPPED by user".format(teacher))
            elif eff <= max_all:
                resolved.append("{}: Resolved  ({} periods  \u2264 {})".format(teacher, eff, max_all))
            else:
                issues.append("{}: still overloaded  ({}/{})  — over by {}".format(
                    teacher, eff, max_all, eff - max_all))
        return {
            'overloaded':   overloaded,
            'issues':       issues,
            'resolved':     resolved,
            'can_proceed':  (not overloaded) or (not issues),
        }

    def get_class_ct_info(self, cn, teacher, teacher_subject):
        """Return class-teacher info and parallel-conflict details for one entry.

        Adapted from the original _get_class_ct_info to use plain dict-based
        class_config_data (no tkinter StringVar).

        Returns a dict:
          ct                  – class teacher name (str)
          ct_subjects         – list[str] of subjects the CT teaches in cn
          is_parallel_with_ct – bool: teacher_subject is parallel to a CT subject
          parallel_ct_subject – str: the CT subject that is parallel ('' if none)
        """
        cd    = self.class_config_data.get(cn, {})
        ct    = cd.get('teacher', '').strip()
        subjs = cd.get('subjects', [])

        ct_subjects = [s['name'] for s in subjs
                       if s.get('teacher', '').strip() == ct]

        is_parallel_with_ct = False
        parallel_ct_subject = ''
        for s in subjs:
            # Primary: teacher teaches teacher_subject, parallel partner is the CT
            if (s.get('teacher', '').strip() == teacher
                    and s['name'] == teacher_subject
                    and s.get('parallel')
                    and s.get('parallel_teacher', '').strip() == ct):
                is_parallel_with_ct = True
                parallel_ct_subject = s.get('parallel_subject', '')
                break
            # Reverse: teacher is the *parallel* teacher; CT teaches the primary
            if (s.get('parallel')
                    and s.get('parallel_teacher', '').strip() == teacher
                    and s.get('parallel_subject', '') == teacher_subject
                    and s.get('teacher', '').strip() == ct):
                is_parallel_with_ct = True
                parallel_ct_subject = s['name']
                break

        return {
            'ct':                   ct,
            'ct_subjects':          ct_subjects,
            'is_parallel_with_ct':  is_parallel_with_ct,
            'parallel_ct_subject':  parallel_ct_subject,
        }


    # =========================================================================
    #  STEP 4 — Timetable Generation Engine
    # =========================================================================


    # =========================================================================
    #  STEP 4 — Timetable Generation  (complete rewrite)
    # =========================================================================

    def run_stage1(self):
        """Run Stage 1 synchronously. Returns status dict."""
        self._progress_log = []
        self._init_gen_state()
        self._run_stage1_phases()
        return getattr(self, '_stage1_status', {})
    # ── Task Analysis page ────────────────────────────────────────────────────
    def _run_task_analysis_allocation(self):
        """
        Orchestrate the full allocation pipeline for the Task Analysis page:
          1. Build all_rows (same logic as _show_task_analysis data phase)
          2. Calculate slots needed per group (_calculate_group_slots)
          3. Allocate slots (_allocate_group_slots)

        Returns (group_slots, group_allocation, all_rows)
        so the caller can pass them directly to _show_task_analysis.
        """
        s3  = getattr(self, 'step3_data', {})
        cfg = self.configuration

        all_classes = []
        for cls in range(6, 13):
            for si in range(cfg['classes'].get(cls, 0)):
                all_classes.append("{}{}".format(cls, chr(65 + si)))

        # Helper: find parallel partner
        def _find_parallel(cn, subject_name):
            cd_subjects = self.class_config_data.get(cn, {}).get('subjects', [])
            for s in cd_subjects:
                if s['name'] == subject_name and s.get('parallel'):
                    ps = (s.get('parallel_subject') or '').strip()
                    pt = (s.get('parallel_teacher') or '').strip()
                    return (ps or '?', pt or '—')
                if (s.get('parallel')
                        and (s.get('parallel_subject') or '').strip() == subject_name
                        and s['name'] != subject_name):
                    return (s['name'], (s.get('teacher') or '').strip() or '—')
            return ('—', '—')

        # Section A — combined groups
        all_rows = []
        group_no = 0
        covered  = set()

        for teacher, s3d in sorted(s3.items()):
            for cb in s3d.get('combines', []):
                classes  = cb.get('classes', [])
                subjects = cb.get('subjects', [])
                if not classes:
                    continue
                group_no += 1
                for j, cn in enumerate(classes):
                    tsub = (subjects[j] if j < len(subjects)
                            else (subjects[0] if subjects else '?'))
                    par_subj, par_teacher = _find_parallel(cn, tsub)
                    all_rows.append({
                        'group': group_no, 'class': cn,
                        'subject': tsub, 'teacher': teacher,
                        'par_subj': par_subj, 'par_teacher': par_teacher,
                        'section': 'A',
                    })
                    covered.add((cn, tsub))
                    if par_subj not in ('—', '?'):
                        covered.add((cn, par_subj))

        # Section B — standalone parallel pairs
        seen_pairs = set()
        for cn in all_classes:
            cd_subjects = self.class_config_data.get(cn, {}).get('subjects', [])
            for s in cd_subjects:
                if not s.get('parallel'):
                    continue
                subj_name   = s['name']
                subj_teach  = (s.get('teacher') or '').strip()
                par_subj    = (s.get('parallel_subject') or '').strip()
                par_teacher = (s.get('parallel_teacher') or '').strip()
                if not par_subj:
                    continue
                if (cn, subj_name) in covered or (cn, par_subj) in covered:
                    continue
                pair_key = frozenset([(cn, subj_name), (cn, par_subj)])
                if pair_key in seen_pairs:
                    continue
                seen_pairs.add(pair_key)
                group_no += 1
                all_rows.append({
                    'group': group_no, 'class': cn,
                    'subject': subj_name, 'teacher': subj_teach,
                    'par_subj': par_subj or '?',
                    'par_teacher': par_teacher or '—',
                    'section': 'B',
                })

        # Section C — consecutive groups
        consec_covered = set(covered)
        for row in all_rows:
            if row['section'] == 'B':
                consec_covered.add((row['class'], row['subject']))
                if row['par_subj'] not in ('—', '?', ''):
                    consec_covered.add((row['class'], row['par_subj']))

        seen_consec = set()
        for cn in all_classes:
            cd_subjects = self.class_config_data.get(cn, {}).get('subjects', [])
            for s in cd_subjects:
                if s.get('consecutive', 'No') != 'Yes':
                    continue
                subj_name  = s['name']
                subj_teach = (s.get('teacher') or '').strip()
                periods    = s.get('periods', '')
                key = (cn, subj_name)
                if key in seen_consec:
                    continue
                seen_consec.add(key)
                group_no += 1
                all_rows.append({
                    'group': group_no, 'class': cn,
                    'subject': subj_name, 'teacher': subj_teach,
                    'par_subj': '—', 'par_teacher': '—',
                    'section': 'C', 'periods': periods,
                })

        # Calculate & allocate
        group_slots      = self._calculate_group_slots(all_rows)
        group_allocation = self._allocate_group_slots(all_rows, group_slots)

        return group_slots, group_allocation, all_rows


    def _build_task_analysis_rows(self):
        """
        Read-only preview helper: builds all_rows WITHOUT allocating slots or
        modifying the grid.  Safe to call on every page render.
        """
        s3  = getattr(self, 'step3_data', {})
        cfg = self.configuration
        all_classes = []
        for cls in range(6, 13):
            for si in range(cfg['classes'].get(cls, 0)):
                all_classes.append("{}{}".format(cls, chr(65 + si)))

        def _find_parallel(cn, subject_name):
            cd_subjects = self.class_config_data.get(cn, {}).get('subjects', [])
            for s in cd_subjects:
                if s['name'] == subject_name and s.get('parallel'):
                    ps = (s.get('parallel_subject') or '').strip()
                    pt = (s.get('parallel_teacher') or '').strip()
                    return (ps or '?', pt or '—')
                if (s.get('parallel')
                        and (s.get('parallel_subject') or '').strip() == subject_name
                        and s['name'] != subject_name):
                    return (s['name'], (s.get('teacher') or '').strip() or '—')
            return ('—', '—')

        all_rows = []; group_no = 0; covered = set()
        for teacher, s3d in sorted(s3.items()):
            for cb in s3d.get('combines', []):
                classes  = cb.get('classes', []); subjects = cb.get('subjects', [])
                if not classes: continue
                group_no += 1
                for j, cn in enumerate(classes):
                    tsub = (subjects[j] if j < len(subjects) else (subjects[0] if subjects else '?'))
                    par_subj, par_teacher = _find_parallel(cn, tsub)
                    all_rows.append({'group': group_no, 'class': cn, 'subject': tsub,
                                     'teacher': teacher, 'par_subj': par_subj,
                                     'par_teacher': par_teacher, 'section': 'A'})
                    covered.add((cn, tsub))
                    if par_subj not in ('—', '?'): covered.add((cn, par_subj))
        seen_pairs = set()
        for cn in all_classes:
            for s in self.class_config_data.get(cn, {}).get('subjects', []):
                if not s.get('parallel'): continue
                subj_name = s['name']; subj_teach = (s.get('teacher') or '').strip()
                par_subj  = (s.get('parallel_subject') or '').strip()
                par_teacher = (s.get('parallel_teacher') or '').strip()
                if not par_subj: continue
                if (cn, subj_name) in covered or (cn, par_subj) in covered: continue
                pair_key = frozenset([(cn, subj_name), (cn, par_subj)])
                if pair_key in seen_pairs: continue
                seen_pairs.add(pair_key); group_no += 1
                all_rows.append({'group': group_no, 'class': cn, 'subject': subj_name,
                                 'teacher': subj_teach, 'par_subj': par_subj or '?',
                                 'par_teacher': par_teacher or '—', 'section': 'B'})
        seen_consec = set()
        for cn in all_classes:
            for s in self.class_config_data.get(cn, {}).get('subjects', []):
                if s.get('consecutive', 'No') != 'Yes': continue
                subj_name = s['name']; key = (cn, subj_name)
                if key in seen_consec: continue
                seen_consec.add(key); group_no += 1
                all_rows.append({'group': group_no, 'class': cn, 'subject': subj_name,
                                 'teacher': (s.get('teacher') or '').strip(),
                                 'par_subj': '—', 'par_teacher': '—',
                                 'section': 'C', 'periods': s.get('periods', '')})
        return all_rows

    def _proceed_to_stage2(self):
        """
        Gate check before entering Stage 2.
        All groups must have been successfully allocated by 'Allocate Periods'.
        If any failed, show a detailed error + suggestion dialog and block.
        If all ok, open the Stage 2 timetable page.
        """
        if not getattr(self, '_last_allocation', None):
            return {'ok': False, 'reason': 'allocation_not_run'}
        failed = {gn: ar for gn, ar in self._last_allocation.items() if not ar.get('ok', False)}
        if failed:
            return {'ok': False, 'reason': 'groups_failed', 'failed': failed,
                    'all_rows': getattr(self, '_last_all_rows', [])}
        return {'ok': True}

    # ── Allocation error dialog ───────────────────────────────────────────────
    def _allocation_suggestion(self, reason, rows, sec):
        """Return a human-readable suggestion string based on the failure reason."""
        reason_l = (reason or '').lower()
        teachers = list(dict.fromkeys(
            t for r in rows
            for t in [r.get('teacher',''), r.get('par_teacher','')]
            if t and t not in ('—','?','')))
        classes  = list(dict.fromkeys(r['class'] for r in rows))

        if 'stage 1 not run' in reason_l:
            return ("Run Stage 1 first (click '▶ Stage 1: Fill CT Slots') "
                    "before allocating periods.")

        if 'not found in engine' in reason_l or 'not in task list' in reason_l:
            return ("The subject name in Step 2 may not exactly match what Step 3 "
                    "recorded. Open Step 2 for class {} and verify the subject "
                    "name spelling matches exactly.".format(classes[0] if classes else '?'))

        if 'teacher' in reason_l and 'busy' in reason_l:
            busy_t = [t for t in teachers if t.lower() in reason_l]
            t_str  = ', '.join(busy_t) if busy_t else ', '.join(teachers)
            return ("Teacher {} is fully occupied at all candidate slots. "
                    "Options:  (a) Reduce total periods for one of their other subjects "
                    "in Step 2,  (b) Remove an unavailability block in Step 3 if one "
                    "was set by mistake,  (c) Reassign this subject to a different "
                    "teacher.".format(t_str))

        if 'occupied' in reason_l or 'class' in reason_l:
            c_str = ', '.join(classes)
            return ("Class {} has no free slots at the required periods. "
                    "Options:  (a) Reduce the period count for another subject "
                    "assigned to this class in Step 2,  (b) Split the combine into "
                    "smaller groups so fewer classes compete for the same "
                    "slots.".format(c_str))

        if sec == 'C':
            return ("No two adjacent free periods found for the consecutive subject. "
                    "Try reducing the period count for another subject in this class, "
                    "or disable the 'Consecutive' flag if back-to-back periods are "
                    "not strictly required.")

        if sec == 'B':
            return ("Both the primary and parallel teachers must be free at the "
                    "same period. Check each teacher's schedule and reduce workload "
                    "or unavailability constraints in Step 3.")

        if sec == 'A':
            return ("All {} classes AND all their teachers must be free at the same "
                    "slot. Reduce the number of classes in this combine group (Step 3) "
                    "or reduce period counts for conflicting subjects "
                    "(Step 2).".format(len(classes)))

        return ("Check that all teachers in this group have free periods available "
                "and that the affected classes have not already been fully "
                "scheduled by Stage 1 / other groups.")

    # ── Task Analysis 2 — allocation engine ─────────────────────────────────
    def _run_ta2_allocation(self):
        """
        Two-phase allocation for Task Analysis 2.

        PHASE 1 — Main Periods  (tasks where periods >= wdays-1, not relaxed):
            Scan periods p=0..ppd-1. At each p, collect all days where the
            class(es) AND teacher(s) are free. If free days >= remaining,
            assign ALL needed slots at that one period. Otherwise fall through.

        PHASE 2 — Filler Periods  (everything else + Phase-1 fall-through):
            Walk d=0..wdays-1, p=0..ppd-1 and place wherever free.

        Returns
        -------
        dict  task_idx -> {
            'phase'      : 'main' | 'filler',
            'placed'     : [(d, p), ...],   # new placements this run
            'remaining'  : int,             # still unplaced after this run
            'fail_reason': str,             # '' if fully placed
        }
        """
        if not hasattr(self, '_gen'):
            return {}

        g     = self._gen
        tasks = g['tasks']
        grid  = g['grid']
        wdays = g['wdays']
        ppd   = g['ppd']
        DAYS  = g['DAYS']

        if not hasattr(self, '_relaxed_main_keys'):
            self._relaxed_main_keys = set()

        results = {}

        # ── helpers ──────────────────────────────────────────────────────────────
        def slot_free(task, d, p):
            for cn in task['cn_list']:
                if grid.get(cn, [[]])[d][p] is not None:
                    return False
            t  = task['teacher']
            pt = task.get('par_teach', '')
            if not g['t_free'](t, d, p) or g['t_unavail'](t, d, p):
                return False
            if pt and pt not in ('', '—', '?'):
                if not g['t_free'](pt, d, p) or g['t_unavail'](pt, d, p):
                    return False
            return True

        def is_main(task):
            key = (frozenset(task['cn_list']), task['subject'])
            if key in self._relaxed_main_keys:
                return False
            return task['periods'] >= wdays - 1   # use original period count

        # ── PHASE 1 — Main Periods (P3: same period every day, Inc order) ─────────
        # P2: if Pref_Per is set, try those periods FIRST then Inc for remainder.
        for task in tasks:
            if task['remaining'] <= 0:
                continue
            if not is_main(task):
                continue

            placed      = []
            needed      = task['remaining']
            fail_reason = ''

            # Build period scan order: Pref_Per first, then Inc
            if task.get('p_pref'):
                pref_ps  = [x - 1 for x in task['p_pref']]
                inc_rest = [p for p in range(ppd) if p not in pref_ps]
                _period_scan = pref_ps + inc_rest
            else:
                _period_scan = list(range(ppd))

            for p in _period_scan:
                if task['remaining'] <= 0:
                    break
                avail = [d for d in range(wdays) if slot_free(task, d, p)]
                if len(avail) >= task['remaining']:
                    for d in avail[:task['remaining']]:
                        self._gen_place(task, d, p)
                        placed.append((d, p))
                    break   # fully placed at this period

            if task['remaining'] > 0:
                # Diagnose which period came closest and what was blocking it
                best_days   = 0
                best_period = -1
                all_busy_teachers = set()
                for p in range(ppd):
                    free_days = [d for d in range(wdays) if slot_free(task, d, p)]
                    if len(free_days) > best_days:
                        best_days   = len(free_days)
                        best_period = p + 1
                    for d in range(wdays):
                        t  = task['teacher']
                        pt = task.get('par_teach', '')
                        if not g['t_free'](t, d, p) or g['t_unavail'](t, d, p):
                            all_busy_teachers.add(t)
                        if pt and pt not in ('', '—', '?'):
                            if not g['t_free'](pt, d, p) or g['t_unavail'](pt, d, p):
                                all_busy_teachers.add(pt)

                busy_t_str = ', '.join(sorted(all_busy_teachers)) or 'teacher(s)'

                if best_days == 0:
                    fail_reason = (
                        'No free slot found at any period across all {} days. '
                        'Teacher(s) {} appear fully booked or all class slots are '
                        'already occupied. Needed {} periods on the SAME period '
                        'each day.'.format(wdays, busy_t_str, needed))
                else:
                    fail_reason = (
                        'Could not find a single period free on all {} required '
                        'days. Best found: Period {} with only {}/{} free days. '
                        'Teacher(s) {} are occupied at conflicting slots. '
                        'Use "Relax to Filler" to allow flexible placement '
                        'across different periods.'.format(
                            needed, best_period, best_days, needed, busy_t_str))

            results[task['idx']] = {
                'phase':       'main',
                'placed':      placed,
                'remaining':   task['remaining'],
                'fail_reason': fail_reason,
            }

        # ── PHASE 2 — Filler Periods (+ relaxed main + main fall-through) ─────────
        for task in tasks:
            if task['remaining'] <= 0:
                if task['idx'] not in results:
                    results[task['idx']] = {
                        'phase':       'main' if is_main(task) else 'filler',
                        'placed':      [],
                        'remaining':   0,
                        'fail_reason': '',
                    }
                continue

            phase  = 'main' if is_main(task) else 'filler'
            placed = list(results.get(task['idx'], {}).get('placed', []))

            # Day-spreading: prefer days where this subject appears least.
            # Pass 1: place at most ceil(n/wdays) per day (capped at 2) to
            # spread periods uniformly; Pass 2 (fallback): no cap, original loop.
            n_total = task['periods']
            nat_max = (n_total + wdays - 1) // wdays
            max_pd  = min(2, nat_max)

            def _subj_count_today(d_):
                return max(
                    (sum(1 for pp in range(ppd)
                         if grid.get(cn_, [[]])[d_][pp] is not None
                         and grid[cn_][d_][pp].get('subject') == task['subject'])
                     for cn_ in task['cn_list'] if cn_ in grid),
                    default=0)

            # Pass 1: place at most max_pd per day, preferring under-populated days
            day_order = sorted(range(wdays), key=_subj_count_today)
            for d in day_order:
                if task['remaining'] <= 0:
                    break
                if _subj_count_today(d) >= max_pd:
                    continue
                for p in range(ppd):
                    if task['remaining'] <= 0:
                        break
                    if slot_free(task, d, p):
                        self._gen_place(task, d, p)
                        placed.append((d, p))

            # Pass 2 (fallback): original loop if any remain unplaced
            for d in range(wdays):
                if task['remaining'] <= 0:
                    break
                for p in range(ppd):
                    if task['remaining'] <= 0:
                        break
                    if slot_free(task, d, p):
                        self._gen_place(task, d, p)
                        placed.append((d, p))

            fail_reason = ''
            if task['remaining'] > 0:
                busy_teachers = set()
                for d in range(wdays):
                    for p in range(ppd):
                        class_ok = all(
                            grid.get(cn, [[]])[d][p] is None
                            for cn in task['cn_list'])
                        if not class_ok:
                            continue
                        t  = task['teacher']
                        pt = task.get('par_teach', '')
                        if not g['t_free'](t, d, p) or g['t_unavail'](t, d, p):
                            busy_teachers.add(t)
                        if pt and pt not in ('', '—', '?'):
                            if not g['t_free'](pt, d, p) or g['t_unavail'](pt, d, p):
                                busy_teachers.add(pt)
                busy_t_str = ', '.join(sorted(busy_teachers)) or 'teacher(s)'
                fail_reason = (
                    '{} slot(s) could not be placed. '
                    'Teacher(s) {} appear fully occupied at all remaining free '
                    'class slots. All {} free grid positions have been '
                    'exhausted.'.format(
                        task['remaining'], busy_t_str,
                        sum(1 for d in range(wdays) for p in range(ppd)
                            if all(grid.get(cn, [[]])[d][p] is None
                                for cn in task['cn_list']))))

            results[task['idx']] = {
                'phase':       phase,
                'placed':      placed,
                'remaining':   task['remaining'],
                'fail_reason': fail_reason,
            }

        return results

    # ── Task Analysis 2 page ─────────────────────────────────────────────────
    def check_ta2_done(self):
        """Return True if stage 2 allocation is complete."""
        return bool(getattr(self, '_last_ta2_allocation', None))

    def run_stage3(self):
        """Run Stage 3 synchronously. Returns status dict."""
        self._progress_log = []
        self._run_stage2_phases()
        return getattr(self, '_stage2_status', {})

    def run_force_fill(self, progress_cb=None):
        """
        Run Force Fill (Min-Conflicts CSP solver) synchronously.
        Stops as soon as all periods are placed (or 1500 iterations max).
        Returns a result dict with:
          - ok:            bool
          - remaining:     int (unplaced period count after force fill)
          - relaxed:       str or None (constraint relaxation notes)
          - overloaded:    list of (teacher, assigned, capacity, excess, unplaced)
          - blocked_only:  list of (teacher, assigned, capacity, unplaced)
          - progress_msgs: list of str
        """
        if self._gen is None:
            return {'ok': False, 'remaining': -1, 'relaxed': None,
                    'overloaded': [], 'blocked_only': [], 'progress_msgs': []}

        progress_msgs = []

        def _cb(msg):
            progress_msgs.append(msg)
            if progress_cb:
                progress_cb(msg)

        relaxed = self._force_fill_backtrack(progress_cb=_cb)

        # Remove any teacher double-bookings that Stage B may have introduced
        self._remove_teacher_conflicts()

        # Refresh timetable snapshot
        self._timetable = self._gen_snapshot_tt()

        g       = self._gen
        tasks   = g['tasks']
        wdays   = g['wdays']
        ppd     = g['ppd']
        remaining = sum(t['remaining'] for t in tasks)
        total_slots = wdays * ppd

        overloaded   = []
        blocked_only = []

        if remaining > 0:
            teacher_assigned = {}
            teacher_unplaced = {}
            for t in tasks:
                for tname in ([t['teacher']] if t['teacher'] else []):
                    teacher_assigned[tname] = teacher_assigned.get(tname, 0) + t['periods']
                    if t['remaining'] > 0:
                        teacher_unplaced[tname] = teacher_unplaced.get(tname, 0) + t['remaining']
                pt = t.get('par_teach', '')
                if pt and pt not in ('', '—', '?'):
                    teacher_assigned[pt] = teacher_assigned.get(pt, 0) + t['periods']
                    if t['remaining'] > 0:
                        teacher_unplaced[pt] = teacher_unplaced.get(pt, 0) + t['remaining']

            for tname, assigned in sorted(teacher_assigned.items()):
                if assigned > total_slots:
                    excess   = assigned - total_slots
                    unp      = teacher_unplaced.get(tname, 0)
                    overloaded.append((tname, assigned, total_slots, excess, unp))

            for tname, unp in sorted(teacher_unplaced.items()):
                if unp > 0 and tname not in {o[0] for o in overloaded}:
                    assigned = teacher_assigned.get(tname, 0)
                    blocked_only.append((tname, assigned, total_slots, unp))

        # Update stage2 status
        self._stage2_status = {
            'unplaced': remaining,
            'ok':       remaining == 0,
            'msg':      ("✅ Force Fill complete — all periods placed!" if remaining == 0
                         else f"⚠ Force Fill done — {remaining} period(s) still unplaced."),
        }

        return {
            'ok':           remaining == 0,
            'remaining':    remaining,
            'relaxed':      relaxed,
            'overloaded':   overloaded,
            'blocked_only': blocked_only,
            'progress_msgs': progress_msgs,
            'wdays':        wdays,
            'ppd':          ppd,
            'total_slots':  total_slots,
        }

    # =========================================================================
    #  FULL AUTO GENERATION  (Step 3 → Final Timetable in one shot)
    # =========================================================================

    def _validate_slot_counts(self):
        """
        Post-generation integrity pass: ensure every subject in every class has
        EXACTLY its configured number of slots placed — no more, no less
        (only stuck-logic reductions of ±1 on Main_sub are acceptable).

        Over-quota: remove the most "replaceable" extra slot (last period in
        Dec order, not HC1, not consec pair, not CT slot).
        Under-quota: logged only — the repair and fill passes handle placement.

        Returns list of violation strings for progress log.
        """
        if self._gen is None:
            return []

        g     = self._gen
        grid  = g['grid']
        tasks = g['tasks']
        wdays = g['wdays']
        ppd   = g['ppd']
        DAYS  = g['DAYS']
        violations = []

        # Build a map: (cn, subject) → task
        task_map = {}
        for t in tasks:
            for cn in t['cn_list']:
                task_map[(cn, t['subject'])] = t

        for cn in g['all_classes']:
            cd = self.class_config_data.get(cn, {})
            ct_t  = cd.get('teacher', '').strip()
            ct_pi = int(cd.get('teacher_period', 0)) - 1  # 0-based

            for s in cd.get('subjects', []):
                subj    = s.get('name', '').strip()
                quota   = s.get('periods', 0)
                if not subj or quota <= 0:
                    continue

                # Count placed slots for this subject in this class
                placed_slots = [
                    (d, p)
                    for d in range(wdays)
                    for p in range(ppd)
                    if grid[cn][d][p] is not None
                    and grid[cn][d][p].get('subject') == subj
                ]
                count = len(placed_slots)

                if count == quota:
                    continue  # ✓ exact

                if count > quota:
                    # Over-quota: remove extras from Dec order, skip CT/HC1 slots
                    task_obj = task_map.get((cn, subj))
                    extra = count - quota
                    # Sort slots: last period first (Dec) so we remove the least critical
                    removable = sorted(
                        [(d, p) for d, p in placed_slots
                         if not (grid[cn][d][p] or {}).get('is_ct')
                         and p != ct_pi],   # never remove CT period slot
                        key=lambda dp: (-dp[1], -dp[0])  # Dec: high p first
                    )
                    removed = 0
                    for d, p in removable:
                        if removed >= extra:
                            break
                        cell = grid[cn][d][p]
                        if cell is None:
                            continue
                        t_idx = g['task_at'][cn][d][p]
                        t_obj = tasks[t_idx] if t_idx is not None and t_idx < len(tasks) else None
                        if t_obj and t_obj.get('is_ct'):
                            continue  # never remove CT slots
                        # Unplace
                        if t_obj:
                            self._gen_unplace(t_obj, d, p)
                        else:
                            # orphan cell (from _fill_freed_slots)
                            t_name = cell.get('teacher', '')
                            pt = cell.get('par_teach', '')
                            grid[cn][d][p] = None
                            g['task_at'][cn][d][p] = None
                            if t_name and t_name not in ('', '—', '?'):
                                g['t_busy'].get(t_name, set()).discard((d, p))
                            if pt and pt not in ('', '—', '?'):
                                g['t_busy'].get(pt, set()).discard((d, p))
                        violations.append(
                            f'OVER-QUOTA removed: {cn} {DAYS[d]} P{p+1} '
                            f'"{subj}" (had {count}, quota={quota})')
                        removed += 1

                elif count < quota:
                    violations.append(
                        f'UNDER-QUOTA: {cn} "{subj}" has {count}/{quota} '
                        f'(fill passes should handle this)')

        return violations

    def _check_ct_violations(self):
        """
        Read-only CT integrity check.  Returns a list of violation dicts:
          {class, day, period_1based, expected_subject, actual_subject/None}

        A violation is any working day where the configured CT period slot for
        a class does NOT contain the CT teacher's subject.
        """
        if self._gen is None:
            return []

        g     = self._gen
        grid  = g['grid']
        wdays = g['wdays']
        ppd   = g['ppd']
        DAYS  = g['DAYS']
        violations = []

        for cn in g['all_classes']:
            cd = self.class_config_data.get(cn, {})
            ct_teacher = cd.get('teacher', '').strip()
            ct_per_raw = cd.get('teacher_period', None)
            if not ct_teacher or not ct_per_raw:
                continue
            ct_p = int(ct_per_raw) - 1   # 0-based
            if ct_p < 0 or ct_p >= ppd:
                continue

            # Find expected CT subject (highest-periods subject by CT teacher)
            ct_subj = None
            ct_max  = 0
            for s in cd.get('subjects', []):
                if s.get('teacher', '').strip() == ct_teacher:
                    if s.get('periods', 0) > ct_max:
                        ct_max  = s['periods']
                        ct_subj = s['name']
            if not ct_subj:
                continue

            # Collect all subjects the CT teacher teaches in this class
            ct_all_subjects = {
                s['name'] for s in cd.get('subjects', [])
                if s.get('teacher', '').strip() == ct_teacher
            }

            for d in range(wdays):
                cell = grid.get(cn, [[]])[d][ct_p] if d < len(grid.get(cn, [])) else None
                actual_subj    = cell.get('subject', '') if cell else None
                actual_teacher = cell.get('teacher', '').strip() if cell else None
                # Valid if: cell is filled AND subject is by the CT teacher
                valid = (
                    actual_subj is not None and
                    actual_teacher == ct_teacher and
                    actual_subj in ct_all_subjects
                )
                if not valid:
                    violations.append({
                        'class':            cn,
                        'day':              DAYS[d],
                        'period_1based':    ct_p + 1,
                        'expected_subject': ct_subj + ' (or any CT-teacher subject)',
                        'actual_subject':   actual_subj or 'FREE',
                        'ct_teacher':       ct_teacher,
                    })

        return violations

    def _repair_ct_periods(self):
        """
        Final safety pass: guarantee every class has its CT subject at the
        configured CT period on EVERY working day.

        For each class:
          1. Find the HC1 task (is_ct=True).
          2. For each working day:
             a. If slot (d, ct_per-1) is None and task.remaining > 0: place it.
             b. If slot holds the WRONG subject (not the CT subject): evict the
                intruder (increment its task.remaining), then place the CT subject.
             c. If slot already has the CT subject: ensure is_ct=True on the cell.

        Never moves HC1 cells that are already correct.
        Rebuilds t_busy after all moves so teacher-wise state stays consistent.

        Returns list of repair records for logging.
        """
        if self._gen is None:
            return []

        g     = self._gen
        grid  = g['grid']
        tasks = g['tasks']
        wdays = g['wdays']
        ppd   = g['ppd']
        DAYS  = g['DAYS']

        repairs = []

        # Build HC1 task index: cn -> task
        # Use class_config_data as ground truth (not is_ct flag which may be stale
        # after overflow demotion).  Find the task by matching CLASS + CT-SUBJECT-NAME.
        hc1_by_class = {}

        # First pass: tasks with is_ct=True (fast path)
        for t in tasks:
            if t.get('is_ct'):
                for cn in t['cn_list']:
                    if cn not in hc1_by_class:
                        hc1_by_class[cn] = t

        # Second pass: any class whose CT task lost is_ct (e.g. overflow demotion edge case)
        for cn in g['all_classes']:
            if cn in hc1_by_class:
                continue
            cd2 = self.class_config_data.get(cn, {})
            ct_t2 = cd2.get('teacher', '').strip()
            ct_p2 = cd2.get('teacher_period', None)
            if not ct_t2 or not ct_p2:
                continue
            # Find expected CT subject (max-periods subject by CT teacher)
            ct_subjs2 = [(s['name'], s.get('periods', 0))
                         for s in cd2.get('subjects', [])
                         if s.get('teacher', '').strip() == ct_t2]
            if not ct_subjs2:
                continue
            ct_subj_name = max(ct_subjs2, key=lambda x: x[1])[0]
            # Find matching task
            for t in tasks:
                if (cn in t['cn_list']
                        and t['teacher'] == ct_t2
                        and t['subject'] == ct_subj_name
                        and t.get('ct_period') == int(ct_p2)):
                    hc1_by_class[cn] = t
                    break

        for cn, hc1_task in hc1_by_class.items():
            ct_p = hc1_task['ct_period'] - 1   # 0-based
            ct_subj = hc1_task['subject']
            ct_teacher = hc1_task['teacher']

            if ct_p < 0 or ct_p >= ppd:
                continue

            for d in range(wdays):
                cell = grid[cn][d][ct_p]

                # Case A: slot has the CT subject — just ensure is_ct flag is set
                if cell is not None and cell.get('subject') == ct_subj:
                    if not cell.get('is_ct'):
                        grid[cn][d][ct_p] = dict(cell, is_ct=True)
                    continue

                # Case B: slot is occupied by wrong subject — evict intruder
                if cell is not None and cell.get('subject') != ct_subj:
                    intruder_idx = g['task_at'][cn][d][ct_p]
                    if intruder_idx is not None:
                        intruder = tasks[intruder_idx]
                        # Never evict another HC1 task (shouldn't happen, but guard)
                        if intruder.get('is_ct'):
                            continue
                        # Evict
                        self._gen_unplace(intruder, d, ct_p)
                        repairs.append(
                            'EVICT: {} {} P{} — "{}" removed to place CT "{}"'.format(
                                cn, DAYS[d], ct_p+1,
                                intruder['subject'], ct_subj))
                    else:
                        # Cell with no task index (written by _fill_freed_slots)
                        i_teacher = cell.get('teacher', '')
                        i_pt = cell.get('par_teach', '')
                        grid[cn][d][ct_p] = None
                        if i_teacher and i_teacher not in ('', '—', '?'):
                            g['t_busy'].get(i_teacher, set()).discard((d, ct_p))
                        if i_pt and i_pt not in ('', '—', '?'):
                            g['t_busy'].get(i_pt, set()).discard((d, ct_p))
                        repairs.append(
                            'CLEAR: {} {} P{} — orphan cell removed to place CT "{}"'.format(
                                cn, DAYS[d], ct_p+1, ct_subj))
                    cell = None  # fall through to Case C

                # Case C: slot is empty and CT task has remaining — place it
                if cell is None and hc1_task['remaining'] > 0:
                    # Free primary teacher if busy at this slot by a non-HC1 task
                    if not g['t_free'](ct_teacher, d, ct_p):
                        for cn2 in g['all_classes']:
                            idx2 = g['task_at'][cn2][d][ct_p]
                            if idx2 is None:
                                continue
                            t2 = tasks[idx2]
                            if (t2['teacher'] == ct_teacher or
                                    (t2.get('par_teach') or '').strip() == ct_teacher):
                                if not t2.get('is_ct'):
                                    self._gen_unplace(t2, d, ct_p)
                                    repairs.append(
                                        'EVICT-T: {} {} P{} — "{}" evicted to free teacher for CT'.format(
                                            cn2, DAYS[d], ct_p+1, t2['subject']))
                                    break

                    if g['t_free'](ct_teacher, d, ct_p):
                        self._gen_place(hc1_task, d, ct_p)
                        repairs.append(
                            'PLACE: {} {} P{} — CT "{}" placed'.format(
                                cn, DAYS[d], ct_p+1, ct_subj))

                # Case C+: slot empty, primary CT fully placed — try secondary
                # CT-teacher subjects (P1a: CT teacher's other subjects fill CT slot)
                if cell is None and hc1_task['remaining'] == 0:
                    cd_cn  = self.class_config_data.get(cn, {})
                    ct_t2  = cd_cn.get('teacher', '').strip()
                    # All tasks for this class with teacher==CT, not is_ct, remaining>0
                    sec_tasks = sorted(
                        [t2 for t2 in tasks
                         if cn in t2['cn_list']
                         and t2['teacher'] == ct_t2
                         and not t2.get('is_ct')
                         and t2['remaining'] > 0],
                        key=lambda t2: -t2['periods']
                    )
                    for sec in sec_tasks:
                        if not g['t_free'](ct_t2, d, ct_p):
                            # Try to evict non-HC1 blocker first
                            for cn3 in g['all_classes']:
                                idx3 = g['task_at'][cn3][d][ct_p]
                                if idx3 is None: continue
                                t3 = tasks[idx3]
                                if (t3['teacher'] == ct_t2 or
                                        (t3.get('par_teach') or '').strip() == ct_t2):
                                    if not t3.get('is_ct'):
                                        self._gen_unplace(t3, d, ct_p)
                                        repairs.append(
                                            'EVICT-T2: {} {} P{} — freed for secondary CT'.format(
                                                cn3, DAYS[d], ct_p+1))
                                        break
                        if not all(grid[c4][d][ct_p] is None for c4 in sec['cn_list']):
                            continue
                        if g['t_free'](ct_t2, d, ct_p):
                            _orig = sec['p_pref']
                            sec['p_pref'] = []
                            self._gen_place(sec, d, ct_p)
                            sec['p_pref'] = _orig
                            # Mark as CT slot fill with secondary subject
                            if grid[cn][d][ct_p] is not None:
                                grid[cn][d][ct_p] = dict(
                                    grid[cn][d][ct_p], is_ct=True, ct_fill_secondary=True)
                            repairs.append(
                                'SECONDARY-CT: {} {} P{} — "{}" (secondary CT subj)'.format(
                                    cn, DAYS[d], ct_p+1, sec['subject']))
                            cell = grid[cn][d][ct_p]  # update local ref
                            break

                # Case E (P1 absolute guarantee): slot STILL empty — force fill
                # with ANY available subject by ANY teacher. Never leave empty.
                if grid[cn][d][ct_p] is None:
                    cd_cn2  = self.class_config_data.get(cn, {})
                    # Pass 1: any CT-teacher subject (even already-maxed)
                    ct_t3 = cd_cn2.get('teacher', '').strip()
                    for any_t in sorted(
                            [t2 for t2 in tasks if cn in t2['cn_list']
                             and t2['teacher'] == ct_t3],
                            key=lambda t2: -t2['periods']):
                        if not g['t_free'](ct_t3, d, ct_p):
                            break
                        if not all(grid[c4][d][ct_p] is None for c4 in any_t['cn_list']):
                            continue
                        _orig2 = any_t['p_pref']
                        any_t['p_pref'] = []
                        any_t['remaining'] += 1   # temporarily bump so _gen_place works
                        self._gen_place(any_t, d, ct_p)
                        any_t['p_pref'] = _orig2
                        if grid[cn][d][ct_p] is not None:
                            grid[cn][d][ct_p] = dict(
                                grid[cn][d][ct_p], is_ct=True, ct_fill_secondary=True)
                        repairs.append(
                            'FORCE-CT: {} {} P{} — "{}" (CT teacher, over-quota)'.format(
                                cn, DAYS[d], ct_p+1, any_t['subject']))
                        break

                    # Pass 2: absolute last resort — any subject by any free teacher
                    if grid[cn][d][ct_p] is None:
                        for any_t2 in sorted(
                                [t2 for t2 in tasks if cn in t2['cn_list']],
                                key=lambda t2: -t2['periods']):
                            if not g['t_free'](any_t2['teacher'], d, ct_p):
                                continue
                            if not all(grid[c4][d][ct_p] is None
                                       for c4 in any_t2['cn_list']):
                                continue
                            _orig3 = any_t2['p_pref']
                            any_t2['p_pref'] = []
                            any_t2['remaining'] = max(1, any_t2['remaining'])
                            self._gen_place(any_t2, d, ct_p)
                            any_t2['p_pref'] = _orig3
                            repairs.append(
                                'LASTRESORT: {} {} P{} — "{}" (any teacher, P1 guarantee)'.format(
                                    cn, DAYS[d], ct_p+1, any_t2['subject']))
                            break

                # Case D: slot has a subject — ensure is_ct flag is correct.
                # If subject belongs to CT teacher, mark is_ct=True.
                if grid[cn][d][ct_p] is not None:
                    cell2   = grid[cn][d][ct_p]
                    cd_cn3  = self.class_config_data.get(cn, {})
                    ct_t4   = cd_cn3.get('teacher', '').strip()
                    ct_subs4 = {_s['name'] for _s in cd_cn3.get('subjects', [])
                                if _s.get('teacher', '').strip() == ct_t4}
                    should_be_ct = (
                        cell2.get('subject') in ct_subs4 and
                        cell2.get('teacher', '').strip() == ct_t4
                    )
                    if should_be_ct and not cell2.get('is_ct'):
                        grid[cn][d][ct_p] = dict(cell2, is_ct=True)

        # Rebuild t_busy to reflect all moves
        g['t_busy'].clear()
        for cn2 in g['all_classes']:
            for d2 in range(wdays):
                for p2 in range(ppd):
                    e = grid[cn2][d2][p2]
                    if e:
                        t  = e.get('teacher', '')
                        pt = e.get('par_teach', '')
                        if t  and t  not in ('', '—', '?'):
                            g['t_busy'].setdefault(t,  set()).add((d2, p2))
                        if pt and pt not in ('', '—', '?'):
                            g['t_busy'].setdefault(pt, set()).add((d2, p2))

        return repairs

    def run_full_generation(self, progress_cb=None):
        """
        Run ALL generation stages automatically without any UI interaction.

        Pipeline:
          1. Stage 1  — HC1 (CT periods) + HC2 (preference/fixed periods)
          2. Task Analysis 1 — place combined/parallel group slots
          3. Task Analysis 2 — place remaining filler slots
          4. Stage 2/3 — SC1 (consecutive), SC2 (daily), fillers + repair loop
          5. Force Fill — backtracking solver (swap chains, shuffle)
          6. Auto Period Reduction — if still stuck, reduce the most-periods
             subject by 1 for the blocked class, then re-run from scratch.
             Repeated up to MAX_REDUCE_ATTEMPTS times.
          7. Post-process — balance teacher free-period distribution across halves.

        Hard Constraints enforced throughout:
          • Teacher cannot be in two places at once.
          • A class cannot have two teachers at the same slot.
          • No class free period (all assigned periods must be placed).

        Soft Constraints (best-effort):
          • Main subjects at the same period each day.
          • Teacher workload distributed evenly — no day with zero teaching
            while another day is completely packed.
          • At least 1 free period in each half per teacher per day.

        Returns dict:
          ok               – bool, True if 0 unplaced periods
          remaining        – int, unplaced period count
          overloaded       – list of (teacher, assigned, cap, excess, unplaced)
          blocked_only     – list of (teacher, assigned, cap, unplaced)
          period_reductions – list of reduction records applied
          progress_log     – list of progress messages
          wdays, ppd, total_slots
        """
        self._progress_log      = []
        self._period_reductions = []

        def _prog(msg):
            # _gen_prog stores (msg, pct) tuples — normalise to plain str
            if isinstance(msg, tuple):
                msg = msg[0]
            msg = str(msg)
            self._progress_log.append(msg)
            if progress_cb:
                progress_cb(msg)

        MAX_REDUCE_ATTEMPTS = 12

        for attempt in range(MAX_REDUCE_ATTEMPTS + 1):
            if attempt > 0:
                _prog(f"\n🔄 Retry attempt {attempt}/{MAX_REDUCE_ATTEMPTS} "
                      f"after period reduction(s)…")

            # ── 1. Init ────────────────────────────────────────────────────
            _prog("⚙  Initialising timetable engine…")
            self._init_gen_state()
            # Only reset relaxation state on first attempt.
            # On retries (after period reduction), keep relaxations from
            # stuck logic Steps 2/3 so they carry forward.
            if attempt == 0:
                self._relaxed_consec_keys = set()
                self._relaxed_main_keys   = set()

            # ── 2. Stage 1: P1_CT (CT periods) + Phase 1b (secondary CT) + P2 (Pref_Per/Day) ──
            _prog("📌 Stage 1 — placing CT & preference periods…")
            self._run_stage1_phases()
            s1_stat = getattr(self, '_stage1_status', {})
            _prog(f"   {s1_stat.get('stage_txt', 'done')}")

            # ── 3. P1 Parallel/Combined — most constrained P1 constraint (TA1) ──────
            # Parallel/combined groups need multiple teachers + classes free
            # simultaneously — place them FIRST after HC1/HC2 (most constrained).
            _prog("🔗 Task Analysis — combined / parallel groups…")
            try:
                grp_slots, allocation, all_rows = self._run_task_analysis_allocation()
                self._last_allocation  = allocation
                self._last_group_slots = grp_slots
                self._last_all_rows    = all_rows
                ok_n   = sum(1 for ar in allocation.values() if ar.get('ok'))
                fail_n = len(allocation) - ok_n
                _prog(f"   Groups: {ok_n} OK / {fail_n} failed")
            except Exception as exc:
                _prog(f"   Task Analysis error: {exc}")

            # ── 4. P1 Parallel/Combined fallthrough (TA2) ────────────────────────────
            _prog("📋 Task Analysis 2 — allocating remaining slots…")
            try:
                ta2_result = self._run_ta2_allocation()
                self._last_ta2_allocation = ta2_result
                ta2_ok = sum(1 for ar in ta2_result.values()
                             if isinstance(ar, dict) and ar.get('remaining', 1) == 0)
                _prog(f"   TA2: {ta2_ok}/{len(ta2_result)} groups fully placed")
            except Exception as exc:
                _prog(f"   TA2 error: {exc}")

            # ── 5. P3 (Main_sub) → P5 (Cons_per) → Filler_sub ──────────────────────
            # SC2/SC1 run INSIDE stage2_phases with correct priority ordering.
            _prog("📅 Stage 2/3 — main subjects, consecutive, fillers…")
            self._run_sc2_phase()    # Inc order — main subjects first
            self._run_sc1_phase()    # Dec order — consecutive pairs
            self._run_stage2_phases()  # fillers + repair
            s3_stat     = getattr(self, '_stage2_status', {})
            unplaced_s3 = s3_stat.get('unplaced', 0)
            _prog(f"   After Stage 2/3: {unplaced_s3} period(s) still unplaced")

            # ── 6. STUCK LOGIC (spec-compliant, in-place, no grid restart) ──
            self._timetable = self._gen_snapshot_tt()
            remaining = sum(t['remaining'] for t in self._gen['tasks'])
            _prog(f"   After Stage 3: {remaining} period(s) unplaced")

            if remaining > 0:
                _prog("🔧 Stuck Logic — applying in-place resolution steps…")
                _reductions_before = len(self._period_reductions)
                try:
                    remaining = self._run_stuck_logic(progress_cb=_prog)
                    # Do NOT call _remove_teacher_conflicts here —
                    # _force_fill_backtrack (called inside stuck logic) already
                    # handles conflicts via Stage C. Calling it here would remove
                    # placed slots and increment remaining with nothing to fill them.
                    self._timetable = self._gen_snapshot_tt()
                    _prog(f"   After Stuck Logic: {remaining} period(s) unplaced")
                except Exception as exc:
                    _prog(f"   Stuck Logic error: {exc}")
                _stuck_made_reductions = len(self._period_reductions) > _reductions_before

            self._timetable = self._gen_snapshot_tt()
            remaining = sum(t['remaining'] for t in self._gen['tasks'])

            if remaining == 0:
                _prog("✅ Complete — all periods placed!")
                break

            # If still stuck after all relaxations: period reduction + grid restart
            if attempt < MAX_REDUCE_ATTEMPTS:
                # If stuck logic already did reductions in Step 4, skip the outer
                # _auto_reduce on this attempt (retry loop will use the reductions
                # already made when re-initing the grid on the next attempt).
                if _stuck_made_reductions:
                    _prog(f"⚠  {remaining} period(s) still unplaced — "
                          f"retrying with period reductions already applied by stuck logic…")
                else:
                    _prog(f"⚠  {remaining} period(s) still unplaced — applying "
                          f"period reduction(s) to free blocked slots…")
                    reductions = self._auto_reduce_stuck_periods()
                    if not reductions:
                        _prog("⚠  No further period reduction possible — "
                              "timetable is as complete as constraints allow.")
                        break
                    for r in reductions:
                        _prog(f"📉  '{r['subject']}' in {r['class']} "
                              f"(teacher: {r['teacher']}): "
                              f"{r['from_periods']} → {r['to_periods']} periods/week")
            else:
                _prog(f"⚠  Max reduction attempts reached — "
                      f"{remaining} period(s) remain unplaced.")

        # ── 7. Post-process: balance teacher free periods ─────────────────
        _prog("🔄 Post-processing — balancing teacher & class distributions…")
        try:
            self._ensure_half_free_periods()
            self._balance_class_subject_distribution()
            self._timetable = self._gen_snapshot_tt()
        except Exception as exc:
            _prog(f"   Post-process warning: {exc}")

        # ── 8. Fill/Validate loop ────────────────────────────────────────────
        # Repeat: fill free slots → validate (remove over-quota) → fill again.
        # After validate removes over-quota slots, fill runs again to place them.
        # Loop until stable (no more changes).
        _prog("📝 Fill & validate loop…")
        for _fv_pass in range(4):
            try:
                self._fill_freed_slots()
            except Exception as exc:
                _prog(f"   Fill warning pass {_fv_pass+1}: {exc}")
            try:
                violations = self._validate_slot_counts()
                over = [v for v in violations if v.startswith('OVER')]
                if not over:
                    _prog(f"   ✅ Pass {_fv_pass+1}: all slot counts correct, no free slots")
                    break
                _prog(f"   Pass {_fv_pass+1}: {len(over)} over-quota slot(s) removed — re-filling…")
            except Exception as exc:
                _prog(f"   Validate warning pass {_fv_pass+1}: {exc}")
                break
        self._timetable = self._gen_snapshot_tt()

        # ── 9. CT Period Integrity Repair ─────────────────────────────────
        # Run AFTER _fill_freed_slots so any wrongly-placed slots are
        # corrected before the timetable is snapshotted for display/export.
        _prog("🔒 CT period integrity check & repair…")
        try:
            ct_repairs = self._repair_ct_periods()
            self._timetable = self._gen_snapshot_tt()
            if ct_repairs:
                _prog(f"   ⚠ {len(ct_repairs)} CT slot(s) repaired:")
                for r in ct_repairs[:8]:
                    _prog(f"     • {r}")
            else:
                _prog("   ✅ All CT periods verified — no violations found")
        except Exception as exc:
            _prog(f"   CT repair warning: {exc}")

        # ── Final snapshot after all repairs ──────────────────────────────
        # Take snapshot here so self._timetable always reflects the final
        # repaired grid (CT integrity pass above may have changed cells).
        self._timetable = self._gen_snapshot_tt()

        # ── Verify CT integrity (for result reporting) ────────────────────
        ct_violations = self._check_ct_violations()

        # ── Recount remaining from actual grid state ─────────────────────
        # After fill/validate, some task.remaining may be stale (fill placed
        # orphan cells without updating task counters). Recount from grid.
        for t in self._gen['tasks']:
            placed_actual = sum(
                1 for d in range(self._gen['wdays'])
                for p in range(self._gen['ppd'])
                for cn in t['cn_list']
                if (cn in self._gen['task_at']
                    and self._gen['task_at'][cn][d][p] == t['idx'])
            ) // max(1, len(t['cn_list']))  # combined tasks count once per slot
            t['remaining'] = max(0, t['periods'] - placed_actual)

        # ── Build result dict ─────────────────────────────────────────────
        final_remaining  = sum(t['remaining'] for t in self._gen['tasks'])
        total_slots      = self._gen['wdays'] * self._gen['ppd']

        teacher_assigned = {}
        teacher_unplaced = {}
        for t in self._gen['tasks']:
            if t['teacher']:
                teacher_assigned[t['teacher']] = (
                    teacher_assigned.get(t['teacher'], 0) + t['periods'])
                if t['remaining'] > 0:
                    teacher_unplaced[t['teacher']] = (
                        teacher_unplaced.get(t['teacher'], 0) + t['remaining'])
            pt = (t.get('par_teach') or '').strip()
            if pt and pt not in ('—', '?'):
                teacher_assigned[pt] = teacher_assigned.get(pt, 0) + t['periods']
                if t['remaining'] > 0:
                    teacher_unplaced[pt] = teacher_unplaced.get(pt, 0) + t['remaining']

        overloaded   = []
        blocked_only = []
        for tname, assigned in sorted(teacher_assigned.items()):
            if assigned > total_slots:
                overloaded.append((tname, assigned, total_slots,
                                   assigned - total_slots,
                                   teacher_unplaced.get(tname, 0)))
            elif tname in teacher_unplaced:
                blocked_only.append((tname, assigned, total_slots,
                                     teacher_unplaced[tname]))

        return {
            'ok':                final_remaining == 0,
            'remaining':         final_remaining,
            'overloaded':        overloaded,
            'blocked_only':      blocked_only,
            'period_reductions': list(self._period_reductions),
            'progress_log':      list(self._progress_log),
            'wdays':             self._gen['wdays'],
            'ppd':               self._gen['ppd'],
            'total_slots':       total_slots,
            'ct_violations':     ct_violations,
        }

    # ── Auto Period Reduction ─────────────────────────────────────────────────

    def _auto_reduce_stuck_periods(self):
        """
        For every still-unplaced task, free a slot in its class(es) by reducing
        the period count of the highest-period already-fully-placed subject in
        that same class by 1.  This is the "deadlock breaker":

          'If Maths (7 periods) in 6A blocks SUPRIYA → reduce Maths to 6.'

        Strategy (smarter than reducing the stuck task itself):
          1. Collect all tasks with remaining > 0  (the "stuck" set).
          2. For each stuck task, look at each of its class(es).
          3. In that class, find the subject that:
               a) is NOT the stuck task itself,
               b) is already FULLY PLACED (remaining == 0),
               c) has more than 1 period,
               d) is not an HC1 / CT task,
               e) has the MOST periods (so reducing it hurts least proportionally).
          4. Reduce that subject by 1 in class_config_data (persists across retries).
          5. Record the reduction and avoid reducing the same (class, subject)
             pair twice in the same call.

        Returns a list of reduction record dicts (may be empty).
        """
        tasks    = self._gen['tasks']
        grid     = self._gen['grid']
        wdays    = self._gen['wdays']
        ppd      = self._gen['ppd']

        stuck_tasks = [t for t in tasks if t['remaining'] > 0]
        if not stuck_tasks:
            return []

        reductions   = []
        reduced_keys = set()   # (cn, subj_name) already reduced this call

        for stuck in stuck_tasks:
            for cn in stuck['cn_list']:
                if cn not in self.class_config_data:
                    continue
                cd_subjects = self.class_config_data[cn].get('subjects', [])

                # Build a candidate list: fully-placed, non-CT, periods > 1
                candidates = []
                for s in cd_subjects:
                    sn = s.get('name', '').strip()
                    if (cn, sn) in reduced_keys:
                        continue
                    if s.get('periods', 0) <= 1:
                        continue
                    # Skip the stuck task's own subject
                    if sn == stuck['subject']:
                        continue
                    # Skip HC1 (CT subjects) — those are fixed
                    # Find the engine task for this (cn, subject) pair
                    eng_task = next(
                        (t for t in tasks
                         if cn in t['cn_list'] and t['subject'] == sn),
                        None)
                    if eng_task and eng_task.get('is_ct'):
                        continue
                    # Must be fully placed (remaining == 0) — we are freeing
                    # one of its placed slots, so the class gets a free period
                    # that the stuck teacher can use
                    if eng_task and eng_task['remaining'] > 0:
                        continue

                    candidates.append(s)

                if not candidates:
                    # Fallback: allow reducing the stuck task's OWN subject if
                    # it has remaining > 0 AND periods > remaining
                    # (i.e. at least one occurrence will stay placed)
                    for s in cd_subjects:
                        sn = s.get('name', '').strip()
                        if (cn, sn) in reduced_keys:
                            continue
                        if s.get('periods', 0) <= max(1, stuck['remaining']):
                            continue
                        if s.get('name', '').strip() != stuck['subject']:
                            continue
                        candidates.append(s)
                    if not candidates:
                        continue

                # Pick subject with the most periods (losing 1 period hurts least)
                best = max(candidates, key=lambda s: s.get('periods', 0))
                old_val       = best['periods']
                best['periods'] = old_val - 1
                key = (cn, best['name'])
                reduced_keys.add(key)
                rec = {
                    'class':        cn,
                    'subject':      best['name'],
                    'teacher':      stuck['teacher'],
                    'from_periods': old_val,
                    'to_periods':   best['periods'],
                }
                self._period_reductions.append(rec)
                reductions.append(rec)
                # One reduction per stuck class is enough for this attempt;
                # break to next stuck task
                break

        return reductions

    def _fill_freed_slots(self):
        """
        Hard Constraint: no class may ever have a free period.

        For every free (None) slot in every class, find the best subject to
        place there.  Candidates are evaluated in three passes with progressively
        relaxed constraints so every slot is filled:

          Pass 1 — Strict:
            • Subject not already on this day in this class
            • Teacher free at (d, pp)
            • Placed-count + 1 ≤ wdays  (≤ 1 per day normally)
            • Prefer Main subjects (periods ≥ wdays) — rank by placed count desc

          Pass 2 — Relax "same-day" rule:
            Same as Pass 1 but allow subjects already on this day (same subject
            twice in one day) — still checks teacher availability.

          Pass 3 — Relax teacher-free check:
            Place the subject with the most placed periods regardless — marks
            the teacher busy so no double-booking is introduced in the grid
            meta-data, even if the teacher is technically overloaded that slot.
            This is a last-resort safety net.

        Parallel subjects: when a subject has a parallel partner, the parallel
        teacher is also marked busy at the chosen slot so the teacher-wise view
        remains consistent.
        """
        g      = self._gen
        grid   = g['grid']
        wdays  = g['wdays']
        ppd    = g['ppd']
        tasks  = g['tasks']

        t_busy_gen = g.setdefault('t_busy', {})

        def _placed_total(cn, subj_name):
            return sum(
                1 for d2 in range(wdays) for pp2 in range(ppd)
                if grid[cn][d2][pp2] is not None
                and grid[cn][d2][pp2].get('subject') == subj_name
            )

        def _teacher_free(teacher, d, p):
            if not teacher or teacher in ('—', '?', ''):
                return True
            return (d, p) not in t_busy_gen.get(teacher, set())

        def _is_main(subj_name, cn):
            """True if subject periods >= wdays (i.e. it appears every day)."""
            cd_subjs = self.class_config_data.get(cn, {}).get('subjects', [])
            for s in cd_subjs:
                if s.get('name', '').strip() == subj_name:
                    return s.get('periods', 0) >= wdays
            return False

        def _place_cell(cn, d, pp, sname, teacher, s_data):
            """Write cell into grid and mark both primary and parallel teacher busy."""
            par_subj    = (s_data.get('parallel_subject') or '').strip()
            par_teacher = (s_data.get('parallel_teacher') or '').strip()
            if not s_data.get('parallel'):
                par_subj = par_teacher = ''
            grid[cn][d][pp] = {
                'type':             'parallel' if par_teacher else 'filler_extra',
                'subject':          sname,
                'teacher':          teacher,
                'par_subj':         par_subj or '—',
                'par_teach':        par_teacher or '—',
                'combined_classes': [],
                'is_ct':            False,
            }
            if teacher and teacher not in ('—', '?', ''):
                t_busy_gen.setdefault(teacher, set()).add((d, pp))
            if par_teacher and par_teacher not in ('—', '?', ''):
                t_busy_gen.setdefault(par_teacher, set()).add((d, pp))

        # Build CT period index per class so we never fill it with a wrong subject
        ct_period_idx = {}  # cn -> 0-based period index of CT period
        for cn in g['all_classes']:
            cd = self.class_config_data.get(cn, {})
            ct_per = cd.get('teacher_period', None)
            if ct_per:
                ct_period_idx[cn] = int(ct_per) - 1  # convert to 0-based

        for cn in g['all_classes']:
            if cn not in grid:
                continue
            cd_subjects = self.class_config_data.get(cn, {}).get('subjects', [])
            if not cd_subjects:
                continue

            for d in range(wdays):
                for pp in range(ppd):
                    if grid[cn][d][pp] is not None:
                        continue  # already filled

                    # CT period slot guard (P1a + P1):
                    # The CT period must hold a subject by the CT teacher.
                    # _fill_freed_slots must NEVER place a non-CT-teacher subject
                    # here — that would corrupt CT integrity.
                    # The _repair_ct_periods pass (step 9) will fill it correctly.
                    if pp == ct_period_idx.get(cn, -1):
                        cd_fill = self.class_config_data.get(cn, {})
                        ct_fill_teacher = cd_fill.get('teacher', '').strip()
                        # Skip this slot — repair pass handles it with correct
                        # priority (primary CT subject first, then secondary).
                        # Only allow _fill_freed_slots to write here if there is
                        # literally NO CT teacher defined (edge case).
                        if ct_fill_teacher:
                            continue  # always reserved for CT repair pass

                    # Subjects already placed today
                    today_subjects = {
                        grid[cn][d][pp2].get('subject', '')
                        for pp2 in range(ppd)
                        if grid[cn][d][pp2] is not None
                    }

                    # Build candidate list for each pass
                    # (current_placed, is_main_flag, sname, teacher, s_data)
                    def _build_candidates(allow_same_day, allow_teacher_busy):
                        cands = []
                        for s in cd_subjects:
                            sname   = s.get('name', '').strip()
                            teacher = (s.get('teacher') or '').strip()
                            if not sname:
                                continue
                            if not allow_same_day and sname in today_subjects:
                                continue
                            # Never place an HC1 (CT-fixed) subject as filler
                            task_obj = next(
                                (t for t in tasks
                                 if cn in t['cn_list'] and t['subject'] == sname
                                 and t.get('is_ct')),
                                None)
                            if task_obj:
                                continue
                            current = _placed_total(cn, sname)
                            # Hard quota: never place more than the configured period count.
                            # Stuck-logic may have reduced a main subject by 1 in
                            # class_config_data, so s['periods'] already reflects that.
                            max_quota = s.get('periods', 0)
                            if current >= max_quota:
                                continue
                            if not allow_teacher_busy:
                                if not _teacher_free(teacher, d, pp):
                                    continue
                                # Also check parallel teacher availability
                                if s.get('parallel'):
                                    par_t = (s.get('parallel_teacher') or '').strip()
                                    if par_t and par_t not in ('', '—', '?'):
                                        if not _teacher_free(par_t, d, pp):
                                            continue
                            main_flag = _is_main(sname, cn)
                            cands.append((current, int(main_flag), sname, teacher, s))
                        return cands

                    placed = False
                    for (same_day_ok, teacher_busy_ok) in [
                        (False, False),   # Pass 1: strict
                        (True,  False),   # Pass 2: allow same day
                        (True,  True),    # Pass 3: allow any
                    ]:
                        cands = _build_candidates(same_day_ok, teacher_busy_ok)
                        if not cands:
                            continue
                        # Sort: prefer main subjects, then by FEWEST already
                        # placed (to spread periods evenly, avoid stacking)
                        cands.sort(key=lambda x: (-x[1], x[0]))
                        _, _, best_subj, best_teacher, best_s = cands[0]
                        _place_cell(cn, d, pp, best_subj, best_teacher, best_s)
                        placed = True
                        break
                    # If placed is still False after all passes, the slot remains
                    # free — this can only happen if the class has zero subjects
                    # configured (a data error caught by Step 2 validation).

    def _reduce_one_period_legacy(self):
        """Legacy single-reduction helper — use _auto_reduce_stuck_periods instead."""
        results = self._auto_reduce_stuck_periods()
        if not results:
            return None
        r = results[0]
        return (f"'{r['subject']}' in {r['class']} "
                f"(teacher: {r['teacher']}): "
                f"{r['from_periods']} → {r['to_periods']} periods/week")

    # ── Post-process: ensure ≥1 free slot in each half per teacher per day ─────

    def _balance_class_subject_distribution(self):
        """
        P1 uniform distribution: for each class, ensure no subject appears
        3+ times on a single day while 0 times on another day.

        Strategy: find (class, subject, day) triples where count >= 2 and there
        exists another day with count == 0, then try to move one occurrence to
        the zero-count day (same subject, free teacher, free class slot).

        Only moves filler/SC2 tasks; never touches HC1/HC2/SC1.
        """
        g      = self._gen
        grid   = g['grid']
        tasks  = g['tasks']
        wdays  = g['wdays']
        ppd    = g['ppd']

        task_idx_map = {t['idx']: t for t in tasks}

        def _is_moveable_cell(d, p, cn):
            idx = g['task_at'][cn][d][p]
            if idx is None:
                return False
            t = task_idx_map.get(idx)
            if t is None:
                return False
            if t['priority'] not in ('filler', 'P3') or t.get('is_ct'):
                return False
            # Never move a secondary CT fill out of the CT period slot
            cell_chk = g['grid'][cn][d][p]
            if cell_chk and cell_chk.get('ct_fill_secondary'):
                return False
            return True

        for cn in g['all_classes']:
            # Build subject day-count map
            for _pass in range(ppd * wdays):  # bounded
                subj_days = {}
                for d in range(wdays):
                    for p in range(ppd):
                        cell = grid[cn][d][p]
                        if cell:
                            sn = cell.get('subject', '')
                            subj_days.setdefault(sn, [0]*wdays)
                            subj_days[sn][d] += 1

                improved = False
                for sn, day_counts in subj_days.items():
                    max_count = max(day_counts)
                    min_count = min(day_counts)
                    if max_count - min_count < 2:
                        continue
                    heavy_d = day_counts.index(max_count)
                    light_d = day_counts.index(min_count)

                    # Find a moveable period of sn on heavy_d
                    for p_src in range(ppd):
                        cell_src = grid[cn][heavy_d][p_src]
                        if not cell_src or cell_src.get('subject') != sn:
                            continue
                        if not _is_moveable_cell(heavy_d, p_src, cn):
                            continue
                        idx_src = g['task_at'][cn][heavy_d][p_src]
                        task_src = task_idx_map.get(idx_src)
                        if task_src is None:
                            continue

                        # Find a free slot on light_d for this task
                        for p_dst in range(ppd):
                            if grid[cn][light_d][p_dst] is not None:
                                continue
                            if not g['t_free'](task_src['teacher'], light_d, p_dst):
                                continue
                            pt2 = (task_src.get('par_teach') or '').strip()
                            if pt2 and pt2 not in ('—', '?'):
                                if not g['t_free'](pt2, light_d, p_dst):
                                    continue
                            if g['t_unavail'](task_src['teacher'], light_d, p_dst):
                                continue
                            # Check all classes in cn_list are free
                            if not all(grid[c2][light_d][p_dst] is None
                                       for c2 in task_src['cn_list']):
                                continue
                            # Execute swap
                            self._gen_unplace(task_src, heavy_d, p_src)
                            self._gen_place(task_src, light_d, p_dst)
                            improved = True
                            break
                        if improved:
                            break
                    if improved:
                        break
                if not improved:
                    break

    def _ensure_half_free_periods(self):
        """
        For every teacher on every day, ensure they have at least 1 free period
        in both the first half (periods 1..half1) and the second half
        (periods half1+1..ppd).

        If a half is completely occupied, try to move the least-constrained
        (filler / SC2) period from that half to a free slot in the other half.
        Only moves are attempted that do NOT violate hard constraints
        (teacher double-booking, class double-booking, HC1/HC2 pinned slots).

        Additionally, tries to even out teaching load across days so no single
        day is completely full while another is almost empty.
        """
        g      = self._gen
        tasks  = g['tasks']
        grid   = g['grid']
        wdays  = g['wdays']
        ppd    = g['ppd']
        half1  = g['half1']
        DAYS   = g['DAYS']

        first_half  = list(range(0, half1))
        second_half = list(range(half1, ppd))

        # Build teacher set
        all_teachers = set()
        for t in tasks:
            if t['teacher']:
                all_teachers.add(t['teacher'])
            pt = (t.get('par_teach') or '').strip()
            if pt and pt not in ('—', '?'):
                all_teachers.add(pt)

        def _teacher_busy_at(teacher, d, p):
            """Check all classes to see if teacher is placed at (d,p)."""
            for cn in g['all_classes']:
                cell = grid[cn][d][p]
                if cell is None:
                    continue
                if (cell.get('teacher') == teacher or
                        (cell.get('par_teach') or '').strip() == teacher):
                    return True
            return False

        def _find_task_at(d, p):
            """Return the task object placed at (d,p) in any class, or None."""
            for cn in g['all_classes']:
                idx = g['task_at'][cn][d][p]
                if idx is not None:
                    return tasks[idx]
            return None

        def _is_moveable(task):
            """Only move filler / SC2 tasks (not HC1, HC2, SC1, or CT slot fills)."""
            if task['priority'] not in ('filler', 'P3'):
                return False
            if task.get('is_ct'):
                return False
            # Never move a task that was placed as a secondary CT slot fill
            # (it is holding down the CT period for that day)
            g2 = self._gen
            for cn2 in task['cn_list']:
                for d2 in range(g2['wdays']):
                    cd2 = self.class_config_data.get(cn2, {})
                    ct_p2 = int(cd2.get('teacher_period', 0)) - 1
                    if ct_p2 < 0:
                        continue
                    cell2 = g2['grid'][cn2][d2][ct_p2]
                    if (cell2 is not None and
                            cell2.get('ct_fill_secondary') and
                            g2['task_at'][cn2][d2][ct_p2] == task['idx']):
                        return False
            return True

        def _try_move(task, d_src, p_src, target_periods, teacher):
            """Try to move one occurrence of task from (d_src, p_src) to any
            (d_src, p_dst) where p_dst is in target_periods and the slot is free.
            Returns True if successful."""
            for p_dst in target_periods:
                if p_dst == p_src:
                    continue
                # Target cell must be empty for all task classes
                if not all(grid[cn][d_src][p_dst] is None
                           for cn in task['cn_list']):
                    continue
                # Teacher(s) must be free at target
                if not g['t_free'](task['teacher'], d_src, p_dst):
                    continue
                pt2 = (task.get('par_teach') or '').strip()
                if pt2 and pt2 not in ('—', '?') and not g['t_free'](pt2, d_src, p_dst):
                    continue
                # Execute move
                self._gen_unplace(task, d_src, p_src)
                self._gen_place(task, d_src, p_dst)
                return True
            return False

        for teacher in all_teachers:
            for d in range(wdays):
                busy_h1 = [p for p in first_half  if _teacher_busy_at(teacher, d, p)]
                busy_h2 = [p for p in second_half if _teacher_busy_at(teacher, d, p)]
                free_h1 = [p for p in first_half  if p not in busy_h1]
                free_h2 = [p for p in second_half if p not in busy_h2]

                # P1: teacher must have ≥1 free in each half.
                # Case A: first half fully occupied (0 free) — move one period out
                if not free_h1 and len(free_h2) >= 1:
                    for p_src in busy_h1:
                        task = _find_task_at(d, p_src)
                        if task and _is_moveable(task):
                            if _try_move(task, d, p_src, second_half, teacher):
                                break

                # Case B: second half fully occupied (0 free) — move one period out
                if not free_h2 and len(free_h1) >= 1:
                    for p_src in busy_h2:
                        task = _find_task_at(d, p_src)
                        if task and _is_moveable(task):
                            if _try_move(task, d, p_src, first_half, teacher):
                                break

        # ── P1 uniform teacher day-load: balance across days ──────────────
        # A teacher should not be completely packed on one day while another
        # day is almost empty. Move one filler/SC2 period from busiest day
        # to least busy day if the difference is > 1 and slot is available.
        for teacher in all_teachers:
            for _bal_pass in range(wdays * ppd):  # bounded iterations
                day_load = [
                    (sum(1 for p in range(ppd)
                         if _teacher_busy_at(teacher, d, p)), d)
                    for d in range(wdays)
                ]
                max_load, max_d = max(day_load, key=lambda x: x[0])
                min_load, min_d = min(day_load, key=lambda x: x[0])
                if max_load - min_load <= 1:
                    break   # balanced enough
                # Find a moveable task on max_d that can go to min_d
                moved = False
                for p_src in range(ppd):
                    if not _teacher_busy_at(teacher, max_d, p_src):
                        continue
                    task = _find_task_at(max_d, p_src)
                    if not (task and _is_moveable(task)):
                        continue
                    # Try moving to min_d at any free period
                    for p_dst in range(ppd):
                        if not all(grid[cn][min_d][p_dst] is None
                                   for cn in task['cn_list']):
                            continue
                        if not g['t_free'](task['teacher'], min_d, p_dst):
                            continue
                        pt2 = (task.get('par_teach') or '').strip()
                        if pt2 and pt2 not in ('—', '?') and not g['t_free'](pt2, min_d, p_dst):
                            continue
                        # Check unavailability (P4)
                        if g['t_unavail'](task['teacher'], min_d, p_dst):
                            continue
                        self._gen_unplace(task, max_d, p_src)
                        self._gen_place(task, min_d, p_dst)
                        moved = True
                        break
                    if moved:
                        break
                if not moved:
                    break

    # ── Task Analysis ────────────────────────────────────────────────────────
    def _calculate_group_slots(self, all_rows):
        """
        For every group in all_rows, determine how many slots that group needs.

        Strategy (per the spec):
          • Take the FIRST row of the group (any class in the group will do,
            as Step 3 guarantees all classes in a combined group share the
            same period count for their combined subject).
          • Look up class_config_data[cn]['subjects'] for that class.
          • Find the entry whose 'name' matches the row's subject.
          • Its 'periods' value is the number of slots required.
          • Parallel subjects share the same slot, so no extra count needed.

        Returns dict:
          { group_no: {'slots': int,  'ok': True} }          — success
          { group_no: {'slots': None, 'ok': False,
                       'reason': '<short reason>'} }          — failure
        """
        # Collect first row per group
        group_first = {}
        for row in all_rows:
            g = row['group']
            if g not in group_first:
                group_first[g] = row

        result = {}
        for g, row in group_first.items():
            cn   = row['class']
            subj = row['subject']

            # ── Guard: class config missing ───────────────────────────────
            if cn not in self.class_config_data:
                result[g] = {'slots': None, 'ok': False,
                             'reason': 'No config for {}'.format(cn)}
                continue

            cd_subjects = self.class_config_data[cn].get('subjects', [])
            if not cd_subjects:
                result[g] = {'slots': None, 'ok': False,
                             'reason': 'No subjects in {}'.format(cn)}
                continue

            # ── Search for the subject by name ────────────────────────────
            periods = None
            for s in cd_subjects:
                if s.get('name', '').strip() == subj:
                    periods = s.get('periods')
                    break

            # ── Fallback: subject may be the PARALLEL side of another entry
            if periods is None:
                for s in cd_subjects:
                    if (s.get('parallel')
                            and s.get('parallel_subject', '').strip() == subj):
                        # Parallel subjects share the same slot as the
                        # primary subject; use the primary's period count.
                        periods = s.get('periods')
                        break

            # ── Evaluate result ───────────────────────────────────────────
            if periods is None:
                result[g] = {'slots': None, 'ok': False,
                             'reason': '"{}" not found in {}'.format(subj, cn)}
            else:
                try:
                    result[g] = {'slots': int(periods), 'ok': True}
                except (ValueError, TypeError):
                    result[g] = {'slots': None, 'ok': False,
                                 'reason': 'Bad period value "{}"'.format(periods)}

        return result

    # ── Slot allocation engine ────────────────────────────────────────────────
    def _allocate_group_slots(self, all_rows, group_slots):
        """
        Allocate timetable slots to every group following RULE1 and RULE2.

        RULE1 – Slot priority order:
            Start from the last period of every day, work backwards.
            i.e. for period p from (ppd-1) down to 0, try all days at that p.

        RULE2 – Only fill what Stage 1 has not already covered:
            remaining = task['remaining']  (already decremented by Stage 1)

        Processing order:   C (consecutive)  →  B (standalone parallel)  →  A (combined)

        Returns
        -------
        dict:  { group_no → alloc_result }

        alloc_result (success):
            {'ok': True, 'total': int, 's1_placed': int,
             'new_placed': int, 'slots': [(d, p), ...]}

        alloc_result (failure / partial):
            {'ok': False, 'total': int, 's1_placed': int,
             'new_placed': int, 'slots': [(d, p), ...],   # partial placements
             'reason': str}
        """
        # ── Guard: Stage 1 must have been run ────────────────────────────────
        if not hasattr(self, '_gen'):
            dummy = {'ok': False, 'total': 0, 's1_placed': 0,
                     'new_placed': 0, 'slots': [],
                     'reason': 'Stage 1 not run yet — generate timetable first'}
            return {row['group']: dummy for row in all_rows}

        g      = self._gen
        grid   = g['grid']
        t_busy = g['t_busy']
        ppd    = g['ppd']
        wdays  = g['wdays']
        DAYS   = g['DAYS']

        # ── Helpers ──────────────────────────────────────────────────────────
        def slot_is_free_for_classes(cn_list, d, p):
            return all(grid.get(cn, [[]])[d][p] is None
                       for cn in cn_list if cn in grid)

        def teacher_free(t, d, p):
            if not t or t in ('—', '?', ''):
                return True
            return ((d, p) not in t_busy.get(t, set())
                    and not g['t_unavail'](t, d, p))

        def all_teachers_free(teachers, d, p):
            return all(teacher_free(t, d, p) for t in teachers)

        def mark_teachers_busy(teachers, d, p):
            for t in teachers:
                if t and t not in ('—', '?', ''):
                    t_busy.setdefault(t, set()).add((d, p))

        def place_slot(task, extra_par_teachers, d, p, class_info_map=None):
            """Place slot then apply per-class cell corrections.

            class_info_map: {cn -> {'type', 'par_subj', 'par_teach'}}
            Built entirely from Task Analysis rows so every class in the group
            gets the correct cell type and parallel-teacher regardless of what
            the engine task stored (which may be incomplete when a combined group
            has mixed parallel/non-parallel classes, e.g. Group 5: 12A no-par,
            12B has CS/Rajender).

            After _gen_place writes a shared cell to all cn in task['cn_list'],
            we overwrite each class's cell individually with correct data.
            """
            self._gen_place(task, d, p)

            # ── Fix each class's cell: type, par_subj, par_teach, teacher ───
            # Always apply — covers every class in combined/parallel groups.
            # primary_teacher fixes 7B/7C cells that still hold Anita (first class)
            # from the shared engine cell created by _gen_place.
            if class_info_map:
                for cn, info in class_info_map.items():
                    if cn in grid and grid[cn][d][p] is not None:
                        patch = {
                            'type':      info['type'],
                            'par_subj':  info['par_subj'],
                            'par_teach': info['par_teach'],
                        }
                        pt = info.get('primary_teacher', '').strip()
                        if pt and pt not in ('—', '?'):
                            patch['teacher'] = pt
                        grid[cn][d][p] = dict(grid[cn][d][p], **patch)

            # ── Mark ALL parallel teachers busy ──────────────────────────────
            # _gen_place marks task['teacher'] + task['par_teach'].
            # Also mark every par_teacher from the rows that is not yet marked.
            all_extra = set(extra_par_teachers)
            if class_info_map:
                for info in class_info_map.values():
                    pt = info.get('par_teach', '') or ''
                    if pt and pt not in ('—', '?', ''):
                        all_extra.add(pt)
            engine_par = (task.get('par_teach') or '').strip()
            for t in all_extra:
                if t and t not in ('—', '?', '') and t != engine_par:
                    t_busy.setdefault(t, set()).add((d, p))

        # ── Build task lookup — index by BOTH primary and parallel identity ─────
        #
        # A task in the engine is always created for the PRIMARY subject (e.g. SKT)
        # with par_subj/par_teach pointing to the parallel subject (e.g. Urdu/Irfan).
        # But a Task Analysis row (Section A/B) may identify a group by the PARALLEL
        # subject (Urdu/Irfan) because that is what is stored in step3_data combines.
        # We must find the task no matter which side the row uses as its identity.
        #
        # task_by_primary  — keyed by (frozenset(cn_list), subject,  teacher)
        # task_by_parallel — keyed by (frozenset(cn_list), par_subj, par_teach)

        task_by_primary  = {}
        task_by_parallel = {}
        for _t in g['tasks']:
            pk = (frozenset(_t['cn_list']), _t['subject'], _t['teacher'])
            task_by_primary[pk] = _t
            ps = (_t.get('par_subj') or '').strip()
            pt = (_t.get('par_teach') or '').strip()
            if ps and pt and ps not in ('—', '?'):
                sk = (frozenset(_t['cn_list']), ps, pt)
                task_by_parallel[sk] = _t

        task_lookup = task_by_primary   # alias used in fallback below

        # ── Organise rows by group ────────────────────────────────────────────
        group_rows    = {}
        group_section = {}
        for row in all_rows:
            gn = row['group']
            if gn not in group_rows:
                group_rows[gn]    = []
                group_section[gn] = row['section']
            group_rows[gn].append(row)

        # ── Process in order: C → B → A ──────────────────────────────────────
        result = {}

        for sec in ('C', 'B', 'A'):
            for gn, rows in sorted(group_rows.items()):
                if group_section[gn] != sec:
                    continue

                gs = group_slots.get(gn)
                if gs is None or not gs['ok']:
                    result[gn] = {
                        'ok': False, 'total': 0, 's1_placed': 0,
                        'new_placed': 0, 'slots': [],
                        'reason': (gs['reason'] if gs else 'Slot count unknown'),
                    }
                    continue

                total_periods  = gs['slots']
                first_row      = rows[0]
                primary_subj   = first_row['subject']
                primary_teach  = first_row['teacher']

                # All classes in this group (in display order, deduplicated)
                all_cn = list(dict.fromkeys(r['class'] for r in rows))

                # ── Find the matching task — 4-pass lookup ────────────────────
                #
                # Pass 1: exact match on (cn_list, subject, teacher)
                # Pass 2: exact match on (cn_list, par_subj, par_teach)  ← key fix
                # Pass 3: loose match on subject/teacher ignoring cn_list size
                # Pass 4: loose match on par_subj/par_teach ignoring cn_list size
                #
                # The row's primary_subj may be the PARALLEL side in the engine
                # (e.g. row says Urdu/Irfan but engine task is SKT/Anita with
                # par_subj=Urdu, par_teach=Irfan).  All passes are checked so we
                # always find the real task regardless of which side is "primary".

                cn_fs = frozenset(all_cn)
                task  = (task_by_primary.get((cn_fs, primary_subj, primary_teach))
                         or task_by_parallel.get((cn_fs, primary_subj, primary_teach)))

                if task is None:
                    # Pass 3 & 4: relax the cn_list requirement (subset match)
                    for t_obj in g['tasks']:
                        cn_overlap = bool(frozenset(t_obj['cn_list']) & cn_fs)
                        if not cn_overlap:
                            continue
                        via_primary  = (t_obj['subject']           == primary_subj
                                        and t_obj['teacher']       == primary_teach)
                        via_parallel = (t_obj.get('par_subj', '')  == primary_subj
                                        and t_obj.get('par_teach', '') == primary_teach)
                        if via_primary or via_parallel:
                            task = t_obj
                            break

                if task is None:
                    # Debug info: show a sample of task subjects to help diagnose
                    sample = [(t['subject'], t['teacher'], t.get('par_subj',''),
                               t.get('par_teach',''), list(t['cn_list']))
                              for t in g['tasks'][:8]]
                    result[gn] = {
                        'ok': False, 'total': total_periods,
                        's1_placed': 0, 'new_placed': 0, 'slots': [],
                        'reason': (
                            'Task not found — "{}"/{} not in engine '
                            '(checked primary + parallel sides). '
                            'Verify subject name matches Step 2 exactly.'.format(
                                primary_subj, primary_teach)),
                    }
                    continue

                # ── Once found, identify which side the row matched ───────────────
                # "found_via_parallel" = the row's primary is the engine's par_subj.
                # This is important for building the all_teachers_needed list correctly:
                # both the task's teacher AND par_teach must be checked as busy.
                found_via_parallel = (
                    task.get('par_subj', '').strip() == primary_subj
                    and task.get('par_teach', '').strip() == primary_teach
                )

                # Stage-1-placed count
                s1_placed = task['periods'] - task['remaining']
                remaining = task['remaining']

                if remaining <= 0:
                    result[gn] = {
                        'ok': True, 'total': total_periods,
                        's1_placed': s1_placed, 'new_placed': 0,
                        'slots': [],
                    }
                    continue

                # ── Collect ALL teachers that must be free at the chosen slot ──
                #
                # We collect teachers from two sources and union them:
                #   Source 1 — the engine task itself: task['teacher'] + task['par_teach']
                #   Source 2 — the Task Analysis rows: each row's teacher + par_teacher
                #              (combined groups have a different par_teacher per class)
                #
                # This ensures that even when the task was found via its parallel side
                # (found_via_parallel=True), we still check both Anita (SKT) and Irfan
                # (Urdu) for availability before claiming the slot.

                task_teachers = []
                if task['teacher']:
                    task_teachers.append(task['teacher'])
                if task.get('par_teach', '') and task['par_teach'] not in ('—','?',''):
                    task_teachers.append(task['par_teach'])

                row_teachers = []
                for row in rows:
                    for fld in ('teacher', 'par_teacher'):
                        t = row.get(fld, '')
                        if t and t not in ('—', '?', ''):
                            row_teachers.append(t)

                all_teachers_needed = list(
                    dict.fromkeys(task_teachers + row_teachers))

                # Extra par teachers: those in all_teachers_needed that the task's
                # _gen_place does NOT already mark busy (it only marks teacher + par_teach)
                engine_marks = set(filter(None, [task['teacher'],
                                                  task.get('par_teach', '')]))
                extra_par = [t for t in all_teachers_needed
                             if t not in engine_marks]

                # ── Per-class info map: type + par_subj + par_teach ──────────
                #
                # Builds {cn → {type, par_subj, par_teach}} entirely from the
                # Task Analysis rows.  This is the AUTHORITATIVE source — the
                # engine task may be incomplete (e.g. Group 5: engine task was
                # built from 12A which has no parallel, so par_teach='' even
                # though 12B has CS/Rajender).  We apply this map to EVERY class
                # unconditionally so each class gets exactly the right cell.
                #
                # Rules for cell type:
                #   Section A + has parallel  → 'combined_parallel'
                #   Section A + no parallel   → 'combined'
                #   Section B                 → 'parallel'
                #   Section C                 → 'normal'
                # ── Per-class info map: type + par_subj + par_teach + primary_teacher ─
                #
                # Section A (combined) rows always show the COMBINE perspective:
                #   row['subject']     = combine subject  (Urdu/Irfan)
                #   row['teacher']     = combine teacher  (Irfan) — same for every row
                #   row['par_subj']    = class primary subject (SKT)
                #   row['par_teacher'] = class primary teacher (Anita/Neha/Mamta)
                #
                # When found_via_parallel=True (Urdu row → engine SKT task):
                #   cell.teacher    = engine's task['teacher'] = Anita (only 7A correct!)
                #   cell.par_teach  = Irfan ✓
                #   cell.subject    = SKT ✓
                # → primary_teacher must come from row['par_teacher'] (Neha for 7B etc.)
                #
                # When found_via_parallel=False (engine task subject matches row):
                #   row['teacher'] maps directly to cell.teacher
                class_info_map = {}
                for row in rows:
                    ps = (row.get('par_subj')    or '').strip()
                    pt = (row.get('par_teacher') or '').strip()
                    has_par = bool(ps and pt and ps not in ('—', '?') and pt not in ('—', '?'))
                    if sec == 'A':
                        cell_type = 'combined_parallel' if has_par else 'combined'
                    elif sec == 'B':
                        cell_type = 'parallel'
                    else:
                        cell_type = 'normal'
                    # Determine the per-class PRIMARY teacher for the cell
                    if found_via_parallel:
                        # row['par_teacher'] = SKT teacher of this class (Anita/Neha/Mamta)
                        primary_teacher = (row.get('par_teacher') or '').strip()
                    else:
                        # row['teacher'] is already the cell's primary teacher
                        primary_teacher = (row.get('teacher') or '').strip()
                    class_info_map[row['class']] = {
                        'type':           cell_type,
                        'par_subj':       ps if has_par else '',
                        'par_teach':      pt if has_par else '',
                        'primary_teacher': primary_teacher,
                    }

                # ── Placement logic ───────────────────────────────────────────
                placed_slots  = []   # (d, p) pairs successfully placed
                last_fail_why = ''

                # Check if this group's consecutive constraint has been relaxed
                _relaxed = getattr(self, '_relaxed_consec_keys', set())
                _group_relaxed = (sec == 'C' and rows and
                                  (rows[0]['class'], rows[0]['subject']) in _relaxed)

                if sec == 'C' and not _group_relaxed:
                    # ── Consecutive: find adjacent pairs (p_start, p_start+1) ──
                    # RULE1 for pairs: start from (ppd-2, ppd-1) and go backwards.
                    for p_start in range(ppd - 2, -1, -1):
                        if len(placed_slots) >= remaining:
                            break
                        p1, p2 = p_start, p_start + 1
                        for d in range(wdays):
                            if len(placed_slots) >= remaining:
                                break
                            # Need both slots free for all classes + all teachers
                            cls_ok = (slot_is_free_for_classes(all_cn, d, p1)
                                      and slot_is_free_for_classes(all_cn, d, p2))
                            tch_ok = (all_teachers_free(all_teachers_needed, d, p1)
                                      and all_teachers_free(all_teachers_needed, d, p2))
                            if cls_ok and tch_ok:
                                # Place both (or only p1 if remaining == 1)
                                if remaining - len(placed_slots) >= 2:
                                    place_slot(task, extra_par, d, p1, class_info_map)
                                    place_slot(task, extra_par, d, p2, class_info_map)
                                    placed_slots.extend([(d, p1), (d, p2)])
                                else:
                                    place_slot(task, extra_par, d, p1, class_info_map)
                                    placed_slots.append((d, p1))
                            else:
                                if not cls_ok:
                                    busy_cn = [cn for cn in all_cn
                                               if cn in grid and (
                                                   grid[cn][d][p1] is not None
                                                   or grid[cn][d][p2] is not None)]
                                    last_fail_why = (
                                        '{} P{}-P{}: class {} occupied'.format(
                                            DAYS[d], p1+1, p2+1,
                                            ', '.join(busy_cn)))
                                else:
                                    busy_t = [t for t in all_teachers_needed
                                              if not teacher_free(t, d, p1)
                                              or not teacher_free(t, d, p2)]
                                    last_fail_why = (
                                        '{} P{}-P{}: teacher {} busy'.format(
                                            DAYS[d], p1+1, p2+1,
                                            ', '.join(busy_t)))

                else:
                    # ── Section B / A (or relaxed-C): period-major + day preference ─
                    # Algorithm:
                    #   Pass 1 (Distribute): period-major but SKIP days where this
                    #     subject already appears ceil(n/wdays) times.  This spreads
                    #     periods across as many days as possible.
                    #   Pass 2 (Fallback):   original period-major with no cap.
                    #     Activates only for any periods left unplaced after Pass 1
                    #     (teacher availability forces all remaining onto fewer days).
                    #
                    # Effect: good distribution when teacher has spread availability;
                    # graceful degradation (minimal stacking) when teacher-constrained.

                    _nat = max(1, min(2, (total_periods + wdays - 1) // wdays))

                    def _subj_on_day(d_):
                        return max(
                            (sum(1 for pp in range(ppd)
                                 if grid.get(cn_, [[]])[d_][pp] is not None
                                 and grid[cn_][d_][pp].get('subject') == primary_subj)
                             for cn_ in all_cn if cn_ in grid),
                            default=0)

                    # Pass 1: distribute — P2 Pref_Per/Pref_Day first, then Dec
                    # Rule 2 (Dec): for parallel/combined follow last→first period.
                    _p_pref_task = task.get('p_pref', [])
                    _d_pref_task = task.get('d_pref', [])
                    if _p_pref_task:
                        _pref_idxs = [x-1 for x in _p_pref_task]
                        _dec_rest  = [p for p in range(ppd-1, -1, -1) if p not in _pref_idxs]
                        _p_order   = _pref_idxs + _dec_rest
                    else:
                        _p_order   = list(range(ppd - 1, -1, -1))
                    for p in _p_order:
                        if len(placed_slots) >= remaining:
                            break
                        # Sort days: P2 Pref_Day first, then fewest subject occurrences
                        if _d_pref_task:
                            _pref_d  = [i for i, dn in enumerate(DAYS) if dn in _d_pref_task]
                            _rest_d  = sorted([i for i in range(wdays) if i not in _pref_d],
                                              key=lambda d_: _subj_on_day(d_))
                            day_order_p1 = _pref_d + _rest_d
                        else:
                            day_order_p1 = sorted(range(wdays),
                                                  key=lambda d_: _subj_on_day(d_))
                        for d in day_order_p1:
                            if len(placed_slots) >= remaining:
                                break
                            if _subj_on_day(d) >= _nat:
                                continue   # skip: already at cap for this day
                            cls_ok = slot_is_free_for_classes(all_cn, d, p)
                            tch_ok = all_teachers_free(all_teachers_needed, d, p)
                            if cls_ok and tch_ok:
                                place_slot(task, extra_par, d, p, class_info_map)
                                placed_slots.append((d, p))
                            else:
                                if not cls_ok:
                                    busy_cn = [cn for cn in all_cn
                                               if cn in grid
                                               and grid[cn][d][p] is not None]
                                    occupant = (grid[busy_cn[0]][d][p]
                                                if busy_cn else {})
                                    last_fail_why = (
                                        '{} P{}: {} occupied by "{}"'.format(
                                            DAYS[d], p+1,
                                            ', '.join(busy_cn),
                                            occupant.get('subject', '?')))
                                else:
                                    busy_t = [t for t in all_teachers_needed
                                              if not teacher_free(t, d, p)]
                                    last_fail_why = (
                                        '{} P{}: teacher {} busy'.format(
                                            DAYS[d], p+1,
                                            ', '.join(busy_t)))

                    # Pass 2 (fallback): period-major, no daily cap
                    if len(placed_slots) < remaining:
                        for p in range(ppd - 1, -1, -1):
                            if len(placed_slots) >= remaining:
                                break
                            for d in range(wdays):
                                if len(placed_slots) >= remaining:
                                    break
                                cls_ok = slot_is_free_for_classes(all_cn, d, p)
                                tch_ok = all_teachers_free(all_teachers_needed, d, p)
                                if cls_ok and tch_ok:
                                    place_slot(task, extra_par, d, p, class_info_map)
                                    placed_slots.append((d, p))
                                else:
                                    if not cls_ok:
                                        busy_cn = [cn for cn in all_cn
                                                   if cn in grid
                                                   and grid[cn][d][p] is not None]
                                        occupant = (grid[busy_cn[0]][d][p]
                                                    if busy_cn else {})
                                        last_fail_why = (
                                            '{} P{}: {} occupied by "{}"'.format(
                                                DAYS[d], p+1,
                                                ', '.join(busy_cn),
                                                occupant.get('subject', '?')))
                                    else:
                                        busy_t = [t for t in all_teachers_needed
                                                  if not teacher_free(t, d, p)]
                                        last_fail_why = (
                                            '{} P{}: teacher {} busy'.format(
                                                DAYS[d], p+1,
                                                ', '.join(busy_t)))

                # ── Build result ──────────────────────────────────────────────
                new_placed = len(placed_slots)
                if new_placed >= remaining:
                    result[gn] = {
                        'ok': True,
                        'total': total_periods,
                        's1_placed': s1_placed,
                        'new_placed': new_placed,
                        'slots': placed_slots,
                    }
                else:
                    still_short = remaining - new_placed
                    if new_placed == 0:
                        reason = ('No free slots found. '
                                  + (last_fail_why or 'All slots occupied'))
                    else:
                        reason = ('{} slot(s) still unplaced. '
                                  'Last conflict: {}'.format(
                                      still_short,
                                      last_fail_why or 'Unknown'))
                    result[gn] = {
                        'ok': False,
                        'total': total_periods,
                        's1_placed': s1_placed,
                        'new_placed': new_placed,
                        'slots': placed_slots,
                        'reason': reason,
                    }

        return result

    # ─────────────────────────────────────────────────────────────────────────
    def _run_stage2(self):
        """Legacy Stage 2 entry — now redirects to Stage 3 (filler phases)."""
        self._run_stage3()

    # =========================================================================
    #  CORE GENERATION ENGINE  (split into init + stage1 + stage2)
    # =========================================================================

    def _init_gen_state(self):
        """
        Build the grid, task list and all helper closures.
        Called once before Stage 1. Results stored on self._gen.
        """
        cfg   = self.configuration
        ppd   = cfg['periods_per_day']
        wdays = cfg['working_days']
        half1 = cfg['periods_first_half']
        DAYS  = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'][:wdays]

        all_classes = []
        for cls in range(6, 13):
            for si in range(cfg['classes'][cls]):
                all_classes.append("{}{}".format(cls, chr(65 + si)))

        grid    = {cn: [[None]*ppd for _ in range(wdays)] for cn in all_classes}
        task_at = {cn: [[None]*ppd for _ in range(wdays)] for cn in all_classes}
        t_busy  = {}

        def t_free(t, d, p):
            return not t or (d, p) not in t_busy.get(t, set())
        def t_mark(t, d, p):
            if t: t_busy.setdefault(t, set()).add((d, p))
        def t_unmark(t, d, p):
            if t: t_busy.get(t, set()).discard((d, p))

        unavail = getattr(self, 'step3_unavailability', {})
        def t_unavail(t, d, p):
            u = unavail.get(t, {})
            if not u: return False
            return DAYS[d] in u.get('days', []) and (p+1) in u.get('periods', [])

        # ── combine lookup ───────────────────────────────────────────────────────
        # Pass 1: map (cn, combine_subject) → combined_class_list
        #   e.g. Step-3 combine for Irfan/Urdu across 7A+7B+7C
        #   gives: ('7A','Urdu')→['7A','7B','7C'], etc.
        s3 = getattr(self, 'step3_data', {})
        cn_subj_combined = {}
        for _teacher, s3d in s3.items():
            for cb in s3d.get('combines', []):
                classes  = sorted(cb.get('classes', []))
                subjects = cb.get('subjects', [])
                if len(classes) >= 2 and subjects:
                    for cn in classes:
                        cn_subj_combined[(cn, subjects[0])] = classes

        # Pass 2: if a primary subject's par_subj is a combine subject, also
        # map that primary subject to the same combined class list.
        #
        # Example: Step-3 combine subject = 'Urdu' → ('7A','Urdu')=['7A','7B','7C']
        # Class config 7A: primary='SKT', par_subj='Urdu'.
        # After pass 2: ('7A','SKT')=['7A','7B','7C'] is also added.
        # This ensures that when the engine iterates 7A's 'SKT' subject it
        # correctly receives cn_list=['7A','7B','7C'] instead of just ['7A'].
        for cn in all_classes:
            cd_subjects = self.class_config_data.get(cn, {}).get('subjects', [])
            for s in cd_subjects:
                primary_subj = s.get('name', '').strip()
                par_subj     = (s.get('parallel_subject') or '').strip()
                # If the parallel-subject of this entry is already a combine key,
                # map this primary subject to the same combined class list.
                if par_subj and (cn, par_subj) in cn_subj_combined:
                    if (cn, primary_subj) not in cn_subj_combined:
                        cn_subj_combined[(cn, primary_subj)] = cn_subj_combined[(cn, par_subj)]
                # Also handle the reverse: if this primary IS the combine subject
                # and it has a par_subj, ensure par_subj is also combined.
                if (cn, primary_subj) in cn_subj_combined and par_subj:
                    if (cn, par_subj) not in cn_subj_combined:
                        cn_subj_combined[(cn, par_subj)] = cn_subj_combined[(cn, primary_subj)]

        # build tasks
        tasks = []
        # Key: (frozenset(cn_list), subject_name)
        # Deduplicates combined groups regardless of which class is processed first
        # or which teacher is recorded — the combine is one task for the whole group.
        seen_combined = set()
        for cn in all_classes:
            if cn not in self.class_config_data:
                continue
            cd     = self.class_config_data[cn]
            ct     = cd.get('teacher', '').strip()
            ct_per = cd.get('teacher_period', 1)

            # Only ONE subject per class is the "CT subject" — the subject
            # whose teacher matches the class teacher AND has the MAXIMUM S_slots.
            # P1a: If CT teaches multiple subjects, use the one with the most
            # periods for the CT fixed period.  Others are scheduled freely.
            ct_subject_name = None
            _ct_max_per = 0
            for _s in cd['subjects']:
                if _s.get('teacher', '').strip() == ct:
                    if _s.get('periods', 0) > _ct_max_per:
                        _ct_max_per     = _s['periods']
                        ct_subject_name = _s['name']

            ct_subject_assigned = False

            for s in cd['subjects']:
                subj = s['name']
                t    = s['teacher'].strip()
                n    = s['periods']

                cn_list = cn_subj_combined.get((cn, subj), [cn])
                if len(cn_list) > 1:
                    # Use frozenset + subject as dedup key (teacher intentionally
                    # excluded — different classes have different par_teachers but
                    # represent the same combined group)
                    key = (frozenset(cn_list), subj)
                    if key in seen_combined:
                        # BUG FIX: if the skipped subject would have been the CT
                        # subject for this class (first subject by CT teacher),
                        # mark ct_subject_assigned=True so subsequent subjects by
                        # the same CT teacher are NOT erroneously promoted to HC1.
                        # Without this, combining a class's CT subject causes the
                        # next CT-teacher subject to claim period 2, colliding with
                        # the combined task already occupying that cell.
                        if t == ct and not ct_subject_assigned:
                            ct_subject_assigned = True
                        continue
                    seen_combined.add(key)

                # is_ct: True only for the CT teacher's subject with MAXIMUM periods
                # (P1a). Additional subjects by the same CT teacher are regular tasks.
                if t == ct and subj == ct_subject_name and not ct_subject_assigned:
                    is_ct = True
                    ct_subject_assigned = True
                else:
                    is_ct = False

                par    = bool(s.get('parallel', False))
                pt     = s.get('parallel_teacher', '').strip() if par else ''
                ps     = s.get('parallel_subject', '').strip() if par else ''
                consec = (s.get('consecutive', 'No') == 'Yes')
                # If the user has relaxed this group's consecutive constraint,
                # override so the engine also treats it as non-consecutive
                if consec and (cn, subj) in getattr(self, '_relaxed_consec_keys', set()):
                    consec = False
                p_pref = list(s.get('periods_pref', []))
                d_pref = list(s.get('days_pref', []))

                if len(cn_list) > 1 and par:
                    ttype = 'combined_parallel'
                elif len(cn_list) > 1:
                    ttype = 'combined'
                elif par:
                    ttype = 'parallel'
                else:
                    ttype = 'normal'

                if is_ct:
                    priority = 'P1_CT'
                elif p_pref or d_pref:
                    priority = 'P2'
                elif consec:
                    priority = 'P5'
                elif n >= wdays:
                    priority = 'P3'
                elif len(cn_list) > 1 and n >= max(2, wdays - len(cn_list)):
                    # Combined tasks need ALL classes free simultaneously.
                    # Treat as P3 (same weight as Main_sub) so Task Analysis
                    # places them before fillers can claim all the slots.
                    priority = 'P3'
                else:
                    priority = 'filler'

                tasks.append({
                    'idx':       len(tasks),
                    'cn_list':   cn_list,
                    'subject':   subj,
                    'teacher':   t,
                    'par_subj':  ps,
                    'par_teach': pt,
                    'periods':   n,
                    'remaining': n,
                    'is_ct':     is_ct,
                    'ct_period': ct_per if is_ct else None,
                    'p_pref':    p_pref,
                    'd_pref':    d_pref,
                    'consec':    consec,
                    'daily':     (n >= wdays),
                    'priority':  priority,
                    'type':      ttype,
                    'rx_sc1':    False,
                    'rx_sc3':    False,
                    'rx_sc2':    False,
                })

        total_atoms = sum(t['periods'] for t in tasks)

        # Build CT period protection maps:
        # ct_period_map: cn → 0-based CT period index (-1 if no CT)
        # hc1_task_idx: cn → task idx of the HC1 task for that class
        ct_period_map = {}
        hc1_task_idx  = {}
        for cn in all_classes:
            cd2 = self.class_config_data.get(cn, {})
            ct_per_v = cd2.get('teacher_period', None)
            if ct_per_v:
                ct_period_map[cn] = int(ct_per_v) - 1
            else:
                ct_period_map[cn] = -1
        for t in tasks:
            if t.get('is_ct'):
                for cn2 in t['cn_list']:
                    hc1_task_idx[cn2] = t['idx']

        # Store everything on self so stages can share state
        self._gen = {
            'cfg': cfg, 'ppd': ppd, 'wdays': wdays, 'half1': half1,
            'DAYS': DAYS, 'all_classes': all_classes,
            'grid': grid, 'task_at': task_at, 't_busy': t_busy,
            'tasks': tasks, 'total_atoms': total_atoms,
            't_free': t_free, 't_mark': t_mark, 't_unmark': t_unmark,
            't_unavail': t_unavail,
            'ct_period_map': ct_period_map,
            'hc1_task_idx':  hc1_task_idx,
        }

    # ── Shared helpers (use self._gen) ────────────────────────────────────────

    def _gen_can_place(self, task, d, p,
                       ignore_sc1=False, ignore_sc3=False, ignore_sc2=False):
        g      = self._gen
        DAYS   = g['DAYS']; ppd = g['ppd']
        grid   = g['grid']
        t_free = g['t_free']; t_unavail = g['t_unavail']
        t      = task['teacher']; pt = task['par_teach']
        p1     = p + 1  # 1-based

        if task['is_ct'] and p1 != task['ct_period']:   # P1_CT: fixed CT period
            return False

        # CT PERIOD PROTECTION (P1a): no non-HC1 task may occupy the
        # CT period slot of any class in this task's cn_list.
        # This is the primary guard that prevents SC1/SC2/fillers from
        # ever taking the slot — repair and Phase 1b own it exclusively.
        if not task.get('is_ct'):
            ct_map = g.get('ct_period_map', {})
            hc1_idx_map = g.get('hc1_task_idx', {})
            for _cn in task['cn_list']:
                ct_p_0 = ct_map.get(_cn, -1)
                if ct_p_0 < 0:
                    continue
                if p != ct_p_0:
                    continue  # not the CT period — allow
                # This slot IS the CT period for _cn.
                # Allow ONLY if this task is the HC1 task for _cn.
                hc1_idx = hc1_idx_map.get(_cn, -1)
                if hc1_idx < 0 or task['idx'] != hc1_idx:
                    return False  # block all non-HC1 tasks at CT period

        if task['p_pref'] and not task['is_ct']:  # P2: respect Pref_Per
            if p1 not in task['p_pref']:
                return False
        if task['d_pref']:
            if DAYS[d] not in task['d_pref']:
                return False
        for cn in task['cn_list']:
            if grid[cn][d][p] is not None:
                return False
        if not t_free(t, d, p): return False
        if pt and not t_free(pt, d, p): return False
        if not (ignore_sc3 or task['rx_sc3']):
            if t_unavail(t, d, p): return False
            if pt and t_unavail(pt, d, p): return False
        # FIX BUG 5: consecutive tasks may go at ANY adjacent pair (p, p+1),
        # not locked to only the last two slots.  The partner slot p+1 must
        # also be empty and teacher-free.
        if task['consec'] and not (ignore_sc1 or task['rx_sc1']):
            if p >= ppd - 1:
                return False   # no room for the partner slot
            for cn in task['cn_list']:
                if grid[cn][d][p + 1] is not None:
                    return False
            if not t_free(t, d, p + 1): return False
            if pt and not t_free(pt, d, p + 1): return False
            if not (ignore_sc3 or task['rx_sc3']):
                if t_unavail(t, d, p + 1): return False
                if pt and t_unavail(pt, d, p + 1): return False
            # A consecutive pair places 2 periods of the same subject on this
            # day.  Block it if the subject already appears here (would be 3+).
            for cn in task['cn_list']:
                already_today = sum(
                    1 for pp in range(ppd)
                    if grid[cn][d][pp] is not None
                    and grid[cn][d][pp].get('subject') == task['subject']
                )
                if already_today > 0:
                    return False
        if not task['consec']:
            # Hard cap: no subject may appear more than 2 periods on the same day.
            wdays_g = g['wdays']
            n_total = task['periods']
            natural_max = (n_total + wdays_g - 1) // wdays_g
            max_per_day = min(2, natural_max)
            for cn in task['cn_list']:
                count_today = sum(
                    1 for pp in range(ppd)
                    if (existing := grid[cn][d][pp]) is not None
                    and existing.get('subject') == task['subject']
                )
                if count_today >= max_per_day:
                    return False
        else:
            # Consecutive task with ignore_sc1=True (placing second slot of a pair):
            # cap is 2 per day (the pair itself), not natural_max which could be 1.
            for cn in task['cn_list']:
                count_today = sum(
                    1 for pp in range(ppd)
                    if (existing := grid[cn][d][pp]) is not None
                    and existing.get('subject') == task['subject']
                )
                if count_today >= 2:   # never more than one pair per day
                    return False

            # P3: Main_sub (SC2/daily) must use the SAME period on all days.
            # If this task is already placed on some other day, find which period
            # it used and enforce that p must match.
            if task.get('daily') and task['priority'] == 'P3':
                committed_p = None
                for d2 in range(wdays_g):
                    if d2 == d:
                        continue
                    for cn2 in task['cn_list']:
                        idx2 = g['task_at'][cn2][d2][p] if p < ppd else None
                        if idx2 == task['idx']:
                            committed_p = p  # already at p on another day — ok
                            break
                        # Check all periods for a placed occurrence on d2
                        for pp2 in range(ppd):
                            if g['task_at'][cn2][d2][pp2] == task['idx']:
                                committed_p = pp2
                                break
                        if committed_p is not None:
                            break
                    if committed_p is not None:
                        break
                if committed_p is not None and committed_p != p:
                    return False  # must stay at the committed period

        return True

    def _gen_count_valid_slots(self, task,
                               ignore_sc1=False, ignore_sc3=False,
                               ignore_sc2=False):
        """Count valid (d,p) placements for *task* right now (MRV helper)."""
        g = self._gen
        return sum(
            1
            for d in range(g['wdays'])
            for p in range(g['ppd'])
            if self._gen_can_place(task, d, p, ignore_sc1, ignore_sc3, ignore_sc2)
        )

    def _gen_make_cell(self, task):
        return {
            'type':      task['type'],
            'subject':   task['subject'],
            'teacher':   task['teacher'],
            'par_subj':  task['par_subj'],
            'par_teach': task['par_teach'],
            'combined_classes': task['cn_list'] if len(task['cn_list'])>1 else [],
            'is_ct':     task['is_ct'],
        }

    def _gen_place(self, task, d, p):
        g = self._gen
        # Quota guard: never place more than the task's period budget.
        # remaining > 0 means budget is available; remaining == 0 means full.
        if task['remaining'] <= 0:
            return   # already at quota — silently skip
        cell = self._gen_make_cell(task)
        for cn in task['cn_list']:
            g['grid'][cn][d][p]    = cell
            g['task_at'][cn][d][p] = task['idx']
        g['t_mark'](task['teacher'], d, p)
        if task['par_teach']:
            g['t_mark'](task['par_teach'], d, p)
        task['remaining'] -= 1

    def _gen_unplace(self, task, d, p):
        g = self._gen
        for cn in task['cn_list']:
            g['grid'][cn][d][p]    = None
            g['task_at'][cn][d][p] = None
        g['t_unmark'](task['teacher'], d, p)
        if task['par_teach']:
            g['t_unmark'](task['par_teach'], d, p)
        task['remaining'] += 1

    def _gen_prog(self, msg, extra_pct=0):
        g = self._gen
        done = g['total_atoms'] - sum(t['remaining'] for t in g['tasks'])
        pct  = min(97, int(100 * done / max(g['total_atoms'], 1))) + extra_pct
        self._progress_log.append((msg, min(97, pct)))

    def _gen_snapshot_tt(self):
        """Return a tt-dict from current gen state (for display)."""
        g = self._gen
        unplaced = sum(t['remaining'] for t in g['tasks'])
        return {
            'grid':        g['grid'],
            'days':        g['DAYS'],
            'ppd':         g['ppd'],
            'half1':       g['half1'],
            'all_classes': g['all_classes'],
            'tasks':       g['tasks'],
            'unplaced':    unplaced,
        }

    # ── STAGE 1: HC1 (CT fixed periods) + HC2 (preference-constrained) ────────

    def _run_stage1_phases(self):
        """
        Stage 1 — place ALL CT (HC1) and ALL fixed/preference (HC2) periods.

        Logic is intentionally dead-simple:
          • For HC1: the CT period is fixed (same period-index, every working day).
            Just write into the grid cell if it is None.  No teacher conflict check
            is needed — Step 2 already guarantees these slots are conflict-free.
          • For HC2: iterate through the preferred (day, period) combinations in
            the order they were specified and fill the required number of slots.
            Again only check that the cell is still empty.

        Any period that STILL could not be placed (cell was already occupied) is
        reported as an issue with an exact explanation of what was blocking it.
        """
        g     = self._gen
        tasks = g['tasks']
        grid  = g['grid']
        wdays = g['wdays']
        ppd   = g['ppd']
        DAYS  = g['DAYS']

        s1_issues = []   # list of human-readable problem strings (HC1/HC2 only)

        # ══════════════════════════════════════════════════════════════════════
        # PHASE 1 — HC1: Class-teacher subject periods
        #
        # The CT teacher's subject S IS the CT period — they are the same thing.
        # Subject S has n periods configured in Step 2.  Those n periods are placed
        # at ct_per (same period-index) across n different working days.
        # No extra "CT admin" slots are added — the subject count n is the total.
        #
        # PARALLEL TEACHER HANDLING:
        # Some CT subjects are parallel (e.g. SKT+URDU where IRFAN teaches URDU to
        # the same class at the same slot).  When two classes share the same CT
        # period AND the same parallel teacher, the parallel teacher can only attend
        # one class at a time.  We check t_free for the par_teach before placing and
        # fall back to the NEXT available period for that day if the par_teach is busy.
        # This avoids the double-booking that _remove_teacher_conflicts would later
        # undo, leaving the HC1 task with remaining > 0.
        # ══════════════════════════════════════════════════════════════════════
        self._gen_prog("Stage 1 · Phase 1 (P1_CT) — Placing Class Teacher fixed periods…")

        # ── Helper: evict a lower-priority task from (d, p) to free the slot ────
        # Used when the parallel teacher of a CT subject is blocked at ct_period
        # by a non-HC1 task that was placed earlier.  We unplace that task so the
        # CT constraint can be honoured; Stage 2 will re-place the evicted task.
        def _try_evict_for_parallel(par_teacher, d, p):
            """
            If par_teacher is busy at (d,p) due to a non-HC1 task, unplace that
            task so the CT subject can be placed at the fixed ct_period.
            Returns True if the slot was freed, False otherwise.
            """
            if g['t_free'](par_teacher, d, p):
                return True   # already free
            # Find which class has this teacher at (d,p)
            for cn2 in g['all_classes']:
                idx2 = g['task_at'][cn2][d][p]
                if idx2 is None:
                    continue
                blocker = tasks[idx2]
                # Only evict non-HC1 tasks (never move CT periods)
                if blocker['priority'] == 'P1_CT':
                    return False
                if (blocker.get('teacher') == par_teacher or
                        (blocker.get('par_teach') or '').strip() == par_teacher):
                    self._gen_unplace(blocker, d, p)
                    return True
            return False

        # Process HC1 tasks: parallel-CT tasks FIRST (most constrained), then
        # non-parallel CT tasks.  This ensures the parallel teacher is committed
        # to the correct slot before non-parallel CT tasks are placed.
        hc1_tasks = [t for t in tasks if t['priority'] == 'P1_CT']
        hc1_tasks.sort(key=lambda t: (
            0 if (t.get('par_teach') or '').strip() not in ('', '—', '?') else 1
        ))

        for task in hc1_tasks:
            p_idx = task['ct_period'] - 1          # 0-based period index (fixed)
            pt    = (task.get('par_teach') or '').strip()
            has_par = bool(pt and pt not in ('—', '?'))

            # HC1 places at most wdays occurrences (one per working day) at p_idx.
            # Any overflow (periods > wdays) is demoted to filler AFTER this loop.
            days_to_place = min(task['remaining'], wdays)

            for d in range(wdays):
                if task['remaining'] <= 0:
                    break

                # ── Step 1: try to place at the FIXED CT period (p_idx) ───────
                # For parallel-CT subjects: if the parallel teacher is busy at
                # p_idx due to a lower-priority task, evict that task first so
                # the CT constraint is never broken.
                cells_free_at_pidx = all(
                    grid[cn][d][p_idx] is None for cn in task['cn_list'])
                primary_free = g['t_free'](task['teacher'], d, p_idx)

                if cells_free_at_pidx and primary_free:
                    par_ok = True
                    if has_par and not g['t_free'](pt, d, p_idx):
                        # Parallel teacher blocked — try to evict the blocker
                        par_ok = _try_evict_for_parallel(pt, d, p_idx)

                    if par_ok:
                        self._gen_place(task, d, p_idx)
                        continue   # placed at fixed CT period ✓

                    # ── Parallel teacher still blocked but CT is MANDATORY ────
                    # The primary CT teacher is free and the cell is empty, but
                    # the parallel teacher (e.g. IRFAN/URDU) cannot be freed.
                    # Requirement: CT period must ALWAYS be placed.  Place the
                    # primary teacher's subject SOLO (without the parallel teacher)
                    # so the class-teacher duty is never missed.  The parallel
                    # teacher's subject will be handled separately by Stage 2/3.
                    if has_par:
                        _sv_pt   = task['par_teach']
                        _sv_ps   = task['par_subj']
                        _sv_type = task['type']
                        task['par_teach'] = ''
                        task['par_subj']  = ''
                        task['type']      = (
                            'combined' if len(task['cn_list']) > 1 else 'normal')
                        self._gen_place(task, d, p_idx)
                        task['par_teach'] = _sv_pt
                        task['par_subj']  = _sv_ps
                        task['type']      = _sv_type
                        continue   # placed solo at fixed CT period ✓

                # ── Step 2: CT period blocked by grid cell or primary teacher ──
                # Report the issue (this should not happen if Step 2 was correct).
                if not cells_free_at_pidx:
                    cn_blk  = next((cn for cn in task['cn_list']
                                    if grid[cn][d][p_idx] is not None), None)
                    cell_blk = grid[cn_blk][d][p_idx] if cn_blk else {}
                    s1_issues.append(
                        "HC1 — CT subject '{}' (teacher: {}, class: {}) "
                        "could NOT be placed on {} at Period {} — "
                        "cell already occupied by '{}' (teacher: {}).".format(
                            task['subject'], task['teacher'],
                            ', '.join(task['cn_list']),
                            DAYS[d], p_idx + 1,
                            cell_blk.get('subject', '?'),
                            cell_blk.get('teacher', '?')))
                elif not primary_free:
                    s1_issues.append(
                        "HC1 — CT subject '{}' (teacher: {}, class: {}) "
                        "could NOT be placed on {} at Period {} — "
                        "primary teacher '{}' is already busy at that slot.".format(
                            task['subject'], task['teacher'],
                            ', '.join(task['cn_list']),
                            DAYS[d], p_idx + 1, task['teacher']))
                # Do NOT fall back to another period for CT subjects — the CT
                # period is a hard constraint.  Force Fill will handle any
                # remaining unplaced occurrences.

        # ── HC1 overflow: P1b — CT subject has more periods than WDays ──────────
        # The CT slots (one per day = wdays) are already placed above.
        # Any REMAINING slots (overflow = periods - wdays) must be released from
        # the CT-period pin so Stage 2/3 can place them freely elsewhere.
        #
        # CRITICAL: We keep is_ct=True and priority='P1_CT' on the TASK OBJECT
        # so that (a) _full_shuffle_and_place never unplaces the already-placed
        # CT slots, and (b) _repair_ct_periods can still find this task.
        #
        # To allow the overflow slots to go anywhere, we create a NEW SEPARATE
        # 'overflow' task with is_ct=False and the correct priority.
        # The original task stays pinned and tracks only the 6 placed CT slots.
        #
        #   P1b(a): user set Pref_Per → overflow task gets HC2 priority
        #   P1b(b): no Pref_Per       → overflow task gets 'filler' priority
        for task in list(tasks):   # list() so we can append while iterating
            if task['priority'] != 'P1_CT' or task['remaining'] <= 0:
                continue
            if task['periods'] > wdays:
                overflow_n = task['remaining']
                # Freeze the HC1 task: it has placed exactly wdays slots
                task['periods']   = wdays          # treat as if only wdays periods
                task['remaining'] = 0              # all CT slots placed
                # Create the overflow filler task (a shallow copy, modified)
                ov_task = dict(task)
                ov_task['idx']       = len(tasks)  # new unique index
                ov_task['periods']   = overflow_n
                ov_task['remaining'] = overflow_n
                ov_task['is_ct']     = False       # overflow is NOT a CT slot
                ov_task['ct_period'] = None
                if task['p_pref']:
                    # P1b(a): Pref_Per given — honour it for overflow slots
                    ov_task['priority'] = 'P2'
                else:
                    # P1b(b): no preference — place freely as filler
                    ov_task['priority'] = 'filler'
                ov_task['daily'] = False
                tasks.append(ov_task)

        # ══════════════════════════════════════════════════════════════════════
        # PHASE 1b — Fill remaining empty CT-period slots with secondary
        # CT-teacher subjects (P1a rule).
        #
        # When the primary (max-S_slots) CT subject has fewer periods than
        # WDays, some days have an empty CT-period slot after Phase 1.
        # Fill them with the CT teacher's other subjects (sorted by periods
        # descending).  The slot must never be left empty (P1).
        # ══════════════════════════════════════════════════════════════════════
        self._gen_prog("Stage 1 · Phase 1b (P1_CT) — Secondary CT-teacher subjects for remaining CT slots…")

        for cn in g['all_classes']:
            cd = self.class_config_data.get(cn, {})
            ct_teacher = cd.get('teacher', '').strip()
            ct_per_raw = cd.get('teacher_period', None)
            if not ct_teacher or not ct_per_raw:
                continue
            ct_p = int(ct_per_raw) - 1   # 0-based
            if ct_p < 0 or ct_p >= ppd:
                continue

            # Collect all tasks for this class whose teacher == CT teacher,
            # excluding the primary HC1 task (already handled in Phase 1).
            # Sort by periods descending (most-used subject first).
            secondary_ct_tasks = sorted(
                [
                    t for t in tasks
                    if cn in t['cn_list']
                    and t['teacher'] == ct_teacher
                    and not t.get('is_ct')
                    and t['remaining'] > 0
                ],
                key=lambda t: -t['periods']
            )

            if not secondary_ct_tasks:
                continue

            # For each working day where CT period slot is still empty,
            # try to place a secondary CT-teacher subject there.
            for d in range(wdays):
                if grid[cn][d][ct_p] is not None:
                    continue   # already filled (by primary CT or another class)

                placed_secondary = False
                for sec_task in secondary_ct_tasks:
                    if sec_task['remaining'] <= 0:
                        continue
                    if not g['t_free'](ct_teacher, d, ct_p):
                        continue   # teacher busy elsewhere
                    if g['t_unavail'](ct_teacher, d, ct_p):
                        continue   # P4: unavailable
                    # Check all classes in the task's cn_list are free
                    if not all(grid[c2][d][ct_p] is None for c2 in sec_task['cn_list']):
                        continue
                    # Temporarily pin this task to the CT period for this placement
                    # (it is NOT an HC1 task — just filling the CT slot)
                    _orig_p_pref = sec_task['p_pref']
                    sec_task['p_pref'] = []   # allow any period for this call
                    self._gen_place(sec_task, d, ct_p)
                    sec_task['p_pref'] = _orig_p_pref
                    # Mark the cell as a CT-slot fill (not is_ct, but using CT teacher)
                    grid[cn][d][ct_p] = dict(grid[cn][d][ct_p],
                                             is_ct=True,
                                             ct_fill_secondary=True)
                    placed_secondary = True
                    break

                if placed_secondary:
                    # Refresh secondary_ct_tasks list (remaining counts changed)
                    secondary_ct_tasks = sorted(
                        [
                            t for t in tasks
                            if cn in t['cn_list']
                            and t['teacher'] == ct_teacher
                            and not t.get('is_ct')
                            and t['remaining'] > 0
                        ],
                        key=lambda t: -t['periods']
                    )

        # ══════════════════════════════════════════════════════════════════════
        # PHASE 2 — HC2: Fixed / preference-constrained subjects
        # Rule: iterate through every (day, period) combination that matches the
        # subject's day and period preferences and fill until remaining == 0.
        # Only block = cell already occupied.
        # ══════════════════════════════════════════════════════════════════════
        self._gen_prog("Stage 1 · Phase 2 (P2) — Placing Pref_Per/Pref_Day subjects…")

        # Sort most-constrained first (fewest allowed slots → place first)
        hc2_tasks = sorted(
            [t for t in tasks if t['priority'] == 'P2'],
            key=lambda t: (len(t['p_pref']) or ppd) * (len(t['d_pref']) or wdays))

        for task in hc2_tasks:
            if task['remaining'] <= 0:
                continue

            # Build the ordered list of preferred (day, period) slots
            pref_p = [x - 1 for x in task['p_pref']] if task['p_pref'] else list(range(ppd))
            pref_d = (
                [DAYS.index(x) for x in task['d_pref'] if x in DAYS]
                if task['d_pref'] else list(range(wdays)))
            slots = [(d, p) for d in pref_d for p in pref_p]

            blocked_slots = []   # (day_name, period_1based, blocker_subject, blocker_teacher)

            for d, p in slots:
                if task['remaining'] <= 0:
                    break
                blocked_by = None
                for cn in task['cn_list']:
                    existing = grid[cn][d][p]
                    if existing is not None:
                        blocked_by = (cn, existing)
                        break

                if blocked_by is not None:
                    cn_blk, cell_blk = blocked_by
                    blocked_slots.append((DAYS[d], p + 1,
                                          cell_blk.get('subject', '?'),
                                          cell_blk.get('teacher', '?')))
                    continue

                # Check teacher availability: both busy-check (t_free) and
                # P4 unavailability (t_unavail) must pass before placing.
                t_blk = None
                _t_free   = g['t_free']
                _t_unavail = g['t_unavail']
                if not _t_free(task['teacher'], d, p):
                    t_blk = task['teacher']
                elif _t_unavail(task['teacher'], d, p):
                    t_blk = task['teacher']  # P4: teacher marked unavailable
                else:
                    _pt = (task.get('par_teach') or '').strip()
                    if _pt and _pt not in ('—', '?'):
                        if not _t_free(_pt, d, p) or _t_unavail(_pt, d, p):
                            t_blk = _pt

                if t_blk is None:
                    self._gen_place(task, d, p)
                else:
                    blocked_slots.append((DAYS[d], p + 1,
                                          'TEACHER BUSY/UNAVAILABLE', t_blk))

            if task['remaining'] > 0:
                # Still unplaced periods — report each blocked slot
                classes_str = ', '.join(task['cn_list'])
                for day_name, per_1b, blk_subj, blk_teach in blocked_slots:
                    s1_issues.append(
                        "HC2 — Subject '{}' (teacher: {}) for class {} could NOT be "
                        "placed on {} at Period {} — slot occupied by '{}' "
                        "(teacher: {}).".format(
                            task['subject'], task['teacher'], classes_str,
                            day_name, per_1b, blk_subj, blk_teach))
                if not blocked_slots:
                    # No preferred slot was even available in the preference list
                    s1_issues.append(
                        "HC2 — Subject '{}' (teacher: {}) for class {} has {} "
                        "period(s) unplaced — no matching preferred slot "
                        "exists in the grid (preferences: days={}, periods={}).".format(
                            task['subject'], task['teacher'],
                            ', '.join(task['cn_list']),
                            task['remaining'],
                            task['d_pref'] or 'Any',
                            task['p_pref'] or 'Any'))

        # ══════════════════════════════════════════════════════════════════════
        # REPORTING
        # ══════════════════════════════════════════════════════════════════════
        hc1_placed = sum(t['periods'] - t['remaining'] for t in tasks if t['priority'] == 'P1_CT')
        hc2_placed = sum(t['periods'] - t['remaining'] for t in tasks if t['priority'] == 'P2')
        hc1_fail   = sum(t['remaining']                for t in tasks if t['priority'] == 'P1_CT')
        hc2_fail   = sum(t['remaining']                for t in tasks if t['priority'] == 'P2')
        other_rem  = sum(t['remaining'] for t in tasks if t['priority'] not in ('P1_CT', 'P2'))

        has_issues = bool(s1_issues)
        stage_bg   = "#c0392b" if has_issues else "#1a7a1a"

        if has_issues:
            stage_txt = ("  ⚠ Stage 1 — {} CT period(s) placed, "
                         "{} preference period(s) placed  |  "
                         "{} issue(s) — see status bar".format(
                            hc1_placed, hc2_placed, len(s1_issues)))
        else:
            stage_txt = ("  ✓ Stage 1 complete — {} CT period(s) placed, "
                         "{} preference period(s) placed — "
                         "no issues".format(hc1_placed, hc2_placed))

        if has_issues:
            issue_lines = "\n".join("  ⚠ {}".format(i) for i in s1_issues)
            status = (
                "⚠ Stage 1 complete with {} issue(s) — these should not occur "
                "if Step 2 was fully validated.\n\n"
                "ISSUES:\n{}\n\n"
                "Click 'Task Analysis →' to review groups, then proceed to Stage 2 "
                "({} more period(s) to place).".format(
                    len(s1_issues), issue_lines, other_rem + hc1_fail + hc2_fail))
        else:
            status = (
                "✅ Stage 1 complete — all CT and fixed/preference periods placed "
                "with zero issues.\n"
                "Click '📋 Task Analysis →' to review parallel groups before Stage 2 "
                "({} period(s) remaining).".format(other_rem))

        self._gen_stage = 1
        tt = self._gen_snapshot_tt()
        self._timetable = tt
        self._stage1_status = {
            'stage_txt': stage_txt,
            'stage_bg':  stage_bg,
            'status':    status,
            'has_issues': has_issues,
        }

    def _run_sc2_phase(self):
        """
        Place all P3 (Main_sub) subjects: same period on every working day (Inc order).
        P3 = subjects where S_slots >= WDays (taught every day at same period).
        Called BEFORE Task Analysis and SC1 so main subjects lock their period first.
        After placing wdays slots on one period, any overflow becomes filler.
        """
        g      = self._gen
        tasks  = g["tasks"]
        grid   = g["grid"]
        wdays  = g["wdays"]
        ppd    = g["ppd"]
        DAYS   = g["DAYS"]

        self._gen_prog("Stage 2a · SC2 — Main subjects (same period every day, Inc)…")
        sc2_tasks = sorted([t for t in tasks if t["priority"] == "P3"
                             and t["remaining"] > 0],
                           key=lambda t: (-len(t["cn_list"]), -t["periods"]))

        teacher_sc2_periods: dict = {}

        for task in sc2_tasks:
            if task["remaining"] <= 0:
                continue
            t_name = task["teacher"]
            used_periods = teacher_sc2_periods.get(t_name, set())
            period_order = [p for p in range(ppd) if p not in used_periods]
            period_order += [p for p in range(ppd) if p in used_periods]
            slots_needed = min(task["remaining"], wdays)
            placed = False
            for p in period_order:
                avail = [d for d in range(wdays)
                         if self._gen_can_place(task, d, p)]
                if len(avail) >= slots_needed:
                    for d in avail[:slots_needed]:
                        self._gen_place(task, d, p)
                    teacher_sc2_periods.setdefault(t_name, set()).add(p)
                    placed = True
                    break
            if not placed:
                best_p, best_avail = -1, []
                for p in range(ppd):
                    avail = [d for d in range(wdays)
                             if self._gen_can_place(task, d, p)]
                    if len(avail) > len(best_avail):
                        best_p, best_avail = p, avail
                if best_p >= 0:
                    for d in best_avail[:task["remaining"]]:
                        self._gen_place(task, d, best_p)
                    teacher_sc2_periods.setdefault(t_name, set()).add(best_p)

        # Demote any remaining SC2 overflow to filler
        for task in tasks:
            if task["priority"] == "P3" and task["remaining"] > 0:
                task["priority"] = "filler"
                task["daily"]    = False

    def _run_sc1_phase(self):
        """
        Place all P5 (Cons_per) subjects using Dec order (last period → first).
        P5 = subjects that must be taught in consecutive double-periods. (last period → first),
        respecting Pref_Per / Pref_Day if set.
        Called BEFORE Task Analysis so consecutive pairs get Dec-order slots first.
        """
        g      = self._gen
        tasks  = g["tasks"]
        wdays  = g["wdays"]
        ppd    = g["ppd"]
        DAYS   = g["DAYS"]

        self._gen_prog("Stage 2b · SC1 — Consecutive subjects (Dec order)…")
        sc1_tasks = sorted([t for t in tasks if t["priority"] == "P5"
                             and t["remaining"] > 0],
                           key=lambda t: -t["periods"])

        for task in sc1_tasks:
            if task["remaining"] <= 0:
                continue
            if task["p_pref"]:
                pref_starts = [x - 1 for x in task["p_pref"] if x - 1 < ppd - 1]
                dec_rest    = [p for p in range(ppd - 2, -1, -1) if p not in pref_starts]
                period_order = pref_starts + dec_rest
            else:
                period_order = list(range(ppd - 2, -1, -1))
            if task["d_pref"]:
                pref_days  = [i for i, dn in enumerate(DAYS) if dn in task["d_pref"]]
                rest_days  = [i for i in range(wdays) if i not in pref_days]
                day_order  = pref_days + rest_days
            else:
                day_order = list(range(wdays))
            for d in day_order:
                if task["remaining"] <= 0:
                    break
                for p_start in period_order:
                    if task["remaining"] <= 0:
                        break
                    if self._gen_can_place(task, d, p_start):
                        self._gen_place(task, d, p_start)
                        # Second slot of the pair: use ignore_sc1=True so
                        # _gen_can_place doesn't demand a *third* adjacent slot.
                        # We already know p_start+1 is free from the first check.
                        if (task["remaining"] > 0
                                and self._gen_can_place(task, d, p_start + 1,
                                                        ignore_sc1=True)):
                            self._gen_place(task, d, p_start + 1)
                        break

    def _run_stage2_phases(self):
        g     = self._gen
        tasks = g["tasks"]
        grid  = g["grid"]
        wdays = g["wdays"]
        ppd   = g["ppd"]
        DAYS  = g["DAYS"]

        # SC1 (consecutive) and SC2 (main) already ran as dedicated phases
        # before Task Analysis in the pipeline (_run_sc2_phase + _run_sc1_phase).
        # Any remaining SC1/SC2 subjects here are fall-through cases handled
        # below in the filler + repair loop.

        # ── Phase 4 — P3 fallthrough: any Main_sub not yet placed by _run_sc2_phase ─
        # (e.g. tasks created mid-generation after initial SC2 phase ran)
        self._gen_prog("Stage 3 · Phase 4 (P3) — Main_sub fallthrough (Inc order)…")
        sc2_tasks = sorted([t for t in tasks if t["priority"] == "P3"
                             and t["remaining"] > 0],
                           key=lambda t: (-len(t["cn_list"]), -t["periods"]))

        # Track which periods each teacher has committed to (for conflict avoidance)
        teacher_sc2_periods: dict = {}  # teacher -> set of periods (0-based) used

        for task in sc2_tasks:
            if task["remaining"] <= 0:
                continue
            t_name = task["teacher"]
            used_periods = teacher_sc2_periods.get(t_name, set())

            placed = False
            # Inc order: period 0 first; skip periods already used by same teacher
            period_order = [p for p in range(ppd) if p not in used_periods]
            period_order += [p for p in range(ppd) if p in used_periods]

            slots_needed = min(task["remaining"], wdays)  # at most one per day

            for p in period_order:
                avail = [d for d in range(wdays)
                         if self._gen_can_place(task, d, p)]
                if len(avail) >= slots_needed:
                    for d in avail[:slots_needed]:
                        self._gen_place(task, d, p)
                    teacher_sc2_periods.setdefault(t_name, set()).add(p)
                    placed = True
                    break
            if not placed:
                # Fallback: place on a single best period with most free days
                best_p, best_avail = -1, []
                for p in range(ppd):
                    avail = [d for d in range(wdays)
                             if self._gen_can_place(task, d, p)]
                    if len(avail) > len(best_avail):
                        best_p, best_avail = p, avail
                if best_p >= 0:
                    for d in best_avail[:task["remaining"]]:
                        self._gen_place(task, d, best_p)
                    teacher_sc2_periods.setdefault(t_name, set()).add(best_p)

        # SC2 overflow → Filler demotion is handled in _run_sc2_phase (earlier).

        # ── Phase 5 — Filler_sub ──────────────────────────────────────────────
        # All remaining tasks (Filler_sub) placed with uniform distribution.
        # Sort: combined > busiest teacher > most periods (most constrained first).
        # Days sorted by subject count ascending → uniform class distribution (P1).
        # Teacher days sorted by load ascending → uniform teacher distribution (P1).
        self._gen_prog("Stage 3 · Phase 5 (Filler_sub) — Filling remaining slots…")
        remaining = [t for t in tasks if t["remaining"] > 0]
        teacher_busy_count = {tn: len(bs) for tn, bs in g['t_busy'].items()}

        def _p5_sort_key(t):
            tb = teacher_busy_count.get(t['teacher'], 0)
            pt2 = (t.get('par_teach') or '').strip()
            if pt2 and pt2 not in ('—', '?'):
                tb = max(tb, teacher_busy_count.get(pt2, 0))
            return (-len(t["cn_list"]), tb, -t["periods"])

        remaining.sort(key=_p5_sort_key)

        def _teacher_day_load(teacher, d):
            """Count periods teacher is teaching on day d (for even distribution)."""
            return sum(
                1 for cn2 in g['all_classes']
                for pp in range(ppd)
                if grid[cn2][d][pp] is not None and (
                    grid[cn2][d][pp].get('teacher') == teacher or
                    (grid[cn2][d][pp].get('par_teach') or '').strip() == teacher
                )
            )

        def _subject_day_count(task_, d_):
            """Times task subject already placed on day d_ for any class in cn_list."""
            return max(
                (sum(1 for pp in range(ppd)
                     if grid[cn_][d_][pp] is not None
                     and grid[cn_][d_][pp].get('subject') == task_['subject'])
                 for cn_ in task_['cn_list']),
                default=0,
            )

        for task in remaining:
            t_name = task['teacher']
            # P1 uniform distribution: prefer days where subject appears least,
            # then by teacher load (ascending) to spread teaching evenly.
            # P2: if Pref_Day set, those days come first.
            if task['d_pref']:
                pref_d   = [i for i, dn in enumerate(DAYS) if dn in task['d_pref']]
                other_d  = sorted([i for i in range(wdays) if i not in pref_d],
                                  key=lambda d: (_subject_day_count(task, d),
                                                 _teacher_day_load(t_name, d), d))
                day_order = pref_d + other_d
            else:
                day_order = sorted(
                    range(wdays),
                    key=lambda d: (_subject_day_count(task, d),
                                   _teacher_day_load(t_name, d), d),
                )
            for d in day_order:
                if task["remaining"] <= 0:
                    break
                period_order = list(range(ppd))
                for p in period_order:
                    if task["remaining"] <= 0:
                        break
                    if self._gen_can_place(task, d, p):
                        self._gen_place(task, d, p)

        # ── Repair loop ────────────────────────────────────────────────────
        # Direct placement + one-level swap only.  No chain swap (too slow).
        # Relax constraints progressively; stop when no progress possible.
        relax_level = 0
        for rep in range(80):
            remaining_tasks = [t for t in tasks if t["remaining"] > 0]
            if not remaining_tasks:
                break
            self._gen_prog("Stage 3 · Repair {}: {} unplaced, relax={}".format(
                rep + 1,
                sum(t["remaining"] for t in remaining_tasks),
                relax_level))

            ix_sc1 = relax_level >= 1
            ix_sc3 = relax_level >= 2
            ix_sc2 = relax_level >= 3
            if ix_sc1:
                for t in tasks: t["rx_sc1"] = True
            if ix_sc3:
                for t in tasks: t["rx_sc3"] = True
            if ix_sc2:
                for t in tasks: t["rx_sc2"] = True

            progress = False

            for task in sorted(remaining_tasks, key=lambda t: (-len(t["cn_list"]), -t["remaining"])):
                if task["remaining"] <= 0:
                    continue
                pt = task["par_teach"]

                # ── Direct placement ──────────────────────────────────────
                for d in range(wdays):
                    if task["remaining"] <= 0: break
                    for p in range(ppd):
                        if task["remaining"] <= 0: break
                        if self._gen_can_place(task, d, p, ix_sc1, ix_sc3, ix_sc2):
                            self._gen_place(task, d, p)
                            progress = True
                            # For consecutive tasks: immediately place second slot
                            # of the pair using ignore_sc1=True (pair partner check done above)
                            if (task.get("consec") and not ix_sc1
                                    and task["remaining"] > 0
                                    and p + 1 < ppd
                                    and self._gen_can_place(task, d, p + 1, ignore_sc1=True,
                                                            ignore_sc3=ix_sc3,
                                                            ignore_sc2=ix_sc2)):
                                self._gen_place(task, d, p + 1)
                            break   # move to next day after placing (pair or single)

                if task["remaining"] <= 0:
                    continue

                # ── One-level swap: try to displace one blocker ───────────
                # Stop as soon as we place one slot (don't exhaust all slots).
                swap_done = False
                for d in range(wdays):
                    if task["remaining"] <= 0 or swap_done: break
                    for p in range(ppd):
                        if task["remaining"] <= 0 or swap_done: break

                        # Teacher must be free at (d,p)
                        if not (g["t_free"](task["teacher"], d, p) and
                                (ix_sc3 or not g["t_unavail"](task["teacher"], d, p))):
                            continue
                        if pt and not (g["t_free"](pt, d, p) and
                                       (ix_sc3 or not g["t_unavail"](pt, d, p))):
                            continue

                        # Hard constraint checks
                        if task["is_ct"] and (p + 1) != task["ct_period"]:
                            continue
                        if task["p_pref"] and not task["is_ct"]:
                            if (p + 1) not in task["p_pref"]: continue
                        if task["d_pref"] and DAYS[d] not in task["d_pref"]:
                            continue
                        if not task["consec"]:
                            _natural = (task["periods"] + wdays - 1) // wdays
                            _max_pd = min(2, _natural)
                            dup = False
                            for cn in task["cn_list"]:
                                _count = sum(
                                    1 for pp in range(ppd)
                                    if g["grid"][cn][d][pp] is not None
                                    and g["grid"][cn][d][pp].get("subject") == task["subject"]
                                )
                                if _count >= _max_pd:
                                    dup = True; break
                            if dup: continue

                        # Find what is blocking
                        blocking_idx = None
                        for cn in task["cn_list"]:
                            if g["grid"][cn][d][p] is not None:
                                blocking_idx = g["task_at"][cn][d][p]
                                break

                        if blocking_idx is None:
                            self._gen_place(task, d, p)
                            progress = True
                            swap_done = True
                            break

                        if blocking_idx >= len(tasks): continue
                        blocker = tasks[blocking_idx]
                        if blocker["priority"] in ("P1_CT", "P2"): continue
                        # CRITICAL: even if the task was demoted from HC1 to filler
                        # (e.g. CT subject with periods > WDays), the placed CELL
                        # still carries is_ct=True and must NEVER be evicted.
                        _blk_cell = g["grid"][task["cn_list"][0]][d][p]
                        if _blk_cell and _blk_cell.get("is_ct"):
                            continue  # protect CT cell regardless of task priority

                        # Try to move blocker to its first available free slot
                        for d2 in range(wdays):
                            moved = False
                            for p2 in range(ppd):
                                if (d2, p2) == (d, p): continue
                                if not self._gen_can_place(
                                        blocker, d2, p2, ix_sc1, ix_sc3, ix_sc2):
                                    continue
                                self._gen_unplace(blocker, d, p)
                                slot_clear = all(
                                    g["grid"][cn][d][p] is None
                                    for cn in task["cn_list"])
                                t_now = (g["t_free"](task["teacher"], d, p) and
                                         (not pt or g["t_free"](pt, d, p)))
                                if slot_clear and t_now:
                                    self._gen_place(task, d, p)
                                    self._gen_place(blocker, d2, p2)
                                    progress = True
                                    swap_done = True
                                    moved = True
                                    break
                                else:
                                    self._gen_place(blocker, d, p)
                            if moved or swap_done: break
                        if swap_done: break

            if not progress:
                relax_level += 1
                if relax_level > 4:
                    break

        # ── Remove any residual teacher double-bookings ────────────────────────
        # Safety net: ensures the grid never has the same teacher in two classes
        # at the same slot (which would be physically impossible).
        self._remove_teacher_conflicts()

        # ── Store result ────────────────────────────────────────────────────
        unplaced = sum(t["remaining"] for t in tasks)
        tt = self._gen_snapshot_tt()
        self._timetable = tt
        self._gen_stage = 3
        self._stage2_status = {
            'unplaced':  unplaced,
            'ok':        unplaced == 0,
            'msg':       ("✅ Complete timetable generated — all periods placed!" if unplaced == 0
                          else "⚠ {} period(s) still unplaced — constraints may be too tight.".format(unplaced)),
            'stage_msg': ("Stage 3 complete ✓ — Full timetable generated!" if unplaced == 0
                          else "Stage 3 done — {} period(s) unplaced".format(unplaced)),
        }

    def _build_timetable(self):
        """Legacy single-shot builder — kept for backward compat. Not used in staged flow."""
        self._init_gen_state()
        self._run_stage1_phases()
        self._run_stage2_phases()
        return self._gen_snapshot_tt()

    # =========================================================================
    #  DISPLAY
    # =========================================================================

    def _run_stuck_logic(self, progress_cb=None):
        """
        Spec-compliant STUCK LOGIC — applies relaxations then calls
        _force_fill_backtrack (which IS the Shuffle: Stage A greedy+swap,
        Stage B full-grid-shuffle, Stage C CSP).

        Step 1: Shuffle (call _force_fill_backtrack as-is)
        Step 2: Relax P5 (mark stuck consec tasks rx_sc1=True) + Shuffle
        Step 3: Relax P4 (mark stuck teachers rx_sc3=True) + Shuffle
        Step 4: Reduce max-S_slots Main_sub by 1, fill freed slot, + Shuffle
                Repeated until stable or no further reduction possible.

        Never breaks P1–P3.
        """
        def _prog(msg):
            if progress_cb:
                progress_cb(msg)

        g     = self._gen
        tasks = g['tasks']
        wdays = g['wdays']
        ppd   = g['ppd']
        DAYS  = g['DAYS']

        if not hasattr(self, '_relaxed_consec_keys'):
            self._relaxed_consec_keys = set()

        def _unplaced():
            return sum(t['remaining'] for t in tasks)

        def _shuffle():
            """Run _force_fill_backtrack (Stage A greedy+swap, Stage B full
            shuffle, Stage C CSP). This IS the spec's Shuffle.
            Note: _force_fill_backtrack already handles teacher conflicts
            via Stage C — do NOT call _remove_teacher_conflicts here as it
            would increment remaining for removed tasks with nothing to re-place them."""
            try:
                self._force_fill_backtrack(progress_cb=_prog)
            except Exception as exc:
                _prog(f"  Shuffle error: {exc}")
            return _unplaced()

        # ── STEP 1: Shuffle ──────────────────────────────────────────────────
        _prog("Stuck Logic Step 1 — Shuffle (re-order Filler_sub)…")
        up = _shuffle()
        _prog(f"  → {up} unplaced after Step 1")
        if up == 0:
            return 0

        # ── STEP 2: Relax P5 (Cons_per → Filler_sub) + Shuffle ──────────────
        _prog("Stuck Logic Step 2 — Relax P5 (treat Cons_per as Filler_sub) + Shuffle…")
        relaxed_any = False
        for t in tasks:
            if t.get('consec') and t['remaining'] > 0:
                t['rx_sc1'] = True
                for cn_i in t['cn_list']:
                    self._relaxed_consec_keys.add((cn_i, t['subject']))
                relaxed_any = True
        if relaxed_any:
            up = _shuffle()
            _prog(f"  → {up} unplaced after Step 2")
            if up == 0:
                return 0

        # ── STEP 3: Relax P4 (teacher unavailability) + Shuffle ─────────────
        _prog("Stuck Logic Step 3 — Relax P4 (bypass teacher unavailability) + Shuffle…")
        unav_relaxed = set()
        for t in tasks:
            if t['remaining'] > 0:
                t['rx_sc3'] = True
                if t['teacher']:
                    unav_relaxed.add(t['teacher'])
        if unav_relaxed:
            _prog(f"  Bypassing unavailability for: {', '.join(sorted(unav_relaxed))}")
            up = _shuffle()
            _prog(f"  → {up} unplaced after Step 3")
            if up == 0:
                return 0

        # ── STEP 4: Reduce max-S_slots Main_sub + fill freed slot + Shuffle ──
        # For each stuck class: find Main_sub with most S_slots, reduce by 1,
        # unplace one grid slot, fill freed slot with another placeable Main_sub.
        # Repeat until no improvement.
        _prog("Stuck Logic Step 4 — Reduce max S_slots of Main_sub (P3) by 1 + Shuffle…")
        reduced_pairs = set()  # (cn, subject) pairs already reduced
        MAX_ROUNDS = 20

        for _round in range(MAX_ROUNDS):
            if _unplaced() == 0:
                break

            stuck_tasks = [t for t in tasks if t['remaining'] > 0]
            if not stuck_tasks:
                break

            made_reduction = False
            for stuck in stuck_tasks:
                if _unplaced() == 0:
                    break

                for cn in stuck['cn_list']:
                    if cn not in self.class_config_data:
                        continue

                    # Find the Main_sub (SC2/filler) with the most S_slots
                    # that is fully placed, not HC1/CT, periods > 1
                    candidates = sorted(
                        [t for t in tasks
                         if cn in t['cn_list']
                         and t['priority'] in ('P3', 'filler')
                         and not t.get('is_ct')
                         and t['remaining'] == 0
                         and t['periods'] > 1
                         and t['subject'] != stuck['subject']
                         and (cn, t['subject']) not in reduced_pairs],
                        key=lambda t: -t['periods']
                    )
                    if not candidates:
                        continue

                    target = candidates[0]
                    reduced_pairs.add((cn, target['subject']))

                    # Find one placed grid slot to unplace (Dec order — last slot)
                    freed = None
                    for d in range(wdays - 1, -1, -1):
                        for p in range(ppd - 1, -1, -1):
                            if (g['task_at'][cn][d][p] == target['idx']
                                    and g['grid'][cn][d][p] is not None
                                    and not g['grid'][cn][d][p].get('is_ct')):
                                freed = (d, p)
                                break
                        if freed:
                            break

                    if not freed:
                        continue

                    # Unplace that slot
                    fd, fp = freed
                    self._gen_unplace(target, fd, fp)

                    # Reduce config so the slot stays freed across retries
                    for s in self.class_config_data[cn].get('subjects', []):
                        if (s.get('name', '').strip() == target['subject']
                                and s['periods'] > 1):
                            s['periods'] -= 1
                            self._period_reductions.append({
                                'class': cn, 'subject': target['subject'],
                                'teacher': target['teacher'],
                                'from_periods': s['periods'] + 1,
                                'to_periods':   s['periods'],
                            })
                            break

                    _prog(f"  Freed: {cn} {DAYS[fd]} P{fp+1} from '{target['subject']}'")

                    # Try to fill freed slot with stuck task or any needy Main_sub
                    for fill_t in ([stuck] + [t for t in tasks
                                              if cn in t['cn_list']
                                              and t['remaining'] > 0
                                              and not t.get('is_ct')]):
                        if g['grid'][cn][fd][fp] is not None:
                            break
                        if not g['t_free'](fill_t['teacher'], fd, fp):
                            continue
                        _orig_pref = fill_t['p_pref']
                        fill_t['p_pref'] = []
                        if self._gen_can_place(fill_t, fd, fp,
                                               ignore_sc1=True, ignore_sc3=True):
                            self._gen_place(fill_t, fd, fp)
                            fill_t['p_pref'] = _orig_pref
                            break
                        fill_t['p_pref'] = _orig_pref

                    up = _shuffle()
                    _prog(f"  → {up} unplaced after round {_round+1} reduction")
                    made_reduction = True
                    if up == 0:
                        return 0
                    break  # one reduction per stuck task per round

            if not made_reduction:
                break

        return _unplaced()


    def _force_fill_backtrack(self, progress_cb=None):
        """
        Multi-stage guaranteed timetable completion.

        Stage A: Greedy + swap with progressive constraint relaxation.
        Stage B: Full grid shuffle — unplace ALL non-HC1 tasks, re-sort by
                 difficulty (combined > parallel > daily > filler), re-place
                 with teacher-free-slot checking.  Repeated up to MAX_SHUFFLES
                 times with random tiebreaking so the same conflict-prone order
                 is not repeated.
        Stage C: Min-Conflicts CSP — move the worst-conflicting task's slot to
                 the position that minimises teacher double-bookings.

        Guarantees:
          • A teacher can never appear in two classes at the same slot after
            Stage B/C because we only place into teacher-free empty cells.
          • HC1 (CT) periods are never touched.
          • All remaining periods are placed — if the problem is infeasible
            (teacher has more periods than available slots) some tasks may
            stay partially unplaced, but teacher double-bookings are zero.

        progress_cb(msg) is called periodically to update the UI label.
        """
        import random as _rnd
        _det_rng = _rnd.Random(42)   # isolated deterministic RNG for all FF operations

        def _det_shuffle(lst):
            _det_rng.shuffle(lst)

        def _prog(msg):
            if progress_cb:
                progress_cb(msg)

        g      = self._gen
        tasks  = g['tasks']
        grid   = g['grid']
        wdays  = g['wdays']
        ppd    = g['ppd']

        if not hasattr(self, '_relaxed_main_keys'):
            self._relaxed_main_keys = set()
        if not hasattr(self, '_relaxed_consec_keys'):
            self._relaxed_consec_keys = set()

        relaxed_notes = []
        # Priority weights: lower = harder to displace (higher spec priority)
        # P1_CT=CT periods, P2=Pref_Per/Day, P3=Main_sub same-period,
        # P5=Consecutive, filler=everything else
        PRIO_W = {'P1_CT': 0, 'P2': 1, 'P3': 2, 'P5': 3, 'filler': 4}

        def _prio(t):
            return PRIO_W.get(t['priority'], 4)

        def _unplaced():
            return sum(t['remaining'] for t in tasks)

        def _can(task, d, p, ign_sc1=False, ign_sc3=False):
            return self._gen_can_place(task, d, p,
                                       ignore_sc1=ign_sc1,
                                       ignore_sc3=ign_sc3)

        # ── Helper: count free slots for a task ───────────────────────────────
        def _free_slots_for(task, ign_sc1=False, ign_sc3=False):
            return sum(1 for d in range(wdays) for p in range(ppd)
                       if _can(task, d, p, ign_sc1, ign_sc3))

        # ─────────────────────────────────────────────────────────────────────
        # STAGE A: greedy + swap with progressive constraint relaxation
        # ─────────────────────────────────────────────────────────────────────
        def _teacher_day_slots(teacher, day):
            """Count how many periods teacher already has on this day."""
            return sum(
                1 for cn2 in g['all_classes']
                for pp2 in range(ppd)
                if grid[cn2][day][pp2] is not None and (
                    grid[cn2][day][pp2].get('teacher') == teacher or
                    (grid[cn2][day][pp2].get('par_teach') or '').strip() == teacher)
            )

        def _greedy_pass(ign_sc1=False, ign_sc3=False):
            remaining_tasks = [t for t in tasks if t['remaining'] > 0]
            # MRV: place most constrained first
            remaining_tasks.sort(key=lambda t: _free_slots_for(t, ign_sc1, ign_sc3))
            for task in remaining_tasks:
                # Sort days by teacher load (least busy day first) for uniform distribution
                day_order = sorted(range(wdays),
                                   key=lambda d: _teacher_day_slots(task['teacher'], d))
                for d in day_order:
                    if task['remaining'] <= 0:
                        break
                    for p in range(ppd):
                        if task['remaining'] <= 0:
                            break
                        if _can(task, d, p, ign_sc1, ign_sc3):
                            self._gen_place(task, d, p)
                            # Consecutive: immediately place second slot of pair
                            if (task.get('consec') and not ign_sc1
                                    and task['remaining'] > 0
                                    and p + 1 < ppd
                                    and _can(task, d, p + 1, True, ign_sc3)):
                                self._gen_place(task, d, p + 1)
                            break   # one pair per day

        def _swap_pass(ign_sc1=False, ign_sc3=False):
            """Try to displace lower-priority tasks to make room for higher-priority ones."""
            for task in sorted(tasks, key=lambda t: (_prio(t), -t['remaining'])):
                if task['remaining'] <= 0 or _prio(task) == 0:
                    continue
                for d in range(wdays):
                    if task['remaining'] <= 0:
                        break
                    for p in range(ppd):
                        if task['remaining'] <= 0:
                            break
                        tname = task['teacher']
                        pt    = task.get('par_teach', '')
                        t_ok  = g['t_free'](tname, d, p)
                        if not ign_sc3:
                            t_ok = t_ok and not g['t_unavail'](tname, d, p)
                        if pt and pt not in ('', '—', '?'):
                            t_ok = t_ok and g['t_free'](pt, d, p)
                        if not t_ok:
                            continue
                        # Find what's blocking
                        bidx = None
                        for cn in task['cn_list']:
                            if grid[cn][d][p] is not None:
                                bidx = g['task_at'][cn][d][p]
                                break
                        if bidx is None:
                            if _can(task, d, p, ign_sc1, ign_sc3):
                                self._gen_place(task, d, p)
                            continue
                        blocker = tasks[bidx]
                        # Only displace tasks of lower or equal priority
                        # (same-priority filler-vs-filler swaps are needed for deadlock breaking)
                        if _prio(blocker) < _prio(task):
                            continue
                        # Try to relocate blocker
                        for d2 in range(wdays):
                            moved = False
                            for p2 in range(ppd):
                                if (d2, p2) == (d, p):
                                    continue
                                if not _can(blocker, d2, p2, ign_sc1, ign_sc3):
                                    continue
                                self._gen_unplace(blocker, d, p)
                                clr = all(grid[cn][d][p] is None
                                          for cn in task['cn_list'])
                                tok = (g['t_free'](tname, d, p)
                                       and (not pt or pt in ('','—','?')
                                            or g['t_free'](pt, d, p)))
                                if clr and tok and _can(task, d, p, ign_sc1, ign_sc3):
                                    self._gen_place(blocker, d2, p2)
                                    self._gen_place(task, d, p)
                                    moved = True
                                    break
                                else:
                                    self._gen_place(blocker, d, p)
                            if moved:
                                break

        def _run_stage_a(ign_sc1=False, ign_sc3=False, rounds=6):
            for _ in range(rounds):
                if _unplaced() == 0:
                    return
                _greedy_pass(ign_sc1, ign_sc3)
            # More swap+greedy iterations to break filler-vs-filler deadlocks
            for _ in range(rounds * 2):
                if _unplaced() == 0:
                    return
                _swap_pass(ign_sc1, ign_sc3)
                _greedy_pass(ign_sc1, ign_sc3)

        _prog("Stage A — greedy placement…")
        _run_stage_a()
        if _unplaced() == 0:
            _prog("")
            return None

        # Relax consecutive
        _prog("Stage A — relaxing consecutive…")
        consec_items = []
        for t in tasks:
            if t['consec'] and t['remaining'] > 0:
                t['rx_sc1'] = True
                for cn_i in t['cn_list']:
                    self._relaxed_consec_keys.add((cn_i, t['subject']))
                consec_items.append("  • {} — {}".format(
                    '+'.join(t['cn_list']), t['subject']))
        if consec_items:
            relaxed_notes.append(
                "Consecutive constraint relaxed for:\n" + '\n'.join(consec_items))
        _run_stage_a(ign_sc1=True)
        if _unplaced() == 0:
            _prog("")
            return '\n\n'.join(relaxed_notes)

        # Relax unavailability
        _prog("Stage A — relaxing unavailability…")
        unav_set = set()
        for t in tasks:
            if t['remaining'] > 0:
                t['rx_sc3'] = True
                if t['teacher']:
                    unav_set.add(t['teacher'])
                pt = t.get('par_teach', '')
                if pt and pt not in ('', '—', '?'):
                    unav_set.add(pt)
        if unav_set:
            relaxed_notes.append(
                "Teacher unavailability bypassed for:\n"
                + '\n'.join("  • {}".format(x) for x in sorted(unav_set)))
        _run_stage_a(ign_sc1=True, ign_sc3=True)
        if _unplaced() == 0:
            _prog("")
            return '\n\n'.join(relaxed_notes)

        # P2 (Pref_Per/Pref_Day) and P3 (Main_sub same period) are NEVER relaxed.
        # Per spec: 'Don't break P1 to P3 priority at any cost.'

        # ─────────────────────────────────────────────────────────────────────
        # STAGE B: Full-grid shuffle — the "Stuck" logic
        #
        # Key insight: if greedy got stuck, the placement ORDER matters.
        # Solution: unplace ALL non-HC1 tasks and re-place them sorted by
        # difficulty (combined > parallel > single-class, then by period count).
        # Repeat with randomised tiebreaking to escape different local minima.
        # ─────────────────────────────────────────────────────────────────────
        relaxed_notes.append(
            "Full-grid shuffle applied: re-ordering all tasks by constraint difficulty.")

        # Mark everything relaxed for Stage B onwards
        for t in tasks:
            t['rx_sc1'] = True
            t['rx_sc3'] = True

        def _difficulty_key(t, rng):
            """Sort key: most constrained first + random tiebreaker.
            Priority order:
              1. Combined classes (need all classes free simultaneously)
              2. Parallel tasks (two teachers must be free)  
              3. Teachers with most periods remaining (least slack)
              4. Most periods in this task
            Tiebreak: random (for shuffle diversity)."""
            n_classes = len(t['cn_list'])
            is_par    = 1 if t.get('par_teach', '') not in ('', '—', '?') else 0
            # Teacher tightness: total remaining periods / available slots
            # (lower = more constrained)
            t_remaining_total = sum(
                tt['remaining'] for tt in tasks
                if tt['teacher'] == t['teacher'] or
                (tt.get('par_teach') or '').strip() == t['teacher'])
            pt3 = (t.get('par_teach') or '').strip()
            if pt3 and pt3 not in ('—', '?'):
                pt3_total = sum(
                    tt['remaining'] for tt in tasks
                    if tt['teacher'] == pt3 or
                    (tt.get('par_teach') or '').strip() == pt3)
                t_remaining_total = max(t_remaining_total, pt3_total)
            return (-n_classes, -is_par, -t_remaining_total, -t['periods'], rng)

        def _full_shuffle_and_place(seed=None):
            """Unplace all non-HC1/non-HC2 tasks, re-sort, re-place. Returns unplaced count."""
            rng_source = _rnd.Random(seed if seed is not None else 42)

            # Step 1: unplace non-HC1, non-HC2, non-SC1 tasks only.
            # HC2 (preference-constrained) and SC1 (same-period daily) tasks are kept
            # in place because re-placing them without their constraints enforced corrupts
            # their carefully chosen slots (e.g. SOL at Wed P7/P8).
            # PRESERVE: never unplace higher-priority tasks during Stage B shuffle.
            # P1_CT=CT periods, P2=Pref_Per/Day, P3=Main_sub, P5=Cons_per
            # Only Filler_sub is unplaced and re-ordered.
            PRESERVE = ('P1_CT', 'P2', 'P3', 'P5')  # P1_CT=CT, P2=Pref, P3=Main, P5=Consec  # P1_CT=CT, P2=Pref, P3=Main, P5=Consec
            for t in tasks:
                if t['priority'] in PRESERVE:
                    continue
                for d in range(wdays):
                    for p in range(ppd):
                        for cn in t['cn_list']:
                            if (g['task_at'][cn][d][p] == t['idx']
                                    and grid[cn][d][p] is not None):
                                self._gen_unplace(t, d, p)
                                break   # one unplace per (d,p)

            # Verify: reset remaining from scratch based on grid state
            for t in tasks:
                if t['priority'] in PRESERVE:
                    continue
                placed = 0
                for d in range(wdays):
                    for p in range(ppd):
                        for cn_r in t['cn_list']:
                            if (cn_r in g['task_at']
                                    and g['task_at'][cn_r][d][p] == t['idx']):
                                placed += 1
                                break  # count once per slot
                t['remaining'] = max(0, t['periods'] - placed)

            # Step 2: re-sort by difficulty (only unpreserved tasks need sorting)
            sortable = [t for t in tasks if t['priority'] not in PRESERVE]
            sortable.sort(key=lambda t: _difficulty_key(t, rng_source.random()))

            # Step 3: place each task greedily in teacher-free empty slots only
            for t in sortable:
                if t['remaining'] <= 0:
                    continue
                t_name = t['teacher']
                pt2    = (t.get('par_teach') or '').strip()

                # Collect candidate slots: teacher-free + class-free
                candidates = []
                for d in range(wdays):
                    for p in range(ppd):
                        cells_free = all(
                            grid[cn][d][p] is None for cn in t['cn_list'])
                        if not cells_free:
                            continue
                        if not g['t_free'](t_name, d, p):
                            continue
                        if pt2 and pt2 not in ('—', '?') and not g['t_free'](pt2, d, p):
                            continue
                        candidates.append((d, p))

                # Shuffle to avoid same-order bias (use seeded rng for determinism)
                rng_source.shuffle(candidates)

                # Place needed slots (consecutive: place pair atomically)
                for d, p in candidates:
                    if t['remaining'] <= 0:
                        break
                    # Re-check (prior placements in this same pass may have changed state)
                    cells_free = all(grid[cn][d][p] is None for cn in t['cn_list'])
                    if not cells_free:
                        continue
                    if not g['t_free'](t_name, d, p):
                        continue
                    if pt2 and pt2 not in ('—', '?') and not g['t_free'](pt2, d, p):
                        continue
                    self._gen_place(t, d, p)
                    # Consecutive: immediately place second slot
                    if (t.get('consec') and t['remaining'] > 0 and p + 1 < g['ppd']):
                        cells2 = all(grid[cn][d][p+1] is None for cn in t['cn_list'])
                        t2_free = (g['t_free'](t_name, d, p+1)
                                   and (not pt2 or pt2 in ('—','?')
                                        or g['t_free'](pt2, d, p+1)))
                        if cells2 and t2_free:
                            self._gen_place(t, d, p+1)

            return sum(t['remaining'] for t in tasks)

        MAX_SHUFFLES = 40   # more attempts for tight 100%-filled grids
        best_unplaced = _unplaced()
        best_snap     = self._ft_snapshot()

        for sh_idx in range(MAX_SHUFFLES):
            if best_unplaced == 0:
                break
            _prog("Stage B — shuffle {} / {} (best={} unplaced)…".format(
                sh_idx + 1, MAX_SHUFFLES, best_unplaced))

            up = _full_shuffle_and_place(seed=sh_idx * 7919)
            if up < best_unplaced:
                best_unplaced = up
                best_snap     = self._ft_snapshot()
            if best_unplaced == 0:
                break

            # Restore best snapshot before next attempt
            if up > best_unplaced:
                self._ft_restore(best_snap)

        # Restore the best result found
        if best_unplaced > 0:
            self._ft_restore(best_snap)

        # ── Deep chain-swap for completely deadlocked tasks ──────────────────
        # When a task's teacher is busy at every slot where the class is free,
        # and the single-level deep-swap found no alternative for the blocker,
        # we need a 2-level chain: move blocker→alternative, evict alt→free spot.
        # We do this with a BFS-like approach: for each stuck task, try evicting
        # the blocker, and for each slot the blocker can't reach, evict THAT slot's
        # occupant if it's a filler/SC2 task, then check if stuck task can be placed.
        if _unplaced() > 0:
            for _chain_pass in range(300):
                if _unplaced() == 0:
                    break
                improved = False

                for stuck_task in sorted(
                    [t for t in tasks if t['remaining'] > 0],
                    key=lambda t: (t['priority'] not in ('P1_CT','P2','P5'), -t['remaining'])
                ):
                    if stuck_task['remaining'] <= 0:
                        continue
                    s_teacher = stuck_task['teacher']
                    cn_list   = stuck_task['cn_list']

                    # Find slots where class is free but teacher is busy
                    for d in range(wdays):
                        if stuck_task['remaining'] <= 0:
                            break
                        for p in range(ppd):
                            if stuck_task['remaining'] <= 0:
                                break

                            # Class(es) must be free here
                            cells_free = all(
                                g['grid'][cn][d][p] is None for cn in cn_list)
                            if not cells_free:
                                continue

                            # Teacher must be busy (otherwise place directly)
                            if g['t_free'](s_teacher, d, p):
                                pt_s = (stuck_task.get('par_teach') or '').strip()
                                if not (pt_s and pt_s not in ('—','?') and
                                        not g['t_free'](pt_s, d, p)):
                                    self._gen_place(stuck_task, d, p)
                                    improved = True
                                    break
                                continue

                            # Find the Level-1 blocker (what is teacher doing?)
                            l1_blocker = None
                            l1_d, l1_p = d, p
                            for cn2 in g['all_classes']:
                                idx2 = g['task_at'][cn2][l1_d][l1_p]
                                if idx2 is None:
                                    continue
                                t2 = tasks[idx2]
                                if (t2['teacher'] == s_teacher or
                                        (t2.get('par_teach') or '').strip() == s_teacher):
                                    if t2['priority'] not in ('P1_CT', 'P2'):
                                        l1_blocker = t2
                                    break
                            if l1_blocker is None:
                                continue

                            # Try Level-1 swap: move l1_blocker to alt slot
                            l1_moved = False
                            for d2 in range(wdays):
                                if l1_moved:
                                    break
                                for p2 in range(ppd):
                                    if (d2, p2) == (l1_d, l1_p):
                                        continue
                                    if self._gen_can_place(l1_blocker, d2, p2,
                                                           ignore_sc1=True,
                                                           ignore_sc3=True):
                                        # Simple L1 swap
                                        self._gen_unplace(l1_blocker, l1_d, l1_p)
                                        if self._gen_can_place(stuck_task, d, p,
                                                               ignore_sc1=True,
                                                               ignore_sc3=True):
                                            self._gen_place(stuck_task, d, p)
                                            self._gen_place(l1_blocker, d2, p2)
                                            improved = True
                                            l1_moved = True
                                            break
                                        else:
                                            self._gen_place(l1_blocker, l1_d, l1_p)

                            # Special case: blocker is FULLY PLACED (remaining=0).
                            # We can remove one occurrence if:
                            #   (a) The day has 2+ occurrences (just removing an extra), or
                            #   (b) Subject has more periods than wdays (one day has a duplicate), or
                            #   (c) Blocker is a low-priority filler with >1 period placed
                            #       (losing one period is acceptable for flexible fillers).
                            if not l1_moved and l1_blocker['remaining'] == 0 and l1_blocker['periods'] > 1:
                                cn0 = l1_blocker['cn_list'][0]
                                same_day_count = sum(
                                    1 for pp2 in range(ppd)
                                    if g['task_at'][cn0][l1_d][pp2] == l1_blocker['idx']
                                )
                                placed_count = l1_blocker['periods'] - l1_blocker['remaining']
                                can_remove = (
                                    same_day_count >= 2        # extra occurrence today
                                    or l1_blocker['periods'] > wdays  # over-wdays: has duplicates
                                    or (l1_blocker['priority'] == 'filler' and placed_count > 1)
                                )
                                if can_remove:
                                    self._gen_unplace(l1_blocker, l1_d, l1_p)
                                    can_stuck = self._gen_can_place(stuck_task, d, p,
                                                                     ignore_sc1=True,
                                                                     ignore_sc3=True)
                                    if can_stuck:
                                        self._gen_place(stuck_task, d, p)
                                        improved = True
                                        l1_moved = True
                                    else:
                                        self._gen_place(l1_blocker, l1_d, l1_p)
                            if l1_moved:
                                break
                            # whatever blocks L1 from a candidate slot.
                            for d2 in range(wdays):
                                if l1_moved:
                                    break
                                for p2 in range(ppd):
                                    if (d2, p2) == (l1_d, l1_p):
                                        continue
                                    # L1 blocker's teacher must be free at (d2,p2)
                                    l1_t = l1_blocker['teacher']
                                    l1_pt = (l1_blocker.get('par_teach') or '').strip()
                                    if not g['t_free'](l1_t, d2, p2):
                                        continue
                                    if l1_pt and l1_pt not in ('—','?') and not g['t_free'](l1_pt, d2, p2):
                                        continue
                                    # But l1_blocker's class(es) are occupied at (d2,p2)
                                    l2_idx_set = set()
                                    for cnB in l1_blocker['cn_list']:
                                        l2_idx = g['task_at'][cnB][d2][p2]
                                        if l2_idx is not None:
                                            l2_idx_set.add(l2_idx)
                                    if not l2_idx_set or len(l2_idx_set) > 1:
                                        continue
                                    l2_blocker = tasks[next(iter(l2_idx_set))]
                                    if l2_blocker['priority'] in ('P1_CT', 'P2'):
                                        continue
                                    if l2_blocker['idx'] == l1_blocker['idx']:
                                        continue

                                    # Try to find a spot for l2_blocker
                                    for d3 in range(wdays):
                                        if l1_moved:
                                            break
                                        for p3 in range(ppd):
                                            if (d3, p3) in ((l1_d, l1_p), (d2, p2)):
                                                continue
                                            if not self._gen_can_place(l2_blocker, d3, p3,
                                                                        ignore_sc1=True,
                                                                        ignore_sc3=True):
                                                continue
                                            # Execute the 2-level chain
                                            self._gen_unplace(l1_blocker, l1_d, l1_p)
                                            self._gen_unplace(l2_blocker, d2, p2)
                                            ok_stuck = self._gen_can_place(stuck_task, d, p,
                                                                            ignore_sc1=True,
                                                                            ignore_sc3=True)
                                            ok_l1    = self._gen_can_place(l1_blocker, d2, p2,
                                                                            ignore_sc1=True,
                                                                            ignore_sc3=True)
                                            if ok_stuck and ok_l1:
                                                self._gen_place(stuck_task, d, p)
                                                self._gen_place(l1_blocker, d2, p2)
                                                self._gen_place(l2_blocker, d3, p3)
                                                improved = True
                                                l1_moved = True
                                                break
                                            else:
                                                # Undo
                                                self._gen_place(l2_blocker, d2, p2)
                                                self._gen_place(l1_blocker, l1_d, l1_p)

                    if improved:
                        break   # restart outer loop after any improvement

                if not improved:
                    break

        # ── Cross-class teacher-swap pass ──────────────────────────────────
        # Pattern: stuck task's class has 1 free slot, but teacher is busy there
        # in a DIFFERENT class (call it the "blocker class").  The blocker task has
        # no empty alternative slot because that class is also 100% full.
        # Solution: swap the blocker WITH an already-placed task in the blocker class
        # to free the teacher at the stuck slot.  We look for a placed task in the
        # blocker class whose teacher IS free at the stuck slot.
        if _unplaced() > 0:
            import time as _time
            _xc_deadline = _time.time() + 5.0   # max 5 seconds for this pass
            for _xc_pass in range(20):   # cap iterations
                if _unplaced() == 0 or _time.time() > _xc_deadline:
                    break
                if _unplaced() == 0:
                    break
                improved = False

                for stuck_task in [t for t in tasks if t['remaining'] > 0]:
                    if stuck_task['remaining'] <= 0:
                        continue
                    s_t = stuck_task['teacher']
                    s_cn = stuck_task['cn_list']

                    # Enumerate free slots in the stuck class(es)
                    free_here = [
                        (d, p)
                        for d in range(wdays)
                        for p in range(ppd)
                        if all(grid[cn][d][p] is None for cn in s_cn)
                    ]
                    if not free_here:
                        continue

                    placed_it = False
                    for (d, p) in free_here:
                        if stuck_task['remaining'] <= 0:
                            break
                        if g['t_free'](s_t, d, p):
                            self._gen_place(stuck_task, d, p)
                            improved = True
                            placed_it = True
                            break

                        # Teacher is busy at (d,p) in some other class — find it
                        busy_cn, busy_task = None, None
                        for cn2 in g['all_classes']:
                            if cn2 in s_cn:
                                continue
                            e2 = grid[cn2][d][p]
                            if not e2:
                                continue
                            if e2.get('teacher') == s_t or e2.get('par_teach','') == s_t:
                                idx2 = g['task_at'][cn2][d][p]
                                if idx2 is not None:
                                    bt = tasks[idx2]
                                    if bt['priority'] not in ('P1_CT', 'P2'):
                                        busy_cn, busy_task = cn2, bt
                                        break

                        if busy_cn is None:
                            continue  # teacher blocked by HC1/HC2, can't swap

                        # Try to find a swap partner in busy_cn at any slot:
                        # A placed task X in busy_cn at slot (d3,p3) such that:
                        #   - X can move to (d,p)  [frees (d3,p3) in busy_cn]
                        #   - busy_task can move to (d3,p3)  [frees teacher at (d,p)]
                        #   - stuck_task can then be placed at (d,p)
                        for d3 in range(wdays):
                            if placed_it:
                                break
                            for p3 in range(ppd):
                                if placed_it or (d3, p3) == (d, p):
                                    continue
                                swap_idx = g['task_at'][busy_cn][d3][p3]
                                if swap_idx is None:
                                    continue
                                swap_task = tasks[swap_idx]
                                if swap_task['priority'] in ('P1_CT', 'P2'):
                                    continue
                                if swap_task['idx'] == busy_task['idx']:
                                    continue
                                # Check: can swap_task go to (d,p)?
                                if not self._gen_can_place(swap_task, d, p,
                                                           ignore_sc1=True, ignore_sc3=True):
                                    continue
                                # Check: can busy_task go to (d3,p3)?
                                if not self._gen_can_place(busy_task, d3, p3,
                                                           ignore_sc1=True, ignore_sc3=True):
                                    continue
                                # Execute 3-way rotation:
                                # unplace swap_task from (d3,p3) and busy_task from (d,p)
                                self._gen_unplace(swap_task, d3, p3)
                                self._gen_unplace(busy_task, d, p)
                                # Verify stuck_task can go to (d,p)
                                if self._gen_can_place(stuck_task, d, p,
                                                       ignore_sc1=True, ignore_sc3=True):
                                    self._gen_place(stuck_task, d, p)
                                    self._gen_place(busy_task, d3, p3)
                                    self._gen_place(swap_task, d, p)
                                    improved = True
                                    placed_it = True
                                    break
                                else:
                                    # Undo
                                    self._gen_place(busy_task, d, p)
                                    self._gen_place(swap_task, d3, p3)

                    if improved:
                        break

                if not improved:
                    break

        if _unplaced() == 0:
            _prog("")
            return '\n\n'.join(relaxed_notes)

        # ─────────────────────────────────────────────────────────────────────
        # STAGE C: Min-Conflicts CSP for any remaining teacher double-bookings
        # and unplaced periods after Stage B.
        #
        # At this point Stage B placed everything teacher-conflict-free, so
        # Stage C mainly handles the (rare) case where teacher capacity is truly
        # exhausted — it tries to reshuffle to pack the remaining periods in.
        # ─────────────────────────────────────────────────────────────────────
        _prog("Stage C — min-conflicts solver…")
        relaxed_notes.append(
            "Min-Conflicts solver applied: soft constraints overridden "
            "to guarantee complete placement.")

        # Force-complete: stuff remaining tasks into ANY teacher-free empty cell
        for task in sorted(tasks, key=lambda t: (-len(t['cn_list']), -t['remaining'])):
            if task['remaining'] <= 0:
                continue
            t_name = task['teacher']
            pt_c   = (task.get('par_teach') or '').strip()
            for d in range(wdays):
                if task['remaining'] <= 0:
                    break
                for p in range(ppd):
                    if task['remaining'] <= 0:
                        break
                    cells_free = all(grid[cn][d][p] is None for cn in task['cn_list'])
                    if not cells_free:
                        continue
                    # Never overwrite HC1
                    hc1 = any(
                        g['task_at'][cn][d][p] is not None and
                        tasks[g['task_at'][cn][d][p]]['priority'] == 'P1_CT'
                        for cn in task['cn_list'])
                    if hc1:
                        continue
                    if not g['t_free'](t_name, d, p):
                        continue
                    if pt_c and pt_c not in ('—', '?') and not g['t_free'](pt_c, d, p):
                        continue
                    self._gen_place(task, d, p)

        if _unplaced() == 0:
            _prog("")
            return '\n\n'.join(relaxed_notes)

        # ── Min-conflicts repair for genuine teacher double-bookings ──────────
        def _slot_conflicts(tname, pt, d, p, own_idx):
            score = 0
            for cn2 in g['all_classes']:
                idx2 = g['task_at'][cn2][d][p]
                if idx2 is None or idx2 == own_idx:
                    continue
                other = tasks[idx2]
                if other['teacher'] == tname:
                    score += 1
                if pt and pt not in ('', '—', '?'):
                    if other['teacher'] == pt or other.get('par_teach', '') == pt:
                        score += 1
            return score

        def _build_task_slots():
            ts = {t['idx']: [] for t in tasks}
            for cn in g['all_classes']:
                for d in range(wdays):
                    for p in range(ppd):
                        idx = g['task_at'][cn][d][p]
                        if idx is not None and (d, p) not in ts[idx]:
                            ts[idx].append((d, p))
            return ts

        def _total_conflicts(task_slots):
            total = 0
            for t in tasks:
                if t['priority'] == 'P1_CT':
                    continue
                pt = t.get('par_teach', '')
                for d, p in task_slots[t['idx']]:
                    total += _slot_conflicts(t['teacher'], pt, d, p, t['idx'])
            return total

        MAX_ITER      = 3000
        RESTART_EVERY = 100
        best_conflicts    = None
        no_improve_count  = 0

        for _iter in range(MAX_ITER):
            task_slots = _build_task_slots()
            total_conf = _total_conflicts(task_slots)

            if _iter % 20 == 0:
                _prog("Stage C — conflicts: {}  (iter {}/{})".format(
                    total_conf, _iter, MAX_ITER))

            if total_conf == 0:
                break

            if best_conflicts is None or total_conf < best_conflicts:
                best_conflicts   = total_conf
                no_improve_count = 0
            else:
                no_improve_count += 1

            if no_improve_count >= RESTART_EVERY:
                _prog("Stage C — restart (stuck at {} conflicts)…".format(total_conf))
                _det_rng.seed(42 + no_improve_count)   # deterministic restarts
                non_hc1 = [t for t in tasks if t['priority'] != 'P1_CT']
                _det_shuffle(non_hc1)
                for t in non_hc1:
                    slots = task_slots[t['idx']]
                    for d, p in slots[:]:
                        for cn in t['cn_list']:
                            grid[cn][d][p] = None
                            g['task_at'][cn][d][p] = None
                        g['t_unmark'](t['teacher'], d, p)
                        pt2 = t.get('par_teach', '')
                        if pt2 and pt2 not in ('', '—', '?'):
                            g['t_unmark'](pt2, d, p)
                        t['remaining'] += 1
                    t_name_r = t['teacher']
                    pt2_r    = (t.get('par_teach') or '').strip()
                    free_slots = [
                        (d2, p2)
                        for d2 in range(wdays) for p2 in range(ppd)
                        if all(grid[cn][d2][p2] is None for cn in t['cn_list'])
                        and g['t_free'](t_name_r, d2, p2)
                        and (not pt2_r or pt2_r in ('—', '?')
                             or g['t_free'](pt2_r, d2, p2))
                    ]
                    _det_shuffle(free_slots)
                    for d, p in free_slots:
                        if t['remaining'] <= 0:
                            break
                        self._gen_place(t, d, p)
                no_improve_count = 0
                best_conflicts   = None
                continue

            # Move the worst-conflicting task's worst slot
            conflicted = []
            for t in tasks:
                if t['priority'] == 'P1_CT':
                    continue
                sc = sum(_slot_conflicts(
                    t['teacher'], t.get('par_teach', ''), d, p, t['idx'])
                    for d, p in task_slots[t['idx']])
                if sc > 0:
                    conflicted.append((t, sc))
            if not conflicted:
                break

            target, _ = max(conflicted, key=lambda x: x[1])
            t_slots = task_slots[target['idx']]
            if not t_slots:
                continue

            worst_d, worst_p = max(
                t_slots,
                key=lambda dp: _slot_conflicts(
                    target['teacher'], target.get('par_teach', ''),
                    dp[0], dp[1], target['idx']))

            for cn in target['cn_list']:
                grid[cn][worst_d][worst_p] = None
                g['task_at'][cn][worst_d][worst_p] = None
            g['t_unmark'](target['teacher'], worst_d, worst_p)
            pt = target.get('par_teach', '')
            if pt and pt not in ('', '—', '?'):
                g['t_unmark'](pt, worst_d, worst_p)
            target['remaining'] += 1

            best_score = None
            best_d, best_p = None, None
            for d in range(wdays):
                for p in range(ppd):
                    cells_free = all(grid[cn][d][p] is None for cn in target['cn_list'])
                    if not cells_free:
                        continue
                    sc = _slot_conflicts(target['teacher'], pt, d, p, target['idx'])
                    if best_score is None or sc < best_score:
                        best_score = sc
                        best_d, best_p = d, p
                    if best_score == 0:
                        break
                if best_score == 0:
                    break

            if best_d is None or best_score >= _slot_conflicts(
                    target['teacher'], pt, worst_d, worst_p, target['idx']):
                best_d, best_p = worst_d, worst_p

            self._gen_place(target, best_d, best_p)

        _prog("")
        return '\n\n'.join(relaxed_notes) if relaxed_notes else None

    # ── Snapshot / Restore (for undo-on-no-improvement) ──────────────────────

    def _ft_snapshot(self):
        """
        Deep-copy all mutable gen state so that any action can be fully undone.

        Captures:
          - grid cells (dict of CN → list-of-lists of cell dicts)
          - task_at   (dict of CN → list-of-lists of idx or None)
          - t_busy    (dict of teacher → set of (d, p))
          - per-task mutable fields
          - relaxed key sets
        """
        import copy
        g = self._gen

        # Grid: each cell is either None or a small dict — shallow copy of the dict
        # is sufficient because cells are replaced wholesale (never mutated in place).
        grid_snap = {
            cn: [[g['grid'][cn][d][p] for p in range(g['ppd'])]
                 for d in range(g['wdays'])]
            for cn in g['all_classes']
        }
        task_at_snap = {
            cn: [[g['task_at'][cn][d][p] for p in range(g['ppd'])]
                 for d in range(g['wdays'])]
            for cn in g['all_classes']
        }
        t_busy_snap = {t: set(s) for t, s in g['t_busy'].items()}

        tasks_snap = [
            {
                'idx':       task['idx'],
                'remaining': task['remaining'],
                'rx_sc1':    task['rx_sc1'],
                'rx_sc2':    task['rx_sc2'],
                'rx_sc3':    task['rx_sc3'],
                'p_pref':    list(task['p_pref']),
                'd_pref':    list(task['d_pref']),
                'daily':     task['daily'],
                'priority':  task['priority'],
                'consec':    task['consec'],
            }
            for task in g['tasks']
        ]

        return {
            'grid':               grid_snap,
            'task_at':            task_at_snap,
            't_busy':             t_busy_snap,
            'tasks':              tasks_snap,
            'relaxed_consec':     set(self._relaxed_consec_keys),
            'relaxed_main':       set(getattr(self, '_relaxed_main_keys', set())),
        }

    def _ft_restore(self, snap):
        """Restore gen state from a snapshot produced by _ft_snapshot."""
        g = self._gen

        # Restore grid and task_at
        for cn in g['all_classes']:
            for d in range(g['wdays']):
                for p in range(g['ppd']):
                    g['grid'][cn][d][p]    = snap['grid'][cn][d][p]
                    g['task_at'][cn][d][p] = snap['task_at'][cn][d][p]

        # Restore t_busy
        g['t_busy'].clear()
        for t, s in snap['t_busy'].items():
            g['t_busy'][t] = set(s)

        # Restore per-task fields
        task_map = {t['idx']: t for t in g['tasks']}
        for ts in snap['tasks']:
            task = task_map.get(ts['idx'])
            if task is None:
                continue
            task['remaining'] = ts['remaining']
            task['rx_sc1']    = ts['rx_sc1']
            task['rx_sc2']    = ts['rx_sc2']
            task['rx_sc3']    = ts['rx_sc3']
            task['p_pref']    = list(ts['p_pref'])
            task['d_pref']    = list(ts['d_pref'])
            task['daily']     = ts['daily']
            task['priority']  = ts['priority']
            task['consec']    = ts['consec']

        # Restore relaxed-key sets
        self._relaxed_consec_keys = set(snap['relaxed_consec'])
        if not hasattr(self, '_relaxed_main_keys'):
            self._relaxed_main_keys = set()
        self._relaxed_main_keys.clear()
        self._relaxed_main_keys.update(snap['relaxed_main'])

    def _ft_targetable(self):
        """Return tasks that are candidates for re-allocation (filler / consec / parallel)."""
        result = []
        for t in self._gen['tasks']:
            p = t['priority']
            if p in ('filler', 'P5') or t['type'] in ('parallel', 'combined_parallel'):
                result.append(t)
        return result

    def _ft_teacher_free_slots(self):
        """Return {teacher: count_of_free_usable_slots}."""
        g = self._gen
        wdays = g['wdays']; ppd = g['ppd']
        t_busy = g['t_busy']
        t_unavail = g['t_unavail']
        teachers = set()
        for task in g['tasks']:
            if task['teacher']:
                teachers.add(task['teacher'])
            if task['par_teach']:
                teachers.add(task['par_teach'])
        result = {}
        for t in teachers:
            busy = t_busy.get(t, set())
            free = 0
            for d in range(wdays):
                for p in range(ppd):
                    if (d, p) not in busy and not t_unavail(t, d, p):
                        free += 1
            result[t] = free
        return result

    def _ft_try_place_task(self, task, ignore_sc1=False, ignore_sc3=False):
        """Greedily place all remaining slots of *task*. Returns count newly placed."""
        g = self._gen
        grid  = g['grid']
        wdays = g['wdays']; ppd = g['ppd']
        placed = 0

        def _subj_on_day(d_):
            return max(
                (sum(1 for pp in range(ppd)
                     if grid[cn_][d_][pp] is not None
                     and grid[cn_][d_][pp].get('subject') == task['subject'])
                 for cn_ in task['cn_list']),
                default=0,
            )

        # Sort days so those without the subject yet come first
        day_order = sorted(range(wdays), key=lambda d_: (_subj_on_day(d_), d_))
        for d in day_order:
            if task['remaining'] == 0:
                break
            for p in range(ppd):
                if task['remaining'] == 0:
                    break
                if self._gen_can_place(task, d, p,
                                       ignore_sc1=ignore_sc1,
                                       ignore_sc3=ignore_sc3):
                    self._gen_place(task, d, p)
                    placed += 1
        return placed

    def _ft_unplace_task(self, task):
        """Remove all placed slots of *task* from the grid and return count removed."""
        g = self._gen
        grid = g['grid']; wdays = g['wdays']; ppd = g['ppd']
        removed = 0
        for d in range(wdays):
            for p in range(ppd):
                for cn in task['cn_list']:
                    if g['task_at'][cn][d][p] == task['idx']:
                        self._gen_unplace(task, d, p)
                        removed += 1
                        break   # one unplace per (d,p) slot is enough
        return removed

    # ── Post-generation conflict cleanup ─────────────────────────────────────

    def _remove_teacher_conflicts(self):
        """
        Post-generation safety pass: detect and remove teacher double-bookings.

        A double-booking occurs when the same teacher appears in two or more
        DIFFERENT classes at the same (day, period) slot — which is physically
        impossible.

        Combined-class tasks (combined_classes field is set) legitimately write
        the same teacher to multiple class cells and are skipped.

        After removal, t_busy is rebuilt from scratch so subsequent operations
        see a consistent state.

        Returns: list of (teacher, removed_class, kept_class, day_idx, period_idx)
        """
        if self._gen is None:
            return []

        g     = self._gen
        grid  = g['grid']
        tasks = g['tasks']
        wdays = g['wdays']
        ppd   = g['ppd']

        conflicts_removed = []

        for d in range(wdays):
            for p in range(ppd):
                teacher_entries = {}   # tname -> list of cn
                for cn in g['all_classes']:
                    e = grid[cn][d][p]
                    if e is None:
                        continue
                    for tname in [e.get('teacher', ''), e.get('par_teach', '')]:
                        if not tname or tname in ('', '—', '?'):
                            continue
                        teacher_entries.setdefault(tname, []).append(cn)

                for tname, classes in teacher_entries.items():
                    if len(classes) <= 1:
                        continue

                    # Group by combined_classes set — same group = legitimate
                    groups = {}
                    for cn in classes:
                        e  = grid[cn][d][p]
                        cc = frozenset(e.get('combined_classes', [])) if e else frozenset()
                        key = cc if len(cc) > 1 else frozenset([cn])
                        groups.setdefault(key, []).append(cn)

                    if len(groups) <= 1:
                        continue   # all same combined group — no real conflict

                    # Multiple groups = real conflict.
                    # ALWAYS prefer keeping the HC1/CT group — a teacher
                    # double-booking must never erase a CT period.
                    order = {cn: i for i, cn in enumerate(g['all_classes'])}

                    def _group_has_hc1(grp_keys):
                        for ck in grp_keys:
                            for cn2 in groups.get(ck, []):
                                idx2 = g['task_at'][cn2][d][p]
                                if idx2 is not None and tasks[idx2].get('is_ct'):
                                    return True
                        return False

                    # Sort: HC1 groups first, then by size desc, then by class order
                    sorted_groups = sorted(
                        groups.items(),
                        key=lambda kv: (
                            0 if any(                     # HC1 group wins
                                g['task_at'].get(cn2, [[None]*ppd]*wdays)[d][p] is not None
                                and tasks[g['task_at'][cn2][d][p]].get('is_ct')
                                for cn2 in kv[1]
                                if cn2 in g['task_at']
                                   and d < len(g['task_at'][cn2])
                                   and g['task_at'][cn2][d][p] is not None
                            ) else 1,
                            -len(kv[1]),
                            min(order.get(c, 999) for c in kv[1])
                        )
                    )
                    keep_classes = sorted_groups[0][1]

                    for _, grp_classes in sorted_groups[1:]:
                        for cn in grp_classes:
                            task_idx = g['task_at'][cn][d][p]
                            grid[cn][d][p]         = None
                            g['task_at'][cn][d][p] = None
                            if task_idx is not None and task_idx < len(tasks):
                                tasks[task_idx]['remaining'] += 1
                            conflicts_removed.append(
                                (tname, cn, keep_classes[0], d, p))

        # Rebuild t_busy from the clean grid
        g['t_busy'].clear()
        for cn in g['all_classes']:
            for d in range(wdays):
                for p in range(ppd):
                    e = grid[cn][d][p]
                    if e:
                        t  = e.get('teacher',   '')
                        pt = e.get('par_teach', '')
                        if t  and t  not in ('', '—', '?'):
                            g['t_busy'].setdefault(t,  set()).add((d, p))
                        if pt and pt not in ('', '—', '?'):
                            g['t_busy'].setdefault(pt, set()).add((d, p))

        return conflicts_removed

    # ── Task A: Allocate ──────────────────────────────────────────────────────

    def _task_allocate(self):
        """
        Task A — Smart allocation of unplaced filler/consecutive/parallel periods.

        1. Compute free-slot counts for every teacher.
        2. Allocate tasks whose teacher has fewest free slots first (most constrained).
        3. Second pass to catch anything still unplaced.

        Returns a summary string (or None).
        """
        targets = self._ft_targetable()
        if not targets:
            return None

        from collections import defaultdict
        t_tasks = defaultdict(list)
        for task in targets:
            if task['remaining'] > 0:
                t_tasks[task['teacher']].append(task)

        if not t_tasks:
            return None

        free_counts = self._ft_teacher_free_slots()
        ordered_teachers = sorted(t_tasks.keys(), key=lambda t: free_counts.get(t, 0))

        for teacher in ordered_teachers:
            for task in t_tasks[teacher]:
                if task['remaining'] > 0:
                    self._ft_try_place_task(task)

        # Second pass
        for task in targets:
            if task['remaining'] > 0:
                self._ft_try_place_task(task)

        return None

    # ── Task S: Shuffle ───────────────────────────────────────────────────────

    def _task_shuffle(self):
        """
        Task S — Unplace all targetable tasks, re-sort by constraint difficulty, re-place.

        Sort order (most constrained first):
          1. Consecutive (SC1)
          2. Combined parallel
          3. Parallel
          4. Daily filler
          5. Standard filler

        Returns a summary string (or None).
        """
        targets = self._ft_targetable()
        if not targets:
            return None

        for task in targets:
            self._ft_unplace_task(task)

        def _sort_key(t):
            if t['consec']:            return (0, -t['periods'])
            if t['type'] == 'combined_parallel': return (1, -t['periods'])
            if t['type'] == 'parallel':          return (2, -t['periods'])
            if t.get('daily'):                   return (3, -t['periods'])
            return (4, -t['periods'])

        free_counts = self._ft_teacher_free_slots()
        targets_sorted = sorted(
            targets,
            key=lambda t: (_sort_key(t), free_counts.get(t['teacher'], 0)))

        for task in targets_sorted:
            if task['remaining'] > 0:
                self._ft_try_place_task(task)

        # Final sweep
        for task in targets_sorted:
            if task['remaining'] > 0:
                self._ft_try_place_task(task)

        return None

    # ── Task C: Relax Consecutive ─────────────────────────────────────────────

    def _task_relax_consecutive(self):
        """
        Task C — Relax consecutive constraints for unplaced SC1 tasks.

        For each task with consec=True and remaining > 0:
          - Set rx_sc1=True (bypass the consecutive-placement rule)
          - Try placing freely
        Returns a summary string listing which rules were relaxed.
        """
        relaxed = []
        for task in self._gen['tasks']:
            if task['consec'] and task['remaining'] > 0:
                task['rx_sc1'] = True
                placed = self._ft_try_place_task(task, ignore_sc1=True)
                if placed > 0:
                    cn = '+'.join(task['cn_list'])
                    relaxed.append((cn, task['subject']))
                    for cn_i in task['cn_list']:
                        self._relaxed_consec_keys.add((cn_i, task['subject']))

        # Second pass
        for task in self._gen['tasks']:
            if task.get('rx_sc1') and task['remaining'] > 0:
                self._ft_try_place_task(task, ignore_sc1=True)

        if relaxed:
            return ("Consecutive rules relaxed for:\n\n" +
                    "\n".join("• {} → {}".format(cn, subj)
                              for cn, subj in sorted(relaxed)))
        return None

    # ── Task M: Relax Main Periods ────────────────────────────────────────────

    def _task_relax_main_periods(self):
        """
        Task M — Convert unplaced main-period tasks to filler, then run Task A.

        Targets: HC2 tasks (period/day preferences) and SC2 tasks (daily) that
        still have remaining > 0.  CT periods (HC1) are never touched.

        Returns a summary string listing which tasks were converted.
        """
        if not hasattr(self, '_relaxed_main_keys'):
            self._relaxed_main_keys = set()

        relaxed = []
        for task in self._gen['tasks']:
            if task['remaining'] == 0 or task['is_ct']:
                continue
            if not (task['p_pref'] or task['d_pref'] or task.get('daily')):
                continue

            task['p_pref']   = []
            task['d_pref']   = []
            task['daily']    = False
            task['priority'] = 'filler'
            cn = '+'.join(task['cn_list'])
            relaxed.append((cn, task['subject']))
            self._relaxed_main_keys.add((frozenset(task['cn_list']), task['subject']))

        if relaxed:
            self._task_allocate()
            return ("Main periods converted to filler for:\n\n" +
                    "\n".join("• {} → {}".format(cn, subj)
                              for cn, subj in sorted(relaxed)))
        return None

    # ── Task UN: Relax Unavailability ─────────────────────────────────────────

    def _task_relax_unavailability(self):
        """
        Task UN — Override unavailability rules for teachers with unplaced periods,
        then run Task A.

        Returns a summary string listing which teachers' unavailability was bypassed.
        """
        unplaced_teachers = set()
        for task in self._gen['tasks']:
            if task['remaining'] > 0:
                if task['teacher']:
                    unplaced_teachers.add(task['teacher'])
                if task['par_teach']:
                    unplaced_teachers.add(task['par_teach'])

        if not unplaced_teachers:
            return None

        # Mark rx_sc3 on all tasks for those teachers
        for task in self._gen['tasks']:
            if task['remaining'] > 0:
                if (task['teacher'] in unplaced_teachers or
                        task['par_teach'] in unplaced_teachers):
                    task['rx_sc3'] = True

        # Run allocation with unavailability bypassed
        from collections import defaultdict
        targets = self._ft_targetable()
        free_counts = self._ft_teacher_free_slots()
        t_tasks = defaultdict(list)
        for task in targets:
            if task['remaining'] > 0:
                t_tasks[task['teacher']].append(task)

        ordered = sorted(t_tasks.keys(), key=lambda t: free_counts.get(t, 0))
        for teacher in ordered:
            for task in t_tasks[teacher]:
                if task['remaining'] > 0:
                    self._ft_try_place_task(task,
                                            ignore_sc3=task.get('rx_sc3', False))

        # Also try non-targetable tasks with rx_sc3
        for task in self._gen['tasks']:
            if task['remaining'] > 0 and task.get('rx_sc3'):
                self._ft_try_place_task(task, ignore_sc3=True)

        return ("Unavailability rules bypassed for:\n\n" +
                "\n".join("• {}".format(t) for t in sorted(unplaced_teachers)))

    # ── Class timetable view ──────────────────────────────────────────────

    def _get_combined_par_display(self, cn, e):
        """For a combined_parallel cell, return (line1, line2) where:
            line1 = "combined_subject / class_subject"   e.g. "URDU / SKT"
            line2 = "combined_teacher / class_teacher"   e.g. "Irfan / Anita"

        Uses step3_data as the ground truth for which teacher+subject is the
        combined one, then looks up class_config_data[cn] for the per-class entry.
        """
        cc = e.get('combined_classes', [])

        # ── Step 1: find combined teacher + subject from step3_data ──────────
        combined_teacher = ''
        combined_subj    = ''
        s3 = getattr(self, 'step3_data', {})
        for _t, s3d in s3.items():
            for cb in s3d.get('combines', []):
                if set(cb.get('classes', [])) == set(cc):
                    combined_teacher = _t
                    combined_subj    = cb.get('subjects', [''])[0] if cb.get('subjects') else ''
                    break
            if combined_teacher:
                break

        # ── Step 2: look up this class's entry to find the class-specific side ─
        class_subj    = ''
        class_teacher = ''
        if combined_subj and cn in self.class_config_data:
            for _s in self.class_config_data[cn].get('subjects', []):
                sname = _s.get('name', '').strip()
                pname = (_s.get('parallel_subject') or '').strip()
                if sname == combined_subj:
                    # primary = combined, parallel = class-specific
                    class_subj    = pname
                    class_teacher = (_s.get('parallel_teacher') or '').strip()
                    break
                elif pname == combined_subj:
                    # parallel = combined, primary = class-specific
                    class_subj    = sname
                    class_teacher = _s.get('teacher', '').strip()
                    break

        # ── Fallback: use cell data if step3_data lookup failed ──────────────
        if not combined_subj:
            combined_subj    = e.get('subject', '')
            combined_teacher = e.get('teacher', '')
            class_subj       = e.get('par_subj', '')
            class_teacher    = e.get('par_teach', '')

        return (
            "{} / {}".format(combined_subj, class_subj),
            "{} / {}".format(combined_teacher, class_teacher),
        )


    def get_excel_bytes(self, mode):
        """Generate Excel workbook and return raw bytes (for Streamlit download)."""
        import io
        buf = io.BytesIO()
        self._write_excel_buf(buf, mode)
        return buf.getvalue()


    def _write_excel_buf(self, filename, mode):
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")

        from collections import defaultdict

        tt          = self._timetable
        days        = tt['days']
        ppd         = tt['ppd']
        half1       = tt['half1']
        grid        = tt['grid']
        all_classes = tt['all_classes']

        def _fill(h): return PatternFill("solid", fgColor=h.lstrip("#"))
        def _font(bold=False, sz=9, col="000000"):
            return Font(bold=bold, size=sz, color=col.lstrip("#"), name="Arial")
        def _border():
            s = Side(style="thin", color="AAAAAA")
            return Border(left=s, right=s, top=s, bottom=s)
        def _align(h="center", wrap=True):
            return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

        HDR_F  = _fill("#2c3e50"); HDR_N  = _font(True, 10, "FFFFFF")
        DAY_F  = _fill("#34495e"); DAY_N  = _font(True,  9, "FFFFFF")
        SUB_F  = _fill("#d5e8d4")
        COMB_F = _fill("#dae8fc")
        PAR_F  = _fill("#ffe6cc")
        CPAF   = _fill("#f8cecc")
        FREE_F = _fill("#f5f5f5")
        WHT_F  = _fill("#FFFFFF")
        SUM_F  = _fill("#eaf2ff")
        CT_H_F = _fill("#1a5276")
        WRN_F  = _fill("#fdebd0")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # ── shared: build teacher grid ───────────────────────────────────────
        def _build_tg():
            tg = {}

            def _add(tname, tcls, tsubj, tct, _d=None, _p=None):
                """Insert one entry into the teacher grid, merging if already occupied."""
                if not tname or tname in ('—', '?'):
                    return
                _dd = d if _d is None else _d
                _pp = p if _p is None else _p
                tg.setdefault(tname, [[None]*ppd for _ in range(len(days))])
                existing = tg[tname][_dd][_pp]
                if existing is not None:
                    merged = existing['class']
                    if tcls not in merged.split('+'):
                        merged += '+' + tcls
                    tg[tname][_dd][_pp] = {
                        'class':   merged,
                        'subject': (existing['subject'] + '/' + tsubj
                                    if tsubj != existing['subject'] else tsubj),
                        'is_ct':   tct or existing.get('is_ct', False),
                    }
                else:
                    tg[tname][_dd][_pp] = {
                        'class': tcls, 'subject': tsubj, 'is_ct': tct}

            def _cp_teachers(cn_local, e_local, cc_local):
                """Return (comb_teacher, comb_subj, class_teacher, class_subj) for a
                combined_parallel cell, using step3_data + class_config_data as the
                authoritative source (same logic as _get_combined_par_display)."""
                comb_teacher = comb_subj = cls_teacher = cls_subj = ''

                s3 = getattr(self, 'step3_data', {})
                for _t, s3d in s3.items():
                    for cb in s3d.get('combines', []):
                        if set(cb.get('classes', [])) == set(cc_local):
                            comb_teacher = _t
                            comb_subj    = (cb.get('subjects', [''])[0]
                                            if cb.get('subjects') else '')
                            break
                    if comb_teacher:
                        break

                if comb_subj and cn_local in self.class_config_data:
                    for _s in self.class_config_data[cn_local].get('subjects', []):
                        sname  = _s.get('name', '').strip()
                        pname  = (_s.get('parallel_subject') or '').strip()
                        if sname == comb_subj:
                            cls_subj    = pname
                            cls_teacher = (_s.get('parallel_teacher') or '').strip()
                            break
                        elif pname == comb_subj:
                            cls_subj    = sname
                            cls_teacher = _s.get('teacher', '').strip()
                            break

                # Fallback to cell data
                if not comb_subj:
                    comb_teacher = e_local.get('teacher', '')
                    comb_subj    = e_local.get('subject', '')
                    cls_teacher  = e_local.get('par_teach', '')
                    cls_subj     = e_local.get('par_subj', '')

                return comb_teacher, comb_subj, cls_teacher, cls_subj

            for cn in all_classes:
                for d in range(len(days)):
                    for p in range(ppd):
                        e = grid.get(cn, [[]])[d][p] \
                            if d < len(grid.get(cn, [])) else None
                        if not e:
                            continue
                        etype = e.get('type', 'normal')
                        cc    = e.get('combined_classes', [])
                        is_cp = etype == 'combined_parallel'
                        is_c  = bool(cc) and etype == 'combined'

                        if is_cp:
                            comb_t, comb_s, cls_t, cls_s = _cp_teachers(cn, e, cc)
                            cls_label = '+'.join(cc) if cc else cn
                            # Write combined-teacher entry, upgrading if already partial.
                            # cc[0]'s cell may have been retyped to 'normal' by place_slot,
                            # so it would have written teacher→'11A' (just one class).
                            # When we later encounter the combined cell at cc[1+], we must
                            # upgrade the label to the full '11A+11B' string.
                            # Determine is_ct: true if combined teacher is CT for any class
                            _comb_is_ct = e.get('is_ct', False)
                            if comb_t and comb_t not in ('—', '?', ''):
                                tg.setdefault(comb_t, [[None]*ppd for _ in range(len(days))])
                                existing = tg[comb_t][d][p]
                                if existing is None:
                                    tg[comb_t][d][p] = {
                                        'class': cls_label, 'subject': comb_s,
                                        'is_ct': _comb_is_ct}
                                else:
                                    # Upgrade class label: merge any missing classes
                                    existing_classes = set(existing['class'].split('+'))
                                    full_classes     = set(cc) if cc else {cn}
                                    merged = '+'.join(sorted(existing_classes | full_classes))
                                    tg[comb_t][d][p] = dict(existing, **{
                                        'class':   merged,
                                        'subject': comb_s or existing['subject'],
                                        'is_ct':   _comb_is_ct or existing.get('is_ct', False),
                                    })
                            # Per-class parallel teacher always gets their own entry
                            if cls_t and cls_t not in ('—', '?', ''):
                                _cls_is_ct = e.get('is_ct', False)
                                _add(cls_t, cn, cls_s, _cls_is_ct)

                        elif is_c:
                            t_name    = e.get('teacher', '')
                            cls_label = '+'.join(cc) if cc else cn
                            if t_name and t_name not in ('—', '?', ''):
                                tg.setdefault(t_name, [[None]*ppd for _ in range(len(days))])
                                existing = tg[t_name][d][p]
                                if existing is None:
                                    tg[t_name][d][p] = {
                                        'class':   cls_label,
                                        'subject': e.get('subject', ''),
                                        'is_ct':   e.get('is_ct', False)}
                                else:
                                    # Upgrade partial class label
                                    existing_classes = set(existing['class'].split('+'))
                                    full_classes     = set(cc) if cc else {cn}
                                    merged = '+'.join(sorted(existing_classes | full_classes))
                                    tg[t_name][d][p] = dict(existing, **{'class': merged})

                        else:
                            # Cross-check is_ct using class config (belt-and-suspenders)
                            _cfg_cn  = self.class_config_data.get(cn, {})
                            _ct_t    = _cfg_cn.get('teacher', '').strip()
                            _ct_pi   = int(_cfg_cn.get('teacher_period', 0)) - 1
                            _ct_subs = {_s['name'] for _s in _cfg_cn.get('subjects', [])
                                        if _s.get('teacher', '').strip() == _ct_t}
                            _ict = (e.get('is_ct', False) or
                                    (p == _ct_pi and
                                     e.get('subject', '') in _ct_subs and
                                     e.get('teacher', '').strip() == _ct_t))
                            _add(e.get('teacher'), cn, e.get('subject', ''), _ict)
                            pt = e.get('par_teach', '')
                            if pt and pt not in ('—', '?', ''):
                                _add(pt, cn, e.get('par_subj', ''), False)

            return tg

        def _sv(val):
            """Safely extract string from a StringVar or plain string."""
            if hasattr(val, 'get'):
                return val.get()
            return val or ''

        def _ct_map():
            ct = {}
            for cn in all_classes:
                cfg = self.class_config_data.get(cn, {})
                t = cfg.get('teacher', '').strip()
                if t:
                    ct.setdefault(t, []).append(cn)
            return ct

        # ─────────────────────────────────────────────────────────────────────
        # 1. CLASSWISE TIMETABLE — one sheet per class
        # ─────────────────────────────────────────────────────────────────────
        if mode == "class":
            for cn in all_classes:
                ws = wb.create_sheet(cn)
                cfg     = self.class_config_data.get(cn, {})
                ct_name = cfg.get('teacher', '').strip()
                ct_per  = str(cfg.get('teacher_period', ''))
                hdr_txt = "Class: {}   |   Class Teacher: {}{}".format(
                    cn, ct_name or '—',
                    "   |   CT Period: {}".format(ct_per) if ct_per else '')

                ws.merge_cells(start_row=1, start_column=1,
                               end_row=1, end_column=ppd+1)
                c = ws.cell(1, 1, hdr_txt)
                c.fill = CT_H_F; c.font = _font(True, 11, "FFFFFF")
                c.alignment = _align(); c.border = _border()
                ws.row_dimensions[1].height = 20

                ws.cell(2, 1, "Day")
                ws.cell(2, 1).fill = HDR_F; ws.cell(2, 1).font = HDR_N
                ws.cell(2, 1).alignment = _align(); ws.cell(2, 1).border = _border()
                for p in range(ppd):
                    h = ws.cell(2, p+2, "P{} {}".format(
                        p+1, "①" if p < half1 else "②"))
                    h.fill = HDR_F; h.font = HDR_N
                    h.alignment = _align(); h.border = _border()
                ws.row_dimensions[2].height = 16

                # Cross-check is_ct by config (belt-and-suspenders for this class)
                _pcls_ct_t    = cfg.get('teacher', '').strip()
                _pcls_ct_pi   = int(cfg.get('teacher_period', 0)) - 1
                _pcls_ct_subs = {_s['name'] for _s in cfg.get('subjects', [])
                                 if _s.get('teacher', '').strip() == _pcls_ct_t}

                def _pcls_is_ct(e_cell, p_idx):
                    if e_cell is None: return False
                    if e_cell.get('is_ct'): return True
                    return (p_idx == _pcls_ct_pi and
                            e_cell.get('subject', '') in _pcls_ct_subs)

                for d, dname in enumerate(days):
                    r = 3 + d
                    ws.row_dimensions[r].height = 48
                    dc = ws.cell(r, 1, dname)
                    dc.fill = DAY_F; dc.font = DAY_N
                    dc.alignment = _align(); dc.border = _border()
                    for p in range(ppd):
                        e = grid.get(cn, [[]])[d][p] \
                            if d < len(grid.get(cn, [])) else None
                        if e is None:
                            txt = "FREE"; fill = FREE_F
                        else:
                            etype  = e.get('type', 'normal')
                            ict    = _pcls_is_ct(e, p)  # config-based check
                            if etype == 'combined_parallel':
                                l1, l2 = self._get_combined_par_display(cn, e)
                                txt = "{}\n{}".format(l1, l2); fill = CPAF
                            elif etype == 'parallel':
                                txt = "{} / {}\n{} / {}".format(
                                    e['subject'], e.get('par_subj',''),
                                    e['teacher'],  e.get('par_teach',''))
                                fill = PAR_F
                            elif etype == 'combined':
                                cc = e.get('combined_classes', [])
                                mark = " ★" if ict else ""
                                txt = "{}{}[{}]\n{}".format(
                                    e['subject'], mark, '+'.join(cc), e['teacher'])
                                fill = COMB_F
                            else:
                                mark = " ★" if ict else ""
                                txt  = "{}{}\n{}".format(e['subject'], mark, e['teacher'])
                                fill = SUB_F if ict else WHT_F
                        c = ws.cell(r, p+2, txt)
                        c.fill = fill; c.alignment = _align()
                        c.border = _border(); c.font = _font(sz=8)

                # Summary
                sr = 3 + len(days) + 1
                ws.merge_cells(start_row=sr, start_column=1,
                               end_row=sr, end_column=ppd+1)
                c = ws.cell(sr, 1, "Summary — {}".format(cn))
                c.fill = HDR_F; c.font = HDR_N
                c.alignment = _align("left"); c.border = _border()

                smry = defaultdict(int)
                for d in range(len(days)):
                    for p in range(ppd):
                        e = grid.get(cn, [[]])[d][p] \
                            if d < len(grid.get(cn, [])) else None
                        if not e: continue
                        etype = e.get('type', 'normal')
                        if etype == 'combined_parallel':
                            l1, l2 = self._get_combined_par_display(cn, e)
                            for ln in (l1, l2):
                                parts = ln.split('\n')
                                smry[(parts[0].strip(),
                                      parts[1].strip() if len(parts) > 1 else '')] += 1
                        elif etype == 'parallel':
                            smry[(e['subject'], e['teacher'])] += 1
                            smry[(e.get('par_subj',''), e.get('par_teach',''))] += 1
                        else:
                            smry[(e['subject'], e['teacher'])] += 1

                hdr_r = sr + 1
                for col, txt in enumerate(["Subject", "Teacher", "Periods/Week"], 1):
                    c = ws.cell(hdr_r, col, txt)
                    c.fill = HDR_F; c.font = HDR_N
                    c.alignment = _align(); c.border = _border()

                for i, ((subj, teach), cnt) in enumerate(
                        sorted(smry.items())):
                    row = hdr_r + 1 + i
                    for col, val in enumerate([subj, teach, cnt], 1):
                        c = ws.cell(row, col, val)
                        c.fill = SUM_F if i % 2 == 0 else WHT_F
                        c.alignment = _align(); c.border = _border()
                        c.font = _font(sz=9)

                ws.column_dimensions["A"].width = 12
                for p in range(ppd):
                    ws.column_dimensions[get_column_letter(p+2)].width = 20

        # ─────────────────────────────────────────────────────────────────────
        # 2. TEACHERWISE TIMETABLE — one sheet per teacher
        # ─────────────────────────────────────────────────────────────────────
        elif mode == "teacher":
            tg    = _build_tg()
            ct_mp = _ct_map()

            for teacher in sorted(tg.keys()):
                ws    = wb.create_sheet(teacher[:31])
                tdata = tg[teacher]
                ctc   = ct_mp.get(teacher, [])
                hdr_txt = "Teacher: {}   |   Class Teacher of: {}".format(
                    teacher, ', '.join(ctc) if ctc else '—')

                ws.merge_cells(start_row=1, start_column=1,
                               end_row=1, end_column=ppd+1)
                c = ws.cell(1, 1, hdr_txt)
                c.fill = CT_H_F; c.font = _font(True, 11, "FFFFFF")
                c.alignment = _align(); c.border = _border()
                ws.row_dimensions[1].height = 20

                ws.cell(2, 1, "Day")
                ws.cell(2, 1).fill = HDR_F; ws.cell(2, 1).font = HDR_N
                ws.cell(2, 1).alignment = _align(); ws.cell(2, 1).border = _border()
                for p in range(ppd):
                    h = ws.cell(2, p+2, "P{} {}".format(
                        p+1, "①" if p < half1 else "②"))
                    h.fill = HDR_F; h.font = HDR_N
                    h.alignment = _align(); h.border = _border()
                ws.row_dimensions[2].height = 16

                for d, dname in enumerate(days):
                    r = 3 + d
                    ws.row_dimensions[r].height = 48
                    dc = ws.cell(r, 1, dname)
                    dc.fill = DAY_F; dc.font = DAY_N
                    dc.alignment = _align(); dc.border = _border()
                    for p in range(ppd):
                        e = tdata[d][p] if d < len(tdata) else None
                        if e is None:
                            txt = "FREE"; fill = FREE_F
                        else:
                            txt  = "{}\n{}".format(e['class'], e['subject'])
                            fill = SUB_F if e.get('is_ct') else WHT_F
                        c = ws.cell(r, p+2, txt)
                        c.fill = fill; c.alignment = _align()
                        c.border = _border(); c.font = _font(sz=8)

                # Summary: class → subject → count
                sr    = 3 + len(days) + 1
                ws.merge_cells(start_row=sr, start_column=1,
                               end_row=sr, end_column=ppd+1)
                c = ws.cell(sr, 1, "Summary — {}".format(teacher))
                c.fill = HDR_F; c.font = HDR_N
                c.alignment = _align("left"); c.border = _border()

                smry  = defaultdict(lambda: defaultdict(int))
                total = 0
                for d in range(len(days)):
                    for p in range(ppd):
                        e = tdata[d][p] if d < len(tdata) else None
                        if e:
                            smry[e['class']][e['subject']] += 1
                            total += 1

                hdr_r = sr + 1
                for col, txt in enumerate(["Class", "Subject", "Periods/Week"], 1):
                    c = ws.cell(hdr_r, col, txt)
                    c.fill = HDR_F; c.font = HDR_N
                    c.alignment = _align(); c.border = _border()

                row = hdr_r + 1
                for cls in sorted(smry.keys()):
                    for subj, cnt in sorted(smry[cls].items()):
                        for col, val in enumerate([cls, subj, cnt], 1):
                            c = ws.cell(row, col, val)
                            c.fill = SUM_F if row % 2 == 0 else WHT_F
                            c.alignment = _align(); c.border = _border()
                            c.font = _font(sz=9)
                        row += 1

                for col, val in enumerate(["", "TOTAL", total], 1):
                    c = ws.cell(row, col, val)
                    c.fill = _fill("#d4e6f1"); c.font = _font(True, 9)
                    c.alignment = _align(); c.border = _border()

                ws.column_dimensions["A"].width = 12
                for p in range(ppd):
                    ws.column_dimensions[get_column_letter(p+2)].width = 20

        # ─────────────────────────────────────────────────────────────────────
        # 3. CLASS TEACHER LIST
        # ─────────────────────────────────────────────────────────────────────
        elif mode == "ct_list":
            ws = wb.create_sheet("Class Teacher List")
            ws.merge_cells("A1:C1")
            c = ws["A1"]; c.value = "Class Teacher List"
            c.fill = HDR_F; c.font = _font(True, 13, "FFFFFF")
            c.alignment = _align(); c.border = _border()
            ws.row_dimensions[1].height = 22

            for col, txt in enumerate(["Class", "Class Teacher", "CT Period"], 1):
                c = ws.cell(2, col, txt)
                c.fill = DAY_F; c.font = DAY_N
                c.alignment = _align(); c.border = _border()

            for i, cn in enumerate(all_classes):
                cfg     = self.class_config_data.get(cn, {})
                ct_name = cfg.get('teacher', '').strip() or '—'
                ct_per  = str(cfg.get('teacher_period', '')) or '—'
                row = 3 + i
                for col, val in enumerate([cn, ct_name, ct_per], 1):
                    c = ws.cell(row, col, val)
                    c.fill = SUM_F if i % 2 == 0 else WHT_F
                    c.alignment = _align(); c.border = _border()
                    c.font = _font(sz=10)

            ws.column_dimensions["A"].width = 14
            ws.column_dimensions["B"].width = 28
            ws.column_dimensions["C"].width = 12

        # ─────────────────────────────────────────────────────────────────────
        # 4. TEACHER WORKLOAD LIST
        # ─────────────────────────────────────────────────────────────────────
        elif mode == "workload":
            tg    = _build_tg()
            ct_mp = _ct_map()
            ws    = wb.create_sheet("Teacher Workload")

            ws.merge_cells("A1:E1")
            c = ws["A1"]; c.value = "Teacher Workload List"
            c.fill = HDR_F; c.font = _font(True, 13, "FFFFFF")
            c.alignment = _align(); c.border = _border()
            ws.row_dimensions[1].height = 22

            for col, txt in enumerate(
                    ["Teacher", "Subject", "Class", "Periods/Week", "Total Periods"], 1):
                c = ws.cell(2, col, txt)
                c.fill = DAY_F; c.font = DAY_N
                c.alignment = _align(); c.border = _border()

            row = 3
            grand_total = 0
            for teacher in sorted(tg.keys()):
                tdata = tg[teacher]
                smry  = defaultdict(lambda: defaultdict(int))
                for d in range(len(days)):
                    for p in range(ppd):
                        e = tdata[d][p] if d < len(tdata) else None
                        if e:
                            smry[e['subject']][e['class']] += 1

                total = sum(c for cd in smry.values() for c in cd.values())
                grand_total += total
                ctc = ct_mp.get(teacher, [])
                start_row = row

                for si, subj in enumerate(sorted(smry.keys())):
                    for cls, cnt in sorted(smry[subj].items()):
                        fill = SUM_F if row % 2 == 0 else WHT_F
                        c = ws.cell(row, 1, teacher if row == start_row else "")
                        c.fill = WRN_F if ctc else fill
                        c.font = _font(True if row == start_row else False, 9)
                        c.alignment = _align(); c.border = _border()

                        for col, val in enumerate([subj, cls, cnt], 2):
                            c2 = ws.cell(row, col, val)
                            c2.fill = fill
                            c2.alignment = _align(); c2.border = _border()
                            c2.font = _font(sz=9)

                        c5 = ws.cell(row, 5, total if row == start_row else "")
                        c5.fill = _fill("#d4e6f1") if row == start_row else fill
                        c5.font = _font(True if row == start_row else False, 9)
                        c5.alignment = _align(); c5.border = _border()
                        row += 1

                span = row - start_row
                if span > 1:
                    ws.merge_cells(start_row=start_row, start_column=1,
                                   end_row=row-1, end_column=1)

            for col, val in enumerate(["", "", "", "GRAND TOTAL", grand_total], 1):
                c = ws.cell(row, col, val)
                c.fill = HDR_F; c.font = _font(True, 10, "FFFFFF")
                c.alignment = _align(); c.border = _border()

            ws.column_dimensions["A"].width = 22
            ws.column_dimensions["B"].width = 22
            ws.column_dimensions["C"].width = 16
            ws.column_dimensions["D"].width = 16
            ws.column_dimensions["E"].width = 16

        # ─────────────────────────────────────────────────────────────────────
        # 5. ONE-SHEET TEACHERWISE
        #    TeacherName | DAY | P1 | P2 | … | Pn   (CLASS/SUBJECT per cell)
        # ─────────────────────────────────────────────────────────────────────
        elif mode == "one_sheet":
            tg = _build_tg()
            ws = wb.create_sheet("Teacherwise Timetable")
            ws.row_dimensions[1].height = 18

            ws.cell(1, 1, "Teacher")
            ws.cell(1, 2, "Day")
            for col in (1, 2):
                ws.cell(1, col).fill = HDR_F; ws.cell(1, col).font = HDR_N
                ws.cell(1, col).alignment = _align(); ws.cell(1, col).border = _border()
            for p in range(ppd):
                c = ws.cell(1, p+3, str(p+1))
                c.fill = HDR_F; c.font = HDR_N
                c.alignment = _align(); c.border = _border()

            row = 2
            for teacher in sorted(tg.keys()):
                tdata   = tg[teacher]
                t_start = row
                for d, dname in enumerate(days):
                    c = ws.cell(row, 1, teacher if d == 0 else "")
                    c.fill = WRN_F; c.alignment = _align()
                    c.font = _font(True if d == 0 else False, 9)
                    c.border = _border()

                    c2 = ws.cell(row, 2, dname)
                    c2.fill = DAY_F; c2.font = DAY_N
                    c2.alignment = _align(); c2.border = _border()

                    for p in range(ppd):
                        e = tdata[d][p] if d < len(tdata) else None
                        if e:
                            txt  = "{}/{}".format(e['class'], e['subject'])
                            fill = SUB_F if e.get('is_ct') else WHT_F
                        else:
                            txt  = ""; fill = FREE_F
                        c3 = ws.cell(row, p+3, txt)
                        c3.fill = fill; c3.alignment = _align()
                        c3.border = _border(); c3.font = _font(sz=8)
                    row += 1

                if len(days) > 1:
                    ws.merge_cells(start_row=t_start, start_column=1,
                                   end_row=row-1, end_column=1)

            ws.column_dimensions["A"].width = 22
            ws.column_dimensions["B"].width = 10
            for p in range(ppd):
                ws.column_dimensions[get_column_letter(p+3)].width = 18

        # ─────────────────────────────────────────────────────────────────────
        # 6. CONSOLIDATED CLASS VIEW
        #    Single sheet — all classes stacked one below the other in order:
        #    6A → 6B → 6C → … → 12G
        #    Each class block: class-header row + period-header row + day rows
        #    Colour legend: ★=CT(green)  combined(blue)  parallel(orange)
        #                   combined+parallel(pink)  free(light grey)
        # ─────────────────────────────────────────────────────────────────────
        elif mode == "consolidated_class":
            ws = wb.create_sheet("All Classes — Consolidated")
            row = 1

            for cn in all_classes:
                cfg_cn  = self.class_config_data.get(cn, {})
                ct_name = cfg_cn.get("teacher", "").strip()
                ct_per  = cfg_cn.get("teacher_period", "")
                hdr_txt = "Class: {}   |   Class Teacher: {}{}".format(
                    cn,
                    ct_name or "—",
                    "   |   CT Period: {}".format(ct_per) if ct_per else "",
                )

                # ── Class header row (spans Day + all periods) ──────────────
                ws.merge_cells(
                    start_row=row, start_column=1,
                    end_row=row,   end_column=ppd + 1,
                )
                c = ws.cell(row, 1, hdr_txt)
                c.fill = CT_H_F; c.font = _font(True, 11, "FFFFFF")
                c.alignment = _align(); c.border = _border()
                ws.row_dimensions[row].height = 20
                row += 1

                # ── Period header row ───────────────────────────────────────
                dc = ws.cell(row, 1, "Day")
                dc.fill = HDR_F; dc.font = HDR_N
                dc.alignment = _align(); dc.border = _border()
                for p in range(ppd):
                    pc = ws.cell(row, p + 2, "P{}{}".format(
                        p + 1, "①" if p < half1 else "②"))
                    pc.fill = HDR_F; pc.font = HDR_N
                    pc.alignment = _align(); pc.border = _border()
                ws.row_dimensions[row].height = 16
                row += 1

                # ── Day rows ────────────────────────────────────────────────
                # Build a set of CT subject names for this class (for is_ct override)
                _ct_subj_names = set()
                _ct_teacher_cn = cfg_cn.get("teacher", "").strip()
                for _s in cfg_cn.get("subjects", []):
                    if _s.get("teacher", "").strip() == _ct_teacher_cn:
                        _ct_subj_names.add(_s["name"])
                _ct_p_idx = int(ct_per) - 1 if ct_per else -1  # 0-based

                def _is_ct_cell(e_cell, p_idx):
                    """True if this cell is a CT period by config, not just by flag."""
                    if e_cell is None:
                        return False
                    if e_cell.get("is_ct"):
                        return True
                    # Cross-check: correct period AND subject belongs to CT teacher
                    if (p_idx == _ct_p_idx
                            and e_cell.get("subject", "") in _ct_subj_names):
                        return True
                    return False

                for d, dname in enumerate(days):
                    ws.row_dimensions[row].height = 48
                    dc2 = ws.cell(row, 1, dname)
                    dc2.fill = DAY_F; dc2.font = DAY_N
                    dc2.alignment = _align(); dc2.border = _border()

                    for p in range(ppd):
                        e = (grid.get(cn, [[]])[d][p]
                             if d < len(grid.get(cn, [])) else None)
                        if e is None:
                            txt  = "FREE"
                            cell_fill = FREE_F
                        else:
                            etype  = e.get("type", "normal")
                            is_ct_ = _is_ct_cell(e, p)  # authoritative check
                            if etype == "combined_parallel":
                                l1, l2 = self._get_combined_par_display(cn, e)
                                txt = "{}\n{}".format(l1, l2)
                                cell_fill = CPAF
                            elif etype == "parallel":
                                txt = "{} / {}\n{} / {}".format(
                                    e["subject"], e.get("par_subj", ""),
                                    e["teacher"],  e.get("par_teach", ""))
                                cell_fill = PAR_F
                            elif etype == "combined":
                                cc_list = e.get("combined_classes", [])
                                mark = " ★" if is_ct_ else ""
                                txt  = "{}{}[{}]\n{}".format(
                                    e["subject"], mark,
                                    "+".join(cc_list), e["teacher"])
                                cell_fill = COMB_F
                            else:
                                mark = " ★" if is_ct_ else ""
                                txt  = "{}{}\n{}".format(
                                    e["subject"], mark, e["teacher"])
                                cell_fill = SUB_F if is_ct_ else WHT_F

                        c2 = ws.cell(row, p + 2, txt)
                        c2.fill = cell_fill; c2.alignment = _align()
                        c2.border = _border(); c2.font = _font(sz=8)
                    row += 1

                # ── Thin spacer between classes ─────────────────────────────
                for col in range(1, ppd + 2):
                    ws.cell(row, col).fill = SUM_F
                    ws.cell(row, col).border = _border()
                ws.row_dimensions[row].height = 5
                row += 1

            # Column widths
            ws.column_dimensions["A"].width = 8
            for p in range(ppd):
                ws.column_dimensions[get_column_letter(p + 2)].width = 20

            # Legend row at the very bottom
            row += 1
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row,   end_column=ppd + 1)
            leg = ws.cell(row, 1,
                "Legend:  ★ = Class Teacher period  |  "
                "Blue = Combined classes  |  "
                "Orange = Parallel teaching  |  "
                "Pink = Combined + Parallel  |  "
                "Light grey = Free")
            leg.fill = SUM_F; leg.font = _font(False, 8, "555555")
            leg.alignment = _align("left"); leg.border = _border()
            ws.row_dimensions[row].height = 14

        wb.save(filename)  # filename can be a file path or BytesIO

