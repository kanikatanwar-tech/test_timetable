"""
streamlit_app.py — Timetable Generator (Streamlit) v4.2

BUGS FIXED vs v1:
  BUG-1 (engine): 11 tkinter StringVar .get() remnants → replaced with plain dict access
  BUG-2 (upload hang): file-uploader fires on every rerun → guarded by SHA-1 hash
         dedup; widget keys deleted after JSON load so widgets reinitialise correctly
  BUG-3 (manual entry stuck): number_input widget keys now ARE the session_state keys
         (ni_ppd etc.) — no shadow s1_ppd copies that get silently overwritten
  LOGGING: Python logging to console + in-memory buffer (sidebar Debug Log)
           Format: [LEVEL] filename:function:lineno — message
"""

import io
import json
import logging
import traceback
import hashlib
from datetime import datetime

import streamlit as st
from engine import TimetableEngine


# ─────────────────────────────────────────────────────────────────────────────
#  LOGGING
# ─────────────────────────────────────────────────────────────────────────────
class _MemHandler(logging.Handler):
    """Keeps last N log lines in memory for the sidebar debug console."""
    MAX = 500
    def __init__(self):
        super().__init__()
        self.lines = []
    def emit(self, record):
        self.lines.append(self.format(record))
        if len(self.lines) > self.MAX:
            self.lines = self.lines[-self.MAX:]

_LOG_FMT = "[%(levelname)s] %(filename)s:%(funcName)s:%(lineno)d — %(message)s"
_mem_handler = _MemHandler()
_mem_handler.setFormatter(logging.Formatter(_LOG_FMT))

logging.basicConfig(level=logging.DEBUG, format=_LOG_FMT,
                    handlers=[logging.StreamHandler(), _mem_handler], force=True)
log = logging.getLogger("timetable")
log.info("streamlit_app module loaded")


# ─────────────────────────────────────────────────────────────────────────────
#  Page config
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Timetable Generator", page_icon="📅",
                   layout="wide", initial_sidebar_state="collapsed")


# ─────────────────────────────────────────────────────────────────────────────
#  Session-state bootstrap
# ─────────────────────────────────────────────────────────────────────────────

# The version string must match TimetableEngine.ENGINE_VERSION in engine.py.
# This is used to detect pickled engine instances from old deployments.
_ENGINE_VERSION = "4.2"


def _engine_is_stale(obj) -> bool:
    """Return True if obj is missing, from an old class, or lacks required methods.

    Deliberately avoids isinstance() because on Streamlit Cloud a module reload
    can produce a new class object, making isinstance() return False for valid
    engines from the same session.  Comparing ENGINE_VERSION strings is immune
    to this problem.
    """
    if obj is None:
        return True
    # Version mismatch = engine is from an older deployment
    if getattr(obj, "ENGINE_VERSION", None) != _ENGINE_VERSION:
        return True
    # Belt-and-suspenders: critical runtime method must exist
    if not callable(getattr(obj, "run_full_generation", None)):
        return True
    return False


def _get_eng() -> TimetableEngine:
    """Always return the live engine from session state.

    Never use a long-lived module-level reference because Streamlit Cloud can
    occasionally swap the session-state engine (e.g. stale-engine replacement)
    AFTER a module-level alias was last set.
    """
    return st.session_state["engine"]


def _init_state():
    defaults = {
        "page":              "step1",
        # Step-1 — widget keys ARE the canonical values (no shadow copies)
        "ni_ppd":            7,
        "ni_wdays":          6,
        "ni_fhalf":          4,
        "ni_shalf":          3,
        "s1_teachers":       [],
        "s1_teacher_fname":  "",
        "s1_sections":       {cls: 4 for cls in range(6, 13)},
        # Upload dedup: hash of the last bytes processed per uploader key
        "_upload_hash":      {},
        # Staged config upload: bytes held until user clicks Load Config
        "_s1_pending_raw":   None,
        "_s1_pending_hash":  None,
        "_s1_pending_name":  None,
        # Step-4 (kept for internal engine use)
        "s4_stage":          0,
        "s4_s1_status":      None,
        "s4_s3_status":      None,
        "s4_ff_result":      None,
        # Auto-generation result
        "gen_result":        None,
        # Task-analysis
        "ta_allocation":     None,
        "ta_group_slots":    None,
        "ta_all_rows":       None,
        "ta2_allocation":    None,
        # Relaxed keys
        "relaxed_consec":    set(),
        "relaxed_main":      set(),
        # Validation caches
"s2_validation_result": None,
        # Step 2 — errors the user has explicitly chosen to ignore
        "s2_ignored_errors":    set(),  # set of string keys
        "s3_validation_result": None,
        # Step gate-keeping — each step must be validated before the next is shown
        "s1_validated":      False,
        "s2_validated":      False,
        "s3_validated":      False,
        # Popup error triggers
        "_s2_show_error_popup": False,
        "_s3_show_error_popup": False,
        # Pending notifications (cleared each render)
        "_notify":           [],
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

    # ── Engine bootstrap / stale-instance guard ───────────────────────────────
    # Must run AFTER the defaults loop so "engine" key exists even on first run.
    if "engine" not in st.session_state:
        st.session_state["engine"] = TimetableEngine()
        log.info("_init_state: fresh engine created")
    elif _engine_is_stale(st.session_state["engine"]):
        # Engine from an old deployment was pickled into session state.
        # Wipe it and start over.  Also reset page + validated flags so the
        # user lands on Step 1 (not on Step 2/3/4 with an empty engine).
        log.warning("_init_state: stale engine detected — resetting")
        st.session_state["engine"]       = TimetableEngine()
        st.session_state["page"]         = "step1"
        st.session_state["s1_validated"] = False
        st.session_state["s2_validated"] = False
        st.session_state["s3_validated"] = False
        st.session_state["gen_result"]   = None

    log.debug("_init_state: done  page=%s  s1v=%s",
              st.session_state.get("page"), st.session_state.get("s1_validated"))


_init_state()


# ─────────────────────────────────────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────────────────────────────────────
DAY_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

def _nav(page: str):
    log.info("_nav: → %s", page)
    st.session_state.page = page
    st.rerun()

def _notify(msg: str, kind: str = "info"):
    log.debug("_notify [%s]: %s", kind, msg)
    st.session_state["_notify"].append((kind, msg))

def _show_notifications():
    for kind, msg in st.session_state.get("_notify", []):
        if kind == "success": st.success(msg)
        elif kind == "error":   st.error(msg)
        elif kind == "warning": st.warning(msg)
        else:                   st.info(msg)
    st.session_state["_notify"] = []

def _file_hash(data: bytes) -> str:
    return hashlib.sha1(data).hexdigest()

def _already_processed(key: str, data: bytes) -> bool:
    """Return True if we already handled this exact file upload (same bytes)."""
    h = _file_hash(data)
    if st.session_state["_upload_hash"].get(key) == h:
        log.debug("_already_processed: key=%s hash=%s already done", key, h[:8])
        return True
    st.session_state["_upload_hash"][key] = h
    log.debug("_already_processed: key=%s hash=%s NEW", key, h[:8])
    return False

def _all_classes():
    cfg = _get_eng().configuration
    return [
        f"{cls}{chr(65+si)}"
        for cls in range(6, 13)
        for si in range(int(cfg.get("classes", {}).get(cls, 0)))
    ]

def _json_download(data: dict, label: str, filename: str):
    st.download_button(label=label, data=json.dumps(data, indent=2),
                       file_name=filename, mime="application/json",
                       use_container_width=True)

def _excel_download(mode: str, label: str):
    log.info("_excel_download: mode=%s", mode)
    try:
        xbytes = _get_eng().get_excel_bytes(mode)
        fname  = f"{mode}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        st.download_button(label=label, data=xbytes, file_name=fname,
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)
        log.info("_excel_download: ready %s (%d bytes)", fname, len(xbytes))
    except Exception as ex:
        log.error("_excel_download: %s\n%s", ex, traceback.format_exc())
        st.error(f"Export error: {ex}")

PAGE_LABELS = {
    "step1":           "Step 1 — Basic Config",
    "step2":           "Step 2 — Class Assignments",
    "step3":           "Step 3 — Teacher Settings",
    "generate":        "Step 4 — Generate",
    "final_timetable": "Final Timetable",
}

def _header(title: str, sub: str = ""):
    st.markdown(f"## {title}")
    if sub: st.caption(sub)
    st.divider()


@st.dialog("❌ Wrong File Uploaded")
def _upload_error_dialog(msg: str):
    """Modal popup for wrong/invalid file uploads."""
    st.error(msg)
    st.markdown("Please upload the **correct file** for this step.")
    if st.button("OK", type="primary", use_container_width=True):
        st.rerun()


@st.dialog("⚠ Cannot Combine")
def _combine_error_dialog(msg: str):
    """Modal popup for combine validation errors in Step 3."""
    st.error(msg)
    if st.button("OK", type="primary", use_container_width=True):
        st.rerun()

@st.dialog("❌ Step 1 — Validation Errors")
def _step1_error_dialog(errors: list):
    """Modal popup listing all Step 1 validation errors with fix advice."""
    for err in errors:
        st.error(err["msg"])
        if err.get("fix"):
            st.info(f"💡 **How to fix:** {err['fix']}")
    st.divider()
    st.caption("Fix all errors above, then click **Continue to Step 2** again.")
    if st.button("Close", type="primary", use_container_width=True):
        st.rerun()


# ── Stable string keys for per-error ignore tracking ────────────────────────
def _hc_key(c: dict) -> str:
    """Deterministic key for a hard-conflict entry."""
    return "hc:{}:{}:{}".format(
        c.get("teacher",""), c.get("slot_a","")[:60], c.get("slot_b","")[:60])

def _wcc_key(c: dict) -> str:
    """Deterministic key for a within-class-conflict entry."""
    return "wcc:{}:{}:{}".format(
        c.get("class",""), c.get("day",""), c.get("reason","")[:80])


@st.dialog("❌ Step 2 — Validation Errors", width="large")
def _step2_error_dialog(vr: dict):
    """Modal popup summarising Step 2 validation errors.

    Period mismatches (①) are BLOCKING — the timetable cannot be complete if
    the total periods do not match ppd × wdays, so they cannot be ignored.

    Teacher conflicts (②) and within-class conflicts (③) CAN be ignored: the
    engine will do its best to resolve them; the user accepts the risk that
    some preferences may not be honoured.
    """
    period_errors  = vr.get("period_errors", [])
    hard_conflicts = vr.get("hard_conflicts", [])
    wcc            = vr.get("within_class_conflicts", [])

    ignored = st.session_state.get("s2_ignored_errors", set())

    # Errors not yet ignored
    active_hc  = [c for c in hard_conflicts
                  if _hc_key(c)  not in ignored]
    active_wcc = [c for c in wcc
                  if _wcc_key(c) not in ignored]
    total_blocking = len(period_errors) + len(active_hc) + len(active_wcc)

    if period_errors:
        st.error(f"**{len(period_errors)} period mismatch(es)** must be fixed — "
                 "they cannot be ignored.")
    if active_hc or active_wcc:
        st.warning(f"**{len(active_hc) + len(active_wcc)} conflict(s)** remain. "
                   "You can ignore individual ones below and proceed.")
    if not period_errors and not active_hc and not active_wcc:
        st.success("All remaining errors have been ignored — you can now proceed.")

    st.divider()

    # ── ① Period Mismatches (BLOCKING, cannot ignore) ──────────────────────
    if period_errors:
        st.markdown(f"**① Period Mismatches — {len(period_errors)} class(es)  🔒 Cannot be ignored**")
        for cn, msg in period_errors[:6]:
            st.error(f"**{cn}:** {msg}")
            st.caption("💡 Add/remove subjects so total periods = periods/day × working days.")
        if len(period_errors) > 6:
            st.caption(f"… and {len(period_errors)-6} more — scroll Step 2 for all details.")

    # ── ② Teacher Conflicts (can be ignored per-item) ─────────────────────
    if hard_conflicts:
        st.markdown(f"**② Teacher Conflicts — {len(hard_conflicts)} total "
                    f"({len(active_hc)} active, "
                    f"{len(hard_conflicts)-len(active_hc)} ignored)**")
        st.caption("✅ = ignored (engine will place best-effort).  "
                   "Check the box to ignore an individual conflict.")
        for i, c in enumerate(hard_conflicts):
            key = _hc_key(c)
            already = key in ignored
            col1, col2 = st.columns([0.08, 0.92])
            with col1:
                new_val = st.checkbox("", value=already,
                                      key=f"dlg_ign_hc_{i}",
                                      help="Ignore this conflict and proceed")
            with col2:
                reason = c["reason"]
                if already:
                    st.success(f"✅ **{c['teacher']}** (ignored) — "
                               f"{reason[:120]}{'...' if len(reason)>120 else ''}")
                else:
                    st.warning(f"**{c['teacher']}:** "
                               f"{reason[:140]}{'...' if len(reason)>140 else ''}")
            if new_val and not already:
                ignored.add(key)
                st.session_state["s2_ignored_errors"] = ignored
                st.rerun()
            elif not new_val and already:
                ignored.discard(key)
                st.session_state["s2_ignored_errors"] = ignored
                st.rerun()

    # ── ③ Within-Class Conflicts (can be ignored per-item) ────────────────
    if wcc:
        st.markdown(f"**③ Within-Class Conflicts — {len(wcc)} total "
                    f"({len(active_wcc)} active, "
                    f"{len(wcc)-len(active_wcc)} ignored)**")
        st.caption("Check the box to ignore an individual conflict.")
        for i, c in enumerate(wcc):
            key = _wcc_key(c)
            already = key in ignored
            col1, col2 = st.columns([0.08, 0.92])
            with col1:
                new_val = st.checkbox("", value=already,
                                      key=f"dlg_ign_wcc_{i}",
                                      help="Ignore this conflict and proceed")
            with col2:
                reason = c["reason"]
                label  = f"{c['class']} / {c['day']}"
                if already:
                    st.success(f"✅ **{label}** (ignored) — "
                               f"{reason[:120]}{'...' if len(reason)>120 else ''}")
                else:
                    st.warning(f"**{label}:** "
                               f"{reason[:140]}{'...' if len(reason)>140 else ''}")
            if new_val and not already:
                ignored.add(key)
                st.session_state["s2_ignored_errors"] = ignored
                st.rerun()
            elif not new_val and already:
                ignored.discard(key)
                st.session_state["s2_ignored_errors"] = ignored
                st.rerun()

    st.divider()
    btn_col1, btn_col2 = st.columns(2)
    with btn_col1:
        if st.button("← Close & Fix Issues", use_container_width=True):
            st.rerun()
    with btn_col2:
        can_proceed = not period_errors and not active_hc and not active_wcc
        if st.button("Proceed with Ignored Errors →",
                     type="primary",
                     disabled=not can_proceed,
                     use_container_width=True,
                     help=("All blocking errors resolved — click to proceed to Step 3."
                           if can_proceed else
                           "Fix or ignore all remaining errors first.")):
            st.session_state["s2_validated"] = True
            st.session_state["s3_validated"] = False
            _notify("✅ Proceeding with some errors ignored — "
                    "engine will place best-effort for ignored conflicts.", "warning")
            st.rerun()


@st.dialog("❌ Step 3 — Validation Errors")
def _step3_error_dialog(vr: dict):
    """Modal popup for Step 3 overload errors with fix advice."""
    issues = vr.get("issues", [])
    st.error(f"**{len(issues)} overloaded teacher(s) must be resolved before generating.**")
    st.divider()
    for ln in issues[:8]:
        st.warning(ln)
    if len(issues) > 8:
        st.caption(f"... and {len(issues)-8} more.")
    st.divider()
    st.info(
        "**How to fix:**\n"
        "- Use **Combine Classes** tab to merge identical lessons across sections — "
        "this reduces the teacher's effective period count.\n"
        "- Or click **⏭ Skip** on the teacher card to bypass the overload check (use with caution)."
    )
    if st.button("Close & Fix Issues", type="primary", use_container_width=True):
        st.rerun()


def _show_upload_error_if_any(key: str):
    """If session state has an upload error for `key`, show the dialog and clear it."""
    err = st.session_state.pop(key, None)
    if err:
        _upload_error_dialog(err)


# ═════════════════════════════════════════════════════════════════════════════
#  STEP 1
# ═════════════════════════════════════════════════════════════════════════════
def page_step1():
    log.info("page_step1: render")
    _show_upload_error_if_any("_s1_upload_err")   # modal popup for wrong-file uploads
    # Show validation error dialog if triggered by continue button
    if st.session_state.get("_s1_val_errors"):
        _step1_error_dialog(st.session_state.pop("_s1_val_errors"))
    _header("📋 Step 1: Basic Configuration",
            "Set periods, working days, upload teachers and define class sections.")
    _show_notifications()

    # ── Save / Load ───────────────────────────────────────────────────────────
    with st.expander("💾 Save / Load Configuration", expanded=False):
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("**Save current config**")
            if st.button("⬇ Prepare Download", key="s1_dl_btn", use_container_width=True):
                log.info("page_step1: preparing download")
                if not st.session_state.s1_teachers:
                    _notify("Upload a teacher file first.", "warning"); st.rerun()
                data = {
                    "periods_per_day":     st.session_state.get("ni_ppd", 7),
                    "working_days":        st.session_state.get("ni_wdays", 6),
                    "periods_first_half":  st.session_state.get("ni_fhalf", 4),
                    "periods_second_half": st.session_state.get("ni_shalf", 3),
                    "teacher_file_path":   st.session_state.s1_teacher_fname,
                    "teacher_names":       st.session_state.s1_teachers,
                    "classes":             {str(k): v for k, v in
                                            st.session_state.s1_sections.items()},
                    "saved_at":            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                }
                st.session_state["_s1_dl_data"] = data
            if "_s1_dl_data" in st.session_state:
                _json_download(st.session_state["_s1_dl_data"], "📥 Click to Download",
                               f"Config_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
        with c2:
            st.markdown("**Load saved config**")
            up = st.file_uploader("Upload Config JSON", type=["json"], key="s1_load_json")
            if up is not None:
                # Stage 1: read bytes once and store — do NOT process yet.
                # up.read() returns empty on the second rerun of the same file,
                # so we only read when we actually have bytes.
                raw = up.read()
                if len(raw) > 0:
                    h = _file_hash(raw)
                    if st.session_state.get("_s1_pending_hash") != h:
                        # New file — cache it and wait for button click
                        st.session_state["_s1_pending_raw"]  = raw
                        st.session_state["_s1_pending_hash"] = h
                        st.session_state["_s1_pending_name"] = up.name
                        log.debug("page_step1: cached new file '%s' (%d bytes)", up.name, len(raw))

            pending = st.session_state.get("_s1_pending_raw")
            if pending:
                pname = st.session_state.get("_s1_pending_name", "config.json")
                st.info(f"📄 Ready to load: **{pname}** ({len(pending):,} bytes)")
                if st.button("📂 Load Config", key="s1_load_btn", type="primary",
                             use_container_width=True):
                    log.info("page_step1: Load Config button clicked for '%s'", pname)
                    _load_step1_config(pending)
            else:
                st.caption("Upload a JSON config file, then click Load Config.")

    # ── Periods & Days ────────────────────────────────────────────────────────
    with st.container(border=True):
        st.markdown("**1. Periods & Working Days**")
        c1, c2 = st.columns(2)
        with c1:
            # FIX: widget key == session_state key → ONE source of truth
            ppd   = st.number_input("Periods per day",   1, 20,
                                    st.session_state.get("ni_ppd",   7), key="ni_ppd")
            wdays = st.number_input("Working days/week", 1, 7,
                                    st.session_state.get("ni_wdays", 6), key="ni_wdays")
        with c2:
            fhalf = st.number_input("Periods — first half",  1, 20,
                                    st.session_state.get("ni_fhalf", 4), key="ni_fhalf")
            shalf = st.number_input("Periods — second half", 1, 20,
                                    st.session_state.get("ni_shalf", 3), key="ni_shalf")
        if fhalf + shalf == ppd:
            st.success(f"✓ Valid: {fhalf} + {shalf} = {ppd}")
        else:
            st.error(f"✗ {fhalf} + {shalf} = {fhalf+shalf}, need {ppd}")

    # ── Teachers ──────────────────────────────────────────────────────────────
    with st.container(border=True):
        st.markdown("**2. Teachers — Upload Excel File**")
        st.caption("Column A only: one name per row. "
                   "Header 'Teacher Name' is automatically skipped.")
        col_up, col_info = st.columns([3, 1])
        with col_up:
            tf = st.file_uploader("Teacher Excel (.xlsx/.xls)", type=["xlsx","xls"],
                                   key="s1_teacher_up")
            if tf is not None:
                raw_tf = tf.read()
                log.debug("page_step1: teacher upload size=%d", len(raw_tf))
                if len(raw_tf) == 0:
                    log.warning("page_step1: teacher 0 bytes, skipping")
                elif not _already_processed("s1_teacher_up", raw_tf):
                    _load_teacher_bytes(raw_tf, tf.name)
        with col_info:
            if st.session_state.s1_teachers:
                st.success(f"✓ {len(st.session_state.s1_teachers)} teachers")
                if st.button("👁 Preview", key="s1_preview"):
                    st.session_state["_s1_show_t"] = not st.session_state.get("_s1_show_t", False)
            else:
                st.warning("No teachers loaded")
        if st.session_state.get("_s1_show_t") and st.session_state.s1_teachers:
            st.dataframe({"Teacher Name": st.session_state.s1_teachers},
                         use_container_width=True, height=200)

    # ── Classes ───────────────────────────────────────────────────────────────
    with st.container(border=True):
        st.markdown("**3. Classes 6–12 — Number of Sections**")
        cols = st.columns(7)
        for idx, cls in enumerate(range(6, 13)):
            with cols[idx]:
                cur = st.session_state.s1_sections.get(cls, 4)
                val = st.number_input(f"Class {cls}", 0, 50, cur, key=f"ni_cls_{cls}")
                st.session_state.s1_sections[cls] = val

    # ── Navigation ────────────────────────────────────────────────────────────
    cb, cc = st.columns([1, 3])
    with cc:
        if st.button("✓ Continue to Step 2 →", type="primary", use_container_width=True):
            _step1_save_and_continue()
    with cb:
        if st.button("⟲ Reset", use_container_width=True):
            log.info("page_step1: reset")
            for k in ["ni_ppd","ni_wdays","ni_fhalf","ni_shalf",
                      "s1_teachers","s1_teacher_fname","_s1_dl_data","_s1_show_t",
                      "_s1_pending_raw","_s1_pending_hash","_s1_pending_name"]:
                st.session_state.pop(k, None)
            st.session_state.s1_sections  = {cls: 4 for cls in range(6, 13)}
            st.session_state["_upload_hash"] = {}
            st.rerun()


def _load_step1_config(raw: bytes):
    log.info("_load_step1_config: %d bytes", len(raw))
    try:
        d = json.loads(raw.decode("utf-8"))
    except json.JSONDecodeError as ex:
        log.error("_load_step1_config: JSON parse error: %s", ex)
        msg = f"Invalid JSON file: {ex}"
        st.session_state["_s1_upload_err"] = msg
        for k in ("_s1_pending_raw", "_s1_pending_hash", "_s1_pending_name"):
            st.session_state.pop(k, None)
        st.rerun()
        return

    required_keys = {"periods_per_day", "working_days", "teacher_names"}
    if not required_keys.issubset(d.keys()):
        wrong_hint = ""
        if "assignments" in d:
            wrong_hint = " — this looks like a Step 2 Assignments file"
        elif "step3_data" in d or "step3_unavailability" in d:
            wrong_hint = " — this looks like a Step 3 config file"
        msg = (f"Wrong file type{wrong_hint}.\n\n"
               f"A Step 1 config file must contain: "
               f"{', '.join(sorted(required_keys))}.")
        log.error("_load_step1_config: %s", msg)
        st.session_state["_s1_upload_err"] = msg
        for k in ("_s1_pending_raw", "_s1_pending_hash", "_s1_pending_name"):
            st.session_state.pop(k, None)
        st.rerun()
        return

    try:
        ppd   = int(d.get("periods_per_day",    7))
        wdays = int(d.get("working_days",        6))
        fhalf = int(d.get("periods_first_half",  4))
        shalf = int(d.get("periods_second_half", 3))
        teachers   = d.get("teacher_names", [])
        classes_raw = d.get("classes", {})
        log.info("_load_step1_config: ppd=%d wdays=%d fhalf=%d shalf=%d teachers=%d",
                 ppd, wdays, fhalf, shalf, len(teachers))

        keys_to_clear = (
            ["ni_ppd", "ni_wdays", "ni_fhalf", "ni_shalf"]
            + [f"ni_cls_{cls}" for cls in range(6, 13)]
        )
        for wk in keys_to_clear:
            st.session_state.pop(wk, None)

        sections = {int(k): v for k, v in classes_raw.items()}
        st.session_state["ni_ppd"]           = ppd
        st.session_state["ni_wdays"]         = wdays
        st.session_state["ni_fhalf"]         = fhalf
        st.session_state["ni_shalf"]         = shalf
        st.session_state["s1_teachers"]      = teachers
        st.session_state["s1_teacher_fname"] = d.get("teacher_file_path", "")
        st.session_state["s1_sections"]      = sections
        for cls, nsec in sections.items():
            st.session_state[f"ni_cls_{cls}"] = nsec

        for k in ("_s1_pending_raw", "_s1_pending_hash", "_s1_pending_name"):
            st.session_state.pop(k, None)

        _notify(f"✓ Config loaded — {len(teachers)} teachers, "
                f"{ppd} periods/day, {wdays} days/week, "
                f"{sum(sections.values())} total sections.", "success")
        log.info("_load_step1_config: applied OK — sections=%s", sections)
        st.rerun()
    except Exception as ex:
        log.error("_load_step1_config: %s\n%s", ex, traceback.format_exc())
        st.session_state["_s1_upload_err"] = f"Failed to load config: {ex}"
        st.rerun()


def _load_teacher_bytes(raw: bytes, fname: str):
    log.info("_load_teacher_bytes: '%s' %d bytes", fname, len(raw))
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active
        max_col = max(
            (ci for row in ws.iter_rows() for ci, c in enumerate(row, 1) if c.value is not None),
            default=0)
        if max_col > 1:
            log.warning("_load_teacher_bytes: %d columns found, expected 1", max_col)
            st.session_state["_s1_upload_err"] = (
                f"File has data in {max_col} columns — use Column A only.\n\n"
                "Each row in Column A should contain exactly one teacher name.")
            st.rerun(); return
        names = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            v = row[0].value
            if v:
                n = str(v).strip()
                if n and n.lower() != "teacher name":
                    names.append(n)
        if not names:
            log.warning("_load_teacher_bytes: no names found")
            st.session_state["_s1_upload_err"] = (
                "No teacher names found in the file.\n\n"
                "Make sure names are in Column A (Row 1 header 'Teacher Name' is auto-skipped).")
            st.rerun(); return
        seen, dups = set(), []
        for n in names:
            if n in seen: dups.append(n)
            seen.add(n)
        if dups:
            log.warning("_load_teacher_bytes: duplicates: %s", dups)
            st.session_state["_s1_upload_err"] = (
                "Duplicate teacher names found:\n\n" +
                "\n".join(f"  • {d}" for d in sorted(set(dups))) +
                "\n\nPlease remove duplicates and re-upload.")
            st.rerun(); return
        names.sort()
        st.session_state["s1_teachers"]      = names
        st.session_state["s1_teacher_fname"] = fname
        _notify(f"✓ {len(names)} teachers loaded (A→Z).", "success")
        log.info("_load_teacher_bytes: %d teachers loaded", len(names))
        st.rerun()
    except Exception as ex:
        log.error("_load_teacher_bytes: %s\n%s", ex, traceback.format_exc())
        st.session_state["_s1_upload_err"] = f"Error reading file '{fname}': {ex}"
        st.rerun()


def _step1_save_and_continue():
    log.info("_step1_save_and_continue: called")
    ppd   = int(st.session_state.get("ni_ppd",   7))
    wdays = int(st.session_state.get("ni_wdays", 6))
    fhalf = int(st.session_state.get("ni_fhalf", 4))
    shalf = int(st.session_state.get("ni_shalf", 3))
    teachers = st.session_state.s1_teachers
    log.debug("_step1_save_and_continue: ppd=%d wdays=%d fhalf=%d shalf=%d teachers=%d",
              ppd, wdays, fhalf, shalf, len(teachers))

    errors = []
    if ppd <= 0 or wdays <= 0:
        errors.append({"msg": "Periods per day and working days must be ≥ 1.",
                        "fix": "Set valid values in the Periods & Working Days section."})
    if fhalf + shalf != ppd:
        errors.append({"msg": f"Halves mismatch: {fhalf} + {shalf} = {fhalf+shalf}, but need {ppd}.",
                        "fix": "Adjust 'Periods — first half' or 'Periods — second half' so they add up to Periods per day."})
    if not teachers:
        errors.append({"msg": "No teacher list loaded.",
                        "fix": "Upload a teacher Excel file (Column A, one name per row)."})
    if errors:
        log.warning("_step1_save_and_continue: %d validation errors", len(errors))
        st.session_state["_s1_val_errors"] = errors
        st.rerun()
        return

    # Always use int values for section counts — number_input can return float
    sections = {cls: int(v) for cls, v in st.session_state.s1_sections.items()}

    # Write configuration directly into session-state engine
    engine = _get_eng()
    engine.configuration = {
        "periods_per_day":    ppd,
        "working_days":       wdays,
        "periods_first_half": fhalf,
        "periods_second_half":shalf,
        "teacher_file":       st.session_state.s1_teacher_fname,
        "teacher_names":      teachers,
        "classes":            sections,
    }
    log.info("_step1_save_and_continue: configuration set, %d total sections",
             sum(sections.values()))

    for cls in range(6, 13):
        for si in range(sections.get(cls, 0)):
            cn = f"{cls}{chr(65+si)}"
            if cn not in engine.class_config_data:
                engine.class_config_data[cn] = {
                    "subjects": [], "teacher": "",
                    "teacher_period": 1, "editing_index": None,
                }
    log.info("_step1_save_and_continue: %d classes ready → step2",
             len(engine.class_config_data))
    st.session_state["s1_validated"] = True
    st.session_state["s2_validated"]     = False
    st.session_state["s3_validated"]     = False
    st.session_state["s2_ignored_errors"] = set()  # reset on config change
    _nav("step2")


# ═════════════════════════════════════════════════════════════════════════════
#  STEP 2
# ═════════════════════════════════════════════════════════════════════════════
def page_step2():
    # Always fetch the live engine — never rely on a possibly-stale module alias
    eng = _get_eng()

    # Guard: if Step 1 hasn't been completed the engine config will be empty.
    # Show a clear message and a button — DO NOT auto-redirect (_nav inside a
    # guard would unconditionally loop the user back to step1 on every rerun).
    if not eng.configuration:
        st.warning("⚠ Step 1 not yet completed — basic configuration is missing.")
        st.info("Please complete Step 1 (Basic Config) before continuing.")
        if st.button("← Go to Step 1", type="primary", key="s2_guard_back"):
            _nav("step1")
        return

    log.info("page_step2: render (%d classes)", len(_all_classes()))

    # Show validation error popup if triggered
    if st.session_state.get("_s2_val_errors"):
        _step2_error_dialog(st.session_state.pop("_s2_val_errors"))

    cfg       = eng.configuration
    ppd       = cfg["periods_per_day"]
    wdays     = cfg["working_days"]
    required  = ppd * wdays
    teachers  = sorted(cfg["teacher_names"])
    day_names = DAY_NAMES[:wdays]
    all_cn    = _all_classes()

    # ── Wrong-file upload: show as modal dialog ────────────────────────────────
    _show_upload_error_if_any("_s2_upload_err")

    _header("👨‍🏫 Step 2: Configure Each Class",
            "Set class teacher and add subjects with periods, preferences and constraints.")
    _show_notifications()

    # ── Save / Load ───────────────────────────────────────────────────────────
    with st.expander("💾 Save / Load Assignments", expanded=False):
        c1, c2 = st.columns(2)
        with c1:
            if st.button("⬇ Prepare Download", key="s2_dl_btn", use_container_width=True):
                log.info("page_step2: preparing assignments download")
                payload = {
                    cn: {"teacher": cd.get("teacher",""),
                         "teacher_period": cd.get("teacher_period",1),
                         "subjects": cd.get("subjects",[])}
                    for cn, cd in _get_eng().class_config_data.items()
                }
                st.session_state["_s2_dl_data"] = {
                    "assignments": payload,
                    "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                }
            if "_s2_dl_data" in st.session_state:
                _json_download(st.session_state["_s2_dl_data"], "📥 Click to Download",
                               f"Assignments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
        with c2:
            up = st.file_uploader("Upload Assignments JSON", type=["json"], key="s2_load_json")
            if up is not None:
                raw = up.read()
                log.debug("page_step2: upload size=%d", len(raw))
                if len(raw) == 0:
                    log.warning("page_step2: 0 bytes, skipping")
                elif not _already_processed("s2_load_json", raw):
                    _load_step2_assignments(raw)

    # ── Navigation ────────────────────────────────────────────────────────────
    s2_validated = st.session_state.get("s2_validated", False)
    cb, cv, cc = st.columns([1, 2, 2])
    with cb:
        if st.button("← Back to Step 1", use_container_width=True):
            _nav("step1")
    with cv:
        if st.button("🔍 Validate Step 2", use_container_width=True,
                     help="Check period counts and teacher conflicts"):
            _step2_validate_and_continue()
    with cc:
        if s2_validated:
            if st.button("✓ Continue to Step 3 →", type="primary", use_container_width=True):
                _nav("step3")
        else:
            st.button("✓ Continue to Step 3 →", type="primary",
                      use_container_width=True, disabled=True,
                      help="Run Validate Step 2 first and fix any errors")

    if s2_validated:
        st.success("✅ Step 2 validated — click **Continue to Step 3** to proceed.")
    else:
        st.info("ℹ️ Fill in all class subjects, then click **🔍 Validate Step 2** to check for errors.")

    if not all_cn:
        st.warning("No classes found — go back to Step 1.")
        return

    st.divider()

    # ── Build grade → sections map ────────────────────────────────────────────
    all_cn_grouped = {}
    for cn in all_cn:
        all_cn_grouped.setdefault(cn[:-1], []).append(cn)
    grades = sorted(all_cn_grouped.keys(), key=int)

    # ── Single source of truth: s2_active_class ───────────────────────────────
    # Always derived from here; never reset by widget reruns.
    if ("s2_active_class" not in st.session_state
            or st.session_state["s2_active_class"] not in all_cn):
        st.session_state["s2_active_class"] = all_cn[0]

    act = st.session_state["s2_active_class"]

    def _cn_fmt(cn):
        cd = eng.class_config_data.get(cn, {})
        t  = sum(s.get("periods", 0) for s in cd.get("subjects", []))
        ic = "🟢" if t == required else ("🔴" if t > required else "⚪")
        return f"{ic} {cn}  ({t}/{required})"

    sel_col, info_col = st.columns([2, 5])
    with sel_col:
        st.markdown("**Select Class**")

        # ── Grade row ─────────────────────────────────────────────────────────
        # Derive current grade from the active class (not from a separate key).
        # Using a grade-specific widget key ensures the stored value always
        # matches the current grade's sections list.
        cur_grade = act[:-1]
        grade_idx = grades.index(cur_grade) if cur_grade in grades else 0

        chosen_grade = st.radio(
            "Grade", grades,
            index=grade_idx,
            horizontal=True,
            key="s2_grade_radio",
            label_visibility="collapsed",
            format_func=lambda g: f"Class {g}",
        )

        # ── When grade changes, move to first section of new grade ────────────
        if chosen_grade != cur_grade:
            act = all_cn_grouped[chosen_grade][0]
            st.session_state["s2_active_class"] = act
            cur_grade = chosen_grade

        sections = all_cn_grouped[cur_grade]
        sec_idx   = sections.index(act) if act in sections else 0

        # ── Section row ───────────────────────────────────────────────────────
        # KEY INSIGHT: use a GRADE-SPECIFIC key so values from grade 6 never
        # pollute grade 7's radio options (which would cause Streamlit to silently
        # snap back to index 0, making it look like the tab reset).
        chosen_sec = st.radio(
            "Section", sections,
            index=sec_idx,
            horizontal=True,
            key=f"s2_sec_{cur_grade}",
            label_visibility="collapsed",
        )

        # If user clicked a different section, update and rerun
        if chosen_sec != act:
            st.session_state["s2_active_class"] = chosen_sec
            st.rerun()

        cur_sel = st.session_state["s2_active_class"]

    with info_col:
        cd_prev = eng.class_config_data.get(cur_sel, {})
        tot     = sum(s.get("periods", 0) for s in cd_prev.get("subjects", []))
        diff    = required - tot
        note    = "✓ exact" if tot == required else (
            f"need {diff} more" if diff > 0 else f"over by {-diff}")
        ic      = "🟢" if tot == required else ("🔴" if tot > required else "⚪")
        st.markdown(f"### Class {cur_sel}")
        st.caption(f"{ic}  Assigned: **{tot}** / {required}  ({note})  "
                   f"— {len(cd_prev.get('subjects', []))} subject(s)")
        with st.expander("📊 All Classes Progress", expanded=False):
            gcols = st.columns(len(grades))
            for gi, g in enumerate(grades):
                with gcols[gi]:
                    st.markdown(f"**Gr.{g}**")
                    for cn in all_cn_grouped[g]:
                        st.caption(_cn_fmt(cn))

    st.divider()
    _class_config_tab(cur_sel, teachers, ppd, wdays, required, day_names)


def _load_step2_assignments(raw: bytes):
    """Load assignments JSON — shows a modal dialog for wrong file uploads."""
    eng = _get_eng()
    log.info("_load_step2_assignments: %d bytes", len(raw))
    try:
        d = json.loads(raw.decode("utf-8"))
    except json.JSONDecodeError as ex:
        log.error("_load_step2_assignments: JSON parse error: %s", ex)
        st.session_state["_s2_upload_err"] = f"Invalid JSON file: {ex}"
        st.rerun()
        return

    if "assignments" not in d:
        wrong_hint = ""
        if "periods_per_day" in d or "teacher_names" in d:
            wrong_hint = "\n\nThis looks like a **Step 1 config** file."
        elif "step3_data" in d or "step3_unavailability" in d:
            wrong_hint = "\n\nThis looks like a **Step 3 config** file."
        msg = (f"Wrong file type for Step 2.{wrong_hint}\n\n"
               f"Please upload an **Assignments file** (saved from Step 2 → Save/Load).")
        log.error("_load_step2_assignments: wrong file")
        st.session_state["_s2_upload_err"] = msg
        st.rerun()
        return

    n = 0
    for cn, saved in d["assignments"].items():
        if cn in eng.class_config_data:
            eng.class_config_data[cn].update({
                "teacher":        saved.get("teacher", ""),
                "teacher_period": saved.get("teacher_period", 1),
                "subjects":       saved.get("subjects", []),
                "editing_index":  None,
            })
            _purge_form_state(cn)
            n += 1
    _notify(f"✓ Assignments loaded for {n} classes.", "success")
    log.info("_load_step2_assignments: %d classes loaded", n)
    st.rerun()


def _purge_form_state(cn: str):
    """Delete ALL form widget keys for class cn from session state.

    This guarantees the add-form shows empty on the next render,
    regardless of what the user had typed previously.
    Called after: subject add, subject edit, assignment load.
    """
    prefix = f"sf_"
    needle = f"_{cn}_"
    to_del = [k for k in list(st.session_state.keys())
              if k.startswith(prefix) and needle in k]
    for k in to_del:
        del st.session_state[k]
    # Also clear form version so the suffix starts at "a0" again
    st.session_state.pop(f"s2_fv_{cn}", None)
    log.debug("_purge_form_state: %s cleared %d keys", cn, len(to_del))


def _s2_init_form(cn, suffix, prefill: dict):
    """Pre-populate form widget keys with prefill values (only on first use).

    Guard: if init_key already set for this suffix, skip (prevents re-init
    on every rerun while the user is editing the form).
    """
    init_key = f"sf_init_{cn}_{suffix}"
    if st.session_state.get(init_key):
        return
    _cfg  = _get_eng().configuration
    ppd   = _cfg["periods_per_day"]
    wdays = _cfg["working_days"]
    st.session_state[f"sf_name_{cn}_{suffix}"]   = prefill.get("name", "")
    st.session_state[f"sf_teach_{cn}_{suffix}"]  = prefill.get("teacher", "")
    st.session_state[f"sf_per_{cn}_{suffix}"]    = int(prefill.get("periods", 1))
    st.session_state[f"sf_cons_{cn}_{suffix}"]   = prefill.get("consecutive", "No")
    st.session_state[f"sf_pref_{cn}_{suffix}"]   = [int(p) for p in prefill.get("periods_pref", [])]
    st.session_state[f"sf_day_{cn}_{suffix}"]    = list(prefill.get("days_pref", []))
    st.session_state[f"sf_par_{cn}_{suffix}"]    = bool(prefill.get("parallel", False))
    st.session_state[f"sf_psub_{cn}_{suffix}"]   = prefill.get("parallel_subject", "")
    st.session_state[f"sf_pteach_{cn}_{suffix}"] = prefill.get("parallel_teacher", "")
    st.session_state[init_key] = True
    log.debug("_s2_init_form: %s/%s initialised", cn, suffix)


def _class_config_tab(cn, teachers, ppd, wdays, required, day_names):
    eng = _get_eng()
    log.debug("_class_config_tab: %s", cn)
    cd = eng.class_config_data.setdefault(cn, {
        "subjects": [], "teacher": "", "teacher_period": 1, "editing_index": None})
    subjects = cd.get("subjects", [])

    # ── Class Teacher ─────────────────────────────────────────────────────────
    with st.container(border=True):
        st.markdown(f"**Class Teacher — {cn}**")
        c1, c2 = st.columns([3, 1])
        with c1:
            opts   = [""] + teachers
            cur_ct = cd.get("teacher", "")
            sel_ct = st.selectbox("Class Teacher", opts,
                                   index=opts.index(cur_ct) if cur_ct in opts else 0,
                                   key=f"ct_{cn}")
            cd["teacher"] = sel_ct
        with c2:
            cur_per = int(cd.get("teacher_period", 1))
            sel_per = st.number_input("CT Period", 1, ppd, cur_per, key=f"ctp_{cn}")
            cd["teacher_period"] = int(sel_per)

    # ── Subject List ──────────────────────────────────────────────────────────
    if subjects:
        with st.container(border=True):
            st.markdown("**Subjects**")
            hdr = st.columns([0.4, 2.2, 2.2, 0.9, 1.0, 1.0, 0.7, 0.7])
            for h, t in zip(hdr, ["#", "Subject", "Teacher", "Periods",
                                   "Consec", "Parallel", "✏", "🗑"]):
                h.markdown(f"**{t}**")
            for i, s in enumerate(subjects):
                row = st.columns([0.4, 2.2, 2.2, 0.9, 1.0, 1.0, 0.7, 0.7])
                row[0].write(str(i + 1))
                row[1].write(s.get("name", "—"))
                row[2].write(s.get("teacher", "—"))
                row[3].write(str(s.get("periods", "")))
                row[4].write("Yes" if s.get("consecutive") == "Yes" else "—")
                row[5].write("✓" if s.get("parallel") else "—")
                if row[6].button("✏", key=f"edit_{cn}_{i}"):
                    log.debug("_class_config_tab: %s edit %d", cn, i)
                    cd["editing_index"] = i
                    # Remove init flag so edit form re-loads correct prefill
                    st.session_state.pop(f"sf_init_{cn}_e{i}", None)
                    st.rerun()
                if row[7].button("🗑", key=f"del_{cn}_{i}"):
                    log.info("_class_config_tab: %s delete %d '%s'", cn, i, s.get("name", ""))
                    subjects.pop(i)
                    ei = cd.get("editing_index")
                    if ei == i:
                        cd["editing_index"] = None
                    elif ei is not None and ei > i:
                        cd["editing_index"] = ei - 1
                    st.rerun()

    # ── Add / Edit Form ───────────────────────────────────────────────────────
    editing_idx = cd.get("editing_index")
    form_ver    = st.session_state.get(f"s2_fv_{cn}", 0)

    if editing_idx is not None and editing_idx < len(subjects):
        suffix     = f"e{editing_idx}"
        prefill    = subjects[editing_idx]
        form_title = f"✏ Edit Subject #{editing_idx + 1}"
    else:
        suffix     = f"a{form_ver}"
        prefill    = {}
        form_title = "➕ Add Subject"

    _s2_init_form(cn, suffix, prefill)

    with st.container(border=True):
        st.markdown(f"**{form_title}**")

        # ── Error display at TOP of form — no scrolling needed ────────────
        form_err_key = f"s2_form_err_{cn}"
        if form_err_key in st.session_state:
            st.error(st.session_state.pop(form_err_key))

        c1, c2, c3 = st.columns(3)
        with c1:
            name = st.text_input("Subject Name",
                                  key=f"sf_name_{cn}_{suffix}")
        with c2:
            opts  = [""] + teachers
            sel_t = st.selectbox("Teacher", opts,
                                  key=f"sf_teach_{cn}_{suffix}")
        with c3:
            pers = st.number_input("Periods/week", 1, ppd * wdays,
                                    key=f"sf_per_{cn}_{suffix}")
        c4, c5 = st.columns(2)
        with c4:
            consec = st.selectbox("Consecutive?", ["No", "Yes"],
                                   key=f"sf_cons_{cn}_{suffix}")
        with c5:
            p_prefs = st.multiselect("Period prefs (optional)", list(range(1, ppd + 1)),
                                      key=f"sf_pref_{cn}_{suffix}")
        d_prefs = st.multiselect("Day prefs (optional)", day_names,
                                  key=f"sf_day_{cn}_{suffix}")
        par = st.checkbox("Parallel teaching?",
                           key=f"sf_par_{cn}_{suffix}")
        par_subj = par_teach = ""
        if par:
            cp1, cp2 = st.columns(2)
            with cp1:
                par_subj = st.text_input("Parallel subject",
                                          key=f"sf_psub_{cn}_{suffix}")
            with cp2:
                opts2     = [""] + teachers
                par_teach = st.selectbox("Parallel teacher", opts2,
                                          key=f"sf_pteach_{cn}_{suffix}")

        btn1, btn2 = st.columns(2)
        with btn1:
            lbl = "✓ Update" if editing_idx is not None else "✓ Add Subject"
            if st.button(lbl, key=f"sf_save_{cn}_{suffix}", type="primary",
                         use_container_width=True):
                log.info("_class_config_tab: %s save '%s'", cn, name)
                if not name.strip():
                    # Store error in session state → shows at TOP on rerun
                    st.session_state[form_err_key] = "⚠ Subject name is required."
                    st.rerun()
                elif not sel_t:
                    st.session_state[form_err_key] = "⚠ Please select a teacher."
                    st.rerun()
                else:
                    entry = {
                        "name":             name.strip(),
                        "teacher":          sel_t,
                        "periods":          int(pers),
                        "consecutive":      consec,
                        "periods_pref":     p_prefs,
                        "days_pref":        d_prefs,
                        "parallel":         par,
                        "parallel_subject": par_subj.strip(),
                        "parallel_teacher": par_teach,
                    }
                    if editing_idx is not None:
                        subjects[editing_idx] = entry
                        cd["editing_index"] = None
                        # Clean up edit-key init flag
                        st.session_state.pop(f"sf_init_{cn}_e{editing_idx}", None)
                        st.toast(f"✓ '{name.strip()}' updated.", icon="✅")
                    else:
                        subjects.append(entry)
                        # ── Form clearing: increment form_ver so the next render
                        # uses a completely NEW key suffix (e.g. a1, a2 ...).
                        # Streamlit has no stored value for the new keys, so
                        # _s2_init_form will initialise them all to empty.
                        # This is more reliable than deleting existing widget keys
                        # because Streamlit may still read from its internal cache.
                        new_ver = form_ver + 1
                        st.session_state[f"s2_fv_{cn}"] = new_ver
                        st.toast(f"✓ '{name.strip()}' added.", icon="✅")
                    cd["editing_index"] = None
                    st.rerun()
        with btn2:
            if editing_idx is not None:
                if st.button("✕ Cancel", key=f"sf_cancel_{cn}",
                             use_container_width=True):
                    cd["editing_index"] = None
                    st.rerun()


def _step2_validate_and_continue():
    """Run all validation checks matching the original tkinter implementation.

    Section ①  Period counts per class.
    Section ②  Teacher-level hard conflicts (same teacher double-booked across classes).
    Section ③  Within-class slot conflicts (two subjects by *different* teachers
               pinned to overlapping period slots in the same class on the same day).
    """
    import math
    eng = _get_eng()
    log.info("_step2_validate_and_continue: running")
    cfg       = eng.configuration
    ppd       = cfg["periods_per_day"]
    wdays     = cfg["working_days"]
    required  = ppd * wdays
    all_cn    = _all_classes()
    day_names = DAY_NAMES[:wdays]

    # ── Section ①: Period counts ─────────────────────────────────────────────
    period_errors = []   # (cn, msg)
    period_ok     = []   # (cn, msg)
    teacher_slots = {}   # teacher -> [slot_descriptor]

    def _add(teacher, desc):
        if teacher:
            teacher_slots.setdefault(teacher, []).append(desc)

    for cn in all_cn:
        cd    = eng.class_config_data.get(cn, {})
        subjs = cd.get("subjects", [])
        if not subjs:
            period_errors.append((cn, "NO SUBJECTS added"))
            log.warning("_step2_validate: %s no subjects", cn)
            continue

        total = sum(s.get("periods", 0) for s in subjs)
        diff  = total - required
        sign  = "+" if diff > 0 else ""
        if total == required:
            period_ok.append((cn, f"{total}/{required} periods  ({len(subjs)} subjects)"))
        else:
            period_errors.append((cn,
                f"Period mismatch: {total} assigned, need {required}  ({sign}{diff})"))
            log.warning("_step2_validate: %s %d≠%d", cn, total, required)

        ct     = cd.get("teacher", "").strip()
        ct_per = cd.get("teacher_period", 1)
        if ct:
            _add(ct, {
                "class":        cn,
                "label":        f"Class Teacher of {cn} (Period {ct_per}, every day)",
                "fixed_period": ct_per,
                "period_prefs": [ct_per],
                "day_set":      set(day_names),
                "is_class_teacher": True,
                "subj_name":    "",
            })

        for s in subjs:
            t = s.get("teacher", "").strip()
            if t:
                _add(t, {
                    "class":        cn,
                    "label":        f"Subject '{s['name']}' in {cn}  (×{s['periods']} periods/wk)",
                    "fixed_period": None,
                    "period_prefs": list(s.get("periods_pref", [])),
                    "day_set":      set(s.get("days_pref", []) or day_names),
                    "is_class_teacher": False,
                    "subj_name":    s["name"],
                    "consecutive":  s.get("consecutive") == "Yes",
                })
            pt = s.get("parallel_teacher", "").strip() if s.get("parallel") else ""
            if pt:
                _add(pt, {
                    "class":        cn,
                    "label":        f"Parallel teacher for '{s.get('parallel_subject','')}' in {cn}",
                    "fixed_period": None,
                    "period_prefs": list(s.get("periods_pref", [])),
                    "day_set":      set(s.get("days_pref", []) or day_names),
                    "is_class_teacher": False,
                    "subj_name":    s.get("parallel_subject", ""),
                })

    # ── Section ②: Teacher-level hard conflicts (across classes) ────────────
    hard_conflicts = []

    def _period_overlap(prefs_a, prefs_b):
        if not prefs_a or not prefs_b:
            return None
        s = set(prefs_a) & set(prefs_b)
        return s if s else None

    for teacher, slots in teacher_slots.items():
        n = len(slots)
        for i in range(n):
            for j in range(i + 1, n):
                a, b = slots[i], slots[j]
                if a["class"] == b["class"]:
                    continue
                days = a["day_set"] & b["day_set"]
                if not days:
                    continue
                # CT fixed period vs subject period pref
                if a["is_class_teacher"] and b["period_prefs"]:
                    if a["fixed_period"] in b["period_prefs"]:
                        hard_conflicts.append({
                            "teacher": teacher, "slot_a": a["label"], "slot_b": b["label"],
                            "reason": (
                                f"Period {a['fixed_period']} is fixed every day for "
                                f"class-teacher duty in {a['class']}, but also listed as "
                                f"preferred period for '{b.get('subj_name','')}' in "
                                f"{b['class']} on days: {', '.join(sorted(days))}")
                        })
                        continue
                if b["is_class_teacher"] and a["period_prefs"]:
                    if b["fixed_period"] in a["period_prefs"]:
                        hard_conflicts.append({
                            "teacher": teacher, "slot_a": b["label"], "slot_b": a["label"],
                            "reason": (
                                f"Period {b['fixed_period']} is fixed every day for "
                                f"class-teacher duty in {b['class']}, but also listed as "
                                f"preferred period for '{a.get('subj_name','')}' in "
                                f"{a['class']} on days: {', '.join(sorted(days))}")
                        })
                        continue
                if a["is_class_teacher"] and b["is_class_teacher"]:
                    hard_conflicts.append({
                        "teacher": teacher, "slot_a": a["label"], "slot_b": b["label"],
                        "reason": (
                            f"Cannot be Class Teacher of two classes simultaneously. "
                            f"Fixed Period {a['fixed_period']} (all days) for {a['class']} "
                            f"AND Period {b['fixed_period']} (all days) for {b['class']}.")
                    })
                    continue
                pov = _period_overlap(a["period_prefs"], b["period_prefs"])
                if pov is not None:
                    hard_conflicts.append({
                        "teacher": teacher, "slot_a": a["label"], "slot_b": b["label"],
                        "reason": (
                            f"Both require Period(s) {sorted(pov)} on "
                            f"{', '.join(sorted(days))} — teacher cannot be in "
                            f"{a['class']} and {b['class']} simultaneously.")
                    })

    log.info("_step2_validate: hard_conflicts=%d", len(hard_conflicts))

    # ── Section ③: Within-class slot conflicts ────────────────────────────
    within_class_conflicts = []

    for cn in all_cn:
        cd     = eng.class_config_data.get(cn, {})
        subjs  = cd.get("subjects", [])

        items = []
        for s in subjs:
            if not s.get("periods_pref"):
                continue
            period_set = set(s["periods_pref"])
            day_set    = set(s.get("days_pref", []) or day_names)
            n_periods  = s["periods"]
            n_days     = len(day_set)
            need_per_day = math.ceil(n_periods / n_days) if n_days > 0 else 1
            items.append({
                "label":        "Subject '{}' (Period(s) {}, day(s) {}, teacher: {})".format(
                                    s["name"],
                                    sorted(period_set),
                                    sorted(day_set) if s.get("days_pref") else "any",
                                    s.get("teacher", "").strip()),
                "period_set":   period_set,
                "day_set":      day_set,
                "need_per_day": need_per_day,
                "teacher":      s.get("teacher", "").strip(),
            })

        if len(items) < 2:
            continue

        for day in day_names:
            active = [it for it in items if day in it["day_set"]]
            if len(active) < 2:
                continue

            # Pairwise check — only flag DIFFERENT-teacher pairs
            for i in range(len(active)):
                for j in range(i + 1, len(active)):
                    a, b = active[i], active[j]
                    if a["teacher"] == b["teacher"]:
                        continue
                    combined_slots = a["period_set"] | b["period_set"]
                    combined_need  = a["need_per_day"] + b["need_per_day"]
                    if combined_need > len(combined_slots):
                        contested = a["period_set"] & b["period_set"]
                        within_class_conflicts.append({
                            "class": cn, "day": day,
                            "item_a": a["label"], "item_b": b["label"],
                            "reason": (
                                f"On {day}, two subjects by DIFFERENT teachers both need "
                                f"Period(s) {sorted(contested) if contested else sorted(combined_slots)} "
                                f"in class {cn} — combined demand "
                                f"({a['need_per_day']} + {b['need_per_day']} = {combined_need}) "
                                f"exceeds available slots {sorted(combined_slots)} — "
                                f"the class grid cannot fit both.")
                        })

            # Whole-group check: any period claimed by > 1 teacher
            all_pref_periods = set()
            for it in active:
                all_pref_periods |= it["period_set"]
            for p in all_pref_periods:
                teachers_needing_p = {it["teacher"] for it in active if p in it["period_set"]}
                if len(teachers_needing_p) > 1:
                    already = any(c["class"] == cn and c["day"] == day
                                  for c in within_class_conflicts)
                    if not already:
                        within_class_conflicts.append({
                            "class": cn, "day": day,
                            "item_a": f"{len(teachers_needing_p)} subjects from different teachers",
                            "item_b": "",
                            "reason": (
                                f"On {day}, Period {p} in class {cn} is claimed by "
                                f"{len(teachers_needing_p)} different teachers "
                                f"({', '.join(sorted(teachers_needing_p))}) — "
                                f"only one teacher can occupy a period slot.")
                        })

    log.info("_step2_validate: errors=%d hc=%d wcc=%d",
             len(period_errors), len(hard_conflicts), len(within_class_conflicts))

    # Filter out errors the user has already chosen to ignore
    ignored = st.session_state.get("s2_ignored_errors", set())
    active_hc  = [c for c in hard_conflicts
                  if _hc_key(c)  not in ignored]
    active_wcc = [c for c in within_class_conflicts
                  if _wcc_key(c) not in ignored]

    all_clear = not period_errors and not active_hc and not active_wcc

    vr = {
        "ok":                    all_clear,
        "period_errors":         period_errors,
        "period_ok":             period_ok,
        "hard_conflicts":        hard_conflicts,        # full list (for display)
        "within_class_conflicts": within_class_conflicts,  # full list (for display)
        "active_hc":             active_hc,            # non-ignored subset
        "active_wcc":            active_wcc,           # non-ignored subset
        "required":              required,
        "wdays":                 wdays,
        "ppd":                   ppd,
    }
    st.session_state["s2_validation_result"] = vr

    if all_clear:
        st.session_state["s2_validated"] = True
        st.session_state["s3_validated"] = False
        ignored_count = (len(hard_conflicts) - len(active_hc) +
                         len(within_class_conflicts) - len(active_wcc))
        if ignored_count:
            _notify(
                f"✅ Validation passed ({ignored_count} error(s) ignored). "
                "Engine will place ignored conflicts best-effort.", "warning")
        else:
            _notify("✅ All validation checks passed.", "success")
        _nav("step3")
    else:
        total_err = len(period_errors) + len(active_hc) + len(active_wcc)
        st.session_state["s2_validated"] = False
        st.session_state["_s2_val_errors"] = vr
        log.warning("_step2_validate: %d active errors found, showing popup", total_err)
        st.rerun()


def _display_s2_validation(vr):
    """Render the full 3-section validation report matching the original."""
    period_errors  = vr.get("period_errors", [])
    period_ok      = vr.get("period_ok", [])
    hard_conflicts = vr.get("hard_conflicts", [])
    wcc            = vr.get("within_class_conflicts", [])
    required       = vr.get("required", "?")
    wdays          = vr.get("wdays", "?")
    ppd            = vr.get("ppd", "?")
    any_error      = bool(period_errors or hard_conflicts or wcc)

    # Quick-count summary bar
    _ign = st.session_state.get("s2_ignored_errors", set())
    _act_hc  = [c for c in hard_conflicts if _hc_key(c)  not in _ign]
    _act_wcc = [c for c in wcc            if _wcc_key(c) not in _ign]
    _n_ign   = (len(hard_conflicts)-len(_act_hc)) + (len(wcc)-len(_act_wcc))

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        if period_errors:
            st.error(f"**{len(period_errors)}**\nPeriod Errors")
        else:
            st.success(f"**0**\nPeriod Errors")
    with c2:
        if _act_hc:
            st.error(f"**{len(_act_hc)}**\nTeacher Conflicts")
        elif hard_conflicts:
            st.warning(f"**{len(hard_conflicts)} ignored**\nTeacher Conflicts")
        else:
            st.success(f"**0**\nTeacher Conflicts")
    with c3:
        if _act_wcc:
            st.error(f"**{len(_act_wcc)}**\nWithin-Class Conflicts")
        elif wcc:
            st.warning(f"**{len(wcc)} ignored**\nWithin-Class Conflicts")
        else:
            st.success(f"**0**\nWithin-Class Conflicts")
    with c4:
        st.success(f"**{len(period_ok)}**\nClasses OK")
    if _n_ign:
        st.info(f"ℹ {_n_ign} error(s) are being **ignored** — "
                "engine will place best-effort for those conflicts. "
                "Scroll down to review or uncheck them.")

    st.divider()

    # ── Section ①: Period Counts ─────────────────────────────────────────────
    s1_label = (f"① PERIOD COUNT — {len(period_ok)} OK / {len(period_errors)} ERRORS")
    with st.expander(s1_label, expanded=bool(period_errors)):
        st.caption(f"Required per class: **{required}** periods/week  "
                   f"({wdays} days × {ppd} periods)")
        if period_errors:
            st.markdown("**❌ Fix these classes:**")
            for cn, msg in sorted(period_errors):
                st.error(f"**{cn}** — {msg}")
        if period_ok:
            st.markdown("**✓ Classes with correct period count:**")
            cols = st.columns(3)
            for idx, (cn, msg) in enumerate(sorted(period_ok)):
                with cols[idx % 3]:
                    st.success(f"**{cn}**: {msg}")

    # ── Section ②: Teacher Conflicts (across classes) ────────────────────────
    ignored = st.session_state.get("s2_ignored_errors", set())
    active_hc  = vr.get("active_hc",  [c for c in hard_conflicts  if _hc_key(c)  not in ignored])
    active_wcc = vr.get("active_wcc", [c for c in wcc             if _wcc_key(c) not in ignored])
    n_ign_hc   = len(hard_conflicts) - len(active_hc)
    n_ign_wcc  = len(wcc)            - len(active_wcc)

    s2_label = (f"② TEACHER CONFLICTS — {len(hard_conflicts)} found  "
                f"({len(active_hc)} active, {n_ign_hc} ignored)")
    with st.expander(s2_label, expanded=bool(active_hc)):
        if not hard_conflicts:
            st.success("✓ No teacher-level hard conflicts detected.")
        else:
            if active_hc:
                st.error("⚠ These conflicts mean a teacher is assigned to two places at "
                         "the same time. Fix or check **Ignore** to let the engine place best-effort.")
            else:
                st.success(f"✓ All {len(hard_conflicts)} conflict(s) have been ignored.")
            st.caption("🔲 = active (blocking)   ✅ = ignored (engine places best-effort)")
            for i, c in enumerate(hard_conflicts, 1):
                key = _hc_key(c)
                already_ignored = key in ignored
                with st.container(border=True):
                    hdr_col, tog_col = st.columns([0.85, 0.15])
                    with hdr_col:
                        icon = "✅" if already_ignored else "🔲"
                        st.markdown(f"**{icon} [{i}] Teacher: {c['teacher']}**")
                        st.caption(f"A: {c['slot_a']}")
                        st.caption(f"B: {c['slot_b']}")
                        if already_ignored:
                            st.success(f"IGNORED — {c['reason'][:160]}")
                        else:
                            st.warning(f"⚠ {c['reason']}")
                            st.info("💡 **FIX:** Change the period/day preference, "
                                    "or check Ignore →")
                    with tog_col:
                        new_val = st.checkbox("Ignore",
                                              value=already_ignored,
                                              key=f"disp_ign_hc_{i}",
                                              help="Ignore this conflict and allow proceeding")
                    if new_val != already_ignored:
                        if new_val:
                            ignored.add(key)
                        else:
                            ignored.discard(key)
                        st.session_state["s2_ignored_errors"] = ignored
                        # Recompute active sets and update vr
                        vr["active_hc"]  = [c2 for c2 in hard_conflicts
                                             if _hc_key(c2)  not in ignored]
                        vr["active_wcc"] = [c2 for c2 in wcc
                                             if _wcc_key(c2) not in ignored]
                        vr["ok"] = (not period_errors and
                                    not vr["active_hc"] and not vr["active_wcc"])
                        st.session_state["s2_validation_result"] = vr
                        st.rerun()

    # ── Section ③: Within-Class Slot Conflicts ───────────────────────────────
    s3_label = (f"③ WITHIN-CLASS SLOT CONFLICTS — {len(wcc)} found  "
                f"({len(active_wcc)} active, {n_ign_wcc} ignored)")
    with st.expander(s3_label, expanded=bool(active_wcc)):
        if not wcc:
            st.success("✓ No within-class slot conflicts detected.")
        else:
            if active_wcc:
                st.error("⚠ Two or more subjects are pinned to the same slot. "
                         "Fix or check **Ignore** to let the engine place best-effort.")
            else:
                st.success(f"✓ All {len(wcc)} conflict(s) have been ignored.")
            st.caption("🔲 = active (blocking)   ✅ = ignored (engine places best-effort)")
            for i, c in enumerate(wcc, 1):
                key = _wcc_key(c)
                already_ignored = key in ignored
                with st.container(border=True):
                    hdr_col, tog_col = st.columns([0.85, 0.15])
                    with hdr_col:
                        icon = "✅" if already_ignored else "🔲"
                        st.markdown(f"**{icon} [{i}] Class: {c['class']}  —  Day: {c['day']}**")
                        st.caption(f"Item A: {c['item_a']}")
                        if c.get("item_b"):
                            st.caption(f"Item B: {c['item_b']}")
                        if already_ignored:
                            st.success(f"IGNORED — {c['reason'][:160]}")
                        else:
                            st.warning(f"⚠ {c['reason']}")
                            st.info("💡 **FIX:** Adjust period/day preferences, "
                                    "or check Ignore →")
                    with tog_col:
                        new_val = st.checkbox("Ignore",
                                              value=already_ignored,
                                              key=f"disp_ign_wcc_{i}",
                                              help="Ignore this conflict and allow proceeding")
                    if new_val != already_ignored:
                        if new_val:
                            ignored.add(key)
                        else:
                            ignored.discard(key)
                        st.session_state["s2_ignored_errors"] = ignored
                        vr["active_hc"]  = [c2 for c2 in hard_conflicts
                                             if _hc_key(c2)  not in ignored]
                        vr["active_wcc"] = [c2 for c2 in wcc
                                             if _wcc_key(c2) not in ignored]
                        vr["ok"] = (not period_errors and
                                    not vr["active_hc"] and not vr["active_wcc"])
                        st.session_state["s2_validation_result"] = vr
                        st.rerun()

    # ── Result footer ─────────────────────────────────────────────────────────
    st.divider()
    total_ignored = n_ign_hc + n_ign_wcc
    total_active  = len(period_errors) + len(active_hc) + len(active_wcc)
    if period_errors:
        st.error(f"**RESULT:** {len(period_errors)} period mismatch(es) must be fixed — "
                 "they cannot be ignored. Fix then click **✓ Validate & Complete** again.")
    elif total_active > 0:
        st.warning(f"**RESULT:** {total_active} active error(s). "
                   f"Check **Ignore** on errors you want to skip, "
                   "then click **✓ Validate & Complete** again.")
    else:
        if total_ignored:
            st.warning(f"**RESULT:** All checks passed ({total_ignored} error(s) ignored). "
                       "Click **✓ Validate & Complete** to proceed to Step 3.")
        else:
            st.success("**RESULT:** All checks passed. "
                       "Click **✓ Validate & Complete** to proceed to Step 3.")

    # Inline "Clear all ignored" link at the bottom of the report
    if total_ignored > 0:
        if st.button(f"↺ Clear all {total_ignored} ignored error(s)",
                     key="s2_clear_ignored"):
            st.session_state["s2_ignored_errors"] = set()
            vr["active_hc"]  = list(hard_conflicts)
            vr["active_wcc"] = list(wcc)
            vr["ok"] = (not period_errors and
                        not hard_conflicts and not wcc)
            st.session_state["s2_validation_result"] = vr
            st.rerun()


# ═════════════════════════════════════════════════════════════════════════════
#  STEP 3
# ═════════════════════════════════════════════════════════════════════════════
def page_step3():
    eng = _get_eng()

    if not eng.configuration:
        st.warning("⚠ Step 1 not yet completed — basic configuration is missing.")
        st.info("Please complete Step 1 (Basic Config) before continuing.")
        if st.button("← Go to Step 1", type="primary", key="s3_guard_back"):
            _nav("step1")
        return

    log.info("page_step3: render")
    _header("⚙ Step 3: Teacher Settings",
            "Review workload, define combined classes and mark teacher unavailability.")
    _show_notifications()

    cfg       = eng.configuration
    ppd       = cfg["periods_per_day"]
    wdays     = cfg["working_days"]
    teachers  = sorted(cfg.get("teacher_names",[]))
    day_names = DAY_NAMES[:wdays]
    all_cn    = _all_classes()

    if not hasattr(eng,"step3_data"):           eng.step3_data = {}
    if not hasattr(eng,"step3_unavailability"): eng.step3_unavailability = {}

    # Always keep workload attributes fresh so validate/workload tab are consistent.
    # This mirrors _build_step3_ui() calling _compute_teacher_workload() on every render.
    try:
        eng.prepare_step3_workload()
    except Exception as ex:
        log.warning("page_step3: prepare_step3_workload failed: %s", ex)

    with st.expander("💾 Save / Load Step 3 Config", expanded=False):
        c1,c2 = st.columns(2)
        with c1:
            if st.button("⬇ Prepare Download", key="s3_dl_btn", use_container_width=True):
                log.info("page_step3: preparing download")
                st.session_state["_s3_dl_data"] = {
                    "step3_data": {t:{"skipped":v.get("skipped",False),"combines":v.get("combines",[])}
                                   for t,v in _get_eng().step3_data.items()},
                    "step3_unavailability": {t:{"days":list(v["days"]),"periods":list(v["periods"])}
                                             for t,v in _get_eng().step3_unavailability.items()},
                    "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "step":3,
                }
            if "_s3_dl_data" in st.session_state:
                _json_download(st.session_state["_s3_dl_data"], "📥 Click to Download",
                               f"Step3_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
        with c2:
            up = st.file_uploader("Upload Step 3 JSON", type=["json"], key="s3_load_json")
            if up is not None:
                raw = up.read()
                log.debug("page_step3: upload size=%d", len(raw))
                if len(raw)==0:
                    log.warning("page_step3: 0 bytes, skipping")
                elif not _already_processed("s3_load_json", raw):
                    _load_step3_config(raw)

    # Show step 3 error popup if triggered
    if st.session_state.get("_s3_show_error_popup"):
        st.session_state["_s3_show_error_popup"] = False
        vr_err = st.session_state.get("s3_validation_result")
        if vr_err and not vr_err.get("can_proceed"):
            _step3_error_dialog(vr_err)

    cb, cv, cc = st.columns(3)
    with cb:
        if st.button("← Back to Step 2", use_container_width=True):
            _nav("step2")
    with cv:
        if st.button("🔍 Validate Step 3", use_container_width=True):
            log.info("page_step3: validate")
            try:
                eng.prepare_step3_workload()
                vr = eng.validate_step3()
                st.session_state["s3_validation_result"] = vr
                log.info("page_step3: validate result can_proceed=%s issues=%d",
                         vr["can_proceed"], len(vr["issues"]))
                if vr["can_proceed"]:
                    st.session_state["s3_validated"] = True
                    _notify("✅ All clear — click Generate Timetable to proceed.", "success")
                else:
                    st.session_state["s3_validated"] = False
                    st.session_state["_s3_show_error_popup"] = True
                    log.warning("page_step3: %d overloads, showing popup", len(vr["issues"]))
            except Exception as ex:
                log.error("page_step3 validate: %s\n%s", ex, traceback.format_exc())
                _notify(f"Validation error: {ex}", "error")
            st.rerun()
    with cc:
        s3_validated = st.session_state.get("s3_validated", False)
        if s3_validated:
            if st.button("🚀 Generate Timetable →", type="primary", use_container_width=True):
                log.info("page_step3: proceed to generate")
                st.session_state["s3_validated"] = True
                for k in ("s4_stage","s4_s1_status","ta_allocation","ta2_allocation",
                          "s4_s3_status","s4_ff_result","gen_result"):
                    st.session_state[k] = 0 if k == "s4_stage" else None
                _nav("generate")
        else:
            st.button("🚀 Generate Timetable →", type="primary",
                      use_container_width=True, disabled=True,
                      help="Run Validate Step 3 first and resolve any overloads")

    # Status message
    s3v = st.session_state.get("s3_validated", False)
    if s3v:
        st.success("✅ Step 3 validated — click **Generate Timetable** to proceed.")
    else:
        st.info("ℹ️ Review teacher workloads, set up combines, then click **🔍 Validate Step 3**.")

    vr = st.session_state.get("s3_validation_result")
    if vr and not vr.get("can_proceed") and vr.get("issues"):
        with st.expander("❌ Still Overloaded — expand for details", expanded=False):
            for ln in vr["issues"]:
                st.error(ln)
    if vr and vr.get("resolved"):
        with st.expander("✓ Resolved / Skipped"):
            for ln in vr["resolved"]: st.success(ln)

    st.divider()
    tab_wl, tab_cb, tab_un = st.tabs(
        ["📊 Teacher Workload","🔗 Combine Classes","🚫 Unavailability"])
    with tab_wl: _render_workload(teachers, wdays, ppd)
    with tab_cb: _render_combine_tab(teachers, all_cn)
    with tab_un: _render_unavailability_tab(teachers, day_names, ppd)


def _load_step3_config(raw: bytes):
    log.info("_load_step3_config: %d bytes", len(raw))
    try:
        d = json.loads(raw.decode("utf-8"))
    except json.JSONDecodeError as ex:
        log.error("_load_step3_config: JSON parse error: %s", ex)
        st.toast(f"❌ Invalid JSON file: {ex}", icon="🚫")
        _notify(f"❌ Invalid JSON: {ex}", "error")
        st.rerun()
        return

    # ── FIX 3: Validate file structure ──────────────────────────────────────
    if "step3_data" not in d and "step3_unavailability" not in d:
        wrong_hint = ""
        if "assignments" in d:
            wrong_hint = " (looks like a Step 2 Assignments file)"
        elif "periods_per_day" in d or "teacher_names" in d:
            wrong_hint = " (looks like a Step 1 config file)"
        msg = (f"❌ Wrong file type{wrong_hint} — expected a Step 3 config file with "
               f"'step3_data' or 'step3_unavailability' keys.")
        log.error("_load_step3_config: %s", msg)
        st.toast(msg, icon="🚫")
        _notify(msg, "error")
        st.rerun()
        return

    try:
        _e3 = _get_eng()
        _e3.step3_data = d.get("step3_data", {})
        _e3.step3_unavailability = {
            t: {"days": v.get("days", []), "periods": v.get("periods", [])}
            for t, v in d.get("step3_unavailability", {}).items()
        }
        _notify("✓ Step 3 config loaded.", "success")
        log.info("_load_step3_config: %d teachers, %d unavail",
                 len(_e3.step3_data), len(_e3.step3_unavailability))
        st.rerun()
    except Exception as ex:
        log.error("_load_step3_config: %s\n%s", ex, traceback.format_exc())
        _notify(f"Failed: {ex}", "error")


def _render_workload(teachers, wdays, ppd):
    """Render teacher workload with Edit/Combine and Skip buttons per teacher."""
    eng = _get_eng()
    log.debug("_render_workload: %d teachers", len(teachers))

    max_allowed = (ppd - 2) * wdays
    wl          = getattr(eng, "_step3_teacher_wl", {})

    if not wl:
        n_cls = sum(1 for cn in eng.class_config_data
                    if eng.class_config_data[cn].get('subjects'))
        st.warning(
            f"No teacher assignments found. Classes with subjects: {n_cls}. "
            "Make sure teachers are assigned to subjects in Step 2.")
        return

    st.caption(
        f"Overload threshold: **{max_allowed}** periods/week  "
        f"({wdays} days × ({ppd}−2) periods).  "
        f"All teachers shown — overloaded ones highlighted in red.")

    overloaded_teachers = sorted(t for t in wl if t in eng._step3_overloaded)
    normal_teachers     = sorted(t for t in wl if t not in eng._step3_overloaded)

    sel_teacher = st.session_state.get("s3_selected_teacher", "")

    def _teacher_card(teacher):
        info      = wl.get(teacher, {"total": 0, "entries": []})
        effective = eng._effective_total(teacher)
        s3d       = eng.step3_data.get(teacher, {})
        skipped   = s3d.get("skipped", False)
        n_comb    = len(s3d.get("combines", []))
        is_over   = teacher in eng._step3_overloaded
        still_over = is_over and effective > max_allowed

        # Card colour
        if teacher == sel_teacher:
            border_style = "border:2px solid #2980b9; padding:8px; border-radius:6px; background:#eaf4fb"
        elif skipped:
            border_style = "border:1px solid #ccc; padding:8px; border-radius:6px; background:#e8f5e9"
        elif still_over:
            border_style = "border:1px solid #e74c3c; padding:8px; border-radius:6px; background:#fdecea"
        elif is_over and effective <= max_allowed:
            border_style = "border:1px solid #f39c12; padding:8px; border-radius:6px; background:#fff8e1"
        else:
            border_style = "border:1px solid #ccc; padding:8px; border-radius:6px; background:#fff"

        # Badge
        if skipped:
            badge = "SKIPPED"
        elif still_over:
            badge = "OVERLOADED"
        elif n_comb and not still_over:
            badge = f"{n_comb} combine{'s' if n_comb>1 else ''}"
        else:
            badge = "OK"

        stat = f"Assigned: **{info['total']}**"
        if effective != info["total"]:
            stat += f"  →  Effective: **{effective}**"
        stat += f"  /  Max: **{max_allowed}**"

        with st.container():
            st.markdown(f"<div style='{border_style}'>", unsafe_allow_html=True)
            h1, h2 = st.columns([3, 1])
            with h1:
                st.markdown(f"**{teacher}**")
                st.caption(stat)
            with h2:
                badge_color = "#c0392b" if still_over else ("#27ae60" if not is_over or skipped else "#f39c12")
                st.markdown(
                    f"<span style='color:{badge_color};font-weight:bold;font-size:0.8rem'>{badge}</span>",
                    unsafe_allow_html=True)

            b1, b2 = st.columns(2)
            with b1:
                btn_label = "✏ Close Detail" if teacher == sel_teacher else "✏ Edit / Combine"
                if st.button(btn_label, key=f"ec_{teacher}", use_container_width=True):
                    if teacher == sel_teacher:
                        st.session_state.pop("s3_selected_teacher", None)
                    else:
                        st.session_state["s3_selected_teacher"] = teacher
                        # Clear any stale checkboxes for this teacher
                        for k in list(st.session_state.keys()):
                            if k.startswith(f"s3cb_{teacher}_"):
                                del st.session_state[k]
                    st.rerun()
            with b2:
                skip_lbl = "↩ Un-skip" if skipped else "⏭ Skip"
                if st.button(skip_lbl, key=f"sk_{teacher}", use_container_width=True):
                    eng.step3_data.setdefault(teacher, {"skipped": False, "combines": []})
                    eng.step3_data[teacher]["skipped"] = not skipped
                    eng.prepare_step3_workload()
                    st.rerun()
            st.markdown("</div>", unsafe_allow_html=True)

    if overloaded_teachers:
        st.markdown(f"#### ⚠ Overloaded  (>{max_allowed}/wk)")
        for t in overloaded_teachers:
            _teacher_card(t)

    if normal_teachers:
        st.markdown(f"#### ✓ Within limit  (≤{max_allowed}/wk)")
        for t in normal_teachers:
            _teacher_card(t)

    # ── Detail panel for selected teacher ─────────────────────────────────────
    if sel_teacher and sel_teacher in wl:
        st.divider()
        _render_teacher_combine_detail(sel_teacher)


def _render_teacher_combine_detail(teacher):
    """Full assignment + combine panel for a selected teacher."""
    eng = _get_eng()
    # ── Combine error: show as modal dialog (not top-of-page) ─────────────────
    combine_err = st.session_state.pop("_s3_combine_err", None)
    if combine_err:
        _combine_error_dialog(combine_err)
        return  # dialog is open; rest of page renders behind it
    wl        = getattr(eng, "_step3_teacher_wl", {})
    info      = wl.get(teacher, {"total": 0, "entries": []})
    entries   = info["entries"]
    s3d       = eng.step3_data.setdefault(teacher, {"skipped": False, "combines": []})
    max_all   = getattr(eng, "_step3_max_allowed", 9999)
    effective = eng._effective_total(teacher)
    is_over   = teacher in getattr(eng, "_step3_overloaded", set())
    still_over = is_over and effective > max_all

    # ── Header strip ──────────────────────────────────────────────────────────
    hdr_color = "#c0392b" if still_over else "#1a7a1a"
    st.markdown(
        f"<div style='background:{hdr_color};color:white;padding:10px 16px;"
        f"border-radius:6px;display:flex;justify-content:space-between;align-items:center'>"
        f"<span style='font-size:1.1rem;font-weight:bold'>{teacher}</span>"
        f"<span style='font-size:0.95rem;font-weight:bold'>"
        f"Assigned: {info['total']}  |  Effective: {effective}  |  Max: {max_all}"
        f"</span></div>",
        unsafe_allow_html=True)
    st.write("")

    # Track which entry indices are already in a combine
    combined_indices = set()
    for cb in s3d["combines"]:
        for idx in cb.get("entry_indices", []):
            combined_indices.add(idx)

    left_col, right_col = st.columns([1, 1])

    # ── LEFT: Assignment checkboxes ───────────────────────────────────────────
    with left_col:
        st.markdown("**Assignments — check entries to combine**")
        if not entries:
            st.info("No assignments found for this teacher.")
        for ei, entry in enumerate(entries):
            in_cb = ei in combined_indices
            ct_info = eng.get_class_ct_info(entry["class"], teacher, entry["subject"])
            ct      = ct_info["ct"]
            ct_subs = ct_info["ct_subjects"]
            has_parallel_warn = ct_info["is_parallel_with_ct"]

            with st.container(border=True):
                cc1, cc2 = st.columns([1, 8])
                with cc1:
                    if in_cb:
                        st.markdown("✅")
                    else:
                        st.checkbox(
                            "", key=f"s3cb_{teacher}_{ei}",
                            value=st.session_state.get(f"s3cb_{teacher}_{ei}", False))
                with cc2:
                    label_color = "#1a7a1a" if in_cb else "#1a1a1a"
                    st.markdown(
                        f"<span style='color:{label_color};font-weight:bold'>"
                        f"[{entry['class']}]  {entry['subject']}  ({entry['periods']} periods)"
                        f"{'  ✓ in combine' if in_cb else ''}</span>",
                        unsafe_allow_html=True)
                    ct_subs_str = ", ".join(ct_subs) if ct_subs else "—"
                    st.caption(f"CT: {ct or '—'}   |   CT Subjects: {ct_subs_str}")
                    if has_parallel_warn:
                        st.markdown(
                            f"<span style='color:#c0392b;font-style:italic;font-size:0.82rem'>"
                            f"⚠  '{entry['subject']}' (teacher) ∥ "
                            f"'{ct_info['parallel_ct_subject']}' (CT) — parallel conflict!"
                            f"</span>",
                            unsafe_allow_html=True)

    # ── RIGHT: Combine Checked Entries + existing combines ────────────────────
    with right_col:
        st.markdown("**Combines**")

        if st.button("✓ Combine Checked Entries", key=f"do_combine_{teacher}",
                     type="primary", use_container_width=True):
            # Collect checked indices
            idxs = [ei for ei in range(len(entries))
                    if st.session_state.get(f"s3cb_{teacher}_{ei}", False)
                    and ei not in combined_indices]

            err = None
            if len(idxs) < 2:
                err = "Check at least 2 assignments (that are not already in a combine)."
            else:
                periods_list = [entries[i]["periods"] for i in idxs]
                if len(set(periods_list)) > 1:
                    err = (f"All entries to combine must have the same period count. "
                           f"Selected: {periods_list}")
                else:
                    # Parallel-CT conflict check
                    blocked = []
                    for i in idxs:
                        e = entries[i]
                        ct_i = eng.get_class_ct_info(e["class"], teacher, e["subject"])
                        if ct_i["is_parallel_with_ct"]:
                            blocked.append(
                                f"Class {e['class']}: '{e['subject']}' ({teacher}) ∥ "
                                f"'{ct_i['parallel_ct_subject']}' (CT: {ct_i['ct']})")
                    if blocked:
                        err = (
                            "Cannot combine — parallel-CT conflict detected:\n\n"
                            + "\n".join(f"  • {b}" for b in blocked)
                            + "\n\nThe CT has a fixed period each day; combining would "
                            "force the parallel subject into the same fixed slot.")

            if err:
                st.session_state["_s3_combine_err"] = err
            else:
                s3d["combines"].append({
                    "entry_indices": idxs,
                    "periods_each":  entries[idxs[0]]["periods"],
                    "classes":       [entries[i]["class"]   for i in idxs],
                    "subjects":      [entries[i]["subject"] for i in idxs],
                })
                # Clear checkboxes
                for ei in range(len(entries)):
                    st.session_state.pop(f"s3cb_{teacher}_{ei}", None)
                eng.prepare_step3_workload()
                log.info("_render_teacher_combine_detail: combine added for %s: %s",
                         teacher, [entries[i]["class"] for i in idxs])
                _notify(f"✓ Combine added for {teacher}.", "success")
            st.rerun()

        st.write("")
        if not s3d["combines"]:
            st.info("No combines yet.\nCheck assignments on the left and click Combine.")
        else:
            for ci, cb in enumerate(s3d["combines"]):
                with st.container(border=True):
                    classes_str = "  +  ".join(cb.get("classes", []))
                    st.markdown(f"**Combine {ci+1}:  {classes_str}**")
                    for i, idx in enumerate(cb.get("entry_indices", [])):
                        if idx < len(entries):
                            st.caption(f"• {entries[idx]['label']}")
                        else:
                            cls = cb["classes"][i] if i < len(cb.get("classes",[])) else "?"
                            sub = cb["subjects"][i] if i < len(cb.get("subjects",[])) else "?"
                            st.caption(f"• '{sub}' in {cls}  x{cb.get('periods_each','?')}/wk")
                    saving = (len(cb.get("entry_indices", [])) - 1) * cb.get("periods_each", 0)
                    st.caption(f"💡 Saves {saving} periods/week for {teacher}")
                    if st.button("✕ Remove", key=f"rm_cb_{teacher}_{ci}",
                                 use_container_width=True):
                        s3d["combines"].pop(ci)
                        eng.prepare_step3_workload()
                        log.info("_render_teacher_combine_detail: removed combine %d for %s",
                                 ci, teacher)
                        st.rerun()


def _render_combine_tab(teachers, all_cn):
    """Shows a summary of all existing combines across all teachers."""
    eng = _get_eng()
    log.debug("_render_combine_tab")
    st.info("💡 To **add** combines, go to the **Teacher Workload** tab and click "
            "**Edit / Combine** next to any teacher.")
    st.markdown("---")
    st.markdown("**All Existing Combines**")
    any_cb = False
    for teacher in sorted(eng.step3_data.keys()):
        s3d = eng.step3_data[teacher]
        cbs = s3d.get("combines", [])
        if not cbs:
            continue
        any_cb = True
        wl   = getattr(eng, "_step3_teacher_wl", {})
        info = wl.get(teacher, {"entries": []})
        entries = info["entries"]

        st.markdown(f"**{teacher}**")
        for ci, cb in enumerate(cbs):
            c1, c2 = st.columns([5, 1])
            with c1:
                classes_str  = " + ".join(cb.get("classes", []))
                subjects_str = ", ".join(cb.get("subjects", []))
                saving = (len(cb.get("entry_indices", [])) - 1) * cb.get("periods_each", 0)
                st.write(f"  📌 **{classes_str}**  ·  {subjects_str}  "
                         f"  *(saves {saving} periods/wk)*")
            with c2:
                if st.button("🗑", key=f"del_cbt_{teacher}_{ci}",
                             help="Remove this combine"):
                    cbs.pop(ci)
                    eng.prepare_step3_workload()
                    log.info("_render_combine_tab: del %d for %s", ci, teacher)
                    st.rerun()
    if not any_cb:
        st.info("No combines defined yet.")


def _render_unavailability_tab(teachers, day_names, ppd):
    eng = _get_eng()
    log.debug("_render_unavailability_tab")
    unavail = eng.step3_unavailability
    with st.container(border=True):
        st.markdown("**Add / Update Unavailability**")
        sel_t = st.selectbox("Teacher", [""]+teachers, key="un_teacher")
        c1,c2 = st.columns(2)
        with c1: sel_days    = st.multiselect("Unavailable Days",    day_names,           key="un_days")
        with c2: sel_periods = st.multiselect("Unavailable Periods", list(range(1,ppd+1)), key="un_periods")
        cs,ck,cl = st.columns(3)
        with cs:
            if st.button("✓ Save", key="un_save"):
                if not sel_t: _notify("Select teacher.","warning")
                elif not sel_days or not sel_periods: _notify("Select days and periods.","warning")
                else:
                    ok, msg = eng._check_unavailability_feasible(sel_t, sel_days, sel_periods)
                    if not ok:
                        log.warning("_render_unavailability_tab: feasibility fail %s: %s", sel_t, msg)
                        _notify(f"❌ Feasibility failed: {msg}","error")
                    else:
                        unavail[sel_t] = {"days":sel_days,"periods":sel_periods}
                        log.info("_render_unavailability_tab: saved %s days=%s", sel_t, sel_days)
                        _notify(f"✓ Saved for {sel_t}.","success"); st.rerun()
        with ck:
            if st.button("🔍 Check Only", key="un_check"):
                if sel_t and sel_days and sel_periods:
                    ok,msg = eng._check_unavailability_feasible(sel_t,sel_days,sel_periods)
                    log.info("_render_unavailability_tab: check %s → ok=%s", sel_t, ok)
                    if ok: _notify(f"✓ Feasible: {msg}","success")
                    else:  _notify(f"❌ {msg}","error")
                    st.rerun()
        with cl:
            if st.button("✕ Clear", key="un_clear"): st.rerun()
    _show_notifications()
    st.markdown("**Current Unavailability**")
    if not unavail:
        st.info("No unavailability set.")
    else:
        for teacher, info in sorted(unavail.items()):
            ok,short = eng._check_unavailability_feasible(teacher,info.get("days",[]),info.get("periods",[]))
            days_str = ", ".join(info.get("days",[]))
            pers_str = ", ".join(f"P{p}" for p in sorted(info.get("periods",[])))
            with st.expander(f"{'🟢' if ok else '🔴'} **{teacher}** — {days_str}  |  {pers_str}", expanded=False):
                st.write(f"Days: {days_str}  ·  Periods: {pers_str}")
                if not ok: st.error(short)
                if st.button("🗑 Remove", key=f"del_un_{teacher}"):
                    log.info("_render_unavailability_tab: remove %s", teacher)
                    del unavail[teacher]; st.rerun()


# ═════════════════════════════════════════════════════════════════════════════
#  GENERATE PAGE  (Step 3 → Final Timetable, fully automatic)
# ═════════════════════════════════════════════════════════════════════════════
def page_generate():
    eng = _get_eng()

    if not eng.configuration:
        st.warning("⚠ Step 1 not yet completed — basic configuration is missing.")
        st.info("Please complete Step 1 (Basic Config) before continuing.")
        if st.button("← Go to Step 1", type="primary", key="gen_guard_back"):
            _nav("step1")
        return

    log.info("page_generate: render")
    cfg   = eng.configuration
    ppd   = cfg.get("periods_per_day", "?")
    wdays = cfg.get("working_days", "?")
    _header("🚀 Generating Timetable",
            f"{wdays} days/week · {ppd} periods/day — fully automatic")
    _show_notifications()

    col_back, _spacer = st.columns([1, 2])
    with col_back:
        if st.button("← Back to Step 3"):
            _nav("step3")

    gen_result = st.session_state.get("gen_result")

    # ── RUN ─────────────────────────────────────────────────────────────────
    if gen_result is None:
        st.info("Click **Generate** to produce the complete timetable automatically.\n\n"
                "The engine will run all stages, apply backtracking, and—if needed—"
                "automatically reduce over-constrained subjects by 1 period to resolve "
                "deadlocks.")

        if st.button("⚡ Generate Timetable", type="primary",
                     key="gen_run_btn", use_container_width=True):
            log.info("page_generate: starting full generation")
            progress_placeholder = st.empty()
            progress_log = []

            def _progress(msg):
                # _gen_prog stores (msg, pct) tuples — normalise to str
                if isinstance(msg, tuple):
                    msg = msg[0]
                msg = str(msg)
                progress_log.append(msg)
                if msg.strip():
                    lines = [m for m in progress_log[-6:] if m.strip()]
                    progress_placeholder.info("⏳  " + "\n\n".join(lines))

            with st.spinner("🚀 Generating timetable — please wait…"):
                try:
                    result = eng.run_full_generation(progress_cb=_progress)
                    st.session_state["gen_result"] = result
                    log.info("page_generate: done ok=%s remaining=%s",
                             result["ok"], result["remaining"])
                    if result["ok"]:
                        _notify("✅ Timetable generated — all periods placed!", "success")
                    else:
                        _notify(
                            f"⚠ Timetable generated with {result['remaining']} "
                            f"period(s) unplaced. See details below.", "warning")
                except Exception as ex:
                    log.error("page_generate: %s\n%s", ex, traceback.format_exc())
                    _notify(f"Generation error: {ex}", "error")

            progress_placeholder.empty()
            st.rerun()
        return   # don't render result section yet

    # ── RESULT ──────────────────────────────────────────────────────────────
    remaining   = gen_result.get("remaining", 0)
    overloaded  = gen_result.get("overloaded", [])
    blocked     = gen_result.get("blocked_only", [])
    reductions  = gen_result.get("period_reductions", [])
    prog_log    = gen_result.get("progress_log", [])
    wdays_r     = gen_result.get("wdays", "?")
    ppd_r       = gen_result.get("ppd", "?")
    total_slots = gen_result.get("total_slots", "?")

    if remaining == 0:
        st.success("✅ **Timetable is 100% complete — all periods placed!**")
    else:
        st.error(f"⚠  **{remaining} period(s) could not be placed.**  "
                 f"The timetable is as complete as possible given the constraints.")

    # ── CT violation summary ────────────────────────────────────────────────
    ct_viols = gen_result.get("ct_violations", [])
    if ct_viols:
        with st.expander(
                f"🔴 CT Period Violations — {len(ct_viols)} slot(s) incorrect",
                expanded=True):
            st.error(
                "The engine attempted to repair these CT period violations. "
                "If any remain, re-run generation or adjust Step 2 preferences.")
            for v in ct_viols:
                st.warning(
                    f"**{v['class']}** | {v['day']} P{v['period_1based']} "
                    f"— expected **{v['expected_subject']}** "
                    f"(CT: {v['ct_teacher']}), "
                    f"found: **{v['actual_subject']}**")
    else:
        st.success("🔒 CT periods verified — all class-teacher slots are correct.")

    # Period reductions report
    if reductions:
        with st.expander(f"📉 Period Reductions Applied ({len(reductions)})",
                         expanded=True):
            st.caption(
                "The engine automatically reduced these subjects by 1 period/week "
                "to break deadlocks. You may want to review these in Step 2.")
            for r in reductions:
                st.warning(
                    f"**{r['subject']}** in **{r['class']}** "
                    f"(teacher: {r['teacher']})  —  "
                    f"{r['from_periods']} → {r['to_periods']} periods/week")

    # Blocked / overloaded teachers
    if overloaded:
        st.markdown("#### ❌ Overloaded Teachers")
        st.caption(f"Grid capacity: {wdays_r} × {ppd_r} = {total_slots} slots")
        for tname, assigned, cap, excess, unp in overloaded:
            st.container(border=True)
            col1, col2 = st.columns([2, 3])
            with col1: st.markdown(f"**❌ {tname}**")
            with col2:
                st.markdown(f"Assigned: **{assigned}** | Capacity: **{cap}** "
                            f"| Excess: **+{excess}** | Unplaced: **{unp}**")

    if blocked:
        st.markdown("#### ⚠ Blocked Teachers *(capacity OK, but all slots clash)*")
        for tname, assigned, cap, unp in blocked:
            with st.container(border=True):
                col1, col2 = st.columns([2, 3])
                with col1: st.markdown(f"**⚠ {tname}**")
                with col2:
                    st.markdown(f"Assigned: **{assigned} / {cap}** | Unplaced: **{unp}**")

    # Action buttons
    st.divider()
    col_view, col_rerun = st.columns(2)
    with col_view:
        if st.button("📊 View Final Timetable →", type="primary",
                     use_container_width=True, key="gen_view_btn"):
            _nav("final_timetable")
    with col_rerun:
        if st.button("🔄 Re-run Generation", use_container_width=True,
                     key="gen_rerun_btn"):
            st.session_state["gen_result"] = None
            st.rerun()

    # Progress log expander
    if prog_log:
        with st.expander("📋 Generation Log", expanded=False):
            # Normalise: _gen_prog stores (msg, pct) tuples; keep only the text
            log_lines = [
                (m[0] if isinstance(m, tuple) else str(m))
                for m in prog_log
            ]
            st.code("\n".join(log_lines), language="")


# ═════════════════════════════════════════════════════════════════════════════
#  STEP 4  (manual mode — kept for power users)
# ═════════════════════════════════════════════════════════════════════════════
def page_step4():
    eng = _get_eng()
    log.info("page_step4: render stage=%d", st.session_state.get("s4_stage",0))
    cfg   = eng.configuration
    ppd   = cfg["periods_per_day"]
    wdays = cfg["working_days"]
    _header("📅 Step 4: Generate Timetable",
            f"{wdays} days/week · {ppd} periods/day · {wdays*ppd} slots/week per class")
    _show_notifications()
    if st.button("← Back to Step 3"): _nav("step3")
    st.divider()

    stage   = st.session_state.get("s4_stage",0)
    s1_stat = st.session_state.get("s4_s1_status")

    with st.container(border=True):
        st.markdown("**Stage 1 — HC1/HC2: Place Class-Teacher & Fixed/Preference Periods**")
        if stage == 0:
            if st.button("▶ Run Stage 1", type="primary", key="s4_run_s1"):
                log.info("page_step4: Run Stage 1")
                with st.spinner("Running Stage 1…"):
                    try:
                        result = eng.run_stage1()
                        st.session_state.update({"s4_s1_status":result,"s4_stage":1})
                        eng._relaxed_consec_keys = set()
                        eng._relaxed_main_keys   = set()
                        log.info("page_step4: Stage 1 done has_issues=%s", result.get("has_issues"))
                        _notify("Stage 1 complete.","success")
                    except Exception as ex:
                        log.error("page_step4 stage1: %s\n%s", ex, traceback.format_exc())
                        _notify(f"Stage 1 error: {ex}","error")
                st.rerun()
        else:
            if s1_stat:
                bg   = s1_stat.get("stage_bg","#1a7a1a")
                stxt = s1_stat.get("stage_txt","Stage 1 done")
                st.markdown(f"<div style='background:{bg};color:white;padding:8px 16px;"
                            f"border-radius:4px;font-weight:bold'>{stxt}</div>",
                            unsafe_allow_html=True)
                st.info(s1_stat.get("status",""))
            c1,c2 = st.columns(2)
            with c1:
                if st.button("📋 Task Analysis →", type="primary", key="s4_ta_btn"):
                    log.info("page_step4: → task_analysis")
                    eng._relaxed_consec_keys = set(st.session_state.get("relaxed_consec",set()))
                    eng._relaxed_main_keys   = set(st.session_state.get("relaxed_main",set()))
                    _nav("task_analysis")
            with c2:
                if st.button("↺ Re-run Stage 1", key="s4_rerun_s1"):
                    log.info("page_step4: re-run Stage 1")
                    for k in ("s4_stage","s4_s1_status","ta_allocation","ta2_allocation","s4_s3_status"):
                        st.session_state[k] = 0 if k=="s4_stage" else None
                    st.rerun()

    if stage >= 1 and eng._timetable:
        st.divider()
        st.markdown("**Stage 1 Preview**")
        _render_timetable_tabs(eng._timetable, key_prefix="s4")
        st.divider()
        st.markdown("**Export Stage 1 snapshot**")
        c1,c2 = st.columns(2)
        with c1: _excel_download("class","📥 Class Timetables")
        with c2: _excel_download("teacher","📥 Teacher Timetables")


# ═════════════════════════════════════════════════════════════════════════════
#  TASK ANALYSIS
# ═════════════════════════════════════════════════════════════════════════════
def page_task_analysis():
    eng = _get_eng()
    log.info("page_task_analysis: render")
    _header("📋 Task Analysis","Review groups before Stage 2. Allocate period slots.")
    _show_notifications()
    nav1,nav2 = st.columns(2)
    with nav1:
        if st.button("← Back to Stage 1"): _nav("step4")
    with nav2:
        if st.button("🗓 Allocate Periods", type="primary", key="ta_alloc_btn"):
            log.info("page_task_analysis: allocating")
            with st.spinner("Allocating…"):
                try:
                    slots, allocation, rows = eng._run_task_analysis_allocation()
                    st.session_state.update({"ta_allocation":allocation,"ta_group_slots":slots,"ta_all_rows":rows})
                    eng._last_allocation  = allocation
                    eng._last_group_slots = slots
                    eng._last_all_rows    = rows
                    ok_n = sum(1 for ar in allocation.values() if ar.get("ok"))
                    log.info("page_task_analysis: done %d ok / %d total", ok_n, len(allocation))
                    _notify(f"✓ {ok_n} OK, {len(allocation)-ok_n} failed.",
                            "success" if ok_n==len(allocation) else "warning")
                except Exception as ex:
                    log.error("page_task_analysis: %s\n%s", ex, traceback.format_exc())
                    _notify(f"Allocation error: {ex}","error")
            st.rerun()

    allocation  = st.session_state.get("ta_allocation")
    group_slots = st.session_state.get("ta_group_slots")
    all_rows    = st.session_state.get("ta_all_rows")

    if allocation:
        all_ok = all(ar.get("ok",False) for ar in allocation.values())
        if all_ok:
            if st.button("▶ Proceed to Stage 2 →", type="primary", key="ta_proceed"):
                log.info("page_task_analysis: → task_analysis2")
                _nav("task_analysis2")
        else:
            st.error("Some groups failed — relax constraints or fix Step 2.")

    _show_notifications()
    if all_rows is None:
        st.info("Click **Allocate Periods** to compute slot assignments.")
        try:
            _, _, rows = eng._run_task_analysis_allocation()
            _render_ta_table(rows, None, None)
        except Exception as ex:
            log.warning("page_task_analysis: preview error: %s", ex)
    else:
        _render_ta_table(all_rows, group_slots, allocation)


def _render_ta_table(all_rows, group_slots, allocation):
    eng = _get_eng()
    log.debug("_render_ta_table: %d rows", len(all_rows) if all_rows else 0)
    if not all_rows:
        st.info("No parallel/combined/consecutive groups found."); return
    DAYS_A = ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]
    sections = {"A":"Combined Groups","B":"Standalone Parallel Pairs","C":"Consecutive Groups"}
    for sec, sec_title in sections.items():
        sec_rows = [r for r in all_rows if r.get("section")==sec]
        if not sec_rows: continue
        st.markdown(f"#### {sec_title}")
        groups = {}
        for r in sec_rows: groups.setdefault(r["group"],[]).append(r)
        for gn, grows in sorted(groups.items()):
            first = grows[0]
            cls_s = ", ".join(r["class"] for r in grows)
            slot_info = ""
            if group_slots and gn in group_slots:
                gs = group_slots[gn]
                slot_info = f"  ·  **{gs['slots']} slot(s)**" if gs.get("ok") else f"  ·  ⚠ {gs.get('reason','?')}"
            alloc_info = ""; alloc_ok = False
            if allocation and gn in allocation:
                ar = allocation[gn]; alloc_ok = ar.get("ok",False)
                if alloc_ok:
                    placed = ar.get("placed", ar.get("slots",[]))
                    ps = "  ·  ".join(f"{DAYS_A[d]} P{p+1}" for d,p in sorted(placed)) if placed else "placed"
                    alloc_info = f"✅ {ps}"
                else:
                    alloc_info = f"❌ {ar.get('remaining','?')} unplaced. {ar.get('reason','')}"
            icon = "🟢" if (allocation and alloc_ok) else ("🔴" if allocation else "⚪")
            with st.expander(f"{icon} **Group {gn}** — {cls_s} · {first['subject']} / {first['teacher']}{slot_info}", expanded=False):
                for r in grows:
                    par = (f"  ‖  Parallel: {r['par_subj']} / {r['par_teacher']}"
                           if r.get("par_subj") not in ("—","?","",None) else "")
                    st.write(f"  📌 **{r['class']}** — {r['subject']} / {r['teacher']}{par}")
                if alloc_info:
                    if "✅" in alloc_info: st.success(alloc_info)
                    else:                  st.error(alloc_info)
                if sec=="C":
                    rk = (first["class"], first["subject"])
                    rel = st.session_state.get("relaxed_consec",set())
                    is_rel = rk in rel
                    if st.button("🔓 Un-relax" if is_rel else "🔒 Relax to Filler", key=f"relax_c_{gn}"):
                        for r in grows:
                            k = (r["class"],r["subject"])
                            if is_rel: rel.discard(k); eng._relaxed_consec_keys.discard(k)
                            else:      rel.add(k);     eng._relaxed_consec_keys.add(k)
                        st.session_state["relaxed_consec"] = rel
                        log.info("_render_ta_table: group %s relax=%s", gn, not is_rel)
                        st.rerun()


# ═════════════════════════════════════════════════════════════════════════════
#  TASK ANALYSIS 2
# ═════════════════════════════════════════════════════════════════════════════
def page_task_analysis2():
    eng = _get_eng()
    log.info("page_task_analysis2: render")
    _header("📋 Task Analysis — Stage 2","Allocate remaining slots before Stage 3.")
    _show_notifications()
    eng._relaxed_consec_keys = set(st.session_state.get("relaxed_consec",set()))
    eng._relaxed_main_keys   = set(st.session_state.get("relaxed_main",set()))
    nav1,nav2 = st.columns(2)
    with nav1:
        if st.button("← Back to Task Analysis"): _nav("task_analysis")
    with nav2:
        if st.button("🗓 Allocate Slots", type="primary", key="ta2_alloc_btn"):
            log.info("page_task_analysis2: allocating")
            with st.spinner("Allocating Stage 2 slots…"):
                try:
                    result = eng._run_ta2_allocation()
                    st.session_state["ta2_allocation"] = result
                    eng._last_ta2_allocation = result
                    ok_n = sum(1 for ar in result.values() if isinstance(ar,dict) and ar.get("remaining",1)==0)
                    log.info("page_task_analysis2: done %d ok", ok_n)
                    _notify(f"✓ Stage 2 done: {ok_n} groups OK.","success")
                except Exception as ex:
                    log.error("page_task_analysis2: %s\n%s", ex, traceback.format_exc())
                    _notify(f"Error: {ex}","error")
            st.rerun()

    ta2 = st.session_state.get("ta2_allocation")
    if ta2:
        fails = [k for k,ar in ta2.items() if isinstance(ar,dict) and ar.get("remaining",1)>0]
        if not fails:
            st.success("✅ All groups allocated.")
            if st.button("📅 Proceed to Stage 3 →", type="primary", key="ta2_proceed"):
                log.info("page_task_analysis2: → stage2_page")
                _nav("stage2_page")
        else:
            st.warning(f"⚠ {len(fails)} group(s) have unplaced periods.")
            if st.button("📅 Proceed to Stage 3 anyway →", key="ta2_proceed_any"):
                log.info("page_task_analysis2: proceed with %d failures", len(fails))
                _nav("stage2_page")
        _render_ta2_table(ta2)
    else:
        st.info("Click **Allocate Slots** to compute allocations.")


def _render_ta2_table(result):
    log.debug("_render_ta2_table: %d groups", len(result) if result else 0)
    if not result: return
    DAYS_A = ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]
    def slot_str(placed):
        return "  ·  ".join(f"{DAYS_A[d]} P{p+1}" for d,p in sorted(placed)) if placed else "—"
    for gn, ar in sorted(result.items()):
        if not isinstance(ar,dict): continue
        placed    = ar.get("slots", ar.get("placed",[]))
        remaining = ar.get("remaining",0)
        ok        = remaining == 0
        with st.expander(f"{'✅' if ok else '❌'} Group {gn}", expanded=not ok):
            c1,c2 = st.columns(2)
            with c1:
                st.metric("Total",          ar.get("total","?"))
                st.metric("Stage 1 placed", ar.get("s1_placed","?"))
            with c2:
                st.metric("Stage 2 placed", ar.get("new_placed","?"))
                st.metric("Remaining",      remaining)
            if placed: st.success(f"Slots: {slot_str(placed)}")
            if not ok:
                st.error(f"Reason: {ar.get('reason','unknown')}")
                rel = st.session_state.get("relaxed_main",set()); k_str = str(gn)
                if st.button("🔓 Un-relax" if k_str in rel else "🔒 Relax to Filler", key=f"relax_m_{gn}"):
                    rel.discard(k_str) if k_str in rel else rel.add(k_str)
                    st.session_state["relaxed_main"] = rel
                    log.info("_render_ta2_table: group %s main-relax=%s", gn, k_str in rel)
                    st.rerun()


# ═════════════════════════════════════════════════════════════════════════════
#  STAGE 2 PAGE
# ═════════════════════════════════════════════════════════════════════════════
def _render_force_fill_summary(ff_result):
    """Render the Force Fill result summary — mirrors the tkinter scrollable dialog."""
    remaining   = ff_result.get("remaining", 0)
    relaxed     = ff_result.get("relaxed")
    overloaded  = ff_result.get("overloaded", [])
    blocked     = ff_result.get("blocked_only", [])
    wdays       = ff_result.get("wdays", "?")
    ppd         = ff_result.get("ppd", "?")
    total_slots = ff_result.get("total_slots", "?")

    if remaining == 0:
        st.success("✅ All periods placed — timetable is 100% complete!")
        if relaxed:
            with st.expander("ℹ️ Constraints relaxed during Force Fill", expanded=True):
                st.code(relaxed, language="")
        return

    st.error(f"⚠  {remaining} period(s) could not be placed.")

    if overloaded:
        st.markdown("#### ❌ Overloaded Teachers")
        st.caption(f"Total grid capacity: {wdays} days × {ppd} periods = {total_slots} slots per teacher")
        for tname, assigned, cap, excess, unp in overloaded:
            with st.container(border=True):
                col1, col2 = st.columns([2, 3])
                with col1:
                    st.markdown(f"**❌ {tname}**")
                with col2:
                    st.markdown(
                        f"Assigned: **{assigned}** &nbsp;|&nbsp; "
                        f"Capacity: **{cap}** &nbsp;|&nbsp; "
                        f"Excess: **+{excess}** ← fix in Step 2 &nbsp;|&nbsp; "
                        f"Unplaced: **{unp}**"
                    )

    if blocked:
        st.markdown("#### ⚠ Blocked Teachers *(within capacity, but slots clash)*")
        for tname, assigned, cap, unp in blocked:
            with st.container(border=True):
                col1, col2 = st.columns([2, 3])
                with col1:
                    st.markdown(f"**⚠ {tname}**")
                with col2:
                    st.markdown(
                        f"Assigned: **{assigned} / {cap}** &nbsp;|&nbsp; Unplaced: **{unp}**"
                    )

    if not overloaded and not blocked:
        st.info("Could not identify a specific cause. Please review teacher workloads in Step 2.")

    if relaxed:
        with st.expander("ℹ️ Constraints relaxed during Force Fill", expanded=False):
            st.code(relaxed, language="")


def page_stage2():
    eng = _get_eng()
    log.info("page_stage2: render")
    _header("⚙ Stage 2 — Fill Remaining Periods",
            "Stage 3 of engine: consecutive pairs, daily subjects, fillers and repair.")
    _show_notifications()
    if st.button("← Back to Task Analysis 2"): _nav("task_analysis2")

    s3_status  = st.session_state.get("s4_s3_status")
    ff_result  = st.session_state.get("s4_ff_result")   # Force Fill result

    # ── Phase 1: Run Stage 3 ──────────────────────────────────────────────────
    if s3_status is None:
        st.info("Click **Run Stage 3** to fill all remaining empty slots.")
        if st.button("▶ Run Stage 3", type="primary", key="s2pg_run"):
            log.info("page_stage2: Run Stage 3")
            with st.spinner("Running Stage 3 — filling remaining periods…"):
                try:
                    result = eng.run_stage3()
                    st.session_state.update({"s4_s3_status": result, "s4_stage": 3,
                                             "s4_ff_result": None})
                    log.info("page_stage2: Stage 3 done ok=%s unplaced=%s",
                             result.get("ok"), result.get("unplaced"))
                    _notify(result.get("msg", "Stage 3 complete."),
                            "success" if result.get("ok") else "warning")
                except Exception as ex:
                    log.error("page_stage2 stage3: %s\n%s", ex, traceback.format_exc())
                    _notify(f"Stage 3 error: {ex}", "error")
            st.rerun()
        return

    # ── Phase 2: Stage 3 done — show status + action buttons ─────────────────
    if s3_status.get("ok"):
        st.success(s3_status.get("msg", "✅ Complete!"))
    else:
        st.warning(s3_status.get("msg", ""))

    # Show Force Fill option only when there are still unplaced periods
    unplaced_after_s3 = s3_status.get("unplaced", 0)

    col_view, col_rerun, col_ff = st.columns([2, 1, 2])
    with col_view:
        if st.button("📊 View Final Timetable →", type="primary", key="s2pg_view"):
            log.info("page_stage2: → final_timetable")
            _nav("final_timetable")
    with col_rerun:
        if st.button("↺ Re-run Stage 3", key="s2pg_rerun"):
            log.info("page_stage2: re-run")
            st.session_state["s4_s3_status"] = None
            st.session_state["s4_ff_result"] = None
            st.rerun()
    with col_ff:
        if unplaced_after_s3 > 0:
            if st.button("🔧 Force Fill", type="secondary", key="s2pg_ff",
                         help="Min-Conflicts CSP solver — up to 1500 iterations, "
                              "stops as soon as all periods are placed."):
                log.info("page_stage2: running Force Fill")
                progress_placeholder = st.empty()
                progress_msgs = []

                def _progress(msg):
                    progress_msgs.append(msg)
                    if msg:
                        progress_placeholder.info(f"⏳ {msg}")

                with st.spinner("🔧 Force Fill running…"):
                    try:
                        ff = eng.run_force_fill(progress_cb=_progress)
                        st.session_state["s4_ff_result"] = ff
                        # Update s3_status so unplaced count is refreshed
                        st.session_state["s4_s3_status"] = {
                            **s3_status,
                            "unplaced": ff["remaining"],
                            "ok":       ff["ok"],
                            "msg":      ff.get("msg", ""),
                        }
                        log.info("page_stage2: Force Fill done ok=%s remaining=%s",
                                 ff["ok"], ff["remaining"])
                        _notify(
                            "✅ Force Fill complete — all periods placed!" if ff["ok"]
                            else f"⚠ Force Fill done — {ff['remaining']} period(s) still unplaced.",
                            "success" if ff["ok"] else "warning",
                        )
                    except Exception as ex:
                        log.error("page_stage2 force_fill: %s\n%s", ex, traceback.format_exc())
                        _notify(f"Force Fill error: {ex}", "error")
                progress_placeholder.empty()
                st.rerun()

    # ── Phase 3: Show Force Fill summary if available ─────────────────────────
    if ff_result is not None:
        st.divider()
        st.markdown("### 🔧 Force Fill Summary")
        _render_force_fill_summary(ff_result)


# ═════════════════════════════════════════════════════════════════════════════
#  FINAL TIMETABLE
# ═════════════════════════════════════════════════════════════════════════════
def page_final_timetable():
    eng = _get_eng()

    if not eng.configuration:
        st.warning("⚠ Step 1 not yet completed — basic configuration is missing.")
        if st.button("← Go to Step 1", type="primary", key="ft_guard_back"):
            _nav("step1")
        return

    log.info("page_final_timetable: render")
    _header("📊 Final Timetable","View and export the complete timetable.")
    _show_notifications()

    # Back button
    if st.button("← Back to Generate"):
        _nav("generate")

    # Period reduction notice
    gen_result = st.session_state.get("gen_result") or {}
    reductions = gen_result.get("period_reductions", [])
    if reductions:
        with st.expander(f"📉 {len(reductions)} period reduction(s) were applied to "
                         f"resolve deadlocks — click to review", expanded=False):
            for r in reductions:
                st.warning(
                    f"**{r['subject']}** in **{r['class']}** "
                    f"(teacher: {r['teacher']})  →  "
                    f"{r['from_periods']} → {r['to_periods']} periods/week")
    tt = eng._timetable
    if not tt:
        st.error("No timetable generated."); log.warning("page_final_timetable: no timetable"); return
    with st.container(border=True):
        st.markdown("**📥 Export to Excel**")
        c1,c2,c3 = st.columns(3)
        with c1: _excel_download("class",             "📥 Class Timetables")
        with c2: _excel_download("consolidated_class","📥 Consolidated Class View")
        with c3: _excel_download("teacher",           "📥 Teacher Timetables")
        c4,c5,c6 = st.columns(3)
        with c4: _excel_download("ct_list",  "📥 CT List")
        with c5: _excel_download("workload", "📥 Workload")
        with c6: _excel_download("one_sheet","📥 One-Sheet")
    st.divider()
    tc, tt2, ts = st.tabs(["🏫 Classwise","👨‍🏫 Teacherwise","📋 Summary"])
    with tc:  _render_class_view(tt)
    with tt2: _render_teacher_view(tt)
    with ts:  _render_summary_view(tt)


# ─────────────────────────────────────────────────────────────────────────────
#  Timetable renderers
# ─────────────────────────────────────────────────────────────────────────────
def _render_timetable_tabs(tt, key_prefix="tt"):
    tc, tt2 = st.tabs(["🏫 Class View","👨‍🏫 Teacher View"])
    with tc:  _render_class_view(tt,   key_prefix=key_prefix+"_c")
    with tt2: _render_teacher_view(tt, key_prefix=key_prefix+"_t")


def _render_class_view(tt, key_prefix="cls"):
    import pandas as pd
    log.debug("_render_class_view: key=%s", key_prefix)
    all_classes = tt["all_classes"]; days = tt["days"]; ppd = tt["ppd"]
    half1 = tt["half1"];             grid = tt["grid"]
    sel_cn = st.selectbox("Select Class", all_classes, key=f"{key_prefix}_sel")
    if not sel_cn: return
    header = ["Day"] + [f"P{p+1}{'①' if p<half1 else '②'}" for p in range(ppd)]
    rows = []
    for d, dname in enumerate(days):
        row = [dname]
        for p in range(ppd):
            g = grid.get(sel_cn,[])
            cell = g[d][p] if d<len(g) and g else None
            if cell:
                row.append(f"{'★' if cell.get('is_ct') else ''}{cell.get('subject','')} / {cell.get('teacher','')}")
            else:
                row.append("—")
        rows.append(row)
    st.dataframe(pd.DataFrame(rows, columns=header), use_container_width=True, hide_index=True)


def _render_teacher_view(tt, key_prefix="tch"):
    import pandas as pd
    log.debug("_render_teacher_view: key=%s", key_prefix)
    all_classes = tt["all_classes"]; days = tt["days"]; ppd = tt["ppd"]; grid = tt["grid"]
    tg = {}
    for cn in all_classes:
        g = grid.get(cn,[])
        for d in range(len(days)):
            if d >= len(g): continue
            for p in range(ppd):
                cell = g[d][p]
                if not cell: continue
                etype = cell.get("type", "normal")
                cc    = cell.get("combined_classes", [])
                is_combined = bool(cc) and etype in ("combined", "combined_parallel")

                # Primary teacher: for combined groups only write once (when cn==cc[0])
                tname = cell.get("teacher", "")
                sname = cell.get("subject", "")
                if tname and tname not in ("—", "?"):
                    if is_combined:
                        # Only write once per combined group slot
                        if not cc or cn == cc[0]:
                            cls_label = "+".join(cc)
                            tg.setdefault(tname, [[None]*ppd for _ in range(len(days))])
                            tg[tname][d][p] = {"class": cls_label, "subject": sname,
                                               "is_ct": cell.get("is_ct", False)}
                    else:
                        tg.setdefault(tname, [[None]*ppd for _ in range(len(days))])
                        tg[tname][d][p] = {"class": cn, "subject": sname,
                                           "is_ct": cell.get("is_ct", False)}

                # Parallel teacher (always per-class)
                pt = cell.get("par_teach", "")
                ps = cell.get("par_subj", "")
                if pt and pt not in ("—", "?"):
                    tg.setdefault(pt, [[None]*ppd for _ in range(len(days))])
                    tg[pt][d][p] = {"class": cn, "subject": ps, "is_ct": False}

    tlist = sorted(tg.keys())
    if not tlist: st.info("No teacher data."); return
    sel_t = st.selectbox("Select Teacher", tlist, key=f"{key_prefix}_sel")
    if not sel_t: return
    trows = tg.get(sel_t,[])
    header = ["Day"] + [f"P{p+1}" for p in range(ppd)]
    rows = []
    for d, dname in enumerate(days):
        row = [dname]
        for p in range(ppd):
            cell = trows[d][p] if d<len(trows) else None
            row.append(f"{'★' if cell and cell.get('is_ct') else ''}"
                       f"{cell['class']} / {cell['subject']}" if cell else "FREE")
        rows.append(row)
    st.dataframe(pd.DataFrame(rows, columns=header), use_container_width=True, hide_index=True)


def _render_summary_view(tt):
    log.debug("_render_summary_view")
    tasks=tt.get("tasks",[]); days=tt["days"]; ppd=tt["ppd"]
    half1=tt["half1"]; grid=tt["grid"]; all_classes=tt["all_classes"]
    unplaced = [t for t in tasks if t.get("remaining",0)>0]
    if unplaced:
        st.error(f"**{len(unplaced)} task(s) with unplaced periods:**")
        for t in unplaced:
            st.write(f"  ❌ {'+'.join(t.get('cn_list',[]))} | {t['subject']} | "
                     f"{t['teacher']} — {t['remaining']} unplaced")
    else:
        st.success("✅ All periods placed.")
    st.markdown("**Teacher Free-Period Distribution**")
    all_teachers = sorted({cell.get("teacher","")
                            for cn in all_classes
                            for d_row in grid.get(cn,[])
                            for cell in d_row if cell and cell.get("teacher")})
    for teacher in all_teachers:
        busy = {}
        for cn in all_classes:
            for d, d_row in enumerate(grid.get(cn,[])):
                for p, cell in enumerate(d_row):
                    if cell and (cell.get("teacher")==teacher or cell.get("par_teach")==teacher):
                        busy.setdefault(d,set()).add(p)
        lines = []
        for d in range(len(days)):
            bd  = busy.get(d,set())
            fh1 = half1 - len([x for x in bd if x<half1])
            fh2 = (ppd-half1) - len([x for x in bd if x>=half1])
            if fh1+fh2 == ppd: continue
            lines.append((days[d], fh1, fh2, fh1>=1 and fh2>=1))
        if lines:
            with st.expander(f"**{teacher}**", expanded=False):
                for dname,fh1,fh2,ok in lines:
                    st.write(f"  {'✓' if ok else '⚠'} {dname} — free H1:{fh1}  H2:{fh2}")
    st.divider()
    st.caption("★=CT  ⊕=Combined  ∥=Parallel  ⊕∥=Combined+Parallel")


# ═════════════════════════════════════════════════════════════════════════════
#  SIDEBAR
# ═════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("## 🗓 Timetable Generator")
    st.caption("V4.2 — Streamlit Edition")
    st.divider()

    cur     = st.session_state.page
    s1v     = st.session_state.get("s1_validated", False)
    s2v     = st.session_state.get("s2_validated", False)
    s3v     = st.session_state.get("s3_validated", False)
    gen_done = st.session_state.get("gen_result") is not None

    # Progress indicator
    n_done = sum([s1v, s2v, s3v, gen_done])
    st.progress(n_done / 4, text=f"{n_done} of 4 steps complete")
    st.divider()

    steps = [
        ("step1",           "1️⃣  Basic Config",      True),
        ("step2",           "2️⃣  Class Assignments", s1v),
        ("step3",           "3️⃣  Teacher Settings",  s2v),
        ("generate",        "4️⃣  Generate",          s3v),
        ("final_timetable", "🏁  Final Timetable",    gen_done),
    ]

    for pid, plabel, unlocked in steps:
        active   = (pid == cur)
        locked   = not unlocked and not active
        done_icon = "✅ " if unlocked and not active and pid != "final_timetable" else ""
        lock_icon = "🔒 " if locked else ""
        arrow     = "▶ " if active else ""
        label     = arrow + done_icon + lock_icon + plabel
        if st.button(label, key=f"nav_{pid}",
                     use_container_width=True,
                     disabled=active or locked):
            _nav(pid)

    st.divider()
    _eng_sb = _get_eng()
    if _eng_sb.configuration:
        cfg = _eng_sb.configuration
        st.caption(f"**Config:** {cfg.get('working_days','?')}d × {cfg.get('periods_per_day','?')}p")
        st.caption(f"**Teachers:** {len(cfg.get('teacher_names',[]))}")
        st.caption(f"**Classes:** {sum(cfg.get('classes',{}).values())}")

    st.divider()
    with st.expander("🪵 Debug Log", expanded=False):
        st.caption("Shows last 200 log lines. Format: [LEVEL] file:func:line — msg")
        if st.button("Clear", key="clear_log"):
            _mem_handler.lines.clear(); st.rerun()
        log_text = "\n".join(_mem_handler.lines[-200:])
        st.code(log_text if log_text else "(no entries yet)", language="")


# ═════════════════════════════════════════════════════════════════════════════
#  ROUTER
# ═════════════════════════════════════════════════════════════════════════════
PAGES = {
    "step1":           page_step1,
    "step2":           page_step2,
    "step3":           page_step3,
    "generate":        page_generate,
    "final_timetable": page_final_timetable,
}

log.debug("Router: page=%s", st.session_state.page)
PAGES.get(st.session_state.page, page_step1)()

# ─────────────────────────────────────────────────────────────────────────────
#  FOOTER  (bottom-centre on every page)
# ─────────────────────────────────────────────────────────────────────────────
st.markdown(
    """
    <style>
    .dev-footer {
        position: fixed;
        bottom: 0;
        left: 0;
        width: 100%;
        text-align: center;
        padding: 6px 0 6px 0;
        font-size: 0.78rem;
        color: #888888;
        background: rgba(255,255,255,0.85);
        backdrop-filter: blur(4px);
        border-top: 1px solid #e0e0e0;
        z-index: 9999;
        letter-spacing: 0.01em;
    }
    </style>
    <div class="dev-footer">
        Developed by: <strong>Kanika Tanwar</strong>, TGT CS
    </div>
    """,
    unsafe_allow_html=True,
)