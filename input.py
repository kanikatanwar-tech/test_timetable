"""
input.py — User Input Pages (Steps 1, 2, 3) + Shared Helpers

Contains:
  • Shared utilities: _get_eng, _nav, _notify, _show_notifications,
                      _file_hash, _already_processed, _all_classes,
                      _json_download, _header, _show_upload_error_if_any
  • Error dialog popups for Steps 1, 2, 3
  • page_step1()  — Basic Config (periods, teachers, sections)
  • page_step2()  — Class Assignments (subjects, CT, preferences)
  • page_step3()  — Teacher Settings  (workload, combines, unavailability)
"""

import io
import json
import logging
import math
import traceback
import hashlib
from datetime import datetime

import streamlit as st
from generator import TimetableEngine

# ─────────────────────────────────────────────────────────────────────────────
#  Logger  (same name as streamlit_app so all messages share one stream)
# ─────────────────────────────────────────────────────────────────────────────
log = logging.getLogger("timetable")

# ─────────────────────────────────────────────────────────────────────────────
#  Shared constants
# ─────────────────────────────────────────────────────────────────────────────
DAY_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

# ─────────────────────────────────────────────────────────────────────────────
#  Shared helper utilities
# ─────────────────────────────────────────────────────────────────────────────

def _get_eng() -> TimetableEngine:
    """Always return the live engine from session state."""
    return st.session_state["engine"]


def _nav(page: str):
    log.info("_nav: → %s", page)
    st.session_state.page = page
    st.rerun()


def _notify(msg: str, kind: str = "info"):
    log.debug("_notify [%s]: %s", kind, msg)
    st.session_state["_notify"].append((kind, msg))


def _show_notifications():
    for kind, msg in st.session_state.get("_notify", []):
        if kind == "success": st.success(msg)
        elif kind == "error":   st.error(msg)
        elif kind == "warning": st.warning(msg)
        else:                   st.info(msg)
    st.session_state["_notify"] = []


def _file_hash(data: bytes) -> str:
    return hashlib.sha1(data).hexdigest()


def _already_processed(key: str, data: bytes) -> bool:
    """Return True if we already handled this exact file upload (same bytes)."""
    h = _file_hash(data)
    if st.session_state["_upload_hash"].get(key) == h:
        log.debug("_already_processed: key=%s hash=%s already done", key, h[:8])
        return True
    st.session_state["_upload_hash"][key] = h
    log.debug("_already_processed: key=%s hash=%s NEW", key, h[:8])
    return False


def _all_classes():
    cfg = _get_eng().configuration
    return [
        f"{cls}{chr(65+si)}"
        for cls in range(6, 13)
        for si in range(int(cfg.get("classes", {}).get(cls, 0)))
    ]


def _json_download(data: dict, label: str, filename: str):
    st.download_button(label=label, data=json.dumps(data, indent=2),
                       file_name=filename, mime="application/json",
                       use_container_width=True)


def _header(title: str, sub: str = ""):
    st.markdown(f"## {title}")
    if sub: st.caption(sub)
    st.divider()


def _show_upload_error_if_any(key: str):
    """If session state has an upload error for `key`, show the dialog and clear it."""
    err = st.session_state.pop(key, None)
    if err:
        _upload_error_dialog(err)


# ─────────────────────────────────────────────────────────────────────────────
#  Error / warning dialog popups
# ─────────────────────────────────────────────────────────────────────────────

@st.dialog("❌ Wrong File Uploaded")
def _upload_error_dialog(msg: str):
    """Modal popup for wrong/invalid file uploads."""
    st.error(msg)
    st.markdown("Please upload the **correct file** for this step.")
    if st.button("OK", type="primary", use_container_width=True):
        st.rerun()


@st.dialog("⚠ Cannot Combine")
def _combine_error_dialog(msg: str):
    """Modal popup for combine validation errors in Step 3."""
    st.error(msg)
    if st.button("OK", type="primary", use_container_width=True):
        st.rerun()


@st.dialog("❌ Step 1 — Validation Errors")
def _step1_error_dialog(errors: list):
    """Modal popup listing all Step 1 validation errors with fix advice."""
    for err in errors:
        st.error(err["msg"])
        if err.get("fix"):
            st.info(f"💡 **How to fix:** {err['fix']}")
    st.divider()
    st.caption("Fix all errors above, then click **Continue to Step 2** again.")
    if st.button("Close", type="primary", use_container_width=True):
        st.rerun()


@st.dialog("❌ Step 2 — Validation Errors")
def _step2_error_dialog(vr: dict):
    """Modal popup summarising Step 2 validation errors with fix advice."""
    period_errors  = vr.get("period_errors", [])
    hard_conflicts = vr.get("hard_conflicts", [])
    wcc            = vr.get("within_class_conflicts", [])
    total = len(period_errors) + len(hard_conflicts) + len(wcc)
    st.error(f"**{total} error(s) found.** Fix all issues, then click Validate & Continue again.")
    st.divider()
    if period_errors:
        st.markdown(f"**① Period Mismatches — {len(period_errors)} class(es)**")
        for cn, msg in period_errors[:5]:
            st.warning(f"**{cn}:** {msg}")
            st.caption("💡 Add/remove subjects so total periods = periods/day × working days.")
        if len(period_errors) > 5:
            st.caption(f"… and {len(period_errors)-5} more — scroll Step 2 for all details.")
    if hard_conflicts:
        st.markdown(f"**② Teacher Conflicts — {len(hard_conflicts)} conflict(s)**")
        for c in hard_conflicts[:3]:
            reason = c["reason"]
            st.warning(f"**{c['teacher']}:** {reason[:150]}{'...' if len(reason)>150 else ''}")
            st.caption("💡 Change period/day preference so the teacher is not double-booked.")
        if len(hard_conflicts) > 3:
            st.caption(f"… and {len(hard_conflicts)-3} more — scroll Step 2 for details.")
    if wcc:
        st.markdown(f"**③ Within-Class Slot Conflicts — {len(wcc)} conflict(s)**")
        for c in wcc[:3]:
            reason = c["reason"]
            st.warning(f"**{c['class']} / {c['day']}:** {reason[:150]}{'...' if len(reason)>150 else ''}")
            st.caption("💡 Adjust period/day preferences so subjects don't compete for the same slot.")
        if len(wcc) > 3:
            st.caption(f"… and {len(wcc)-3} more — scroll Step 2 for details.")
    if st.button("Close & Fix Issues", type="primary", use_container_width=True):
        st.rerun()


@st.dialog("❌ Step 3 — Validation Errors")
def _step3_error_dialog(vr: dict):
    """Modal popup for Step 3 overload errors with fix advice."""
    issues = vr.get("issues", [])
    st.error(f"**{len(issues)} overloaded teacher(s) must be resolved before generating.**")
    st.divider()
    for ln in issues[:8]:
        st.warning(ln)
    if len(issues) > 8:
        st.caption(f"... and {len(issues)-8} more.")
    st.divider()
    st.info(
        "**How to fix:**\n"
        "- Use **Combine Classes** tab to merge identical lessons across sections — "
        "this reduces the teacher's effective period count.\n"
        "- Or click **⏭ Skip** on the teacher card to bypass the overload check (use with caution)."
    )
    if st.button("Close & Fix Issues", type="primary", use_container_width=True):
        st.rerun()


# ═════════════════════════════════════════════════════════════════════════════
#  STEP 1 — Basic Configuration
# ═════════════════════════════════════════════════════════════════════════════

def page_step1():
    log.info("page_step1: render")
    _show_upload_error_if_any("_s1_upload_err")
    if st.session_state.get("_s1_val_errors"):
        _step1_error_dialog(st.session_state.pop("_s1_val_errors"))
    _header("📋 Step 1: Basic Configuration",
            "Set periods, working days, upload teachers and define class sections.")
    _show_notifications()

    # ── Save / Load ───────────────────────────────────────────────────────────
    with st.expander("💾 Save / Load Configuration", expanded=False):
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("**Save current config**")
            if st.button("⬇ Prepare Download", key="s1_dl_btn", use_container_width=True):
                log.info("page_step1: preparing download")
                if not st.session_state.s1_teachers:
                    _notify("Upload a teacher file first.", "warning"); st.rerun()
                data = {
                    "periods_per_day":     st.session_state.get("ni_ppd", 7),
                    "working_days":        st.session_state.get("ni_wdays", 6),
                    "periods_first_half":  st.session_state.get("ni_fhalf", 4),
                    "periods_second_half": st.session_state.get("ni_shalf", 3),
                    "teacher_file_path":   st.session_state.s1_teacher_fname,
                    "teacher_names":       st.session_state.s1_teachers,
                    "classes":             {str(k): v for k, v in
                                            st.session_state.s1_sections.items()},
                    "saved_at":            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                }
                st.session_state["_s1_dl_data"] = data
            if "_s1_dl_data" in st.session_state:
                _json_download(st.session_state["_s1_dl_data"], "📥 Click to Download",
                               f"Config_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
        with c2:
            st.markdown("**Load saved config**")
            up = st.file_uploader("Upload Config JSON", type=["json"], key="s1_load_json")
            if up is not None:
                raw = up.read()
                if len(raw) > 0:
                    h = _file_hash(raw)
                    if st.session_state.get("_s1_pending_hash") != h:
                        st.session_state["_s1_pending_raw"]  = raw
                        st.session_state["_s1_pending_hash"] = h
                        st.session_state["_s1_pending_name"] = up.name
                        log.debug("page_step1: cached new file '%s' (%d bytes)", up.name, len(raw))

            pending = st.session_state.get("_s1_pending_raw")
            if pending:
                pname = st.session_state.get("_s1_pending_name", "config.json")
                st.info(f"📄 Ready to load: **{pname}** ({len(pending):,} bytes)")
                if st.button("📂 Load Config", key="s1_load_btn", type="primary",
                             use_container_width=True):
                    log.info("page_step1: Load Config button clicked for '%s'", pname)
                    _load_step1_config(pending)
            else:
                st.caption("Upload a JSON config file, then click Load Config.")

    # ── Periods & Days ────────────────────────────────────────────────────────
    with st.container(border=True):
        st.markdown("**1. Periods & Working Days**")
        c1, c2 = st.columns(2)
        with c1:
            ppd   = st.number_input("Periods per day",   1, 20,
                                    st.session_state.get("ni_ppd",   7), key="ni_ppd")
            wdays = st.number_input("Working days/week", 1, 7,
                                    st.session_state.get("ni_wdays", 6), key="ni_wdays")
        with c2:
            fhalf = st.number_input("Periods — first half",  1, 20,
                                    st.session_state.get("ni_fhalf", 4), key="ni_fhalf")
            shalf = st.number_input("Periods — second half", 1, 20,
                                    st.session_state.get("ni_shalf", 3), key="ni_shalf")
        if fhalf + shalf == ppd:
            st.success(f"✓ Valid: {fhalf} + {shalf} = {ppd}")
        else:
            st.error(f"✗ {fhalf} + {shalf} = {fhalf+shalf}, need {ppd}")

    # ── Teachers ──────────────────────────────────────────────────────────────
    with st.container(border=True):
        st.markdown("**2. Teachers — Upload Excel File**")
        st.caption("Column A only: one name per row. "
                   "Header 'Teacher Name' is automatically skipped.")
        col_up, col_info = st.columns([3, 1])
        with col_up:
            tf = st.file_uploader("Teacher Excel (.xlsx/.xls)", type=["xlsx","xls"],
                                   key="s1_teacher_up")
            if tf is not None:
                raw_tf = tf.read()
                log.debug("page_step1: teacher upload size=%d", len(raw_tf))
                if len(raw_tf) == 0:
                    log.warning("page_step1: teacher 0 bytes, skipping")
                elif not _already_processed("s1_teacher_up", raw_tf):
                    _load_teacher_bytes(raw_tf, tf.name)
        with col_info:
            if st.session_state.s1_teachers:
                st.success(f"✓ {len(st.session_state.s1_teachers)} teachers")
                if st.button("👁 Preview", key="s1_preview"):
                    st.session_state["_s1_show_t"] = not st.session_state.get("_s1_show_t", False)
            else:
                st.warning("No teachers loaded")
        if st.session_state.get("_s1_show_t") and st.session_state.s1_teachers:
            st.dataframe({"Teacher Name": st.session_state.s1_teachers},
                         use_container_width=True, height=200)

    # ── Classes ───────────────────────────────────────────────────────────────
    with st.container(border=True):
        st.markdown("**3. Classes 6–12 — Number of Sections**")
        cols = st.columns(7)
        for idx, cls in enumerate(range(6, 13)):
            with cols[idx]:
                cur = st.session_state.s1_sections.get(cls, 4)
                val = st.number_input(f"Class {cls}", 0, 50, cur, key=f"ni_cls_{cls}")
                st.session_state.s1_sections[cls] = val

    # ── Navigation ────────────────────────────────────────────────────────────
    cb, cc = st.columns([1, 3])
    with cc:
        if st.button("✓ Continue to Step 2 →", type="primary", use_container_width=True):
            _step1_save_and_continue()
    with cb:
        if st.button("⟲ Reset", use_container_width=True):
            log.info("page_step1: reset")
            for k in ["ni_ppd","ni_wdays","ni_fhalf","ni_shalf",
                      "s1_teachers","s1_teacher_fname","_s1_dl_data","_s1_show_t",
                      "_s1_pending_raw","_s1_pending_hash","_s1_pending_name"]:
                st.session_state.pop(k, None)
            st.session_state.s1_sections  = {cls: 4 for cls in range(6, 13)}
            st.session_state["_upload_hash"] = {}
            st.rerun()


def _load_step1_config(raw: bytes):
    log.info("_load_step1_config: %d bytes", len(raw))
    try:
        d = json.loads(raw.decode("utf-8"))
    except json.JSONDecodeError as ex:
        log.error("_load_step1_config: JSON parse error: %s", ex)
        msg = f"Invalid JSON file: {ex}"
        st.session_state["_s1_upload_err"] = msg
        for k in ("_s1_pending_raw", "_s1_pending_hash", "_s1_pending_name"):
            st.session_state.pop(k, None)
        st.rerun()
        return

    required_keys = {"periods_per_day", "working_days", "teacher_names"}
    if not required_keys.issubset(d.keys()):
        wrong_hint = ""
        if "assignments" in d:
            wrong_hint = " — this looks like a Step 2 Assignments file"
        elif "step3_data" in d or "step3_unavailability" in d:
            wrong_hint = " — this looks like a Step 3 config file"
        msg = (f"Wrong file type{wrong_hint}.\n\n"
               f"A Step 1 config file must contain: "
               f"{', '.join(sorted(required_keys))}.")
        log.error("_load_step1_config: %s", msg)
        st.session_state["_s1_upload_err"] = msg
        for k in ("_s1_pending_raw", "_s1_pending_hash", "_s1_pending_name"):
            st.session_state.pop(k, None)
        st.rerun()
        return

    try:
        ppd   = int(d.get("periods_per_day",    7))
        wdays = int(d.get("working_days",        6))
        fhalf = int(d.get("periods_first_half",  4))
        shalf = int(d.get("periods_second_half", 3))
        teachers   = d.get("teacher_names", [])
        classes_raw = d.get("classes", {})
        log.info("_load_step1_config: ppd=%d wdays=%d fhalf=%d shalf=%d teachers=%d",
                 ppd, wdays, fhalf, shalf, len(teachers))

        keys_to_clear = (
            ["ni_ppd", "ni_wdays", "ni_fhalf", "ni_shalf"]
            + [f"ni_cls_{cls}" for cls in range(6, 13)]
        )
        for wk in keys_to_clear:
            st.session_state.pop(wk, None)

        sections = {int(k): v for k, v in classes_raw.items()}
        st.session_state["ni_ppd"]           = ppd
        st.session_state["ni_wdays"]         = wdays
        st.session_state["ni_fhalf"]         = fhalf
        st.session_state["ni_shalf"]         = shalf
        st.session_state["s1_teachers"]      = teachers
        st.session_state["s1_teacher_fname"] = d.get("teacher_file_path", "")
        st.session_state["s1_sections"]      = sections
        for cls, nsec in sections.items():
            st.session_state[f"ni_cls_{cls}"] = nsec

        for k in ("_s1_pending_raw", "_s1_pending_hash", "_s1_pending_name"):
            st.session_state.pop(k, None)

        _notify(f"✓ Config loaded — {len(teachers)} teachers, "
                f"{ppd} periods/day, {wdays} days/week, "
                f"{sum(sections.values())} total sections.", "success")
        log.info("_load_step1_config: applied OK — sections=%s", sections)
        st.rerun()
    except Exception as ex:
        log.error("_load_step1_config: %s\n%s", ex, traceback.format_exc())
        st.session_state["_s1_upload_err"] = f"Failed to load config: {ex}"
        st.rerun()


def _load_teacher_bytes(raw: bytes, fname: str):
    log.info("_load_teacher_bytes: '%s' %d bytes", fname, len(raw))
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active
        max_col = max(
            (ci for row in ws.iter_rows() for ci, c in enumerate(row, 1) if c.value is not None),
            default=0)
        if max_col > 1:
            log.warning("_load_teacher_bytes: %d columns found, expected 1", max_col)
            st.session_state["_s1_upload_err"] = (
                f"File has data in {max_col} columns — use Column A only.\n\n"
                "Each row in Column A should contain exactly one teacher name.")
            st.rerun(); return
        names = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            v = row[0].value
            if v:
                n = str(v).strip()
                if n and n.lower() != "teacher name":
                    names.append(n)
        if not names:
            log.warning("_load_teacher_bytes: no names found")
            st.session_state["_s1_upload_err"] = (
                "No teacher names found in the file.\n\n"
                "Make sure names are in Column A (Row 1 header 'Teacher Name' is auto-skipped).")
            st.rerun(); return
        seen, dups = set(), []
        for n in names:
            if n in seen: dups.append(n)
            seen.add(n)
        if dups:
            log.warning("_load_teacher_bytes: duplicates: %s", dups)
            st.session_state["_s1_upload_err"] = (
                "Duplicate teacher names found:\n\n" +
                "\n".join(f"  • {d}" for d in sorted(set(dups))) +
                "\n\nPlease remove duplicates and re-upload.")
            st.rerun(); return
        names.sort()
        st.session_state["s1_teachers"]      = names
        st.session_state["s1_teacher_fname"] = fname
        _notify(f"✓ {len(names)} teachers loaded (A→Z).", "success")
        log.info("_load_teacher_bytes: %d teachers loaded", len(names))
        st.rerun()
    except Exception as ex:
        log.error("_load_teacher_bytes: %s\n%s", ex, traceback.format_exc())
        st.session_state["_s1_upload_err"] = f"Error reading file '{fname}': {ex}"
        st.rerun()


def _step1_save_and_continue():
    log.info("_step1_save_and_continue: called")
    ppd   = int(st.session_state.get("ni_ppd",   7))
    wdays = int(st.session_state.get("ni_wdays", 6))
    fhalf = int(st.session_state.get("ni_fhalf", 4))
    shalf = int(st.session_state.get("ni_shalf", 3))
    teachers = st.session_state.s1_teachers
    log.debug("_step1_save_and_continue: ppd=%d wdays=%d fhalf=%d shalf=%d teachers=%d",
              ppd, wdays, fhalf, shalf, len(teachers))

    errors = []
    if ppd <= 0 or wdays <= 0:
        errors.append({"msg": "Periods per day and working days must be ≥ 1.",
                        "fix": "Set valid values in the Periods & Working Days section."})
    if fhalf + shalf != ppd:
        errors.append({"msg": f"Halves mismatch: {fhalf} + {shalf} = {fhalf+shalf}, but need {ppd}.",
                        "fix": "Adjust 'Periods — first half' or 'Periods — second half' so they add up to Periods per day."})
    if not teachers:
        errors.append({"msg": "No teacher list loaded.",
                        "fix": "Upload a teacher Excel file (Column A, one name per row)."})
    if errors:
        log.warning("_step1_save_and_continue: %d validation errors", len(errors))
        st.session_state["_s1_val_errors"] = errors
        st.rerun()
        return

    sections = {cls: int(v) for cls, v in st.session_state.s1_sections.items()}

    engine = _get_eng()
    engine.configuration = {
        "periods_per_day":    ppd,
        "working_days":       wdays,
        "periods_first_half": fhalf,
        "periods_second_half":shalf,
        "teacher_file":       st.session_state.s1_teacher_fname,
        "teacher_names":      teachers,
        "classes":            sections,
    }
    log.info("_step1_save_and_continue: configuration set, %d total sections",
             sum(sections.values()))

    for cls in range(6, 13):
        for si in range(sections.get(cls, 0)):
            cn = f"{cls}{chr(65+si)}"
            if cn not in engine.class_config_data:
                engine.class_config_data[cn] = {
                    "subjects": [], "teacher": "",
                    "teacher_period": 1, "editing_index": None,
                }
    log.info("_step1_save_and_continue: %d classes ready → step2",
             len(engine.class_config_data))
    st.session_state["s1_validated"] = True
    st.session_state["s2_validated"] = False
    st.session_state["s3_validated"] = False
    _nav("step2")


# ═════════════════════════════════════════════════════════════════════════════
#  STEP 2 — Class Assignments
# ═════════════════════════════════════════════════════════════════════════════

def page_step2():
    eng = _get_eng()

    if not eng.configuration:
        st.warning("⚠ Step 1 not yet completed — basic configuration is missing.")
        st.info("Please complete Step 1 (Basic Config) before continuing.")
        if st.button("← Go to Step 1", type="primary", key="s2_guard_back"):
            _nav("step1")
        return

    log.info("page_step2: render (%d classes)", len(_all_classes()))

    if st.session_state.get("_s2_val_errors"):
        _step2_error_dialog(st.session_state.pop("_s2_val_errors"))

    cfg       = eng.configuration
    ppd       = cfg["periods_per_day"]
    wdays     = cfg["working_days"]
    required  = ppd * wdays
    teachers  = sorted(cfg["teacher_names"])
    day_names = DAY_NAMES[:wdays]
    all_cn    = _all_classes()

    _show_upload_error_if_any("_s2_upload_err")

    _header("👨‍🏫 Step 2: Configure Each Class",
            "Set class teacher and add subjects with periods, preferences and constraints.")
    _show_notifications()

    # ── Save / Load ───────────────────────────────────────────────────────────
    with st.expander("💾 Save / Load Assignments", expanded=False):
        c1, c2 = st.columns(2)
        with c1:
            if st.button("⬇ Prepare Download", key="s2_dl_btn", use_container_width=True):
                log.info("page_step2: preparing assignments download")
                payload = {
                    cn: {"teacher": cd.get("teacher",""),
                         "teacher_period": cd.get("teacher_period",1),
                         "subjects": cd.get("subjects",[])}
                    for cn, cd in _get_eng().class_config_data.items()
                }
                st.session_state["_s2_dl_data"] = {
                    "assignments": payload,
                    "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                }
            if "_s2_dl_data" in st.session_state:
                _json_download(st.session_state["_s2_dl_data"], "📥 Click to Download",
                               f"Assignments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
        with c2:
            up = st.file_uploader("Upload Assignments JSON", type=["json"], key="s2_load_json")
            if up is not None:
                raw = up.read()
                log.debug("page_step2: upload size=%d", len(raw))
                if len(raw) == 0:
                    log.warning("page_step2: 0 bytes, skipping")
                elif not _already_processed("s2_load_json", raw):
                    _load_step2_assignments(raw)

    # ── Navigation ────────────────────────────────────────────────────────────
    s2_validated = st.session_state.get("s2_validated", False)
    cb, cv, cc = st.columns([1, 2, 2])
    with cb:
        if st.button("← Back to Step 1", use_container_width=True):
            _nav("step1")
    with cv:
        if st.button("🔍 Validate Step 2", use_container_width=True,
                     help="Check period counts and teacher conflicts"):
            _step2_validate_and_continue()
    with cc:
        if s2_validated:
            if st.button("✓ Continue to Step 3 →", type="primary", use_container_width=True):
                _nav("step3")
        else:
            st.button("✓ Continue to Step 3 →", type="primary",
                      use_container_width=True, disabled=True,
                      help="Run Validate Step 2 first and fix any errors")

    if s2_validated:
        st.success("✅ Step 2 validated — click **Continue to Step 3** to proceed.")
    else:
        st.info("ℹ️ Fill in all class subjects, then click **🔍 Validate Step 2** to check for errors.")

    if not all_cn:
        st.warning("No classes found — go back to Step 1.")
        return

    st.divider()

    # ── Build grade → sections map ────────────────────────────────────────────
    all_cn_grouped = {}
    for cn in all_cn:
        all_cn_grouped.setdefault(cn[:-1], []).append(cn)
    grades = sorted(all_cn_grouped.keys(), key=int)

    if ("s2_active_class" not in st.session_state
            or st.session_state["s2_active_class"] not in all_cn):
        st.session_state["s2_active_class"] = all_cn[0]

    act = st.session_state["s2_active_class"]

    def _cn_fmt(cn):
        cd = eng.class_config_data.get(cn, {})
        t  = sum(s.get("periods", 0) for s in cd.get("subjects", []))
        ic = "🟢" if t == required else ("🔴" if t > required else "⚪")
        return f"{ic} {cn}  ({t}/{required})"

    sel_col, info_col = st.columns([2, 5])
    with sel_col:
        st.markdown("**Select Class**")
        cur_grade = act[:-1]
        grade_idx = grades.index(cur_grade) if cur_grade in grades else 0

        chosen_grade = st.radio(
            "Grade", grades,
            index=grade_idx,
            horizontal=True,
            key="s2_grade_radio",
            label_visibility="collapsed",
            format_func=lambda g: f"Class {g}",
        )

        if chosen_grade != cur_grade:
            act = all_cn_grouped[chosen_grade][0]
            st.session_state["s2_active_class"] = act
            cur_grade = chosen_grade

        sections = all_cn_grouped[cur_grade]
        sec_idx   = sections.index(act) if act in sections else 0

        chosen_sec = st.radio(
            "Section", sections,
            index=sec_idx,
            horizontal=True,
            key=f"s2_sec_{cur_grade}",
            label_visibility="collapsed",
        )

        if chosen_sec != act:
            st.session_state["s2_active_class"] = chosen_sec
            st.rerun()

        cur_sel = st.session_state["s2_active_class"]

    with info_col:
        cd_prev = eng.class_config_data.get(cur_sel, {})
        tot     = sum(s.get("periods", 0) for s in cd_prev.get("subjects", []))
        diff    = required - tot
        note    = "✓ exact" if tot == required else (
            f"need {diff} more" if diff > 0 else f"over by {-diff}")
        ic      = "🟢" if tot == required else ("🔴" if tot > required else "⚪")
        st.markdown(f"### Class {cur_sel}")
        st.caption(f"{ic}  Assigned: **{tot}** / {required}  ({note})  "
                   f"— {len(cd_prev.get('subjects', []))} subject(s)")
        with st.expander("📊 All Classes Progress", expanded=False):
            gcols = st.columns(len(grades))
            for gi, g in enumerate(grades):
                with gcols[gi]:
                    st.markdown(f"**Gr.{g}**")
                    for cn in all_cn_grouped[g]:
                        st.caption(_cn_fmt(cn))

    st.divider()
    _class_config_tab(cur_sel, teachers, ppd, wdays, required, day_names)


def _load_step2_assignments(raw: bytes):
    """Load assignments JSON — shows a modal dialog for wrong file uploads."""
    eng = _get_eng()
    log.info("_load_step2_assignments: %d bytes", len(raw))
    try:
        d = json.loads(raw.decode("utf-8"))
    except json.JSONDecodeError as ex:
        log.error("_load_step2_assignments: JSON parse error: %s", ex)
        st.session_state["_s2_upload_err"] = f"Invalid JSON file: {ex}"
        st.rerun()
        return

    if "assignments" not in d:
        wrong_hint = ""
        if "periods_per_day" in d or "teacher_names" in d:
            wrong_hint = "\n\nThis looks like a **Step 1 config** file."
        elif "step3_data" in d or "step3_unavailability" in d:
            wrong_hint = "\n\nThis looks like a **Step 3 config** file."
        msg = (f"Wrong file type for Step 2.{wrong_hint}\n\n"
               f"Please upload an **Assignments file** (saved from Step 2 → Save/Load).")
        log.error("_load_step2_assignments: wrong file")
        st.session_state["_s2_upload_err"] = msg
        st.rerun()
        return

    n = 0
    for cn, saved in d["assignments"].items():
        if cn in eng.class_config_data:
            eng.class_config_data[cn].update({
                "teacher":        saved.get("teacher", ""),
                "teacher_period": saved.get("teacher_period", 1),
                "subjects":       saved.get("subjects", []),
                "editing_index":  None,
            })
            _purge_form_state(cn)
            n += 1
    _notify(f"✓ Assignments loaded for {n} classes.", "success")
    log.info("_load_step2_assignments: %d classes loaded", n)
    st.rerun()


def _purge_form_state(cn: str):
    """Delete ALL form widget keys for class cn from session state."""
    prefix = f"sf_"
    needle = f"_{cn}_"
    to_del = [k for k in list(st.session_state.keys())
              if k.startswith(prefix) and needle in k]
    for k in to_del:
        del st.session_state[k]
    st.session_state.pop(f"s2_fv_{cn}", None)
    log.debug("_purge_form_state: %s cleared %d keys", cn, len(to_del))


def _s2_init_form(cn, suffix, prefill: dict):
    """Pre-populate form widget keys with prefill values (only on first use)."""
    init_key = f"sf_init_{cn}_{suffix}"
    if st.session_state.get(init_key):
        return
    _cfg  = _get_eng().configuration
    ppd   = _cfg["periods_per_day"]
    wdays = _cfg["working_days"]
    st.session_state[f"sf_name_{cn}_{suffix}"]   = prefill.get("name", "")
    st.session_state[f"sf_teach_{cn}_{suffix}"]  = prefill.get("teacher", "")
    st.session_state[f"sf_per_{cn}_{suffix}"]    = int(prefill.get("periods", 1))
    st.session_state[f"sf_cons_{cn}_{suffix}"]   = prefill.get("consecutive", "No")
    st.session_state[f"sf_pref_{cn}_{suffix}"]   = [int(p) for p in prefill.get("periods_pref", [])]
    st.session_state[f"sf_day_{cn}_{suffix}"]    = list(prefill.get("days_pref", []))
    
    # Handle both old format (parallel/parallel_subject/parallel_teacher) and new format (parallel_subjects)
    parallel_subjects = prefill.get("parallel_subjects", [])
    if not parallel_subjects and prefill.get("parallel"):
        # Convert old format to new format
        old_par_subj = prefill.get("parallel_subject", "").strip()
        old_par_teach = prefill.get("parallel_teacher", "")
        if old_par_subj and old_par_teach:
            parallel_subjects = [{"subject": old_par_subj, "teacher": old_par_teach}]
    
    st.session_state[f"sf_parallels_{cn}_{suffix}"] = parallel_subjects
    st.session_state[init_key] = True
    log.debug("_s2_init_form: %s/%s initialised with %d parallel configs", cn, suffix, len(parallel_subjects))


def _class_config_tab(cn, teachers, ppd, wdays, required, day_names):
    eng = _get_eng()
    log.debug("_class_config_tab: %s", cn)
    cd = eng.class_config_data.setdefault(cn, {
        "subjects": [], "teacher": "", "teacher_period": 1, "editing_index": None})
    subjects = cd.get("subjects", [])

    # ── Class Teacher ─────────────────────────────────────────────────────────
    with st.container(border=True):
        st.markdown(f"**Class Teacher — {cn}**")
        c1, c2 = st.columns([3, 1])
        with c1:
            opts   = [""] + teachers
            cur_ct = cd.get("teacher", "")
            sel_ct = st.selectbox("Class Teacher", opts,
                                   index=opts.index(cur_ct) if cur_ct in opts else 0,
                                   key=f"ct_{cn}")
            cd["teacher"] = sel_ct
        with c2:
            cur_per = int(cd.get("teacher_period", 1))
            sel_per = st.number_input("CT Period", 1, ppd, cur_per, key=f"ctp_{cn}")
            cd["teacher_period"] = int(sel_per)

    # ── Subject List ──────────────────────────────────────────────────────────
    if subjects:
        with st.container(border=True):
            st.markdown("**Subjects**")
            hdr = st.columns([0.4, 2.2, 2.2, 0.9, 1.0, 1.0, 0.7, 0.7])
            for h, t in zip(hdr, ["#", "Subject", "Teacher", "Periods",
                                   "Consec", "Parallel", "✏", "🗑"]):
                h.markdown(f"**{t}**")
            for i, s in enumerate(subjects):
                row = st.columns([0.4, 2.2, 2.2, 0.9, 1.0, 1.0, 0.7, 0.7])
                row[0].write(str(i + 1))
                row[1].write(s.get("name", "—"))
                row[2].write(s.get("teacher", "—"))
                row[3].write(str(s.get("periods", "")))
                row[4].write("Yes" if s.get("consecutive") == "Yes" else "—")
                row[5].write("✓" if s.get("parallel") else "—")
                if row[6].button("✏", key=f"edit_{cn}_{i}"):
                    log.debug("_class_config_tab: %s edit %d", cn, i)
                    cd["editing_index"] = i
                    st.session_state.pop(f"sf_init_{cn}_e{i}", None)
                    st.rerun()
                if row[7].button("🗑", key=f"del_{cn}_{i}"):
                    log.info("_class_config_tab: %s delete %d '%s'", cn, i, s.get("name", ""))
                    subjects.pop(i)
                    ei = cd.get("editing_index")
                    if ei == i:
                        cd["editing_index"] = None
                    elif ei is not None and ei > i:
                        cd["editing_index"] = ei - 1
                    st.rerun()

    # ── Add / Edit Form ───────────────────────────────────────────────────────
    editing_idx = cd.get("editing_index")
    form_ver    = st.session_state.get(f"s2_fv_{cn}", 0)

    if editing_idx is not None and editing_idx < len(subjects):
        suffix     = f"e{editing_idx}"
        prefill    = subjects[editing_idx]
        form_title = f"✏ Edit Subject #{editing_idx + 1}"
    else:
        suffix     = f"a{form_ver}"
        prefill    = {}
        form_title = "➕ Add Subject"

    _s2_init_form(cn, suffix, prefill)

    with st.container(border=True):
        st.markdown(f"**{form_title}**")

        form_err_key = f"s2_form_err_{cn}"
        if form_err_key in st.session_state:
            st.error(st.session_state.pop(form_err_key))

        c1, c2, c3 = st.columns(3)
        with c1:
            name = st.text_input("Subject Name",
                                  key=f"sf_name_{cn}_{suffix}")
        with c2:
            opts  = [""] + teachers
            sel_t = st.selectbox("Teacher", opts,
                                  key=f"sf_teach_{cn}_{suffix}")
        with c3:
            pers = st.number_input("Periods/week", 1, ppd * wdays,
                                    key=f"sf_per_{cn}_{suffix}")
        c4, c5 = st.columns(2)
        with c4:
            consec = st.selectbox("Consecutive?", ["No", "Yes"],
                                   key=f"sf_cons_{cn}_{suffix}")
        with c5:
            p_prefs = st.multiselect("Period prefs (optional)", list(range(1, ppd + 1)),
                                      key=f"sf_pref_{cn}_{suffix}")
        d_prefs = st.multiselect("Day prefs (optional)", day_names,
                                  key=f"sf_day_{cn}_{suffix}")
        
        # ── Multiple Parallel Subjects Support ──────────────────────────────
        st.markdown("**Parallel Subjects** (optional — can add multiple)")
        
        # Initialize parallel list in form state if not present
        par_key = f"sf_parallels_{cn}_{suffix}"
        if par_key not in st.session_state:
            st.session_state[par_key] = prefill.get("parallel_subjects", [])
        
        parallels = st.session_state[par_key]
        
        # Display existing parallels with edit/delete options
        for pidx, pconfig in enumerate(list(parallels)):
            pcol1, pcol2, pcol3 = st.columns([2, 2, 0.8])
            with pcol1:
                st.text_input("Parallel Subject", value=pconfig.get("subject", ""), 
                             disabled=True, key=f"par_sub_display_{cn}_{suffix}_{pidx}")
            with pcol2:
                opts2 = [""] + teachers
                st.selectbox("Teacher", options=opts2, 
                            index=opts2.index(pconfig.get("teacher", "")) if pconfig.get("teacher", "") in opts2 else 0,
                            disabled=True, key=f"par_tea_display_{cn}_{suffix}_{pidx}")
            with pcol3:
                if st.button("🗑", key=f"del_par_{cn}_{suffix}_{pidx}", help="Remove"):
                    parallels.pop(pidx)
                    st.rerun()
        
        # Add new parallel section
        st.markdown("##### Add New Parallel Configuration")
        pnew_col1, pnew_col2, pnew_col3 = st.columns([2, 2, 0.8])
        with pnew_col1:
            pnew_subj = st.text_input("Parallel Subject", key=f"sf_psub_new_{cn}_{suffix}")
        with pnew_col2:
            opts2 = [""] + teachers
            pnew_teach = st.selectbox("Parallel Teacher", opts2, key=f"sf_pteach_new_{cn}_{suffix}")
        with pnew_col3:
            if st.button("➕ Add", key=f"add_par_{cn}_{suffix}"):
                if pnew_subj.strip() and pnew_teach:
                    parallels.append({
                        "subject": pnew_subj.strip(),
                        "teacher": pnew_teach,
                    })
                    st.session_state[par_key] = parallels
                    st.rerun()
                elif pnew_subj.strip() and not pnew_teach:
                    st.session_state[form_err_key] = "⚠ Please select a parallel teacher."
                    st.rerun()

        btn1, btn2 = st.columns(2)
        with btn1:
            lbl = "✓ Update" if editing_idx is not None else "✓ Add Subject"
            if st.button(lbl, key=f"sf_save_{cn}_{suffix}", type="primary",
                         use_container_width=True):
                log.info("_class_config_tab: %s save '%s'", cn, name)
                if not name.strip():
                    st.session_state[form_err_key] = "⚠ Subject name is required."
                    st.rerun()
                elif not sel_t:
                    st.session_state[form_err_key] = "⚠ Please select a teacher."
                    st.rerun()
                else:
                    entry = {
                        "name":             name.strip(),
                        "teacher":          sel_t,
                        "periods":          int(pers),
                        "consecutive":      consec,
                        "periods_pref":     p_prefs,
                        "days_pref":        d_prefs,
                        "parallel_subjects": st.session_state.get(f"sf_parallels_{cn}_{suffix}", []),
                    }
                    if editing_idx is not None:
                        subjects[editing_idx] = entry
                        cd["editing_index"] = None
                        st.session_state.pop(f"sf_init_{cn}_e{editing_idx}", None)
                        st.toast(f"✓ '{name.strip()}' updated.", icon="✅")
                    else:
                        subjects.append(entry)
                        new_ver = form_ver + 1
                        st.session_state[f"s2_fv_{cn}"] = new_ver
                        st.toast(f"✓ '{name.strip()}' added.", icon="✅")
                    cd["editing_index"] = None
                    st.rerun()
        with btn2:
            if editing_idx is not None:
                if st.button("✕ Cancel", key=f"sf_cancel_{cn}",
                             use_container_width=True):
                    cd["editing_index"] = None
                    st.rerun()


def _step2_validate_and_continue():
    """Run all validation checks — period counts, teacher conflicts, within-class conflicts."""
    eng = _get_eng()
    log.info("_step2_validate_and_continue: running")
    cfg       = eng.configuration
    ppd       = cfg["periods_per_day"]
    wdays     = cfg["working_days"]
    required  = ppd * wdays
    all_cn    = _all_classes()
    day_names = DAY_NAMES[:wdays]

    # ── Section ①: Period counts ─────────────────────────────────────────────
    period_errors = []
    period_ok     = []
    teacher_slots = {}

    def _add(teacher, desc):
        if teacher:
            teacher_slots.setdefault(teacher, []).append(desc)

    for cn in all_cn:
        cd    = eng.class_config_data.get(cn, {})
        subjs = cd.get("subjects", [])
        if not subjs:
            period_errors.append((cn, "NO SUBJECTS added"))
            log.warning("_step2_validate: %s no subjects", cn)
            continue

        total = sum(s.get("periods", 0) for s in subjs)
        diff  = total - required
        sign  = "+" if diff > 0 else ""
        if total == required:
            period_ok.append((cn, f"{total}/{required} periods  ({len(subjs)} subjects)"))
        else:
            period_errors.append((cn,
                f"Period mismatch: {total} assigned, need {required}  ({sign}{diff})"))
            log.warning("_step2_validate: %s %d≠%d", cn, total, required)

        ct     = cd.get("teacher", "").strip()
        ct_per = cd.get("teacher_period", 1)
        if ct:
            _add(ct, {
                "class":        cn,
                "label":        f"Class Teacher of {cn} (Period {ct_per}, every day)",
                "fixed_period": ct_per,
                "period_prefs": [ct_per],
                "day_set":      set(day_names),
                "is_class_teacher": True,
                "subj_name":    "",
            })

        for s in subjs:
            t = s.get("teacher", "").strip()
            if t:
                _add(t, {
                    "class":        cn,
                    "label":        f"Subject '{s['name']}' in {cn}  (×{s['periods']} periods/wk)",
                    "fixed_period": None,
                    "period_prefs": list(s.get("periods_pref", [])),
                    "day_set":      set(s.get("days_pref", []) or day_names),
                    "is_class_teacher": False,
                    "subj_name":    s["name"],
                    "consecutive":  s.get("consecutive") == "Yes",
                })
            pt = s.get("parallel_teacher", "").strip() if s.get("parallel") else ""
            if pt:
                _add(pt, {
                    "class":        cn,
                    "label":        f"Parallel teacher for '{s.get('parallel_subject','')}' in {cn}",
                    "fixed_period": None,
                    "period_prefs": list(s.get("periods_pref", [])),
                    "day_set":      set(s.get("days_pref", []) or day_names),
                    "is_class_teacher": False,
                    "subj_name":    s.get("parallel_subject", ""),
                })

    # ── Section ②: Teacher-level hard conflicts ──────────────────────────────
    hard_conflicts = []

    def _period_overlap(prefs_a, prefs_b):
        if not prefs_a or not prefs_b:
            return None
        s = set(prefs_a) & set(prefs_b)
        return s if s else None

    for teacher, slots in teacher_slots.items():
        n = len(slots)
        for i in range(n):
            for j in range(i + 1, n):
                a, b = slots[i], slots[j]
                if a["class"] == b["class"]:
                    continue
                days = a["day_set"] & b["day_set"]
                if not days:
                    continue
                if a["is_class_teacher"] and b["period_prefs"]:
                    if a["fixed_period"] in b["period_prefs"]:
                        hard_conflicts.append({
                            "teacher": teacher, "slot_a": a["label"], "slot_b": b["label"],
                            "reason": (
                                f"Period {a['fixed_period']} is fixed every day for "
                                f"class-teacher duty in {a['class']}, but also listed as "
                                f"preferred period for '{b.get('subj_name','')}' in "
                                f"{b['class']} on days: {', '.join(sorted(days))}")
                        })
                        continue
                if b["is_class_teacher"] and a["period_prefs"]:
                    if b["fixed_period"] in a["period_prefs"]:
                        hard_conflicts.append({
                            "teacher": teacher, "slot_a": b["label"], "slot_b": a["label"],
                            "reason": (
                                f"Period {b['fixed_period']} is fixed every day for "
                                f"class-teacher duty in {b['class']}, but also listed as "
                                f"preferred period for '{a.get('subj_name','')}' in "
                                f"{a['class']} on days: {', '.join(sorted(days))}")
                        })
                        continue
                if a["is_class_teacher"] and b["is_class_teacher"]:
                    hard_conflicts.append({
                        "teacher": teacher, "slot_a": a["label"], "slot_b": b["label"],
                        "reason": (
                            f"Cannot be Class Teacher of two classes simultaneously. "
                            f"Fixed Period {a['fixed_period']} (all days) for {a['class']} "
                            f"AND Period {b['fixed_period']} (all days) for {b['class']}.")
                    })
                    continue
                pov = _period_overlap(a["period_prefs"], b["period_prefs"])
                if pov is not None:
                    hard_conflicts.append({
                        "teacher": teacher, "slot_a": a["label"], "slot_b": b["label"],
                        "reason": (
                            f"Both require Period(s) {sorted(pov)} on "
                            f"{', '.join(sorted(days))} — teacher cannot be in "
                            f"{a['class']} and {b['class']} simultaneously.")
                    })

    log.info("_step2_validate: hard_conflicts=%d", len(hard_conflicts))

    # ── Section ③: Within-class slot conflicts ────────────────────────────
    within_class_conflicts = []

    for cn in all_cn:
        cd     = eng.class_config_data.get(cn, {})
        subjs  = cd.get("subjects", [])

        items = []
        for s in subjs:
            if not s.get("periods_pref"):
                continue
            period_set = set(s["periods_pref"])
            day_set    = set(s.get("days_pref", []) or day_names)
            n_periods  = s["periods"]
            n_days     = len(day_set)
            need_per_day = math.ceil(n_periods / n_days) if n_days > 0 else 1
            items.append({
                "label":        "Subject '{}' (Period(s) {}, day(s) {}, teacher: {})".format(
                                    s["name"],
                                    sorted(period_set),
                                    sorted(day_set) if s.get("days_pref") else "any",
                                    s.get("teacher", "").strip()),
                "period_set":   period_set,
                "day_set":      day_set,
                "need_per_day": need_per_day,
                "teacher":      s.get("teacher", "").strip(),
            })

        if len(items) < 2:
            continue

        for day in day_names:
            active = [it for it in items if day in it["day_set"]]
            if len(active) < 2:
                continue

            for i in range(len(active)):
                for j in range(i + 1, len(active)):
                    a, b = active[i], active[j]
                    if a["teacher"] == b["teacher"]:
                        continue
                    combined_slots = a["period_set"] | b["period_set"]
                    combined_need  = a["need_per_day"] + b["need_per_day"]
                    if combined_need > len(combined_slots):
                        contested = a["period_set"] & b["period_set"]
                        within_class_conflicts.append({
                            "class": cn, "day": day,
                            "item_a": a["label"], "item_b": b["label"],
                            "reason": (
                                f"On {day}, two subjects by DIFFERENT teachers both need "
                                f"Period(s) {sorted(contested) if contested else sorted(combined_slots)} "
                                f"in class {cn} — combined demand "
                                f"({a['need_per_day']} + {b['need_per_day']} = {combined_need}) "
                                f"exceeds available slots {sorted(combined_slots)} — "
                                f"the class grid cannot fit both.")
                        })

            all_pref_periods = set()
            for it in active:
                all_pref_periods |= it["period_set"]
            for p in all_pref_periods:
                teachers_needing_p = {it["teacher"] for it in active if p in it["period_set"]}
                if len(teachers_needing_p) > 1:
                    already = any(c["class"] == cn and c["day"] == day
                                  for c in within_class_conflicts)
                    if not already:
                        within_class_conflicts.append({
                            "class": cn, "day": day,
                            "item_a": f"{len(teachers_needing_p)} subjects from different teachers",
                            "item_b": "",
                            "reason": (
                                f"On {day}, Period {p} in class {cn} is claimed by "
                                f"{len(teachers_needing_p)} different teachers "
                                f"({', '.join(sorted(teachers_needing_p))}) — "
                                f"only one teacher can occupy a period slot.")
                        })

    log.info("_step2_validate: errors=%d hc=%d wcc=%d",
             len(period_errors), len(hard_conflicts), len(within_class_conflicts))

    vr = {
        "ok":                    not period_errors and not hard_conflicts and not within_class_conflicts,
        "period_errors":         period_errors,
        "period_ok":             period_ok,
        "hard_conflicts":        hard_conflicts,
        "within_class_conflicts": within_class_conflicts,
        "required":              required,
        "wdays":                 wdays,
        "ppd":                   ppd,
    }
    st.session_state["s2_validation_result"] = vr

    if vr["ok"]:
        st.session_state["s2_validated"] = True
        st.session_state["s3_validated"] = False
        _notify("✅ All validation checks passed — click Continue to Step 3.", "success")
        _nav("step3")
    else:
        total_err = len(period_errors) + len(hard_conflicts) + len(within_class_conflicts)
        st.session_state["s2_validated"] = False
        st.session_state["_s2_val_errors"] = vr
        log.warning("_step2_validate: %d errors found, showing popup", total_err)
        st.rerun()


def _display_s2_validation(vr):
    """Render the full 3-section validation report."""
    period_errors  = vr.get("period_errors", [])
    period_ok      = vr.get("period_ok", [])
    hard_conflicts = vr.get("hard_conflicts", [])
    wcc            = vr.get("within_class_conflicts", [])
    required       = vr.get("required", "?")
    wdays          = vr.get("wdays", "?")
    ppd            = vr.get("ppd", "?")
    any_error      = bool(period_errors or hard_conflicts or wcc)

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        if period_errors: st.error(f"**{len(period_errors)}**\nPeriod Errors")
        else:             st.success(f"**0**\nPeriod Errors")
    with c2:
        if hard_conflicts: st.error(f"**{len(hard_conflicts)}**\nTeacher Conflicts")
        else:              st.success(f"**0**\nTeacher Conflicts")
    with c3:
        if wcc: st.error(f"**{len(wcc)}**\nWithin-Class Conflicts")
        else:   st.success(f"**0**\nWithin-Class Conflicts")
    with c4:
        st.success(f"**{len(period_ok)}**\nClasses OK")

    st.divider()

    s1_label = (f"① PERIOD COUNT — {len(period_ok)} OK / {len(period_errors)} ERRORS")
    with st.expander(s1_label, expanded=bool(period_errors)):
        st.caption(f"Required per class: **{required}** periods/week  "
                   f"({wdays} days × {ppd} periods)")
        if period_errors:
            st.markdown("**❌ Fix these classes:**")
            for cn, msg in sorted(period_errors):
                st.error(f"**{cn}** — {msg}")
        if period_ok:
            st.markdown("**✓ Classes with correct period count:**")
            cols = st.columns(3)
            for idx, (cn, msg) in enumerate(sorted(period_ok)):
                with cols[idx % 3]:
                    st.success(f"**{cn}**: {msg}")

    s2_label = f"② TEACHER CONFLICTS (across classes) — {len(hard_conflicts)} found"
    with st.expander(s2_label, expanded=bool(hard_conflicts)):
        if not hard_conflicts:
            st.success("✓ No teacher-level hard conflicts detected.")
        else:
            st.error("DEFINITE impossibilities — a teacher cannot be in two classes at the same time.")
            for i, c in enumerate(hard_conflicts, 1):
                with st.container(border=True):
                    st.markdown(f"**[{i}] Teacher: {c['teacher']}**")
                    st.caption(f"Assignment A: {c['slot_a']}")
                    st.caption(f"Assignment B: {c['slot_b']}")
                    st.warning(f"⚠ {c['reason']}")
                    st.info("💡 **FIX:** Change the period/day preference for one of the two assignments.")

    s3_label = f"③ WITHIN-CLASS SLOT CONFLICTS — {len(wcc)} found"
    with st.expander(s3_label, expanded=bool(wcc)):
        if not wcc:
            st.success("✓ No within-class slot conflicts detected.")
        else:
            st.error("DEFINITE impossibilities — two or more subjects in the same class are pinned "
                     "to the same period slot(s) on the same day, leaving no valid placement.")
            for i, c in enumerate(wcc, 1):
                with st.container(border=True):
                    st.markdown(f"**[{i}] Class: {c['class']}  —  Day: {c['day']}**")
                    st.caption(f"Item A: {c['item_a']}")
                    if c.get("item_b"):
                        st.caption(f"Item B: {c['item_b']}")
                    st.warning(f"⚠ {c['reason']}")
                    st.info("💡 **FIX:** Adjust the period preference or day preference "
                            "so subjects do not compete for the same slot.")

    st.divider()
    if any_error:
        total_err = len(period_errors) + len(hard_conflicts) + len(wcc)
        st.error(f"**RESULT:** {total_err} error(s) found. Fix all errors above then "
                 f"click **✓ Validate & Complete** again.")
    else:
        st.success("**RESULT:** All checks passed. Click **✓ Validate & Complete** to proceed to Step 3.")


# ═════════════════════════════════════════════════════════════════════════════
#  STEP 3 — Teacher Settings
# ═════════════════════════════════════════════════════════════════════════════

def page_step3():
    eng = _get_eng()

    if not eng.configuration:
        st.warning("⚠ Step 1 not yet completed — basic configuration is missing.")
        st.info("Please complete Step 1 (Basic Config) before continuing.")
        if st.button("← Go to Step 1", type="primary", key="s3_guard_back"):
            _nav("step1")
        return

    log.info("page_step3: render")
    _header("⚙ Step 3: Teacher Settings",
            "Review workload, define combined classes and mark teacher unavailability.")
    _show_notifications()

    cfg       = eng.configuration
    ppd       = cfg["periods_per_day"]
    wdays     = cfg["working_days"]
    teachers  = sorted(cfg.get("teacher_names",[]))
    day_names = DAY_NAMES[:wdays]
    all_cn    = _all_classes()

    if not hasattr(eng,"step3_data"):           eng.step3_data = {}
    if not hasattr(eng,"step3_unavailability"): eng.step3_unavailability = {}

    try:
        eng.prepare_step3_workload()
    except Exception as ex:
        log.warning("page_step3: prepare_step3_workload failed: %s", ex)

    with st.expander("💾 Save / Load Step 3 Config", expanded=False):
        c1,c2 = st.columns(2)
        with c1:
            if st.button("⬇ Prepare Download", key="s3_dl_btn", use_container_width=True):
                log.info("page_step3: preparing download")
                st.session_state["_s3_dl_data"] = {
                    "step3_data": {t:{"skipped":v.get("skipped",False),"combines":v.get("combines",[])}
                                   for t,v in _get_eng().step3_data.items()},
                    "step3_unavailability": {t:{"days":list(v["days"]),"periods":list(v["periods"])}
                                             for t,v in _get_eng().step3_unavailability.items()},
                    "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "step":3,
                }
            if "_s3_dl_data" in st.session_state:
                _json_download(st.session_state["_s3_dl_data"], "📥 Click to Download",
                               f"Step3_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
        with c2:
            up = st.file_uploader("Upload Step 3 JSON", type=["json"], key="s3_load_json")
            if up is not None:
                raw = up.read()
                log.debug("page_step3: upload size=%d", len(raw))
                if len(raw)==0:
                    log.warning("page_step3: 0 bytes, skipping")
                elif not _already_processed("s3_load_json", raw):
                    _load_step3_config(raw)

    if st.session_state.get("_s3_show_error_popup"):
        st.session_state["_s3_show_error_popup"] = False
        vr_err = st.session_state.get("s3_validation_result")
        if vr_err and not vr_err.get("can_proceed"):
            _step3_error_dialog(vr_err)

    cb, cv, cc = st.columns(3)
    with cb:
        if st.button("← Back to Step 2", use_container_width=True):
            _nav("step2")
    with cv:
        if st.button("🔍 Validate Step 3", use_container_width=True):
            log.info("page_step3: validate")
            try:
                eng.prepare_step3_workload()
                vr = eng.validate_step3()
                st.session_state["s3_validation_result"] = vr
                log.info("page_step3: validate result can_proceed=%s issues=%d",
                         vr["can_proceed"], len(vr["issues"]))
                if vr["can_proceed"]:
                    st.session_state["s3_validated"] = True
                    _notify("✅ All clear — click Generate Timetable to proceed.", "success")
                else:
                    st.session_state["s3_validated"] = False
                    st.session_state["_s3_show_error_popup"] = True
                    log.warning("page_step3: %d overloads, showing popup", len(vr["issues"]))
            except Exception as ex:
                log.error("page_step3 validate: %s\n%s", ex, traceback.format_exc())
                _notify(f"Validation error: {ex}", "error")
            st.rerun()
    with cc:
        s3_validated = st.session_state.get("s3_validated", False)
        if s3_validated:
            if st.button("🚀 Generate Timetable →", type="primary", use_container_width=True):
                log.info("page_step3: proceed to generate")
                st.session_state["s3_validated"] = True
                for k in ("s4_stage","s4_s1_status","ta_allocation","ta2_allocation",
                          "s4_s3_status","s4_ff_result","gen_result"):
                    st.session_state[k] = 0 if k == "s4_stage" else None
                _nav("generate")
        else:
            st.button("🚀 Generate Timetable →", type="primary",
                      use_container_width=True, disabled=True,
                      help="Run Validate Step 3 first and resolve any overloads")

    s3v = st.session_state.get("s3_validated", False)
    if s3v:
        st.success("✅ Step 3 validated — click **Generate Timetable** to proceed.")
    else:
        st.info("ℹ️ Review teacher workloads, set up combines, then click **🔍 Validate Step 3**.")

    vr = st.session_state.get("s3_validation_result")
    if vr and not vr.get("can_proceed") and vr.get("issues"):
        with st.expander("❌ Still Overloaded — expand for details", expanded=False):
            for ln in vr["issues"]:
                st.error(ln)
    if vr and vr.get("resolved"):
        with st.expander("✓ Resolved / Skipped"):
            for ln in vr["resolved"]: st.success(ln)

    st.divider()
    tab_wl, tab_cb, tab_un = st.tabs(
        ["📊 Teacher Workload","🔗 Combine Classes","🚫 Unavailability"])
    with tab_wl: _render_workload(teachers, wdays, ppd)
    with tab_cb: _render_combine_tab(teachers, all_cn)
    with tab_un: _render_unavailability_tab(teachers, day_names, ppd)


def _load_step3_config(raw: bytes):
    log.info("_load_step3_config: %d bytes", len(raw))
    try:
        d = json.loads(raw.decode("utf-8"))
    except json.JSONDecodeError as ex:
        log.error("_load_step3_config: JSON parse error: %s", ex)
        st.toast(f"❌ Invalid JSON file: {ex}", icon="🚫")
        _notify(f"❌ Invalid JSON: {ex}", "error")
        st.rerun()
        return

    if "step3_data" not in d and "step3_unavailability" not in d:
        wrong_hint = ""
        if "assignments" in d:
            wrong_hint = " (looks like a Step 2 Assignments file)"
        elif "periods_per_day" in d or "teacher_names" in d:
            wrong_hint = " (looks like a Step 1 config file)"
        msg = (f"❌ Wrong file type{wrong_hint} — expected a Step 3 config file with "
               f"'step3_data' or 'step3_unavailability' keys.")
        log.error("_load_step3_config: %s", msg)
        st.toast(msg, icon="🚫")
        _notify(msg, "error")
        st.rerun()
        return

    try:
        _e3 = _get_eng()
        _e3.step3_data = d.get("step3_data", {})
        _e3.step3_unavailability = {
            t: {"days": v.get("days", []), "periods": v.get("periods", [])}
            for t, v in d.get("step3_unavailability", {}).items()
        }
        _notify("✓ Step 3 config loaded.", "success")
        log.info("_load_step3_config: %d teachers, %d unavail",
                 len(_e3.step3_data), len(_e3.step3_unavailability))
        st.rerun()
    except Exception as ex:
        log.error("_load_step3_config: %s\n%s", ex, traceback.format_exc())
        _notify(f"Failed: {ex}", "error")


def _render_workload(teachers, wdays, ppd):
    """Render teacher workload with Edit/Combine and Skip buttons per teacher."""
    eng = _get_eng()
    log.debug("_render_workload: %d teachers", len(teachers))

    max_allowed = (ppd - 2) * wdays
    wl          = getattr(eng, "_step3_teacher_wl", {})

    if not wl:
        n_cls = sum(1 for cn in eng.class_config_data
                    if eng.class_config_data[cn].get('subjects'))
        st.warning(
            f"No teacher assignments found. Classes with subjects: {n_cls}. "
            "Make sure teachers are assigned to subjects in Step 2.")
        return

    st.caption(
        f"Overload threshold: **{max_allowed}** periods/week  "
        f"({wdays} days × ({ppd}−2) periods).  "
        f"All teachers shown — overloaded ones highlighted in red.")

    overloaded_teachers = sorted(t for t in wl if t in eng._step3_overloaded)
    normal_teachers     = sorted(t for t in wl if t not in eng._step3_overloaded)

    sel_teacher = st.session_state.get("s3_selected_teacher", "")

    def _teacher_card(teacher):
        info      = wl.get(teacher, {"total": 0, "entries": []})
        effective = eng._effective_total(teacher)
        s3d       = eng.step3_data.get(teacher, {})
        skipped   = s3d.get("skipped", False)
        n_comb    = len(s3d.get("combines", []))
        is_over   = teacher in eng._step3_overloaded
        still_over = is_over and effective > max_allowed

        if teacher == sel_teacher:
            border_style = "border:2px solid #2980b9; padding:8px; border-radius:6px; background:#eaf4fb"
        elif skipped:
            border_style = "border:1px solid #ccc; padding:8px; border-radius:6px; background:#e8f5e9"
        elif still_over:
            border_style = "border:1px solid #e74c3c; padding:8px; border-radius:6px; background:#fdecea"
        elif is_over and effective <= max_allowed:
            border_style = "border:1px solid #f39c12; padding:8px; border-radius:6px; background:#fff8e1"
        else:
            border_style = "border:1px solid #ccc; padding:8px; border-radius:6px; background:#fff"

        if skipped:
            badge = "SKIPPED"
        elif still_over:
            badge = "OVERLOADED"
        elif n_comb and not still_over:
            badge = f"{n_comb} combine{'s' if n_comb>1 else ''}"
        else:
            badge = "OK"

        stat = f"Assigned: **{info['total']}**"
        if effective != info["total"]:
            stat += f"  →  Effective: **{effective}**"
        stat += f"  /  Max: **{max_allowed}**"

        with st.container():
            st.markdown(f"<div style='{border_style}'>", unsafe_allow_html=True)
            h1, h2 = st.columns([3, 1])
            with h1:
                st.markdown(f"**{teacher}**")
                st.caption(stat)
            with h2:
                badge_color = "#c0392b" if still_over else ("#27ae60" if not is_over or skipped else "#f39c12")
                st.markdown(
                    f"<span style='color:{badge_color};font-weight:bold;font-size:0.8rem'>{badge}</span>",
                    unsafe_allow_html=True)

            b1, b2 = st.columns(2)
            with b1:
                btn_label = "✏ Close Detail" if teacher == sel_teacher else "✏ Edit / Combine"
                if st.button(btn_label, key=f"ec_{teacher}", use_container_width=True):
                    if teacher == sel_teacher:
                        st.session_state.pop("s3_selected_teacher", None)
                    else:
                        st.session_state["s3_selected_teacher"] = teacher
                        for k in list(st.session_state.keys()):
                            if k.startswith(f"s3cb_{teacher}_"):
                                del st.session_state[k]
                    st.rerun()
            with b2:
                skip_lbl = "↩ Un-skip" if skipped else "⏭ Skip"
                if st.button(skip_lbl, key=f"sk_{teacher}", use_container_width=True):
                    eng.step3_data.setdefault(teacher, {"skipped": False, "combines": []})
                    eng.step3_data[teacher]["skipped"] = not skipped
                    eng.prepare_step3_workload()
                    st.rerun()
            st.markdown("</div>", unsafe_allow_html=True)

    if overloaded_teachers:
        st.markdown(f"#### ⚠ Overloaded  (>{max_allowed}/wk)")
        for t in overloaded_teachers:
            _teacher_card(t)

    if normal_teachers:
        st.markdown(f"#### ✓ Within limit  (≤{max_allowed}/wk)")
        for t in normal_teachers:
            _teacher_card(t)

    if sel_teacher and sel_teacher in wl:
        st.divider()
        _render_teacher_combine_detail(sel_teacher)


def _render_teacher_combine_detail(teacher):
    """Full assignment + combine panel for a selected teacher."""
    eng = _get_eng()
    combine_err = st.session_state.pop("_s3_combine_err", None)
    if combine_err:
        _combine_error_dialog(combine_err)
        return
    wl        = getattr(eng, "_step3_teacher_wl", {})
    info      = wl.get(teacher, {"total": 0, "entries": []})
    entries   = info["entries"]
    s3d       = eng.step3_data.setdefault(teacher, {"skipped": False, "combines": []})
    max_all   = getattr(eng, "_step3_max_allowed", 9999)
    effective = eng._effective_total(teacher)
    is_over   = teacher in getattr(eng, "_step3_overloaded", set())
    still_over = is_over and effective > max_all

    hdr_color = "#c0392b" if still_over else "#1a7a1a"
    st.markdown(
        f"<div style='background:{hdr_color};color:white;padding:10px 16px;"
        f"border-radius:6px;display:flex;justify-content:space-between;align-items:center'>"
        f"<span style='font-size:1.1rem;font-weight:bold'>{teacher}</span>"
        f"<span style='font-size:0.95rem;font-weight:bold'>"
        f"Assigned: {info['total']}  |  Effective: {effective}  |  Max: {max_all}"
        f"</span></div>",
        unsafe_allow_html=True)
    st.write("")

    combined_indices = set()
    for cb in s3d["combines"]:
        for idx in cb.get("entry_indices", []):
            combined_indices.add(idx)

    left_col, right_col = st.columns([1, 1])

    with left_col:
        st.markdown("**Assignments — check entries to combine**")
        if not entries:
            st.info("No assignments found for this teacher.")
        for ei, entry in enumerate(entries):
            in_cb = ei in combined_indices
            ct_info = eng.get_class_ct_info(entry["class"], teacher, entry["subject"])
            ct      = ct_info["ct"]
            ct_subs = ct_info["ct_subjects"]
            has_parallel_warn = ct_info["is_parallel_with_ct"]

            with st.container(border=True):
                cc1, cc2 = st.columns([1, 8])
                with cc1:
                    if in_cb:
                        st.markdown("✅")
                    else:
                        st.checkbox(
                            "", key=f"s3cb_{teacher}_{ei}",
                            value=st.session_state.get(f"s3cb_{teacher}_{ei}", False))
                with cc2:
                    label_color = "#1a7a1a" if in_cb else "#1a1a1a"
                    st.markdown(
                        f"<span style='color:{label_color};font-weight:bold'>"
                        f"[{entry['class']}]  {entry['subject']}  ({entry['periods']} periods)"
                        f"{'  ✓ in combine' if in_cb else ''}</span>",
                        unsafe_allow_html=True)
                    ct_subs_str = ", ".join(ct_subs) if ct_subs else "—"
                    st.caption(f"CT: {ct or '—'}   |   CT Subjects: {ct_subs_str}")
                    if has_parallel_warn:
                        st.markdown(
                            f"<span style='color:#c0392b;font-style:italic;font-size:0.82rem'>"
                            f"⚠  '{entry['subject']}' (teacher) ∥ "
                            f"'{ct_info['parallel_ct_subject']}' (CT) — parallel conflict!"
                            f"</span>",
                            unsafe_allow_html=True)

    with right_col:
        st.markdown("**Combines**")

        if st.button("✓ Combine Checked Entries", key=f"do_combine_{teacher}",
                     type="primary", use_container_width=True):
            idxs = [ei for ei in range(len(entries))
                    if st.session_state.get(f"s3cb_{teacher}_{ei}", False)
                    and ei not in combined_indices]

            err = None
            if len(idxs) < 2:
                err = "Check at least 2 assignments (that are not already in a combine)."
            else:
                periods_list = [entries[i]["periods"] for i in idxs]
                if len(set(periods_list)) > 1:
                    err = (f"All entries to combine must have the same period count. "
                           f"Selected: {periods_list}")
                else:
                    # Check 1: Parallel-CT conflict
                    blocked = []
                    for i in idxs:
                        e = entries[i]
                        ct_i = eng.get_class_ct_info(e["class"], teacher, e["subject"])
                        if ct_i["is_parallel_with_ct"]:
                            blocked.append(
                                f"Class {e['class']}: '{e['subject']}' ({teacher}) ∥ "
                                f"'{ct_i['parallel_ct_subject']}' (CT: {ct_i['ct']})")
                    if blocked:
                        err = (
                            "Cannot combine — parallel-CT conflict detected:\n\n"
                            + "\n".join(f"  • {b}" for b in blocked)
                            + "\n\nThe CT has a fixed period each day; combining would "
                            "force the parallel subject into the same fixed slot.")
                    
                    # Check 2: NEW - Class teacher cannot combine their own class
                    if not err:
                        ct_combined = []
                        class_config = eng.class_config_data
                        for i in idxs:
                            e = entries[i]
                            cn = e["class"]
                            # Check if this teacher is the CT of this class
                            cn_ct = class_config.get(cn, {}).get('teacher', '').strip()
                            if cn_ct == teacher.strip():
                                ct_combined.append(cn)
                        
                        if ct_combined:
                            err = (
                                f"Cannot combine — {teacher} is the class teacher of "
                                f"{', '.join(ct_combined)}:\n\n"
                                f"A class teacher's combined class must have the SAME slot "
                                f"every day in the first half (so it acts as the CT period "
                                f"for that class). Combining classes prevents this.\n\n"
                                f"Solution: Create a separate combined group for non-CT "
                                f"subjects only.")

            if err:
                st.session_state["_s3_combine_err"] = err
            else:
                s3d["combines"].append({
                    "entry_indices": idxs,
                    "periods_each":  entries[idxs[0]]["periods"],
                    "classes":       [entries[i]["class"]   for i in idxs],
                    "subjects":      [entries[i]["subject"] for i in idxs],
                })
                for ei in range(len(entries)):
                    st.session_state.pop(f"s3cb_{teacher}_{ei}", None)
                eng.prepare_step3_workload()
                log.info("_render_teacher_combine_detail: combine added for %s: %s",
                         teacher, [entries[i]["class"] for i in idxs])
                _notify(f"✓ Combine added for {teacher}.", "success")
            st.rerun()

        st.write("")
        if not s3d["combines"]:
            st.info("No combines yet.\nCheck assignments on the left and click Combine.")
        else:
            for ci, cb in enumerate(s3d["combines"]):
                with st.container(border=True):
                    classes_str = "  +  ".join(cb.get("classes", []))
                    st.markdown(f"**Combine {ci+1}:  {classes_str}**")
                    for i, idx in enumerate(cb.get("entry_indices", [])):
                        if idx < len(entries):
                            st.caption(f"• {entries[idx]['label']}")
                        else:
                            cls = cb["classes"][i] if i < len(cb.get("classes",[])) else "?"
                            sub = cb["subjects"][i] if i < len(cb.get("subjects",[])) else "?"
                            st.caption(f"• '{sub}' in {cls}  x{cb.get('periods_each','?')}/wk")
                    saving = (len(cb.get("entry_indices", [])) - 1) * cb.get("periods_each", 0)
                    st.caption(f"💡 Saves {saving} periods/week for {teacher}")
                    if st.button("✕ Remove", key=f"rm_cb_{teacher}_{ci}",
                                 use_container_width=True):
                        s3d["combines"].pop(ci)
                        eng.prepare_step3_workload()
                        log.info("_render_teacher_combine_detail: removed combine %d for %s",
                                 ci, teacher)
                        st.rerun()


def _render_combine_tab(teachers, all_cn):
    """Shows a summary of all existing combines across all teachers."""
    eng = _get_eng()
    log.debug("_render_combine_tab")
    st.info("💡 To **add** combines, go to the **Teacher Workload** tab and click "
            "**Edit / Combine** next to any teacher.")
    st.markdown("---")
    st.markdown("**All Existing Combines**")
    any_cb = False
    for teacher in sorted(eng.step3_data.keys()):
        s3d = eng.step3_data[teacher]
        cbs = s3d.get("combines", [])
        if not cbs:
            continue
        any_cb = True
        wl   = getattr(eng, "_step3_teacher_wl", {})
        info = wl.get(teacher, {"entries": []})
        entries = info["entries"]

        st.markdown(f"**{teacher}**")
        for ci, cb in enumerate(cbs):
            c1, c2 = st.columns([5, 1])
            with c1:
                classes_str  = " + ".join(cb.get("classes", []))
                subjects_str = ", ".join(cb.get("subjects", []))
                saving = (len(cb.get("entry_indices", [])) - 1) * cb.get("periods_each", 0)
                st.write(f"  📌 **{classes_str}**  ·  {subjects_str}  "
                         f"  *(saves {saving} periods/wk)*")
            with c2:
                if st.button("🗑", key=f"del_cbt_{teacher}_{ci}",
                             help="Remove this combine"):
                    cbs.pop(ci)
                    eng.prepare_step3_workload()
                    log.info("_render_combine_tab: del %d for %s", ci, teacher)
                    st.rerun()
    if not any_cb:
        st.info("No combines defined yet.")


def _render_unavailability_tab(teachers, day_names, ppd):
    eng = _get_eng()
    log.debug("_render_unavailability_tab")
    unavail = eng.step3_unavailability
    with st.container(border=True):
        st.markdown("**Add / Update Unavailability**")
        sel_t = st.selectbox("Teacher", [""]+teachers, key="un_teacher")
        c1,c2 = st.columns(2)
        with c1: sel_days    = st.multiselect("Unavailable Days",    day_names,           key="un_days")
        with c2: sel_periods = st.multiselect("Unavailable Periods", list(range(1,ppd+1)), key="un_periods")
        cs,ck,cl = st.columns(3)
        with cs:
            if st.button("✓ Save", key="un_save"):
                if not sel_t: _notify("Select teacher.","warning")
                elif not sel_days or not sel_periods: _notify("Select days and periods.","warning")
                else:
                    ok, msg = eng._check_unavailability_feasible(sel_t, sel_days, sel_periods)
                    if not ok:
                        log.warning("_render_unavailability_tab: feasibility fail %s: %s", sel_t, msg)
                        _notify(f"❌ Feasibility failed: {msg}","error")
                    else:
                        unavail[sel_t] = {"days":sel_days,"periods":sel_periods}
                        log.info("_render_unavailability_tab: saved %s days=%s", sel_t, sel_days)
                        _notify(f"✓ Saved for {sel_t}.","success"); st.rerun()
        with ck:
            if st.button("🔍 Check Only", key="un_check"):
                if sel_t and sel_days and sel_periods:
                    ok,msg = eng._check_unavailability_feasible(sel_t,sel_days,sel_periods)
                    log.info("_render_unavailability_tab: check %s → ok=%s", sel_t, ok)
                    if ok: _notify(f"✓ Feasible: {msg}","success")
                    else:  _notify(f"❌ {msg}","error")
                    st.rerun()
        with cl:
            if st.button("✕ Clear", key="un_clear"): st.rerun()
    _show_notifications()
    st.markdown("**Current Unavailability**")
    if not unavail:
        st.info("No unavailability set.")
    else:
        for teacher, info in sorted(unavail.items()):
            ok,short = eng._check_unavailability_feasible(teacher,info.get("days",[]),info.get("periods",[]))
            days_str = ", ".join(info.get("days",[]))
            pers_str = ", ".join(f"P{p}" for p in sorted(info.get("periods",[])))
            with st.expander(f"{'🟢' if ok else '🔴'} **{teacher}** — {days_str}  |  {pers_str}", expanded=False):
                st.write(f"Days: {days_str}  ·  Periods: {pers_str}")
                if not ok: st.error(short)
                if st.button("🗑 Remove", key=f"del_un_{teacher}"):
                    log.info("_render_unavailability_tab: remove %s", teacher)
                    del unavail[teacher]; st.rerun()
